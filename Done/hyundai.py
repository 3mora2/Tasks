from selenium import webdriver
from time import sleep
from itertools import cycle
from openpyxl import load_workbook, Workbook

book = load_workbook('hyundai.xlsx')
sheet1 = book.active

book_new = Workbook()
sheet = book_new.active
con = 1
sheet.cell(con, 1).value = 'Name'
sheet.cell(con, 2).value = 'Cat1'
sheet.cell(con, 3).value = 'Cat2'
sheet.cell(con, 4).value = 'Cat3'
sheet.cell(con, 5).value = 'Cat4'

sheet.cell(con, 14).value = 'IMg'
sheet.cell(con, 15).value = 'URL'
sheet.cell(con, 6).value = 'PNC'
sheet.cell(con, 7).value = 'Part_Number'
sheet.cell(con, 8).value = 'Qty'
sheet.cell(con, 9).value = 'Part_Name'
sheet.cell(con, 10).value = 'Model_Description'
sheet.cell(con, 11).value = 'Start_Date'
sheet.cell(con, 12).value = 'End_Date'

# url = 'https://hyundai.catalogs-parts.com/#{client:1;page:models;lang:en;catalog:gen}'

driver = webdriver.Chrome()
driver.implicitly_wait(6)
con = 2
for ii in range(2, sheet1.max_row + 1):
    print(ii)
    if ii == 11:
        break
    url = sheet1.cell(ii, 12).value
    driver.get(url)
    driver.get(url)
    name = driver.find_element_by_css_selector('#cat_content > h1').text
    cat1 = driver.find_elements_by_css_selector('#breadcrumbs > li.hidden-xs > a')[1].text
    cat2 = driver.find_elements_by_css_selector('#breadcrumbs > li.hidden-xs > a')[2].text
    cat3 = driver.find_elements_by_css_selector('#breadcrumbs > li.hidden-xs > a')[4].text
    cat4 = driver.find_element_by_css_selector('#breadcrumbs > li.active.hidden-xs').text
    img = '\n'.join([im.get_attribute('src') for im in driver.find_elements_by_css_selector('img')])

    sheet.cell(con, 1).value = name
    sheet.cell(con, 2).value = cat1
    sheet.cell(con, 3).value = cat2
    sheet.cell(con, 4).value = cat3
    sheet.cell(con, 5).value = cat4

    sheet.cell(con, 14).value = img
    sheet.cell(con, 15).value = url
    for el in driver.find_elements_by_css_selector('#cat_content table > tbody > tr')[1:]:
        PNC = el.find_elements_by_css_selector('td')[0].text
        Part_Number = el.find_elements_by_css_selector('td')[1].text
        Qty = el.find_elements_by_css_selector('td')[2].text
        Part_Name = el.find_elements_by_css_selector('td')[3].text
        Model_Description = el.find_elements_by_css_selector('td')[4].text
        Start_Date = el.find_elements_by_css_selector('td')[5].text
        End_Date = el.find_elements_by_css_selector('td')[6].text
        information = el.find_elements_by_css_selector('td')[7].text
        sheet.cell(con, 6).value = PNC
        sheet.cell(con, 7).value = Part_Number
        sheet.cell(con, 8).value = Qty
        sheet.cell(con, 9).value = Part_Name
        sheet.cell(con, 10).value = Model_Description
        sheet.cell(con, 11).value = Start_Date
        sheet.cell(con, 12).value = End_Date
        sheet.cell(con, 13).value = information
        con += 1
    book_new.save('n.xlsx')
