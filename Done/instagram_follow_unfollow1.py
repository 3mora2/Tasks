from datetime import datetime
from time import sleep
from openpyxl import Workbook, load_workbook
import json

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium import webdriver
import threading


class Thread(threading.Thread):
    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        self.p = None
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class Insta:
    def __init__(self):
        self.username = input('- Enter your username : ')  # 'pywpywpywpy'
        self.password = input('- Enter password : ')  # 'pYpY2222'
        url = input('- Enter Username ( example:- username ) : ')  # chhc

        if url.find('instagram.com/') > -1:
            self.user = url.split('com/')[1].split('/')[0]
        else:
            self.user = url

        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet['B1'] = 'user name'
        self.sheet.column_dimensions['B'].width = 21
        self.end_cell = None
        self.driver = webdriver.Chrome()
        print('- Please Wait While Start..........')
        self.login()

    def login(self):
        self.driver.get("https://www.instagram.com/accounts/login/?")
        sleep(2)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, '#loginForm > div > div:nth-child(1) > div > label > input'))).send_keys(self.username)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, '#loginForm > div > div:nth-child(2) > div > label > input'))).send_keys(self.password)
        sleep(2)
        self.driver.find_element_by_css_selector('#loginForm > div > div:nth-child(3) > button').click()
        sleep(4)
        if self.driver.current_url != 'https://www.instagram.com/':
            input('- Please complicate login than press Enter')
        try:
            sleep(4)
            self.driver.find_element_by_class_name('HoLwm').click()
            sleep(2)
        except:
            pass

    def get_followers(self):
        now1 = datetime.now()
        p = 1
        u = True
        end_cursor = None
        #####################################################################
        self.driver.get(f"https://www.instagram.com/{self.user}/?__a=1")
        user_id = json.loads(self.driver.find_element_by_css_selector('body').text)['graphql']['user']['id']
        self.driver.get(f"https://www.instagram.com/graphql/query/?query_id=17851374694183129&id={user_id}&first=100")
        sleep(3)
        resp = json.loads(self.driver.find_element_by_css_selector('body').text)
        print('- Number of Followers {0}'.format(resp['data']['user']['edge_followed_by']['count']))
        print('- Getting followers.....')
        #####################################################################
        self.thr = Thread()
        self.thr.start()
        users = []
        while u:
            try:
                resp = json.loads(self.driver.find_element_by_css_selector('body').text)
                u = resp['data']['user']['edge_followed_by']['page_info']['has_next_page']
                end_cursor = resp['data']['user']['edge_followed_by']['page_info']['end_cursor']
                for i in resp['data']['user']['edge_followed_by']['edges']:
                    if i['node']['username'] not in users:
                        user = i['node']['username']
                        self.sheet.cell(row=p, column=2).value = user
                        print(f'{p} : {user}')
                        p += 1
                self.driver.get(
                    f"https://www.instagram.com/graphql/query/?query_id=17851374694183129&id={user_id}&first=100&after={end_cursor}")
                sleep(3)
            except KeyError:
                print('- Error, Please Wait....')
                sleep(120)
                self.driver.refresh()
            except Exception as e:
                print(e)
            self.book.save(f'{self.user}.xlsx')
            try:
                if self.thr.p == 'pause' or self.thr.p == 'p':
                    start = input('(Type start or s and press Enter)>>>')
                    while True:
                        if start == 'start' or start == 's':
                            self.thr = Thread()
                            self.thr.start()
                            break
                        elif start == 'exit' or start == 'e':
                            exit()
                        else:
                            start = input()
                elif self.thr.p == 'exit' or self.thr.p == 'e':
                    exit()
            except:
                pass

        now = datetime.now()
        end_time = now - now1
        print(end_time)
        print('-- Done --')
        print('- Number of Followers {0}'.format(resp['data']['user']['edge_followed_by']['count']))


class InstaBot:
    def __init__(self):
        self.username = input('Enter your username : ')#'pywpywpywpy'
        self.password = input('Enter password : ')# 'pYpY2222'
        self.number = int(input('Enter number people : '))
        self.sleep_number = int(input('Enter Time (S) : '))
        self.file = input('Enter File name : ')# 'chhc.xlsx'

        self.book = load_workbook(self.file)
        self.sheet = self.book.active
        self.driver = webdriver.Chrome()
        self.login()

    def login(self):
        self.driver.get("https://www.instagram.com/accounts/login/?")
        sleep(2)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, '#loginForm > div > div:nth-child(1) > div > label > input'))).send_keys(self.username)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, '#loginForm > div > div:nth-child(2) > div > label > input'))).send_keys(self.password)
        sleep(2)
        self.driver.find_element_by_css_selector('#loginForm > div > div:nth-child(3) > button').click()
        try:
            sleep(4)
            self.driver.find_element_by_class_name('HoLwm').click()
            sleep(2)
        except:
            pass

    def Make_follow(self):
        self.users = [(self.sheet[f'B{i}'].value, i) for i in range(1, self.number + 1) if
                      self.sheet[f'C{i}'].value != 'Done']
        for user, i in self.users:
            print(i, ' - ', user)
            try:
                self.driver.get(f'https://www.instagram.com/{user}/')
                for button in WebDriverWait(self.driver, 20).until(EC.visibility_of_all_elements_located((
                    By.CSS_SELECTOR,
                    '#react-root > section > main > div > header > section > div > div  button'))):
                    if 'Follow' in button.text:
                        button.click()
                        print('Click Follow')
                        break
            except:
                pass
            self.sheet[f'C{i}'].value = 'Done'
            self.book.save(self.file)
            sleep(self.sleep_number)

    def Make_unfollow(self):
        self.users = [(self.sheet[f'B{i}'].value, i) for i in range(1, self.number + 1) if
                      self.sheet[f'A{i}'].value != 'Done']
        for user, i in self.users:
            try:
                print(i, ' - ', user)
                self.driver.get(f'https://www.instagram.com/{user}/')
                sleep(2)
                try:
                    WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((
                        By.CSS_SELECTOR,
                        '#react-root > section > main > div > header > section > div > div:nth-child(3) > div > span > span.vBF20 > button'))).click()
                except:
                    for button in WebDriverWait(my_bot.driver, 20).until(EC.visibility_of_all_elements_located((
                            By.CSS_SELECTOR,
                            '#react-root > section > main > div > header > section > div > div  button'))):
                        print(button.text)
                        if 'Request' in button.text:
                            button.click()
                            break
                sleep(2)
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((
                By.CSS_SELECTOR, 'body > div.RnEpo.Yx5HN > div > div > div > div.mt3GC > button.aOOlW.-Cab_'))).click()
                print('Click Unfollow')
            except:
                pass
            self.sheet[f'A{i}'].value = 'Done'
            self.book.save(self.file)
            sleep(self.sleep_number)


if __name__ == '__main__':
    while True:
        Input = str(input('''
-----------------------------------------------------
- This script to extract information from instagram -
- 1 - To Get Followers >>> Enter 1                  -
- 2 - To Follow  >>> Enter 2                        -
- 2 - To Unfollow  >>> Enter 3                      -
-----------------------------------------------------
>>> '''))
        if Input == '1':
            m = Insta()
            m.get_followers()
            break
        elif Input == '2':
            my_bot = InstaBot()
            my_bot.Make_follow()
            break
        elif Input == '3':
            my_bot = InstaBot()
            my_bot.Make_unfollow()
            break
