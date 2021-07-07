from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from time import sleep
from openpyxl import Workbook
import subprocess

class Facebook:
    def __init__(self,n):
        # put wmail and password to login
        mail ='moduhondu@gmail.com' #str(input("enter username: "))
        pas = 'ammar2020'#str(input("enter password: "))
        #number of post which you want
        self.num=n
        #to open chrome with remote debugging port
        #subprocess.Popen('"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222')
        chrome_options = Options()
        # disable notifications
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-notifications")
        # make chrom hideen
        #chrome_options.add_argument("--headless")
        #to connect selenium with chrome which aready opened
        #chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        #open chrom
        self.driver = webdriver.Chrome('G:\\programing\\projects\\automation\\chromedriver.exe',chrome_options=chrome_options)
        try:
            #link of ligin
            self.driver.get("https://facebook.com/login/")
        except:
            # if return error it will refresh
            self.driver.refresh()
        sleep(2)
        self.driver.refresh()
        #to make sture you need login
        if self.driver.current_url !='https://www.facebook.com/':
            self.driver.find_element_by_id('email').send_keys(mail)
            sleep(2)
            self.driver.find_element_by_id('pass').send_keys(pas)
            sleep(2)
            self.driver.find_element_by_id('loginbutton').click()
            sleep(2)
        #start
        self.like_commont_share()
        #after it finsh it will clise
        #self.ex_data()
        self.driver.close()
    def like_commont_share(self):
        #open link page
        self.driver.get('https://www.facebook.com/pg/Vodafone.Egypt/posts/?ref=page_internal')
        #post_num=0
        #to select all posts
        all_posts = self.driver.find_elements_by_class_name('_427x')
        self.Make_like_comment(all_posts)
        print('tern')
        #to select part of like and commont and share
        #posts = self.driver.find_elements_by_class_name('_3vuz')
        '''u=True
        while u:
            post_num=len(posts)
            #after get posts send it to (Make_like_comment)
            u=self.Make_like_comment(posts)
            #to make scrolldown
            g=self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            sleep(4)
            # to select part of like and commont and share
            post_new = self.driver.find_elements_by_class_name('_3vuz')
            #delet old post
            del post_new[0:post_num-1]
            posts=post_new'''

    def Make_like_comment(self,post):
        for i in post:
            #to select part of like and commont and share
            bar1 = i.find_element_by_class_name('_3vuz')
            if self.num >0:
                sleep(1)
                print('star')
                print(self.num)
                # select like botton
                hov = bar1.find_element_by_class_name('_3l2t')
                # to make sure it is'n liked
                if hov.get_attribute('aria-pressed') != 'true':
                    sleep(2)
                    #to make love
                    #hover = ActionChains(self.driver).move_to_element(hov)
                    #hover.perform()
                    sleep(2)
                    #lik=self.driver.find_elements_by_class_name('_iuw')
                    #lik[len(lik)-7].click()
                    print('make love')
                    hov.click()
                    try:
                        #some time apper notfication it will close
                        self.driver.find_element_by_xpath('//*[@id="facebook"]/body/div[16]/div[2]/div[1]/div[1]/div/div/div/div/div[1]/div').click()
                    except:
                        pass
                sleep(2)
            ###############################################
            #select comment botton
            comment = bar1.find_element_by_class_name('_42ft').click()
            sleep(3)
            # where it will write
            h = i.find_element_by_class_name('notranslate')
            #text comment
            h.send_keys('vodafone')
            #to put enter and post
            h.send_keys(Keys.ENTER)
            print('comment')
            ##############################################_2nj7
            #to select botton share
            share=bar1.find_element_by_class_name('_2nj7')
            share.click()
            sleep(5)
            box = self.driver.find_elements_by_class_name('__MenuItem')
            box[len(box)-8].find_element_by_tag_name('a').click()
            print('make share')
            sleep(4)
            #######################################################
            #to get information about post
            title_post=i.find_elements_by_class_name('fcg')[0].text
            time_post=i.find_elements_by_class_name('fcg')[2].text
            content_post=i.find_element_by_class_name('userContent').text
            bar_likes_comment=i.find_element_by_class_name('_3vum')
            number_likes=bar_likes_comment.find_elements_by_class_name('_81hb')[0].text
            number_share=bar_likes_comment.find_elements_by_class_name('_42ft')[1].text
            number_comment=bar_likes_comment.find_elements_by_class_name('_42ft')[0].text
            print('title_post :',title_post)
            print('time_post :',time_post)
            print('content_post :',content_post)
            print('number_likes :',number_likes)
            print('number_comment',number_comment)
            print('number_share',number_share)
            self.num-=1
    def ex_data(self):
        self.driver.get('https://www.facebook.com/pg/Vodafone.Egypt/posts/?ref=page_internal')
        # to sellect all posts
        sleep(2)
        all_posts = self.driver.find_elements_by_class_name('_427x')
        for post in all_posts:
            sleep(1)
            # to get number of likes
            o = post.find_element_by_class_name('_3dlf')
            link_likes=o.get_attribute('href')
            number_likes=o.text.split('\n')[0]
            self.driver.execute_script('window.open("");')
            self.driver.switch_to_window(self.driver.window_handles[1])
            self.driver.get(link_likes)
            '''while True:
                list=['1','2','3','4','7','8','16']
                for i in list:
                    try:
                        sleep(2)
                        self.driver.find_element_by_xpath('//*[@id="reaction_profile_pager{0}"]/div/a'.format(i)).click()
                    except Exception as e:#//*[@id="reaction_profile_browser1"]/li[1]/div   //*[@id="reaction_profile_browser7"]/li[1]/div
                        print(e)
                        del list[list.index(i)]
                    sleep(4)
                if len(list)==0:
                    return False'''
            pe=self.driver.find_elements_by_class_name('_5i_q')
            num=0
            for p in pe:
                if num ==50:
                    return False
                pp=self.driver.find_element_by_class_name('_5i_s')
                link_img=pp.find_element_by_class_name('img').get_attribute('src')
                print(link_img)
                #link_profle=p.find_element_by_name('a').get_attribute('href')
                #print(link_profle,link_img,name)
                num+=1

            self.driver.close()
from openpyxl import load_workbook
path = "time.xlsx"
book = load_workbook(path)
sheet = book.active
tim = sheet.cell(row=2, column=1).value
number = sheet.cell(row=2, column=2).value
print(number)
print(tim)
import time
tt=time.localtime()
h1=int(time.strftime("%H"))
m1=int(time.strftime("%M"))
#tim='00:40'
t=tim.split('-')
h2=int(t[0])
m2=int(t[1])
h=h2-h1
m=m2-m1
if m<0:
    m=m*(-1)
s=h*60*60+m*69
print(s)
#sleep(s)
Facebook(7)







'''class InstaBot:
    def __init__(self):#https://www.instagram.com/amrdiab
        #https://www.instagram.com/chhc/followers/
        m = str(input("enter username: "))
        p = str(input("enter password: "))
        #self.url = str(input("enter url : "))
        #options = webdriver.ChromeOptions()
        #options.add_argument("headless")
        self.driver = webdriver.Chrome()#chrome_options=options)
        self.driver.get("https://www.instagram.com/accounts/login/?")
        sleep(2)
        self.driver.find_element_by_xpath('/html/body/div[1]/section/main/div/article/div/div[1]/div/form/div[2]/div/label/input').send_keys(m)
        self.driver.find_element_by_xpath('/html/body/div[1]/section/main/div/article/div/div[1]/div/form/div[3]/div/label/input').send_keys(p)
        self.driver.find_element_by_xpath('/html/body/div[1]/section/main/div/article/div/div[1]/div/form/div[4]/button/div').click()
        sleep(4)
        self.driver.find_element_by_class_name('HoLwm').click()
        sleep(2)

    def get_unfollowers(self):
        self.driver.get('https://www.instagram.com/chhc')
        sleep(2)
        self.driver.find_element_by_xpath("/html/body/div[1]/section/main/div/header/section/ul/li[2]/a").click()
        following = self._get_names()
        ##########################
        
    def _get_names(self):
        num=int(self.driver.find_elements_by_class_name('g47SY')[1].get_attribute('title').replace(',',''))
        sleep(2)
        sugs = self.driver.find_element_by_xpath('/html/body/div[1]/section/footer/div/nav/ul/li[10]/a')
        self.driver.execute_script('arguments[0].scrollIntoView()', sugs)
        sleep(2)
        scroll_box = self.driver.find_element_by_class_name("isgrP")
        text=[]
        while num > len(text):
            sleep(2)
            ht = self.driver.execute_script("""
                arguments[0].scrollTo(0, arguments[0].scrollHeight); 
                return arguments[0].scrollHeight;
                """, scroll_box)
            links = scroll_box.find_elements_by_tag_name('a')
            if len(links) :
                for name in links:
                    try:
                        if name.text:
                            if name.text != '':
                                if name.text not in text:
                                    text.append(name.text)
                                    print(str(text.index(name.text)+1)+'-'+ name.text)
                    except Exception as e:
                        pass


        self.driver.find_element_by_xpath("/html/body/div[4]/div/div[1]/div/div[2]/button").click()
        self.driver.close()
        print('done')
        ######################
        try:
            import openpyxl
            path = "Insta_Names.xlsx"
            book = openpyxl.load_workbook(path)
            sheet = book.active
            end_cell = sheet.max_row
            i=1
            for n in text:
                sheet.cell(row=(end_cell + i), column=1).value = n
                i+=1
            book.save(path)
        except :
            book = Workbook()
            sheet = book.active
            sheet['A1'] = 'names'
            i=2
            for n in text:
                sheet.cell(row=i, column=1).value = n
                i+=1
                print(n)
            book.save('Insta_Names.xlsx')
            
            ###################



my_bot = InstaBot()
my_bot.get_unfollowers()
'''
