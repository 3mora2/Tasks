from openpyxl import load_workbook
from io import BytesIO
from openpyxl.drawing.image import Image
from requests_html import HTMLSession

session = HTMLSession()
headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    "accept-language": "ar,en-US;q=0.9,en;q=0.8",
    "cache-control": "no-cache",
    "pragma": "no-cache",
    "sec-ch-ua": "\" Not;A Brand\";v=\"99\", \"Google Chrome\";v=\"91\", \"Chromium\";v=\"91\"",
    "sec-ch-ua-mobile": "?0",
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "same-origin",
    "sec-fetch-user": "?1",
    "upgrade-insecure-requests": "1"
  }

while True:
    try:
        path = input('- Enter file name : ').strip()
        book = load_workbook(path)
        break
    except FileNotFoundError:
        print('- No such file or directory')
    except Exception as e:
        print(e)

sheet = book.active
sheet.column_dimensions['C'].width = 24

for i in range(2, sheet.max_row + 1):
    if sheet.cell(i, 1).value is not None:
        try:
            r = session.get(f"https://www.jumia.com.eg/ar/catalog/?q={sheet.cell(i, 1).value}", headers=headers)
            main_img = r.html.find('.img-c > img', first=True).attrs['data-src'].replace('300x300', '500x500')
            print(i, main_img, end='\r')
            if main_img:
                res = session.get(main_img)
                image_file = BytesIO(res.content)
                img = Image(image_file)
                img.width = 200
                img.height = 270
                sheet.row_dimensions[i].height = 206
                sheet.add_image(img, f'C{i}')
        except:
            pass

book.save(path.split('.')[0] + '_IMG.xlsx')
# pyinstaller --add-data c:\users\3mora\anaconda3\envs\autoenv\lib\site-packages\pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info