import subprocess
from time import sleep
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime


class CollectPosts(object):
    def __init__(self):
        self.book = Workbook()
        self.sheet = self.book.active
        self.old_data = []
        self.current = 2
        self.lists = []
        chrome_options = Options()
        chrome_options.add_argument("--user-data-dir={}".format(
            r'C:\Users\3mora\AppData\Local\Google\Chrome\User Data'))
        chrome_options.add_argument('--profile-directory=Profile 3')
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)

    def load_old_data(self, file):
        sheet = load_workbook(file).active
        for i in range(2, sheet.max_row+1):
            url = sheet.cell(i, 3).value
            text = sheet.cell(i, 1).value
            if url:
                self.old_data.append(url)
            if text:
                self.old_data.append(text)

    def safe_find_element_by(self, by, elem):
        try:
            return self.driver.find_element(by, elem)
        except NoSuchElementException:
            return None

    def mobile_start(self, url):
        url = url.replace('www.facebook.com', 'm.facebook.com')
        if '?' in url:
            url = url.split('?')[0]
        self.driver.get(url)
        # sleep(10)

    def GetPostsNew(self, n, e):
        for i in range(n):
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            sleep(4)
            for element in self.driver.find_elements_by_css_selector('[data-sigil="m-feed-voice-subtitle"] > a'):
                u = element.get_attribute('href')
                # if u:
                #     u = u.replace('m.facebook.com', 'mbasic.facebook.com')
                if u not in self.lists and u not in self.old_data:
                    self.lists.append(u)
                if u in self.old_data:
                    print('- found')
            if len(self.lists) > e:
                break
            print(len(self.lists))

    def Post(self, url):
        # url = url.replace('m.face', 'mbasic.face')
        print(self.current)
        self.driver.get(url)
        if self.driver.find_element_by_css_selector('head title').text == 'أنت محظور مؤقتاً':
            raise Exception
        try:
            try:
                name = self.driver.find_element_by_css_selector('h3 strong > a').text
                time = self.driver.find_element_by_css_selector('abbr').text
            except:
                name = None
                time = None

            try:
                text = self.driver.find_element_by_css_selector('.story_body_container > div').text
                img = [element.get_attribute('style').split('"')[1] for element in self.driver.find_elements_by_css_selector('.story_body_container > div div i.img')]
            except:
                text = self.driver.find_element_by_css_selector('.msg > div').text
                img = [element.get_attribute('style').split('"')[1] for element in self.driver.find_elements_by_css_selector('div._57-o._57-n i.img') if element.get_attribute('style') != '']
            if text not in self.old_data or url not in self.old_data:
                self.old_data.append(text)
                self.old_data.append(url)
                self.sheet.cell(self.current, 1).value = text
                self.sheet.cell(self.current, 2).value = '\n'.join(img)
                self.sheet.cell(self.current, 3).value = str(url)
                self.sheet.cell(self.current, 4).value = name
                self.sheet.cell(self.current, 5).value = time
                self.current += 1
        except:
            pass


if __name__ == '__main__':
    self = CollectPosts()
    self.load_old_data('old.xlsx')
    self.mobile_start('https://www.facebook.com/IQmedch/')
'''
اريد 500 من هذه الصفحة
https://www.facebook.com/IQmedch/
و 500 من هذه الصفحة
https://www.facebook.com/drmuawiaaloliwi
و 1000 من هذه الصفحة
https://www.facebook.com/Dr.Nairoukh/



for l in self.lists[353:]:
    print(self.lists.index(l))
    Post(self, l)
    sleep(4)
    

https://www.facebook.com/dsaifsamir/
'''