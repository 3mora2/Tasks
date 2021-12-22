from requests_html import HTMLSession
import time

session = HTMLSession()

from openpyxl import Workbook  # الحلاقة وإزالة الشعر

wb = Workbook()
ws = wb.active
ws['A1'] = 'title'
ws['B1'] = 'Price'
ws['C1'] = 'Manufacture'
ws['D1'] = 'volume'
ws['E1'] = 'Bar code'
ws['F1'] = 'hight'
ws['G1'] = 'battary'
ws['H1'] = 'rata'
ws['I1'] = 'seller'
ws['J1'] = 'availability'
ws['K1'] = 'model'
ws['l1'] = 'category'
ws['M1'] = 'Features'
ws['N1'] = 'detalis1'
ws['O1'] = 'detalis2'
ws['P1'] = 'Descriptions'
ws['Q1'] = 'link'
ws['R1'] = 'link_img'

try:  # 10:24

    num = 2
    page = 1
    while page < 161:  # 161
        r = session.get(
            f'https://www.amazon.sa/s?rh=n%3A12462991031%2Cn%3A%2112462992031%2Cn%3A16630488031&page={page}')

        page += 1
        links = ['https://www.amazon.sa' + link.attrs['href'] for link in r.html.xpath('//div/span/div/div/span/a')]
        if len(links) == 0:
            links = [link.attrs['href'] for link in r.html.xpath('//div/div[3]/div[1]/a')]

        p = 0
        for lin in links:
            respons = session.get(lin)
            html = respons.html
            bar_code = lin[lin.find('/dp/') + 4:lin.rfind('/ref')]
            try:
                title = html.find('#productTitle', first=True).text.strip()
            except:
                title = None
            try:
                price = html.find('#priceblock_ourprice', first=True).text.strip()
            except:
                price = None
            try:
                manuf = html.find('#bylineInfo', first=True).text.strip()
            except:
                manuf = None
            try:
                feature = html.find('#feature-bullets', first=True).text.strip()
            except:
                feature = None
            try:
                availability = html.find('#availability', first=True).text.strip()
            except:
                availability = None
            try:
                category = html.find('#wayfinding-breadcrumbs_container', first=True).text.strip()
            except:
                category = None
            try:
                Description = html.find('#productDescription', first=True).text.strip()
            except:
                Description = None

            #########################################################################
            try:
                whigh = [t.find('.value', first=True).text.strip() for t in
                         html.xpath('//div/table/tbody')[0].xpath('//tr') if 'الوزن' in t.text]
                if whigh == 0:
                    volum = None
                else:
                    volum = whigh[0]

                if volum == None:
                    Volume = [t.find('.value')[0].text.strip() for t in html.xpath('//div/table/tbody')[0].xpath('//tr')
                              if 'Volume' in t.text]
                    if Volume == 0:
                        volum = None
                    else:
                        volum = Volume[0]
                if volum == None:
                    g = [t.find('.value')[0].text.strip() for t in html.xpath('//div/table/tbody')[-1].xpath('//tr') if
                         'وزن' in t.text]
                    if g == 0:
                        volum = None
                    else:
                        volum = g[0]
            except:
                try:
                    volum = \
                    [e.xpath('li/text()', first=True) for e in html.xpath('//tr/td/div/ul/li') if 'وزن' in e.text][
                        0].strip()
                except:
                    volum = None
            try:
                high = [e.xpath('li/text()', first=True).strip() for e in html.xpath('//tr/td/div/ul/li') if
                        'أبعاد' in e.text][0]
            except:
                high = None
            try:
                battary = [e.xpath('li/text()', first=True).strip() for e in html.xpath('//tr/td/div/ul/li') if
                           'بطاريات' in e.text][0]
            except:
                battary = None

            try:  # مرجع الشركة
                model = [e.xpath('li/text()', first=True).strip() for e in html.xpath('//tr/td/div/ul/li') if
                         'رقم موديل' in e.text][0].strip()
            except:
                try:  # مرجع الشركة
                    model = [e.xpath('li/text()', first=True).strip() for e in html.xpath('//tr/td/div/ul/li') if
                             'مرجع الشركة' in e.text][0].strip()
                except:
                    try:
                        model = [s.text.split(':')[1].strip() for s in html.xpath('//*[@id="feature-bullets"]/ul/li') if
                                 'رقم الموديل' in s.text][0]
                    except:
                        model = None
            try:
                rata = [e.xpath('li/span/span', first=True).text.strip() for e in html.xpath('//tr/td/div/ul/li') if
                        'مراجعات المستخدمين' in e.text and 'كن أول' not in e.text][0]
            except:
                rata = None

            try:
                seller = html.find('#sellerProfileTriggerId', first=True).text.strip()
            except:
                seller = None

            try:
                pic = ''
                imgs = [t.attrs['src'] for t in html.xpath('//*[@id="altImages"]/ul', first=True).find('img') if
                        '.gif' not in t.attrs['src']]
                for img in imgs:
                    pic = pic + img + '\n'
            except:
                pic = None

            try:
                con = html.xpath('//*[@id="detail_bullets_id"]', first=True).find('tr', first=True).find('ul')[
                    0].text.strip()
            except:
                con = None

            try:
                con1 = html.xpath('//*[@id="prodDetails"]/div', first=True).text.strip()
            except:
                con1 = None

            p += 1
            print('number :', num - 1, 'number page : ', page - 1)
            # print(p,'category',category,'title : ',title,'volum ',volum,'bar code',bar_code,'price : ',price,'manufstastion : ',manuf,'feature : ',feature,'Description : ',Description,'availability',availability,'link_img',link_img,'\n')
            ws[f'A{num}'] = title
            ws[f'B{num}'] = price
            ws[f'C{num}'] = manuf
            ws[f'D{num}'] = volum
            ws[f'E{num}'] = bar_code
            ws[f'F{num}'] = high
            ws[f'G{num}'] = battary
            ws[f'H{num}'] = rata
            ws[f'I{num}'] = seller
            ws[f'J{num}'] = availability
            ws[f'K{num}'] = model
            ws[f'L{num}'] = category

            ws[f'M{num}'] = feature
            ws[f'N{num}'] = con
            ws[f'O{num}'] = con1
            ws[f'P{num}'] = Description
            ws[f'Q{num}'] = lin
            ws[f'R{num}'] = pic

            num += 1
            time.sleep(.5)


except Exception as e:
    print(e)
    print(lin)

wb.save('الحمام والجسم.xlsx')
