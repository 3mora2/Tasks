# from requests_html import HTMLSession
from selenium import webdriver
from openpyxl import load_workbook
from time import sleep
# s = HTMLSession()
driver = webdriver.Chrome()
url = 'https://sa.iherb.com/c/beauty?noi=192'
driver.get(url)
input('finished')
file = 'NIS.xlsx'  # 'beauty.xlsx'
book = load_workbook(file)
sheet = book.active

book2 = load_workbook('NIS - Copy.xlsx')
sheet2 = book2.active
# i = 1
# sheet.cell(i, 1).value = 'UPC Code'
# sheet.cell(i, 2).value = 'Product Code'
# sheet.cell(i, 3).value = 'product id'
# sheet.cell(i, 4).value = 'Product Title'
# sheet.cell(i, 5).value = 'Product Price'
# sheet.cell(i, 6).value = 'Discounted Price'
# sheet.cell(i, 7).value = 'Product State'
# sheet.cell(i, 8).value = 'Product Brand'
# sheet.cell(i, 9).value = 'Product Rate'
# sheet.cell(i, 10).value = 'Cat1'
# sheet.cell(i, 11).value = 'Cat2'
# sheet.cell(i, 12).value = 'Cat3'
# sheet.cell(i, 13).value = 'Shipping Weight kg'
# sheet.cell(i, 14).value = 'Actual Weight kg'
# sheet.cell(i, 15).value = 'Shipping Weight lb'
# sheet.cell(i, 16).value = 'Actual Weight lb'
# sheet.cell(i, 17).value = 'Dimensions in'
# sheet.cell(i, 18).value = 'Dimensions cm'
# sheet.cell(i, 19).value = 'Expiration Date'
# sheet.cell(i, 20).value = 'Product Quantity'
# sheet.cell(i, 21).value = 'BrandCrumbs'
# sheet.cell(i, 22).value = 'Description'
# sheet.cell(i, 23).value = 'All'
# sheet.cell(i, 24).value = 'Product URL'

# 8:53 ---- 461
for i in range(10, sheet.max_row+1):
    link = sheet[f'BL{i}'].value
    if link is not None and sheet[f'A{i}'].value is None:
        print(i)
        # continue
        try:
            print(i)
            # if i == 100:
            #     break
            driver.get(link)
            # sleep(.5)
            try:
                UPC_Code = \
                [li.text.split(':')[-1].strip() for li in driver.find_elements_by_css_selector('#product-specs-list > li') if
                 'UPC' in li.text][0]
            except:
                sleep(5)
                driver.refresh()
                try:
                    UPC_Code = \
                    [li.text.split(':')[-1].strip() for li in
                     driver.find_elements_by_css_selector('#product-specs-list > li') if
                     'UPC' in li.text][0]
                except:
                    UPC_Code = None
            print(UPC_Code)
            Product_Code = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-part-number')
            product_id = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-product-id')
            title = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-product-name')
            price = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-list-price')
            discounted_price = driver.find_element_by_css_selector('#modelProperties').get_attribute(
                'data-discounted-price')
            try:
                state = driver.find_element_by_css_selector('#stock-status > strong').text
            except:
                state = None
            try:
                brand = driver.find_elements_by_css_selector('#brand > a > span > bdi')[-1].text
            except:
                brand = None
            # try:
            #     _ = driver.find_element_by_css_selector('#product-reviews div.top-rating-wrapper > div').location_once_scrolled_into_view
            #     rate = driver.find_element_by_css_selector('#product-reviews div.top-rating-wrapper > div').text
            # except:
            #     rate = None
            try:
                cat1 = driver.find_element_by_css_selector('#breadCrumbs > a:nth-child(6)').text
            except:
                cat1 = None
            try:
                cat2 = driver.find_elements_by_css_selector('#breadCrumbs > a.last')[1].text
            except:
                cat2 = None
            try:
                cat3 = driver.find_elements_by_css_selector('#breadCrumbs > a.last')[2].text
            except:
                cat3 = None
            shipping_weight_kg = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-shipping-weight-kg')
            actual_weight_kg = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-actual-weight-kg')
            shipping_weight_lb = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-shipping-weight-lb')
            actual_weight_lb = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-actual-weight-lb')
            dimensions_in = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-dimensions-in')
            dimensions_cm = driver.find_element_by_css_selector('#modelProperties').get_attribute('data-dimensions-cm')

            try:
                Expiration_Date = \
                [li.text.split(':?')[-1].strip() for li in driver.find_elements_by_css_selector('#product-specs-list > li')
                 if 'Expiration' in li.text][0]
            except:
                Expiration_Date = None
            try:
                quantity = [li.text.split(':')[-1].strip() for li in driver.find_elements_by_css_selector('#product-specs-list > li') if 'Quantity' in li.text][0]
            except:
                quantity = None
            try:

                des = driver.find_element_by_css_selector('div[itemprop="description"]').text
            except:
                des = None
            try:
                all_des = driver.find_element_by_css_selector('section > div.inner-content').text
            except:
                all_des = None
            try:
                photo = [elem.get_attribute('data-large-img') for elem in
                         driver.find_elements_by_css_selector('#product-image > div.thumbnail-container > img')]
            except:
                photo = []

            try:
                brandCrumbs = driver.find_elements_by_css_selector('#breadCrumbs > a.last')[0].text
            except:
                brandCrumbs = None
            sheet2.cell(i, 1).value = UPC_Code
            sheet2.cell(i, 2).value = Product_Code
            sheet2.cell(i, 3).value = product_id
            sheet2.cell(i, 4).value = title
            sheet2.cell(i, 5).value = price
            sheet2.cell(i, 6).value = discounted_price
            sheet2.cell(i, 7).value = state
            sheet2.cell(i, 8).value = brand
            # sheet2.cell(i, 9).value = rate
            sheet2.cell(i, 10).value = cat1
            sheet2.cell(i, 11).value = cat2
            sheet2.cell(i, 12).value = cat3
            sheet2.cell(i, 13).value = shipping_weight_kg
            sheet2.cell(i, 14).value = actual_weight_kg
            sheet2.cell(i, 15).value = shipping_weight_lb
            sheet2.cell(i, 16).value = actual_weight_lb
            sheet2.cell(i, 17).value = dimensions_in
            sheet2.cell(i, 18).value = dimensions_cm
            sheet2.cell(i, 19).value = Expiration_Date
            sheet2.cell(i, 20).value = quantity
            sheet2.cell(i, 21).value = brandCrumbs
            sheet2.cell(i, 22).value = des
            sheet2.cell(i, 23).value = all_des
            #####################################
            sheet.cell(i, 1).value = UPC_Code
            sheet.cell(i, 2).value = product_id
            sheet.cell(i, 4).value = brand
            sheet.cell(i, 5).value = title
            sheet.cell(i, 8).value = title
            try:
                sheet.cell(i, 12).value = shipping_weight_kg.replace('kg', '').strip()
            except:
                sheet.cell(i, 12).value = shipping_weight_kg
            if des is not None:
                try:
                    t = des.split(' ')
                    n = len(t) // 3
                    t1 = ' '.join(t[:n])
                    t2 = ' '.join(t[n:-n])
                    t3 = ' '.join(t[-n:])
                    sheet.cell(i, 30).value = t1
                    sheet.cell(i, 31).value = t2
                    sheet.cell(i, 32).value = t3
                except:
                    pass
            try:
                for co in range(len(photo)):
                    sheet.cell(i, co+35).value = photo[co]
            except:
                pass
            try:
                sheet.cell(i, 53).value = price.replace('SAR', '').strip()
            except:
                sheet.cell(i, 53).value = price

            if i % 50 == 0:
                print('save')
                book.save(file)
                book2.save('NIS - Copy.xlsx')
                sleep(200)

        except Exception as e:
            print(e)

book.save(file)
book2.save('NIS - Copy.xlsx')
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# product = []
# url = 'https://sa.iherb.com/c/beauty?noi=192'
# 'https://sa.iherb.com/pro/countryselected/?isAjax=true&CountryCode=SA&CurrencyCode=SAR&LanguageCode=en-US'
# while True:
#     r = s.get(url)
#     print(url)
#     product += [ele.attrs['href'] for ele in r.html.find('div.product-inner.product-inner-wide > div.absolute-link-wrapper > a')]
#     url = 'https://sa.iherb.com'+r.html.find('.pagination-next')[0].attrs['href']
#     print(len(product))
#     sleep(3)
# for i, item in enumerate(product):
#     print(i+2)
#     ws[f'W{i+2}'] = item
# wb.save('beauty.xlsx')

