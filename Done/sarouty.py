from requests_html import HTMLSession
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from time import sleep

from openpyxl import Workbook

wb = Workbook()
# wb = load_workbook('3tor.xlsx')
ws = wb.active
num = 1
ws[f'A{num}'] = 'Title'
ws[f'B{num}'] = 'City'
ws[f'C{num}'] = 'District'
ws[f'D{num}'] = 'Price'
ws[f'E{num}'] = 'Property_Type'
ws[f'F{num}'] = 'Reference'
ws[f'G{num}'] = 'Area'
ws[f'H{num}'] = 'Rooms'
ws[f'I{num}'] = 'Bathrooms'
ws[f'J{num}'] = 'Des'
ws[f'K{num}'] = 'Pictures'
ws[f'L{num}'] = 'urls'
num += 1

s = HTMLSession()
links = []
'''
https://www.sarouty.ma/fr/broker/guy-hoquet-rabat-77
https://www.sarouty.ma/fr/broker/marrakani-immobilier-2070
https://www.sarouty.ma/fr/broker/luximmo-318
'''
for page in range(1, 10 + 1):
    r = s.get(
        f'https://www.sarouty.ma/fr/broker/luximmo-318?properties[page]={page}&properties[order_by]=-featured')
    for url in r.html.find('.card-list__item a'):
        links.append('https://www.sarouty.ma' + url.attrs['href'])
        print(len(links))

driver = webdriver.Chrome()
for urls in links:
    driver.get(urls)
    City = driver.find_elements_by_css_selector('div.breadcrumb a.breadcrumb__link ')[0].text
    District = driver.find_elements_by_css_selector('div.breadcrumb a.breadcrumb__link ')[1].text
    Title = driver.find_element_by_css_selector('.property-header__title--detail').text
    try:
        Price = driver.find_element_by_css_selector('.facts__price').text
    except:
        Price = None
    try:
        Property_Type = [elem.find_element_by_css_selector('.facts__content').text for elem in
                         driver.find_elements_by_css_selector('.facts__list-item') if 'ype' in elem.text][0]
    except:
        Property_Type = None
    try:
        Reference = [elem.find_element_by_css_selector('.facts__content').text for elem in
                     driver.find_elements_by_css_selector('.facts__list-item') if 'férenc' in elem.text][0]
    except:
        Reference = None
    try:
        Area = [elem.find_element_by_css_selector('.facts__content').text for elem in
                driver.find_elements_by_css_selector('.facts__list-item') if 'urfac' in elem.text][0]
    except:
        Area = None
    try:
        Rooms = [elem.find_element_by_css_selector('.facts__content').text for elem in
                 driver.find_elements_by_css_selector('.facts__list-item') if 'ièce' in elem.text][0]
    except:
        Rooms = None
    try:
        Bathrooms = [elem.find_element_by_css_selector('.facts__content').text for elem in
                     driver.find_elements_by_css_selector('.facts__list-item') if 'de bain' in elem.text][0]
    except:
        Bathrooms = None
    sleep(2)
    try:
        try:
            _ = driver.find_element_by_css_selector('div.gallery__show-all').location_once_scrolled_into_view
            WebDriverWait(driver, 10).until(
                ec.visibility_of_element_located((By.CSS_SELECTOR, "div.gallery__show-all"))).click()
        except:
            _ = driver.find_element_by_css_selector('div.gallery__show-all').location_once_scrolled_into_view
            driver.find_element_by_css_selector('div.gallery__show-all').click()
    except:
        pass
    sleep(1)
    Pictures = ' ,\n'.join([element.get_attribute('src') for element in
                            driver.find_elements_by_css_selector('.gallery__img.progressive-image--loaded ')])
    Des = driver.find_element_by_css_selector('.property-description__text-trim').text
    ws[f'A{num}'] = Title
    ws[f'B{num}'] = City
    ws[f'C{num}'] = District
    ws[f'D{num}'] = Price
    ws[f'E{num}'] = Property_Type
    ws[f'F{num}'] = Reference
    ws[f'G{num}'] = Area
    ws[f'H{num}'] = Rooms
    ws[f'I{num}'] = Bathrooms
    ws[f'J{num}'] = Des
    ws[f'K{num}'] = Pictures
    ws[f'L{num}'] = urls
    num += 1
    print(num - 1)
wb.save('sarouty3.xlsx')
