from requests_html import HTMLSession
from time import sleep

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 'Title'
ws['B1'] = 'Type'
ws['C1'] = 'price'
ws['D1'] = 'ref'
ws['E1'] = 'bedrooms'
ws['F1'] = 'bathrooms'
ws['G1'] = 'Area'
ws['H1'] = 'Description'
ws['I1'] = 'URL'
ws['J1'] = 'Img_url'
num = 2
#####################################################################################
session = HTMLSession()
page = 1
try:
    while page < 49:  #
        r = session.get(f'https://nourreska.com/offres-immobilieres-par-statut/location/page/{page}')
        page += 1
        url = [ur.xpath('//a')[0].attrs['href'] for ur in r.html.find('.rh_figure_property_one')]
        for urll in url:
            respons = session.get(urll)
            ht = respons.html
            try:
                title = ht.find('.rh_page__title', first=True).text
            except:
                title = None

            try:
                Type = ht.find('.property-breadcrumbs', first=True).xpath('//li')[-1].text.split(':')[1].strip()
            except:
                Type = None

            try:
                pic = ht.xpath('//*[@id="property-featured-image"]/a', first=True).attrs['href']
            except:
                pic = ''
                try:
                    main_img = [im.attrs['src'] for im in ht.find('.slides', first=True).find('img')]
                    for img in main_img:
                        pic = pic + img + '\n'
                except Exception as e:
                    print(e)

            try:
                price = ht.find('.price', first=True).text
            except:
                price = None

            try:
                ref = ht.find('.id', first=True).text
            except:
                ref = None

            try:
                bedrooms = ht.find('.prop_bedrooms', first=True).find('.figure', first=True).text
            except:
                bedrooms = None

            try:
                bathrooms = ht.find('.prop_bathrooms', first=True).find('.figure', first=True).text
            except:
                bathrooms = None

            try:
                area = ht.find('.prop_area', first=True).xpath('//div/div')[0].text
            except:
                area = None

            try:
                o = ht.find('.rh_content', first=True).text
                Description = o[:o.rfind('\nArticles similaires')]
            except:
                Description = None

            # print(title,Type,pic,price,ref,bedrooms,bathrooms,area,'\n',Description,'\n',urll,'\n')
            print(num - 1)
            sleep(5)
            ws[f'A{num}'] = title
            ws[f'B{num}'] = Type
            ws[f'C{num}'] = price
            ws[f'D{num}'] = ref
            ws[f'E{num}'] = bedrooms
            ws[f'F{num}'] = bathrooms
            ws[f'G{num}'] = area
            ws[f'H{num}'] = Description
            ws[f'I{num}'] = urll
            ws[f'J{num}'] = pic
            print(len(pic))
            num += 1

except Exception as a:
    print(a)

wb.save("Nourreska-location.xlsx")
