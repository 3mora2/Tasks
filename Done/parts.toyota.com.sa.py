import os
import sys
import warnings
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from time import sleep
from requests_html import HTML
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from webdriver_manager.firefox import GeckoDriverManager

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class Main:
    def __init__(self):
        try:
            self.driver = webdriver.Firefox()
        except:
            self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())

    def wait_load(self):
        sleep(3)
        while True:
            # try:
            #     if 'block' not in self.driver.find_element_by_css_selector('#loading_section').get_attribute('style'):
            #         break
            # except:
            #     pass

            try:
                if 'block' not in self.driver.find_element_by_css_selector('#loading').get_attribute('style'):
                    break
            except:
                break
        sleep(4)

    def close_int(self):
        try:
            self.driver.find_element_by_css_selector(
                '.ins-selectable-element.ins-element-close-button').click()
            print('- close not')
        except:
            pass
        try:
            self.driver.find_element_by_css_selector('#closedialog').click()
        except:
            pass

    def get_product(self, url):
        file_name = '-'.join(url.split('/')[-3:-1])
        book = Workbook()
        sheet = book.active
        n = 2
        same = 0
        old = 0
        prod = []
        self.driver.get(url)
        sleep(3)
        while True:
            if WebDriverWait(self.driver, 60).until(
                    (ec.visibility_of_element_located((By.CSS_SELECTOR, '#filter-result')))).text != '(0)':
                break

        cat = self.driver.find_element_by_css_selector('div.entry-meta.mb-4>ul').text
        while True:

            for element in self.driver.find_elements_by_css_selector('a.catalogue-itemdetails'):
                code = element.get_attribute('data-pn')
                if code not in prod:
                    prod.append(code)
                    sheet.cell(n, 30).value = 'https://parts.toyota.com.sa/spare-parts/AR/' + code
                    sheet.cell(n, 29).value = code
                    sheet.cell(n, 28).value = cat
                    n += 1
                print(prod.__len__(), n)
            if old == prod.__len__():
                same += 1
            else:
                old = prod.__len__()

            if same == 3:
                break
            try:
                self.driver.find_element_by_css_selector('.category-nex').click()
                print('- next')
                self.wait_load()
                sleep(2)
            except Exception as e:
                print(e)
                if len(self.driver.find_elements_by_css_selector(
                        '.ins-selectable-element.ins-element-close-button')) > 0:
                    self.close_int()
                    continue
                elif self.driver.find_element_by_css_selector('#filter-result').text == '0':
                    break
                if input('- if u want break enter (q)') == 'q':
                    break

        book.save(file_name + '.xlsx')

    def safe_find_element_by(self, by, elem):
        try:
            return self.driver.find_element(by, elem)
        except NoSuchElementException:
            return None

    def get_details(self, file):
        self.book = load_workbook(file)
        sheet = self.book.active
        for i in range(2, sheet.max_row + 1):
            url = sheet.cell(i, 30).value
            if url and not sheet.cell(i, 10).value:
                self.driver.delete_all_cookies()
                self.driver.get(url.replace('/AR/', '/EN/'))
                self.wait_load()

                cat = self.safe_find_element_by(By.CSS_SELECTOR, 'div.entry-meta>ul')
                if cat:
                    cat = cat.text
                full_name = self.safe_find_element_by(By.CSS_SELECTOR, 'h1')
                if full_name:
                    full_name = full_name.text
                state = self.safe_find_element_by(By.CSS_SELECTOR, '#normal-price > .row > .text-md-right > p')
                if state:
                    state = state.text
                name = self.safe_find_element_by(By.CSS_SELECTOR, '#normal-price > .row  h6')
                if name:
                    name = name.text
                desk = self.safe_find_element_by(By.CSS_SELECTOR, '.details-block.details-weight')
                if desk:
                    desk = desk.text
                price = self.safe_find_element_by(By.CSS_SELECTOR, '#normal-price > .row > .text-md-right > h3')
                if price:
                    price = price.text
                model = self.safe_find_element_by(By.CSS_SELECTOR, '#application')
                if model:
                    model = model.get_attribute('innerHTML').replace('\n', '').strip().replace('   ', '')

                details = self.safe_find_element_by(By.CSS_SELECTOR, '#description')
                if details:
                    details = details.get_attribute('innerHTML').replace('\n', '').strip().replace('   ', '')

                all_page = self.safe_find_element_by(By.CSS_SELECTOR, '#itemDetails')
                if all_page:
                    all_page = all_page.get_attribute('innerHTML').replace('\n', '').strip().replace('   ', '')

                # sheet.cell(i, 1).value = full_name
                # sheet.cell(i, 2).value = name
                # sheet.cell(i, 3).value = state
                # sheet.cell(i, 4).value = price
                # sheet.cell(i, 5).value = desk
                # sheet.cell(i, 6).value = model
                # sheet.cell(i, 7).value = details
                # sheet.cell(i, 8).value = cat
                # sheet.cell(i, 9).value = all_page

                sheet.cell(i, 10).value = full_name
                sheet.cell(i, 11).value = name
                sheet.cell(i, 12).value = state
                sheet.cell(i, 13).value = price
                sheet.cell(i, 14).value = desk
                sheet.cell(i, 15).value = model
                sheet.cell(i, 16).value = details
                sheet.cell(i, 17).value = cat
                sheet.cell(i, 18).value = all_page
                print(i, name, price)
        self.book.save(file)


# if __name__ == '__main__':
#     self = Main()
    # self.get_details('final.xlsx')#self.book.save('final.xlsx')
    # urls = [
    #         'https://parts.toyota.com.sa/catalogue/264/']
    # for
    # url in urls:
    #     self.get_product(url)

# edit excel
book = load_workbook('final to edit.xlsx')
sheet = book.active
for i in range(2, 10):
    models = sheet[f'I{i}'].value.replace('<th', '<td').replace('th>', 'td>').replace('\n', '').strip()
    html = HTML(html=sheet[f'J{i}'].value)
    desk = html.find('ul')[0].html
    wight = html.find('ul>li')[-1].text.replace('الوزن:', '').replace('kg', '').strip()
    total = desk + '\n' + models
    sheet[f'K{i}'].value = total
    sheet[f'L{i}'].value = wight
    print(i)
book.save('test_edit.xlsx')
