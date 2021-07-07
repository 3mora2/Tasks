import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import pyautogui
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup



wb = Workbook()
ws = wb.active
ws['A1'] = 'Name'
ws['B1'] = 'Price'
ws['C1'] = 'Img'
ws['D1'] = 'Desc link'
ws['E1'] = 'full Desc'
ws['F1'] = 'some desc'
ws['G1'] = 'all desc'
ws['H1'] = 'Code1'
ws['I1'] = 'Code2'
ws['J1'] = 'Code3'
ws['K1'] = 'Code4'
ws['L1'] = 'Code5'
ws['M1'] = 'Code6'
ws['N1'] = 'Code7'
ws['O1'] = 'Code8'
ws['P1'] = 'Code9'
ws['Q1'] = 'Code10'
ws['R1'] = 'Code11'
ws['S1'] = 'Code12'
ws['T1'] = 'Pic'



#ws['B1'] = datetime.datetime.now()
x=2
y=2
z=2
w=2
s=2
p=2
t=2
m=2
n=2
i=2
l=2
o=2
q=2
j=2
xx=2
yy=2
zz=2
ww=2
mm=2
nn=2
ll=2
oo=2
pp=2
qq=2


url='https://deals.souq.com/sa-ar/men-care/c/14878'



driver = webdriver.Chrome()
driver.maximize_window() 
driver.get(url)
time.sleep(5)

for i in range(3):
    pyautogui.scroll(-3000)
    time.sleep(3)


names=driver.find_elements_by_class_name('itemTitle')
prices=driver.find_elements_by_class_name('price')
#images=driver.find_elements_by_tag_name('img')
Desc_links = driver.find_elements_by_class_name('img-link')

for name in names:
    driver.find_elements_by_class_name('itemTitle')

    ws['A{}'.format(x)]=name.text
    x+=1

    print(name.text)

for price in prices:
    driver.find_elements_by_class_name('price')

    ws['B{}'.format(y)]=price.text
    y+=1
    
    print(price.text)

#for image in images:
    #driver.find_elements_by_tag_name('img')

    #ws['C{}'.format(z)]=image.get_attribute('data-src')
    #z+=1
    #print(image.get_attribute('data-src'))

 
drivera = webdriver.Chrome()
drivera.maximize_window()

for Desc_link in Desc_links:

    ws['D{}'.format(w)]=Desc_link.get_attribute('href')
    w+=1
    print(Desc_link.get_attribute('href'))

    dls=Desc_link.get_attribute('href')
    
    drivera.get(dls)
    l=''


    soup = BeautifulSoup(drivera.page_source, 'html.parser')

    time.sleep(3)

    drivera.find_element_by_xpath('//*[@id="specs"]/a').click()

    time.sleep(2)
    
    descs = soup.find(id='description-short')
    ws['E{}'.format(s)]=descs.text
    s+=1
    print(descs.text)
    
    fulls = soup.find(id='specs-short')
    ws['F{}'.format(t)]=fulls.text
    t+=1
    print(fulls.text)
    
    descsall = soup.find(id='specs-full')
    ws['G{}'.format(m)]=descsall.text
    m+=1
    print(descsall.text)

    code1 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[1]') # or [1,2,3]
    ws['H{}'.format(n)]=code1.text
    n+=1
    print(code1.text)

    code2 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[2]') # or [1,2,3]
    ws['I{}'.format(i)]=code2.text
    i+=1
    print(code2.text)

    code3 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[3]') # or [1,2,3]
    ws['J{}'.format(o)]=code3.text
    o+=1
    print(code3.text)

    code4 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[4]') # or [1,2,3]
    ws['K{}'.format(q)]=code4.text
    q+=1
    print(code4.text)

    try:
        code5 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[5]') # or [1,2,3]
        ws['L{}'.format(xx)]=code5.text
        xx+=1
        print(code5.text)
    except:
        ws['L{}'.format(xx)]='No Data'
        xx+=1

    try:
        code6 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[6]') # or [1,2,3]
        ws['M{}'.format(yy)]=code6.text
        yy+=1
        print(code6.text)
    except:
        ws['M{}'.format(yy)]='No Data'
        yy+=1

    try:
        code7 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[7]') # or [1,2,3]
        ws['N{}'.format(zz)]=code7.text
        zz+=1
        print(code7.text)
    except:
        ws['N{}'.format(zz)]='No Data'
        zz+=1

    try:
        code8 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[8]') # or [1,2,3]
        ws['O{}'.format(mm)]=code8.text
        mm+=1
        print(code8.text)
    except:
        ws['O{}'.format(mm)]='No Data'
        mm+=1

    try:
        code9 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[9]') # or [1,2,3]
        ws['P{}'.format(nn)]=code9.text
        nn+=1
        print(code9.text)
    except:
        ws['P{}'.format(nn)]='No Data'
        nn+=1

    try:
        code10 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[10]') # or [1,2,3]
        ws['Q{}'.format(ll)]=code10.text
        ll+=1
        print(code10.text)
    except:
        ws['Q{}'.format(ll)]='No Data'
        ll+=1

    try:
        code11 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[11]') # or [1,2,3]
        ws['R{}'.format(oo)]=code11.text
        oo+=1
        print(code11.text)
    except:
        ws['R{}'.format(oo)]='No Data'
        oo+=1

    try:
        code12 = drivera.find_element_by_xpath('//*[@id="specs-full"]/dl/dd[12]') # or [1,2,3]
        ws['S{}'.format(pp)]=code12.text
        pp+=1
        print(code12.text)
    except:
        ws['S{}'.format(pp)]='No Data'
        pp+=1
        
        
        
    descs_pics = soup.find('div', {'class' : 'slick-track'})
    for descs_pic in descs_pics.find_all('img'):
        l=descs_pic.attrs['src']+'\n'+l

        print(descs_pic.attrs['src'])
  
    ws['T{}'.format(p)]=l
    p+=1
    
        

    #drivera.quit()    
        


        

     

wb.save("العناية الشخصية الرجالية.xlsx")    
    
