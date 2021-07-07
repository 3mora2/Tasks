from msedge.selenium_tools import Edge, EdgeOptions
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import os
from openpyxl import load_workbook
from requests_html import HTMLSession
from time import sleep
from pywinauto.application import Application

session = HTMLSession()


class Main:
    def __init__(self):
        self.page = 1
        self.index = 1
        self.url_login = 'https://egshoping.com/index.php?route=account/login'
        self.url_add = 'https://egshoping.com/index.php?route=seller/account-product/create'

        # self.book = load_workbook('dddd.xlsx')
        self.book = load_workbook('soq_data_part_1.xlsx')
        # self.book = load_workbook('soq_data_part_1 - Copy.xlsx')
        self.sheet = self.book.active

        options = EdgeOptions()
        options.use_chromium = True
        options.add_argument("disable-gpu")
        self.driver = Edge(options=options)

    def check_login(self):
        self.driver.get(self.url_add)
        if 'seller/account-product/create' not in self.driver.current_url:
            self.driver.get(self.url_login)
            self.driver.find_element_by_name('email').send_keys('ahmedali@gmail.com')
            self.driver.find_element_by_name('password').send_keys('Password12')
            self.driver.find_element_by_class_name('button').click()

    def extract_soq(self):
        path = 'photo'
        if not os.path.exists(path):
            os.makedirs(path)
        products_url = []
        self.driver.get('https://egypt.souq.com/eg-en/crazykids/p/')
        box = self.driver.find_element_by_xpath('//*[@id="id_type_item-accordion"]/div/ul')
        page_urls = []
        for page_url in box.find_elements_by_class_name("sk-clr2-eff"):
            page_urls.append(
                page_url.find_element_by_class_name("valign-mid").get_attribute('value') + '?section=2&page=')
        print('number page', len(page_urls))
        for page_ur in page_urls:
            self.page = 1
            while True:
                self.driver.get(page_ur + f'{self.page}')
                self.page += 1
                if len(self.driver.find_elements_by_class_name("itemLink")) > 1:
                    cat = self.driver.find_element_by_xpath(
                        '//*[@id="id_type_item-accordion"]/div/ul/li[1]').get_attribute('data-refinement')
                    for pro in self.driver.find_elements_by_class_name("itemLink"):
                        if pro.get_attribute('href') != '':
                            name = pro.text
                            '''
                            r = session.get(pro.get_attribute('href'))
                            name = r.html.find('.product-title', first=True).find('h1', first=True).text
                            try:
                                desc = r.html.find('#description-full', first=True).text

                            except:
                                desc = r.html.find('#description-short', first=True).text
                            price = r.html.find('h3.price ', first=True).xpath('h3/text()[2]', first=True).strip()
                            photo = ''
                            for i in r.html.find('a div.img-bucket img'):
                                if '.gif' not in i.attrs['data-url']:
                                    do=session.get(i.attrs['data-url'])
                                    with open(f'photo\\{i.attrs["data-url"].split("/item/")[1].replace("/","-")}', 'wb') as fil:
                                        fil.write(do.content)
                                    photo += i.attrs["data-url"].split("/item/")[1].replace("/","-") + '~'

                            ############################################################################
                            res = session.get(pro.get_attribute('href').replace('g-ar', 'g-en'))
                            name_en = res.html.find('.product-title', first=True).find('h1', first=True).text
                            try:
                                desc_en = res.html.find('#description-full', first=True).text

                            except:
                                desc_en = res.html.find('#description-short', first=True).text
                            print(photo)
                            print('*' * 50)

                            # products_url.append(pro.get_attribute('href'))
                            self.sheet[f'A{self.index}'] = name'''
                            self.sheet[f'A{self.index}'] = name
                            # self.sheet[f'C{self.index}'] = cat
                            # self.sheet[f'D{self.index}'] = price
                            # self.sheet[f'E{self.index}'] = desc
                            # self.sheet[f'F{self.index}'] = desc_en
                            # self.sheet[f'H{self.index}'] = pro.get_attribute('href')
                            # self.sheet[f'G{self.index}'] = photo
                            print(self.index)
                            self.index += 1

                else:
                    print('******************* No new *****************', len(products_url))
                    break
                self.book.save('dddd.xlsx')

    def upload(self):
        self.get_exel()
        self.check_login()
        self.driver.get(self.url_add)
        for name, name_en, cat, price, desc, desc_en, phot in zip(self.Name, self.Name_en, self.Categ, self.Pric,
                                                                  self.Des, self.Des_en, self.Photo):
            # print(name, name_en, cat, price)# , desc, desc_en, phot)

            self.driver.find_element_by_xpath('//*[@id="language-tabs"]/a[1]').click()
            # name en
            self.driver.find_element_by_name('languages[1][product_name]').send_keys(name_en)
            sleep(.5)
            # descri en
            WebDriverWait(self.driver, 10).until(
                EC.frame_to_be_available_and_switch_to_it((By.XPATH, '//*[@id="cke_1_contents"]/iframe')))
            sleep(1)
            self.driver.find_element_by_tag_name('body').send_keys(desc_en)
            self.driver.switch_to.default_content()
            sleep(.5)
            ################################################################################################################
            # swith to arbic
            self.driver.find_element_by_xpath('//*[@id="language-tabs"]/a[2]').click()
            # name arbic
            self.driver.find_element_by_name('languages[2][product_name]').send_keys(name)
            sleep(.5)
            # dicrib arbic
            WebDriverWait(self.driver, 10).until(
                EC.frame_to_be_available_and_switch_to_it((By.XPATH, '//*[@id="cke_2_contents"]/iframe')))
            sleep(1)
            self.driver.find_element_by_tag_name('body').send_keys(desc)
            self.driver.switch_to.default_content()
            sleep(.5)
            # tag
            self.driver.find_element_by_name('languages[2][product_tags]').send_keys(name)
            # select
            self.driver.find_element_by_class_name('select2-search__field').send_keys(cat)# اكسسوار موبايل
            self.driver.find_element_by_class_name('select2-search__field').send_keys(Keys.ENTER)
            # price
            self.driver.find_element_by_name('product_price').send_keys(price)
            for ph in phot.split('~'):
                if phot.split('~').index(ph) == 3:
                    break
                print(ph)
                # add photo
                if ph == '':
                    continue
                _ = self.driver.find_element_by_name('ms-file-addimages').location_once_scrolled_into_view
                sleep(.5)
                self.driver.find_element_by_id('ms-file-addimages').click()

                sleep(2)
                try:
                    app = Application().connect(title_re='Open')
                except:
                    self.driver.find_element_by_id('ms-file-addimages').click()
                    sleep(2)
                    app = Application().connect(title_re='Open')

                app.window().Edit.set_edit_text(f'C:\\Users\\3mora\\Dropbox\\wael\\photo1\\{ph}')
                sleep(.5)
                app.window().Button.click()
                try:
                    app.window().Button.click()
                    app.window().Button.click()
                except:
                    pass
            sleep(1)
            # save
            self.driver.find_element_by_id('ms-submit-button').click()

    def get_exel(self):
        # end_cell = self.sheet.max_row
        end_cell = 5
        self.Name = [self.sheet.cell(i, 1).value for i in range(1, end_cell + 1)]
        self.Name_en = [self.sheet.cell(i, 2).value for i in range(1, end_cell + 1)]
        self.Categ = [self.sheet.cell(i, 3).value for i in range(1, end_cell + 1)]
        self.Pric = [self.sheet.cell(i, 4).value for i in range(1, end_cell + 1)]
        self.Des = [self.sheet.cell(i, 5).value for i in range(1, end_cell + 1)]
        self.Des_en = [self.sheet.cell(i, 6).value for i in range(1, end_cell + 1)]
        self.Photo = [self.sheet.cell(i, 7).value for i in range(1, end_cell + 1)]

    def detels_exl(self):
        path = 'photo1'
        if not os.path.exists(path):
            os.makedirs(path)
        self.index = 2165
        end_cell = self.sheet.max_row
        Cat = [self.sheet.cell(i, 3).value for i in range(self.index, end_cell + 1)]
        Url = [self.sheet.cell(i, 8).value for i in range(self.index, end_cell + 1)]
        # print(len(Cat), len(Url))
        for cat, url in zip(Cat, Url):
            # print(cat, url)
            r = session.get(url)
            name = r.html.find('.product-title', first=True).find('h1', first=True).text
            try:
                sell = r.html.find('span.unit-seller-link a b', first=True).text
            except:
                sell = ''

            try:
                desc = r.html.find('#description-full', first=True).text
            except:
                try:
                    desc = r.html.find('#description-short', first=True).text
                except:
                    try:
                        desc = r.html.find('#specs-full',first=True).text
                    except:
                        desc = r.html.find('#specs-short',first=True).text

            price = r.html.find('h3.price ', first=True).xpath('h3/text()[2]', first=True).strip()
            photo = ''
            for i in r.html.find('a div.img-bucket img'):
                if '.gif' not in i.attrs['data-url']:
                    do = session.get(i.attrs['data-url'])
                    with open(f'{path}\\{i.attrs["data-url"].split("/item/")[1].replace("/", "-")}', 'wb') as fil:
                        fil.write(do.content)
                    photo += i.attrs["data-url"].split("/item/")[1].replace("/", "-") + '~'

            ############################################################################
            res = session.get(url.replace('g-ar', 'g-en'))
            name_en = res.html.find('.product-title', first=True).find('h1', first=True).text
            try:
                desc_en = res.html.find('#description-full', first=True).text

            except:
                try:
                    desc_en = r.html.find('#description-short', first=True).text
                except:
                    try:
                        desc_en = r.html.find('#specs-full', first=True).text
                    except:
                        desc_en = r.html.find('#specs-short', first=True).text

            self.sheet[f'A{self.index}'] = name
            self.sheet[f'B{self.index}'] = name_en
            self.sheet[f'C{self.index}'] = cat
            self.sheet[f'D{self.index}'] = price
            self.sheet[f'E{self.index}'] = desc
            self.sheet[f'F{self.index}'] = desc_en
            self.sheet[f'H{self.index}'] = url
            self.sheet[f'G{self.index}'] = photo
            self.sheet[f'J{self.index}'] = sell
            print(self.index, sell)
            self.index += 1
            self.book.save('soq_data_part_1.xlsx')


f = Main()
# f.detels_exl()
f.upload()
