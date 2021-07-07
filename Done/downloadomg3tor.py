import requests_html
import urllib.request
from time import sleep
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

chrome_options = Options()
chrome_options.add_argument("--user-data-dir={}".format(
    r'C:\Users\3mora\AppData\Local\Google\Chrome\User Data'))
driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)


def get_image(url_):
    driver.get(url_)
    sleep(2)
    input('- enter : ')
    # Get the dimensions of the browser and image.
    orig_h = driver.execute_script("return window.outerHeight")
    orig_w = driver.execute_script("return window.outerWidth")
    margin_h = orig_h - driver.execute_script("return window.innerHeight")
    margin_w = orig_w - driver.execute_script("return window.innerWidth")
    new_h = driver.execute_script('return document.getElementsByTagName("img")[0].height')
    new_w = driver.execute_script('return document.getElementsByTagName("img")[0].width')

    # Resize the browser window.
    driver.set_window_size(new_w + margin_w, new_h + margin_h)

    # Get the image by taking a screenshot of the page.
    img_val = driver.get_screenshot_as_png()
    # Set the window size back to what it was.
    driver.set_window_size(orig_w, orig_h)

    return img_val


with open('test.jpg', 'wb') as f:
    f.write(get_image(url_='https://cdn.danube.sa/spree/images/attachments/000/026/896/large/Dove_Beauty_Cream_Bar_Pink_160gm_6281006481312_FOP.jpg'))


# lin = []
# wb = load_workbook('Tassel_Full.xlsx')
# ws = wb.active
# print(ws.max_row)
# for i in range(462, ws.max_row + 1):
#     # if i in [539,1084,1448,1449,1450,1451,1453,1454,1457,1458,1460,1463]:
#     ur = ws[f'M{i}'].value
#     na = ws[f'A{i}'].value
#     if ur is None:
#         continue
#     co = ur.split('\n')
#     # if ws[f'P{i}'].value is not None or ws[f'P{i}'].value != '':
#
#     l = []
#     p = 1
#     for url in co:
#         if co.index(url) == 0:
#             name_img = str(na)
#         else:
#             name_img = str(na) + f'-{p}'
#             p += 1
#         if name_img not in l:
#             l.append(name_img)
#         lin.append(url)
#         print(len(lin))
        #     #if ws[f'P{i}'].value != ' ,\n'.join(l):
        # print(i)
        # # r = requests.get(url)
        # img = get_image(url_=url)  # 463,539,552,667,668,670,702,1084,1462
        # try:
        #     # r = urllib.request.urlopen(url.replace('550x550', '1100x1100'))
        #     sleep(1)
        #     with open('photo//' + name_img + '.jpg', 'wb') as f:
        #         f.write(img)
        #     print(i, ' - ', name_img)
        #     ws[f'P{i}'].value = ' ,\n'.join(l)
        #     wb.save('Tassel_Full.xlsx')
        # except Exception as e:
        #     print('***************************',i)
        #     print(e)
