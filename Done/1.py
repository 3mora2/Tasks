
from requests_html import HTMLSession,HTML,AsyncHTMLSession
session=HTMLSession()
asession = AsyncHTMLSession()
#############################################
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
wb = Workbook()
ws = wb.active

ws['A1'] = 'Name'
ws['B1'] = 'Price'
ws['C1'] = 'categrary'
ws['D1'] = 'Référence'
ws['E1'] = 'details'
ws['F1'] = 'url'
ws['G1'] = 'img url'
############################################
'''wb1 = Workbook()
ws1 = wb1.active
ws1['A1'] = 'img'
ws1.column_dimensions['A'].width = 13
ws1['B1'] = 'Name'
ws1['C1'] = 'Price' '''
############################################
import os
file='photo'
if not os.path.exists(file):
    os.makedirs(file)
def Links():
    from selenium import webdriver
    from time import sleep
    import random
    ########################################
    while True:
        urls=input(' enter url example(http://www.parapharma.ma/20-para-medical) : ')
        if 'parapharma.ma' in urls:
            break
        else:
            print(' ---------------- wrong url------------------ ')
    driver = webdriver.Firefox()
    driver.implicitly_wait(4)
    driver.get(urls)
    sleep(random.randint(3,6))
    try:driver.find_element_by_id('nb_item1').send_keys(50)
    except:pass
    sleep(random.randint(3,6))
    u=True
    links=[]
    while u:
        html = HTML(html=driver.page_source)
        [links.append(i.attrs['href']) for i in html.find('.center_block a.product_img_link')]
        try:
            Next=driver.find_element_by_id('pagination_next')
            
            if Next.get_attribute('class')=='':
                Next.click()
                sleep(random.randint(4,6))
            else:
                u=False
        except:u=False            
    u=False
    driver.quit()
    return links
 
#####################################################################################################
async def extract_data():
    links=Links()
    num=2
    for link in links:
        res=await asession.get(link)
        try:name=res.html.find('#pb-left-column h1',first=True).text
        except:name=None
        try:price=res.html.find('span#our_price_display',first=True).text
        except:price=None
        try:categr=res.html.find('.breadcrumb a')[1].text
        except:categr=None
        try:ref=res.html.find('span.editable',first=True).text
        except:ref=None
        try:img_url=res.html.find('#image-block img',first=True).attrs['src']
        except:img_url=None
        try:detels=res.html.find('#more_info_sheets',first=True).text
        except:detels=None
        #########################################################
        try:
            img_data=await asession.get(img_url)
            with open(file+'\\'+name+'.jpg','wb') as f:
                f.write(img_data.content)
        except:
            pass

        '''########################################################
        ws[f'A{num}'] = name
        ws1[f'B{num}'] = name
        
        ws[f'B{num}'] = price
        ws1[f'C{num}'] = price
        
        ws[f'C{num}'] = categr
        ws[f'D{num}'] = ref
        ws[f'E{num}'] = detels
        ws[f'F{num}'] = link
        ws[f'G{num}'] = img_url
        print('product : '+(num-1))
        num+=1
    wb.save('parapharma.xlsx')
    wb1.save('parapharma_img.xlsx')'''
#########################################################
asession.run(extract_data)
#extract_data(links)
#########################################################

