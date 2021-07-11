from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.webdriver import WebDriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
import os
import sys
import warnings
from random import choice

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class Send:
    driver: WebDriver

    def open(self, path, profile):
        chrome_options = Options()
        chrome_options.add_argument("--user-data-dir={}".format(path))
        chrome_options.add_argument('--profile-directory='+profile)
        self.driver = webdriver.Chrome(ChromeDriverManager(log_level=0).install(), chrome_options=chrome_options)

    def close(self):
        try:
            self.driver.quit()
        except:
            pass

    def send(self, message, url):
        # self.driver.get(f'https://mbasic.facebook.com/messages/thread/{user}/?entrypoint=profile_message_button')
        self.driver.get(url)
        sleep(1)
        self.driver.find_element_by_link_text('Message').click()
        sleep(2)
        self.driver.find_element_by_css_selector('textarea').send_keys(message)
        sleep(1)
        # self.driver.find_element_by_css_selector('input[value="Send"]').click()
        sleep(3)


def get_message(path):
    messages = []
    book = load_workbook(file_name)
    sheet = book.active
    for i in range(2, sheet.max_row+1):
        message = sheet.cell(i, 1).value
        if message:
            messages.append(message)

    return messages


def user_data():
    list_path = []
    book = load_workbook(file_name)
    sheet = book.active
    for i in range(2, sheet.max_row+1):
        path = sheet.cell(i, 1).value
        pro = sheet.cell(i, 2).value
        list_path.append((path, pro))

    return list_path


def main(users_path, path, messages):
    book = load_workbook(path)
    sheet = book.active
    num = 2
    for path, pro in users_path:
        self.open(path, pro)
        for i in range(max_n):
            url = sheet.cell(num, 3).value
            user_id = sheet.cell(num, 2).value
            if user_id:
                self.send(choice(messages), url)
                print(num, i, user_id)
            num += 1
            sleep(time_sleep)
        self.close()


if __name__ == '__main__':
    time_sleep = 10
    max_n = 10
    self = Send()
    file_name = 'user_data.xlsx'
    path_message = 'messages.xlsx'
    lists = user_data()
    list_messages = get_message(path_message)
    if not len(list_messages):
        list_messages = [input('- Enter Massage: '), ]
    main(lists, 'face -2021-07-06-15-10-16.xlsx', list_messages)

    # self.open(r'C:\Users\3mora\AppData\Local\Google\Chrome\User Data', 'Profile 3')
