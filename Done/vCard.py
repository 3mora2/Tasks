from openpyxl import load_workbook

sheet = load_workbook('Final-All.xlsx').active

with open('Final-All.vcf', 'w') as output:
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(i, 1).value and sheet.cell(i, 2).value:
            print(i, sheet.cell(i, 1).value)
            output.write("BEGIN:VCARD\n")
            output.write("VERSION:3.0\n")
            # output.write("N:" + row[1] + ";" + row[0] + ";;;\n")
            output.write("FN:" + str(sheet.cell(i, 1).value) + "\n")
            output.write("TEL;TYPE=HOME,VOICE:+" + str(sheet.cell(i, 2).value) + "\n")
            output.write("END:VCARD\n")