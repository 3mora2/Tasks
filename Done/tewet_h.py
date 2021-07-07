import warnings

warnings.filterwarnings("ignore")
from msedge.selenium_tools import Edge, EdgeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium import webdriver
from time import sleep
import random
from openpyxl import load_workbook

while True:
    path = 'data.xlsx'  # input('Input name of exal example( data.xlsx ) : ')
    if '.xlsx' in path:
        break
try:
    book = load_workbook(path.strip())
    sheet = book.active
    end_cell = sheet.max_row
    Subject = [sheet.cell(i, 4).value for i in range(2, end_cell + 1)]
    Description = [sheet.cell(i, 5).value for i in range(2, end_cell + 1)]
    Name = [sheet.cell(i, 1).value for i in range(2, end_cell + 1)]
    User_name = [sheet.cell(i, 2).value for i in range(2, end_cell + 1)]
    Email = [sheet.cell(i, 3).value for i in range(2, end_cell + 1)]
except:
    exit()

options = EdgeOptions()
options.use_chromium = True
options.add_argument("disable-gpu")
options.add_argument("--user-data-dir=Edge-data")
#driver = Edge(options=options)

chrome_options = Options()
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--disable-notifications")
driver = webdriver.Chrome(chrome_options=chrome_options)



urls = 'https://twitter.com/explore'
url = 'https://help.twitter.com/forms/restore'
while True:
    n = 0
    for name, user_name, email, subject, description in zip(Name, User_name, Email, Subject, Description):
        print('open link')
        sleep(1)
        driver.get(urls)
        sleep(3)
        
        driver.get(url)
        sleep(2)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'reactivation'))).click()
        sleep(1)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'followers_zero'))).click()
        sleep(1)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'followers_waited_48'))).click()
        sleep(1)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'app'))).send_keys(
            driver.find_element_by_css_selector('#app > option:nth-child(4)').text)
        sleep(1)
        for sub in subject:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'subject'))).send_keys(sub)
            sleep(random.randint(2, 6) / 100)
        sleep(1)
        for des in description:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'description'))).send_keys(des)
            sleep(random.randint(2, 6) / 100)
        sleep(1)
        for nam in name:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'name'))).send_keys(nam)
            sleep(random.randint(2, 6) / 100)
        sleep(1)
        for user in user_name:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'user_name'))).send_keys(user)
            sleep(random.randint(2, 6) / 100)
        sleep(1)
        for mail in email:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'email'))).send_keys(mail)
            sleep(random.randint(2, 6) / 100)
        sleep(2)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'submit_button'))).click()
        sleep(4)
        n += 1
        print('Send :', n)
        # driver = Edge(options=options)
        sleep(10)
    driver.delete_all_cookies()
    print('Close and Reopen')
    driver.quit()
    driver = webdriver.Chrome(chrome_options=chrome_options)
        

driver.quit()
"""
driver.find_element_by_id('reactivation').click()
driver.find_element_by_id('followers_zero').click()
driver.find_element_by_id('followers_waited_48').click()
driver.find_element_by_id('app').send_keys(driver.find_element_by_css_selector('#app > option:nth-child(4)').text)
driver.find_element_by_id('subject').send_keys('subject')
driver.find_element_by_id('description').send_keys('description')
driver.find_element_by_id('name').send_keys('name')
driver.find_element_by_id('user_name').send_keys('user_name')
driver.find_element_by_id('email').send_keys('email@gmail.com')
driver.find_element_by_id('submit_button').submit()
"""
