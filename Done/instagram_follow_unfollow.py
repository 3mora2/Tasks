from datetime import datetime
from time import sleep
from openpyxl import Workbook, load_workbook
import requests
import json

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium import webdriver
import threading


class Thread(threading.Thread):
    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        self.p = None
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class Insta:
    def __init__(self):
        fil = input('- Enter file accounts : ')
        book1 = load_workbook(fil)
        sheet1 = book1.active
        user1 = sheet1['A2'].value
        pass1 = sheet1['B2'].value

        user2 = sheet1['A3'].value
        pass2 = sheet1['B3'].value
        # Username: rainmaker_ata
        # Password: Ahmed2030
        url = input('- Enter Username ( example:- username ) : ')  # chhc
        if url.find('instagram.com/') > -1:
            self.user = url.split('com/')[1].split('/')[0]
        else:
            self.user = url
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet['B1'] = 'user name'
        self.sheet.column_dimensions['B'].width = 21
        self.end_cell = None

        print('- Please Wait While Start..........')

        self.session1 = self.login(username=user1, password=pass1)
        sleep(2)
        self.session2 = self.login(username=user2, password=pass2)
        # self.session1 = self.login(username='201028946519', password='ammar2020')
        sleep(3)
        # self.session2 = self.login(username='ebrahemelmorsy22001@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session3 = self.login(username='ebrahemelmorsy22002@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session4 = self.login(username='pywpywpy', password='pYpY2222')
        # sleep(1)
        # # self.session5 = self.login(username='pywpywpywpy', password='pYpY2222')
        # sleep(1)
        # self.session6 = self.login(username='moduhondu@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session7 = self.login(username='uhiotyui6@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session8 = self.login(username='uiuyhjyy@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session9 = self.login(username='jvbjyhg@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session10 = self.login(username='shakermouty@gmail.com', password='ammar2020')
        # sleep(1)

    def login(self, username, password):
        session = requests.Session()
        headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
                   'Accept-Language': 'en-GB,en;q=0.9,en-US;q=0.8,ml;q=0.7',
                   'User-Agent': 'Mozilla/S.0 (X11; Linux x86_64) AppleWebKit/S37.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36'}
        session.headers.update(headers)
        session.get('https://www.instagram.com/web/__mid/')
        csrf_token = session.cookies.get_dict()['csrftoken']
        session.headers.update({'X-CSRFToken': csrf_token})
        enc_password = '#PWD_INSTAGRAM_BROWSER:0:{}:{}'.format(int(datetime.now().timestamp()), password)
        resp = session.post('https://www.instagram.com/accounts/login/ajax/',
                            data={'enc_password': enc_password, 'username': username},
                            allow_redirects=True)  # ,proxies=proxyDict)
        print(resp.status_code)
        ###################################################################################################
        try:
            resp_json = resp.json()
        except Exception as e:
            print("Login error: JSON decode fail, {} - {}.".format(resp.status_code, resp.reason))
        try:
            if resp_json.get('checkpoint_url'):
                print("Login: Checkpoint required. Point your browser to "
                      "https://www.instagram.com{} - "
                      "follow the instructions, then retry.".format(resp_json.get('checkpoint_url')))
            if resp_json['status'] != 'ok':
                if 'message' in resp_json:
                    print("Login error: \"{}\" status, message \"{}\".".format(resp_json['status'],
                                                                               resp_json['message']))
                else:
                    print("Login error: \"{}\" status.".format(resp_json['status']))
            if 'authenticated' not in resp_json:
                if 'message' in resp_json:
                    print("Login error: Unexpected response, \"{}\".".format(resp_json['message']))
                else:
                    print("Login error: Unexpected response, this might indicate a blocked IP.")
            if not resp_json['authenticated']:
                if resp_json['user']:
                    print('Login error: Wrong password.')
                else:
                    print('Login error: User {} does not exist.'.format(username))
        except:
            pass
        return session

    def sessions(self, b):
        if b == 1 or b == 3 or b == 5 or b == 7 or b == 9:
            session = self.session1
        else:
            session = self.session2
        # elif b == 3:
        #     session = self.session3
        # elif b == 4:
        #     session = self.session4
        # # elif b == 5:
        # #     session = self.session5
        # elif b == 6:
        #     session = self.session6
        # elif b == 7:
        #     session = self.session7
        # elif b == 8:
        #     session = self.session8
        # elif b == 9:
        #     session = self.session9
        # elif b == 10 or b is None:
        #     session = self.session10
        # else:
        #     session = self.session2

        return session

    def get_followers(self):
        try:
            now1 = datetime.now()
            p = 1
            b = 1
            u = True
            end_cursor = None
            #####################################################################
            strHtml = self.session2.get(f"https://www.instagram.com/{self.user}/?__a=1")
            sleep(3)
            user_id = strHtml.json()['graphql']['user']['id']
            resp = self.gf(user_id)
            print('- Number of Followers {0}'.format(resp['data']['user']['edge_followed_by']['count']))
            print('- Getting followers.....')
            sleep(4)
            #####################################################################
            self.thr = Thread()
            self.thr.start()
            users = []
            while u:
                try:
                    if b is None:
                        b = 1
                    elif b < 10:
                        b += 1
                    elif b == 10:
                        b = 1

                    if resp != 'errors':
                        if len(resp['data']['user']['edge_followed_by']['edges']) != 0:
                            u = resp['data']['user']['edge_followed_by']['page_info']['has_next_page']
                            end_cursor = resp['data']['user']['edge_followed_by']['page_info']['end_cursor']
                            for i in resp['data']['user']['edge_followed_by']['edges']:
                                if i['node']['username'] not in users:
                                    user = i['node']['username']
                                    self.sheet.cell(row=p, column=2).value = user
                                    print(f'{p} : {user}')
                                    p += 1
                        else:
                            print('- Please Wait....')
                            sleep(5)

                        sleep(1.5)
                        resp = self.gf(user_id=user_id, end_cursor=end_cursor, b=b)
                    else:
                        print('- Error, Please Wait....')
                        sleep(40)
                        if b is None:
                            b = 1
                        elif b < 10:
                            b += 1
                        elif b == 10:
                            b = 1
                        resp = self.gf(user_id=user_id, end_cursor=end_cursor, b=b)
                except Exception as e:
                    print(e)

                self.book.save(f'{self.user}.xlsx')
                try:
                    if self.thr.p == 'pause' or self.thr.p == 'p':
                        start = input('(Type start or s and press Enter)>>>')
                        while True:
                            if start == 'start' or start == 's':
                                self.thr = Thread()
                                self.thr.start()
                                break
                            elif start == 'exit' or start == 'e':
                                exit()
                            else:
                                start = input()
                    elif self.thr.p == 'exit' or self.thr.p == 'e':
                        exit()
                except:
                    pass
            now = datetime.now()
            end_time = now - now1
            print(end_time)
            print('-- Done --')
            print('- Number of Followers {0}'.format(resp['data']['user']['edge_followed_by']['count']))
        except Exception as e:
            print(e)

    def gf(self, user_id, end_cursor=None, b=None):
        session = self.sessions(b)
        if end_cursor is None:
            strHtml = session.get(
                f"https://www.instagram.com/graphql/query/?query_id=17851374694183129&id={user_id}&first=100")
        else:
            strHtml = session.get(
                f"https://www.instagram.com/graphql/query/?query_id=17851374694183129&id={user_id}&first=100&after={end_cursor}")
        if strHtml.ok:
            resp = json.loads(strHtml.text)
            return resp
        else:
            print(strHtml.status_code)
            return 'errors'


class InstaBot:
    def __init__(self):
        self.username = input('Enter your username : ')  # 'pywpywpywpy'
        self.password = input('Enter password : ')  # 'pYpY2222'
        self.number = int(input('Enter number people : '))
        self.sleep_number = int(input('Enter Time (S) : '))
        self.file = input('Enter File name : ')  # 'chhc.xlsx'

        self.book = load_workbook(self.file)
        self.sheet = self.book.active
        self.driver = webdriver.Chrome()
        self.login()

    def login(self):
        self.driver.get("https://www.instagram.com/accounts/login/?")
        sleep(2)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, '#loginForm > div > div:nth-child(1) > div > label > input'))).send_keys(self.username)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, '#loginForm > div > div:nth-child(2) > div > label > input'))).send_keys(self.password)
        sleep(2)
        self.driver.find_element_by_css_selector('#loginForm > div > div:nth-child(3) > button').click()
        try:
            sleep(4)
            self.driver.find_element_by_class_name('HoLwm').click()
            sleep(2)
        except:
            pass

    def Make_follow(self):
        self.users = [(self.sheet[f'B{i}'].value, i) for i in range(1, self.number + 1) if
                      self.sheet[f'C{i}'].value != 'Done']
        for user, i in self.users:
            print(i, ' - ', user)
            try:
                self.driver.get(f'https://www.instagram.com/{user}/')
                for button in WebDriverWait(self.driver, 20).until(EC.visibility_of_all_elements_located((
                        By.CSS_SELECTOR,
                        '#react-root > section > main > div > header > section > div > div  button'))):
                    if 'Follow' in button.text:
                        button.click()
                        print('Click Follow')
                        break
            except:
                pass
            self.sheet[f'C{i}'].value = 'Done'
            self.book.save(self.file)
            sleep(self.sleep_number)

    def Make_unfollow(self):
        self.users = [(self.sheet[f'B{i}'].value, i) for i in range(1, self.number + 1) if
                      self.sheet[f'A{i}'].value != 'Done']
        for user, i in self.users:
            try:
                print(i, ' - ', user)
                self.driver.get(f'https://www.instagram.com/{user}/')
                sleep(2)
                try:
                    WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((
                        By.CSS_SELECTOR,
                        '#react-root > section > main > div > header > section > div > div:nth-child(3) > div > span > span.vBF20 > button'))).click()
                except:
                    for button in WebDriverWait(my_bot.driver, 20).until(EC.visibility_of_all_elements_located((
                            By.CSS_SELECTOR,
                            '#react-root > section > main > div > header > section > div > div  button'))):
                        print(button.text)
                        if 'Request' in button.text:
                            button.click()
                            break

                sleep(2)
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((
                    By.CSS_SELECTOR,
                    'body > div.RnEpo.Yx5HN > div > div > div > div.mt3GC > button.aOOlW.-Cab_'))).click()
                print('Click Unfollow')
            except:
                pass
            self.sheet[f'A{i}'].value = 'Done'
            self.book.save(self.file)
            sleep(self.sleep_number)


if __name__ == '__main__':
    while True:
        Input = str(input('''
-----------------------------------------------------
- This script to extract information from instagram -
- 1 - To Get Followers >>> Enter 1                  -
- 2 - To Follow  >>> Enter 2                        -
- 3 - To Unfollow  >>> Enter 3                      -
-----------------------------------------------------
>>> '''))
        if Input == '1':
            Insta().get_followers()
            break
        elif Input == '2':
            my_bot = InstaBot()
            my_bot.Make_follow()
            break
        elif Input == '3':
            my_bot = InstaBot()
            my_bot.Make_unfollow()
            break
