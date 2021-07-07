import os
import traceback
from time import sleep
from requests_html import HTMLSession
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import json
from openpyxl import load_workbook


class Main:
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-sa/noon-catalog'

    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'a[href*="/en-sa/noon-catalog/preview/"]')
    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')
    DEACTIVATE_SCRIPT = 'if (document.querySelector(\'input[id="offerActive"]\').checked == true){document.querySelector(\'input[id="offerActive"]\').click()};'
    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="offerActive"]\').checked == false){document.querySelector(\'input[id="offerActive"]\').click()};'

    SAVE_CHANGE_SELECTOR = (By.XPATH, '//*[contains(text(),"Save Changes")]')
    SUBMIT_SELECTOR = (By.XPATH,
                       '//div[calss="btnWrapper"]/div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')

    def __init__(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def check(self):
        c = self.driver.get_cookies()
        data, names = test_request(c)
        old_data = self.edit_data()
        current_storage = read_file()
        self.deactivate(data, old_data, names=names, current_storage=current_storage)

    @staticmethod
    def edit_data():
        files = [int(f.split('-')[0]) for f in os.listdir() if f.endswith('new.json')]
        files.sort()
        try:
            with open(f'{files[-1]}-new.json', ) as fl:
                data_old = json.load(fl)

            return [data_old, files[-1]]

        except Exception as e:
            print(e)
            return [dict(), 0]

    def add_new(self, name):
        try:
            self.driver.get(self.URL_CAT)
            sleep(2)
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))
                                                 ).send_keys(name, Keys.ENTER)
            sleep(3)
            WebDriverWait(self.driver, 10).until(
                ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
            return True
        except:
            print(traceback.print_exc())
            return False

    def add_qu(self, q):
        sleep(1)
        try:
            WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
            return True
        except:
            print(traceback.print_exc())
            print('- Cant Change quantity')
            return False

    def deactivate(self, stock, old_data, names, current_storage):
        nn = old_data[1]
        self.login()
        old_data = old_data[0]
        for code in stock:
            value = None
            if code.replace('-', '') in current_storage and stock[code] == 'OutStock' and float(
                    current_storage[code.replace('-', '')]) > 0:
                value = str(current_storage[code.replace('-', '')])
                action = ''

            elif code in old_data:
                if stock[code] == old_data[code]['now']:
                    old_data[code] = {'names': names[code], 'static': None, 'now': stock[code]}
                    continue
                else:
                    if stock[code] == 'InStock':
                        print(code, stock[code], )
                        old_data[code] = {'names': names[code], 'static': stock[code], 'now': stock[code]}
                        action = self.ACTIVE_SCRIPT
                    elif stock[code] == 'OutStock':
                        old_data[code] = {'names': names[code], 'static': stock[code], 'now': stock[code]}
                        print(code, stock[code], )
                        action = self.DEACTIVATE_SCRIPT
                    else:
                        old_data[code] = {'names': names[code], 'static': stock[code], 'now': None}
                        continue

            else:
                if stock[code] == 'InStock':
                    continue
                    # print(code, stock[code], )
                    # old_data[code] = {'static': stock[code], 'now': stock[code]}
                    # action = self.ACTIVE_SCRIPT
                elif stock[code] == 'OutStock':
                    old_data[code] = {'names': names[code], 'static': stock[code], 'now': stock[code]}
                    print(code, stock[code], )
                    action = self.DEACTIVATE_SCRIPT
                else:
                    old_data[code] = {'names': names[code], 'static': stock[code], 'now': None}
                    continue

            name = names[code]
            res = self.add_new(name)

            if not res:
                print('- Try remove ')
                try:
                    name = ','.join(name.split(',')[:-end_word])
                    res = self.add_new(name)
                except:
                    print(traceback.print_exc())
                    old_data[code] = {'names': names[code], 'static': stock[code], 'now': None}
                    continue

            if not res:
                old_data[code] = {'names': names[code], 'static': stock[code], 'now': None}
                continue
            sleep(5)
            try:
                if value:
                    resell = self.add_qu(value)
                    if not resell:
                        old_data[code] = {'names': names[code], 'static': stock[code], 'now': None}
                else:
                    self.driver.execute_script(action)

                WebDriverWait(self.driver, 10).until(
                    (ec.element_to_be_clickable(self.SAVE_CHANGE_SELECTOR))).click()
                sleep(1)
                WebDriverWait(self.driver, 10).until(
                    (ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                sleep(1)
            except:
                print(traceback.print_exc(0))
                print('- Cant......')
                old_data[code] = {'names': names[code], 'static': stock[code], 'now': None}

        with open(f'{nn + 1}-new.json', 'w') as fl:
            json.dump(old_data, fl)


def test_request(c):
    new_edit = dict()
    name_edit = dict()
    cont = 0
    out_ = 0
    in_ = 0
    s = HTMLSession()
    cookie = {'name': 'iher-pref1', 'value': 'storeid=0&sccode=SA&lan=en-US&scurcode=SAR&wp=2&lchg=1&ifv=1&accsave=0'}
    s.cookies.set(cookie['name'], cookie['value'])
    r = s.get('https://sa.iherb.com/')
    if r.ok:
        urls = ['https://sa.iherb.com' + element.attrs['href'] + '?noi=192' for element in
                r.html.find('.category') if '/c/' in element.attrs['href']]

        for url in urls:
            r_ = s.get(url)
            next_pag = True
            while next_pag:
                for ele in r_.html.find('div.product-inner.product-inner-wide'):
                    code = ele.find('div.absolute-link-wrapper > a', first=True).attrs['data-part-number']
                    title = ele.find('div.absolute-link-wrapper > a', first=True).attrs['title']
                    name_edit[code] = title
                    try:
                        static = ele.find('meta+link')[0].attrs['href']
                    except:
                        static = ''
                    if code in new_edit:
                        pass
                    elif 'InStock' in static:
                        new_edit[code] = 'InStock'
                        in_ += 1
                    elif static == '':
                        new_edit[code] = ''
                    else:
                        out_ += 1
                        new_edit[code] = 'OutStock'

                    cont += 1
                print(f'- {cont} , IN : {in_}, OUT : {out_}', end='\r')
                try:
                    url = 'https://sa.iherb.com' + r_.html.find('.pagination-next')[0].attrs['href']
                    r_ = s.get(url)
                    if not r_.ok:
                        sleep(30)
                        r_ = s.get(url)
                except:
                    break

        for elem in new_edit.keys():
            if new_edit[elem] == '':
                r = s.get('https://sa.iherb.com/search?kw=' + elem)
                static = r.html.find('#stock-status')[0].text
                if 'متوفر حاليا' in static or 'In Stock' in static:
                    new_edit[elem] = 'InStock'
                else:
                    new_edit[elem] = 'OutStock'
    else:
        print(r.status_code)

    return new_edit, name_edit


def read_file():
    data = dict()
    sheet = load_workbook(file_name).active
    for i in range(2, sheet.max_row + 1):
        code = sheet.cell(i, 1).value
        value = sheet.cell(i, 2).value
        if code and value:
            data[code] = value

    return data


if __name__ == "__main__":
    while True:
        try:
            file_name = input('- Enter file name : ')
            if os.path.isfile(file_name) and file_name.endswith('.xlsx'):
                break
            print('- No such file or directory')
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)

    email = input('- Enter your email : ')
    password = input('- Enter your password : ')
    end_word = input('- Enter Number Label : ')
    try:
        end_word = int(end_word)
        if end_word > 0:
            pass
        else:
            end_word = 1
    except:
        end_word = 1

    app = Main()
    app.check()
# pyinstaller --add-data C:/Users/3mora/AppData/Local/Programs/Python/Python39/Lib/site-packages/pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info
