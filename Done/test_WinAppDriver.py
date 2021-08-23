# import win32com.shell.shell as shell
# commands = '''
# reg add "HKEY_LOCAL_MACHINE\SOFTWARE\Microsoft\Windows\CurrentVersion\AppModelUnlock" /t REG_DWORD /f /v "AllowDevelopmentWithoutDevLicense" /d "1"
# '''
# commands2 = '''
# reg add "HKEY_LOCAL_MACHINE\SOFTWARE\Microsoft\Windows\CurrentVersion\AppModelUnlock" /t REG_DWORD /f /v "AllowAllTrustedApps" /d "1"
# '''
# shell.ShellExecuteEx(lpVerb='runas', lpFile='C:\WINDOWS\system32\cmd.exe', lpParameters='/c '+commands)

# from appium import webdriver
#
# desired_caps = {}
# desired_caps["app"] = "Microsoft.WindowsCalculator_8wekyb3d8bbwe!App"
# driver = webdriver.Remote(
#     command_executor='http://127.0.0.1:4723',
#     desired_capabilities=desired_caps)
from openpyxl import load_workbook

book = load_workbook('all - Copy.xlsx')
sheet = book.active

for i in range(2, sheet.max_row + 1):
    address = sheet.cell(i, 7).value
    if address:
        address = ' '.join([t for t in address.split(' ') if '+' not in t and not t.replace('،', '').isnumeric()])
        sheet.cell(i, 7).value = address.split('،')[-1].split(',')[-1]

book.save('all - Copy.xlsx')
