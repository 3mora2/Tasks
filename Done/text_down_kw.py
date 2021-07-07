from selenium import webdriver
from time import sleep
from openpyxl import Workbook
book = Workbook()
sheet = book.active

driver = webdriver.Chrome()
driver.get('http://search.kna.kw/web/Retrieval/GeneralSearch.aspx?tid=1')
#iViewer
# http://search.kna.kw/web/Retrieval/

row = 0
for el in driver.find_elements_by_css_selector('tbody > tr'):
    colume = 1
    url = True
    if 'grdPager' not in el.get_attribute('class'):
        if 'grdHeader' in el.get_attribute('class'):
            url = False
        for td in el.find_elements_by_css_selector('td'):
            try:
                t = td.find_element_by_css_selector('a > font').get_attribute('onclick')
                ur = ('http://search.kna.kw/web/Retrieval/' + t[t.find('DocumentsView'):t.find("','")])
                driver.execute_script("window.open('')")
                driver.switch_to_window(driver.window_handles[1])
                driver.get(ur)
                sleep(4)
                name = driver.find_element_by_css_selector('iframe').get_attribute('src').split('download\\')[-1]
                print(name, end=' ')
                driver.switch_to_frame(driver.find_element_by_css_selector('iframe'))
                sleep(1)
                driver.execute_script('document.getElementById("download").click()')
                # driver.find_element_by_css_selector('#download').click()
                driver.close()
                driver.switch_to_window(driver.window_handles[0])
                sheet.cell(row, colume).value = 'download_' + name
            except:
                print(td.text, end='')
                sheet.cell(row, colume).value = td.text

            colume += 1
        print('\n')
        print(row)
        row += 1

book.save('12.xlsx')