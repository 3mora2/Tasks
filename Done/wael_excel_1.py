from openpyxl import load_workbook

bar_price = dict()

book = load_workbook(r'C:\Users\3mora\Downloads\3.xlsx')
sheet = book.active
for i in range(2, sheet.max_row+1):
    barcode = sheet[f'B{i}'].value
    if barcode is None or barcode == '':
        continue
    else:
        bar_price[barcode] = (sheet[f'D{i}'].value, sheet[f'E{i}'].value)

other_book = load_workbook(r'C:\Users\3mora\Downloads\5.xlsx')
other_sheet = other_book.active

for i in range(2, other_sheet.max_row+1):
    barcode = other_sheet[f'A{i}'].value
    if barcode is None or barcode == '':
        continue
    else:
        if barcode in bar_price.keys():
            other_sheet[f'B{i}'].value = bar_price[barcode][0]
            other_sheet[f'C{i}'].value = bar_price[barcode][1]
        else:
            print(barcode, 'not found')

other_book.save(r'C:\Users\3mora\Downloads\5.xlsx')
