import re
import warnings

warnings.filterwarnings("ignore")

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium import webdriver

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

from requests_html import HTMLSession
from io import BytesIO
from time import sleep
import subprocess

session = HTMLSession()


class Main:
    def __init__(self):
        while True:
            self.sear = input('Search :')
            if self.sear != '':
                break
        self.search = self.sear.replace(' ', '+')
        self.book = Workbook()
        self.sheet = self.book.active
        try:
            chrome_options = Options()
            chrome_options.add_argument("--disable-infobars")
            chrome_options.add_argument("--disable-notifications")
            chrome_options.add_argument("--user-data-dir={}".format(
                r'C:\Users\3mora\AppData\Local\Google\Chrome\User Data'))
            self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)

        except:
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
        ##############################################################################################################
        self.driver.maximize_window()

    def open(self):
        # self.sheet['A1'] = 'Image'
        # self.sheet['B1'] = 'Name'
        # self.sheet['C1'] = 'Type'
        # self.sheet['D1'] = 'Rate'
        # self.sheet['E1'] = 'Address'
        # self.sheet['F1'] = 'Plus Code'
        # self.sheet['G1'] = 'Phone'
        # self.sheet['H1'] = 'Website'
        # self.sheet['I1'] = "Note"
        # self.sheet['J1'] = "URL"
        # self.sheet.column_dimensions['A'].width = 13
        # self.sheet.column_dimensions['B'].width = 50
        # self.sheet.column_dimensions['C'].width = 30
        # self.sheet.column_dimensions['D'].width = 10
        # self.sheet.column_dimensions['E'].width = 100
        # self.sheet.column_dimensions['F'].width = 70
        # self.sheet.column_dimensions['G'].width = 25
        # self.sheet.column_dimensions['H'].width = 30
        # self.sheet.column_dimensions['I'].width = 50
        # self.sheet.column_dimensions['J'].width = 120
        # for column in range(1, 11):
        #     try:
        #         self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        #     except:
        #         pass

        self.driver.get(f'https://www.google.com/maps/search/{self.search}')
        print('Program is loading,please wait!')
        sleep(5)
        input('- Enter...')
        urls = []
        while True:
            while True:
                try:
                    self.driver.find_element_by_css_selector('.wo1ice-loading').location_once_scrolled_into_view
                    sleep(6)
                except:
                    break
            sleep(2)
            num_pl = len(self.driver.find_elements_by_css_selector('div.section-layout.section-scrollbox div>a'))
            if num_pl == 0:
                break

            for element in self.driver.find_elements_by_css_selector('div.section-layout.section-scrollbox div>a'):
                if element.get_attribute('href') not in urls:
                    urls.append(element.get_attribute('href'))
                print(len(urls))
            # self.driver.find_element_by_css_selector("[aria-label=\"الصفحة التالية\"]").click()
            self.driver.execute_script('document.querySelector(\'*[aria-label="الصفحة التالية"]\').click()')
            sleep(5)

        num = 2
        for url in urls:
            try:
                self.driver.get(url)
                sleep(4)
                name = WebDriverWait(self.driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, '//h1/span[1]'))).text
                try:
                    ty = self.driver.find_element_by_css_selector('[jsaction="pane.rating.category"]').text
                except:
                    ty = None
                try:
                    phone = \
                        self.driver.find_element_by_xpath(
                            "//div/button[contains(@data-item-id, 'phone:')]").get_attribute(
                            'data-item-id').split('tel:')[-1]
                except:
                    phone = None
                try:
                    address = \
                        self.driver.find_element_by_xpath("//div/button[contains(@data-item-id, 'address')]").get_attribute(
                            'aria-label').split(":")[-1]
                except:
                    address = None
                try:
                    web = \
                        self.driver.find_element_by_xpath(
                            "//div/button[contains(@data-item-id, 'authority')]").get_attribute(
                            'aria-label').split(":")[-1]
                except:
                    web = None
                try:
                    cont = self.driver.find_element_by_css_selector('[data-item-id="oloc"]').get_attribute('aria-label').split(':')[-1]
                except:
                    cont = None

                self.sheet[f'A{num}'] = name
                self.sheet[f'B{num}'] = ty
                self.sheet[f'C{num}'] = address.split('،')[-1]
                self.sheet[f'D{num}'] = phone
                self.sheet[f'E{num}'] = cont
                self.sheet[f'G{num}'] = web
                self.sheet[f'H{num}'] = address
                print(num, name, address, phone, web)
            except Exception as e:
                print(e)
            self.sheet[f'K{num}'] = url
            num += 1
        self.book.save('final_all.xlsx')
        #         address = None
        #         plus_code = None
        #         phone_number = None
        #         website = None
        #         sleep(1)
        #         try:
        #             w = self.driver.find_elements_by_class_name('section-result-content')[
        #                 i].find_elements_by_css_selector(
        #                 'div:nth-child(1) > a.section-result-action.section-result-action-wide')
        #             try:
        #                 if len(w) != 0:
        #                     website = w[0].get_attribute('href')
        #             except:
        #                 pass
        #             self.driver.find_elements_by_class_name('section-result-content')[i].click()
        #             sleep(3)
        #             try:
        #                 main_photo = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((
        #                     By.CSS_SELECTOR,
        #                     'div.section-hero-header-image-hero-container.collapsible-hero-image > button > img'))).get_attribute(
        #                     'src')
        #             except:
        #                 main_photo = None
        #
        #             try:
        #                 name = WebDriverWait(self.driver, 4).until(
        #                     EC.visibility_of_element_located((By.XPATH, '//h1/span[1]'))).text
        #
        #             except:
        #                 name = None
        #
        #             try:
        #                 typ = WebDriverWait(self.driver, 4).until(EC.visibility_of_element_located(
        #                     (By.CSS_SELECTOR, 'span.section-rating-term > span:nth-child(1) > button'))).text
        #             except:
        #                 typ = None
        #
        #             try:
        #                 rat = WebDriverWait(self.driver, 4).until(
        #                     EC.visibility_of_element_located((By.CLASS_NAME, 'section-star-display'))).text
        #             except:
        #                 try:
        #                     rat = self.driver.find_element_by_css_selector(
        #                         '#pane > div > div.widget-pane-content.scrollable-y > div > div > div.jqnFjrOWMVU__root > div > div.jqnFjrOWMVU__right > div.gm2-display-2')
        #                 except:
        #                     rat = None
        #
        #             try:
        #                 note = WebDriverWait(self.driver, 4).until(EC.visibility_of_element_located(
        #                     (By.CSS_SELECTOR, 'div.section-editorial-attribute-container'))).text
        #             except:
        #                 note = None
        #
        #             try:
        #                 _ = self.driver.find_elements_by_class_name('ugiz4pqJLAG__content')[
        #                     -1].location_once_scrolled_into_view
        #             except:
        #                 pass
        #
        #             try:
        #                 for element in self.driver.find_elements_by_xpath(
        #                         '//*[@id="pane"]/div/div[1]/div/div/div/button'):
        #                     if element.get_attribute('data-tooltip') is not None:
        #                         if 'العنوان' in element.get_attribute(
        #                                 'data-tooltip') or 'address' in element.get_attribute('data-tooltip'):
        #                             address = element.text
        #                             break
        #             except:
        #                 pass
        #
        #             try:
        #                 for element in self.driver.find_elements_by_xpath(
        #                         '//*[@id="pane"]/div/div[1]/div/div/div/button'):
        #                     if element.get_attribute('data-tooltip') is not None:
        #                         if 'الموقع المفتوح' in element.get_attribute(
        #                                 'data-tooltip') or 'plus code' in element.get_attribute('data-tooltip'):
        #                             plus_code = element.text
        #                             break
        #             except:
        #                 pass
        #
        #             sleep(1)
        #             try:
        #                 for element in self.driver.find_elements_by_xpath(
        #                         '//*[@id="pane"]/div/div[1]/div/div/div/button'):
        #                     if element.get_attribute('data-tooltip') is not None:
        #                         if 'هاتف' in element.get_attribute(
        #                                 'data-tooltip') or 'phone number' in element.get_attribute('data-tooltip'):
        #                             phone_number = element.text
        #                             break
        #             except:
        #                 pass
        #
        #             try:
        #                 if website is None:
        #                     for element in self.driver.find_elements_by_xpath(
        #                             '//*[@id="pane"]/div/div[1]/div/div/div/button'):
        #                         if element.get_attribute('data-tooltip') is not None:
        #                             if 'الموقع الإلكتروني' in element.get_attribute(
        #                                     'data-tooltip') or 'website' in element.get_attribute('data-tooltip'):
        #                                 website = element.text
        #                                 break
        #             except:
        #                 pass
        #
        #             try:
        #                 if main_photo is not None:
        #                     res = session.get(main_photo)
        #                     image_file = BytesIO(res.content)
        #                     img = Image(image_file)
        #                     img.width = 90
        #                     img.height = 75
        #                     self.sheet.row_dimensions[num].height = 56
        #                     self.sheet.add_image(img, f'A{num}')
        #             except:
        #                 pass
        #             self.sheet[f'B{num}'] = name
        #             self.sheet[f'C{num}'] = typ
        #             self.sheet[f'D{num}'] = rat
        #             self.sheet[f'E{num}'] = address
        #             self.sheet[f'F{num}'] = plus_code
        #             self.sheet[f'G{num}'] = phone_number
        #             self.sheet[f'H{num}'] = website
        #             self.sheet[f'I{num}'] = note
        #             try:
        #                 self.sheet[f'J{num}'] = self.driver.current_url
        #             except:
        #                 pass
        #             try:
        #                 self.book.save(f'{self.sear}.xlsx')
        #             except:
        #                 pass
        #             for column in range(2, 11):
        #                 try:
        #                     self.sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center',
        #                                                                        wrap_text=True)
        #                     self.sheet.cell(num + 1, column).alignment = Alignment(horizontal='center',
        #                                                                            vertical='center', wrap_text=True)
        #                 except:
        #                     pass
        #             print(f'{num - 1} - {name}')
        #             num += 1
        #             try:
        #                 WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
        #                     (By.CLASS_NAME, 'section-back-to-list-button'))).click()
        #             except:
        #                 try:
        #                     WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
        #                         (By.CLASS_NAME, 'section-back-to-list-button'))).click()
        #                 except:
        #                     pass
        #             sleep(2)
        #         except:
        #             pass
        #     try:
        #         self.driver.find_elements_by_class_name('n7lv7yjyC35__button')[1].click()
        #     except:
        #         try:
        #             self.driver.find_elements_by_class_name('n7lv7yjyC35__button')[1].click()
        #         except:
        #             break
        #     sleep(1)
        # self.driver.quit()


if __name__ == '__main__':
    app = Main()
    # app.open()
# num = 2
# urls = []
self = app
# i = 2
# import re
# self = app
# self.book = load_workbook('n.xlsx')
# self.sheet = self.book.active
# for i in range(2, self.sheet.max_row + 1):
#     url = self.sheet.cell(i, 11).value
#     if url:
#         url = url.strip()
#         try:
#             self.driver.get(url.strip())
#             sleep(3)
#             if len(self.driver.find_elements_by_xpath("//div/button[contains(@data-item-id, 'authority')]")):
#                 self.driver.execute_script(
#                     'document.querySelector("[data-item-id=\'authority\'] [dir=\'ltr\']").click();')
#                 sleep(3)
#                 if len(self.driver.window_handles) > 1:
#                     self.driver.close()
#                     self.driver.switch_to.window(self.driver.window_handles[-1])
#                     print(set(re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', self.driver.page_source)), self.driver.current_url)
#                     self.sheet.cell(i, 6).value = '\n'.join(set(re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', self.driver.page_source)))
#                     self.sheet.cell(i, 7).value = self.driver.current_url
#                     self.book.save('n1.xlsx')
#                 else:
#                     pass
#         except Exception as e:
#             print(e)

# for i in range(2, self.sheet.max_row + 1):
#     url = self.sheet.cell(i, 11).value
#     if url:
#         url = url.strip()
#         try:
#             self.driver.get(url)
#             sleep(4)
#             name = WebDriverWait(self.driver, 10).until(
#                 EC.visibility_of_element_located((By.XPATH, '//h1/span[1]'))).text
#             try:
#                 ty = self.driver.find_element_by_css_selector('[jsaction="pane.rating.category"]').text
#             except:
#                 ty = None
#             try:
#                 phone = \
#                     self.driver.find_element_by_xpath(
#                         "//div/button[contains(@data-item-id, 'phone:')]").get_attribute(
#                         'data-item-id').split('tel:')[-1]
#             except:
#                 phone = None
#             try:
#                 address = \
#                     self.driver.find_element_by_xpath("//div/button[contains(@data-item-id, 'address')]").get_attribute(
#                         'aria-label').split(":")[-1]
#             except:
#                 address = None
#             try:
#                 web = \
#                     self.driver.find_element_by_xpath(
#                         "//div/button[contains(@data-item-id, 'authority')]").get_attribute(
#                         'aria-label').split(":")[-1]
#             except:
#                 web = None
#             try:
#                 cont = \
#                 self.driver.find_element_by_css_selector('[data-item-id="oloc"]').get_attribute('aria-label').split(':')[-1]
#             except:
#                 cont = None
#             try:
#                 if len(self.driver.find_elements_by_xpath("//div/button[contains(@data-item-id, 'authority')]")):
#                     self.driver.execute_script(
#                         'document.querySelector("[data-item-id=\'authority\'] [dir=\'ltr\']").click();')
#                     sleep(3)
#                     if len(self.driver.window_handles) > 1:
#                         self.driver.close()
#                         self.driver.switch_to.window(self.driver.window_handles[-1])
#                         print(set(re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', self.driver.page_source)), self.driver.current_url)
#                         self.sheet[f'F{num}'] = '\n'.join(set(re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', self.driver.page_source)))
#
#                         self.sheet[f'I{num}'] = self.driver.current_url
#                     else:
#                         pass
#             except:
#                 pass
#             self.sheet[f'A{num}'] = name
#             self.sheet[f'B{num}'] = ty
#             self.sheet[f'C{num}'] = address.split('،')[-1]
#             self.sheet[f'D{num}'] = phone
#             self.sheet[f'E{num}'] = cont
#
#             self.sheet[f'G{num}'] = web
#             self.sheet[f'H{num}'] = address
#             print(num, name, ty, cont, address, phone, web)
#         except Exception as e:
#             print(e)
#         self.sheet[f'K{num}'] = url
#         num += 1
# self.book.save('final_all.xlsx')

self.book = load_workbook('final_all_.xlsx')
self.sheet = self.book.active
for i in range(2, self.sheet.max_row + 1):
    url = self.sheet.cell(i, 11).value
    if url and not self.sheet.cell(i, 1).value:
        url = url.strip()
        self.driver.get(url)
        try:
            sleep(4)
            name = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//h1/span[1]'))).text
            try:
                ty = self.driver.find_element_by_css_selector('[jsaction="pane.rating.category"]').text
            except:
                ty = None
            try:
                phone = \
                    self.driver.find_element_by_xpath(
                        "//div/button[contains(@data-item-id, 'phone:')]").get_attribute(
                        'data-item-id').split('tel:')[-1]
            except:
                phone = None
            try:
                address = \
                    self.driver.find_element_by_xpath("//div/button[contains(@data-item-id, 'address')]").get_attribute(
                        'aria-label').split(":")[-1]
            except:
                address = None
            try:
                web = \
                    self.driver.find_element_by_xpath(
                        "//div/button[contains(@data-item-id, 'authority')]").get_attribute(
                        'aria-label').split(":")[-1]
            except:
                web = None
            try:
                cont = \
                    self.driver.find_element_by_css_selector('[data-item-id="oloc"]').get_attribute('aria-label').split(
                        ':')[-1]
            except:
                cont = None
            try:
                if len(self.driver.find_elements_by_xpath("//div/button[contains(@data-item-id, 'authority')]")):
                    self.driver.execute_script(
                        'document.querySelector("[data-item-id=\'authority\'] [dir=\'ltr\']").click();')
                    sleep(3)
                    if len(self.driver.window_handles) > 1:
                        self.driver.close()
                        self.driver.switch_to.window(self.driver.window_handles[-1])
                        print(set(re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', self.driver.page_source)),
                              self.driver.current_url)
                        self.sheet[f'F{i}'] = '\n'.join(
                            set(re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', self.driver.page_source)))

                        self.sheet[f'I{i}'] = self.driver.current_url
                    else:
                        pass
            except:
                pass
            self.sheet[f'A{i}'] = name
            self.sheet[f'B{i}'] = ty
            self.sheet[f'C{i}'] = address.split('،')[-1]
            self.sheet[f'D{i}'] = phone
            self.sheet[f'E{i}'] = cont

            self.sheet[f'G{i}'] = web
            self.sheet[f'H{i}'] = address
            print(i, name, ty, cont, address, phone, web)
            self.book.save('final_all_.xlsx')
        except Exception as e:
            print(e)