import os
import traceback
import uuid
from random import randrange
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import subprocess


class Main:
    MAIL_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#username_input')
    PASS_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#Password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#sign_in_button')

    def __init__(self):
        self.book = Workbook()
        self.sheet = self.book.active
        chrome_options = Options()
        try:
            subprocess.Popen(
                '"C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222')
        except:
            subprocess.Popen(
                '"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222')
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        self.driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), chrome_options=chrome_options)

        self.login()

    def language(self):
        if 'EN' not in self.driver.find_element_by_css_selector('div.language-select').text:
            self.driver.find_element_by_css_selector('div.language-select').click()
            sleep(2)
            self.driver.find_element_by_css_selector('.select-language').click()
            sleep(1)
            self.driver.find_element_by_css_selector('[data-val="en-US"]').click()
            sleep(4)
            self.driver.find_element_by_css_selector('button.save-selection').click()
            sleep(2)

    def name(self, i):
        self.sheet.cell(i, 1).value = 'Order Date'
        self.sheet.cell(i, 2).value = 'Order Number'
        self.sheet.cell(i, 3).value = 'Product Number'
        self.sheet.cell(i, 4).value = 'Product Description'
        self.sheet.cell(i, 5).value = 'Item Price'
        self.sheet.cell(i, 6).value = 'Qty.'
        self.sheet.cell(i, 7).value = 'Item Discount'
        self.sheet.cell(i, 8).value = 'Rewards'
        self.sheet.cell(i, 9).value = 'Subtotal'
        self.sheet.cell(i, 10).value = 'Shipping'
        self.sheet.cell(i, 11).value = 'VAT / Duties'
        self.sheet.cell(i, 12).value = 'Order Discount'
        self.sheet.cell(i, 13).value = 'Current Order Total'

    def login(self):
        # try:
        self.driver.get('https://secure.iherb.com/myaccount/orders')
        # self.driver.get(self.driver.find_element_by_css_selector('a.product-link').get_attribute('href'))
        # self.driver.find_element_by_css_selector('div.item > div:nth-child(1) > div > div a').click()
        # self.language()
        # ActionChains(self.driver).move_to_element(self.driver.find_element_by_css_selector('button.tablet-icon-login-link')).perform()
        # self.driver.find_element_by_css_selector('[data-ga-event-label="myaccount-orders-click"]').click()
        # WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
        # WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
        # WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
        # while True:
        #     if 'Login' not in self.driver.current_url:
        #         print('break')
        #         break
        #
        # except:
        #     print(traceback.print_exc())
        #     print('- Login Error')

    def get_urls(self):
        next_ok = True
        urls = list()
        while next_ok:
            sleep(1)
            for elements in self.driver.find_elements_by_css_selector('#OrderHistoryResultsND > div > article'):
                static = elements.find_element_by_css_selector('.order-status-label').text
                if 'Returned' in static or 'تمت اعادته' in static or 'Cancelled' in static or 'تم إلغاءه' in static:
                    print('- Not Add Returned')
                    continue
                number = elements.get_attribute('data-order-number')
                urls.append(number)
                print(len(urls))
                if id_stop.strip() in number:
                    next_ok = False
                    break
            if not next_ok:
                break

            try:
                self.driver.find_element_by_css_selector('.pagination-next').click()
                sleep(1)
            except:
                break

            sleep(1)
            while True:
                if 'loading' not in self.driver.find_element_by_css_selector('.loader').get_attribute('class'):
                    break
                else:
                    print('loading')
                sleep(5)

        return urls[::-1]

    def get_url(self, url):
        try:
            self.driver.get('https://secure.iherb.com/myaccount/ordersummary/?on=' + str(url))
            sleep(3)
            _ = self.driver.find_element_by_css_selector('.orderLocaleDate').text
            return True
        except:
            return False

    def get_product(self):
        try:
            self.language()
        except:
            print('- Cant change language')
        urls = self.get_urls()
        n = 1
        for url in urls:
            self.name(n)
            print(url)
            n += 1
            while True:
                static = self.get_url(url)
                if static:
                    break
                else:
                    sleep(120)

            order_date = self.driver.find_element_by_css_selector('.orderLocaleDate').text
            try:
                rewards = self.driver.find_element_by_xpath('//div[p[contains(text(),"Rewards")]]/p[2]').text
            except:
                rewards = None
            try:
                subtotal = self.driver.find_element_by_xpath('//div[p[text()="Subtotal"]]/p[2]').text
            except:
                subtotal = None
            try:
                shipping = self.driver.find_element_by_xpath('//div[p[text()="Shipping"]]/p[2]').text
            except:
                shipping = None
            try:
                duties = self.driver.find_element_by_xpath('//div[p[contains(text(),"Duties & Taxes")]]/span[2]').text
            except:
                duties = None

            try:
                order_discount = self.driver.find_element_by_css_selector(
                    '.total-savings-title').text.replace('(', '').replace(')', '')
            except:
                order_discount = None
            try:
                total = self.driver.find_element_by_css_selector('.total-row.order-total-row p.val').text
            except:
                total = None

            self.sheet.cell(n, 1).value = order_date
            self.sheet.cell(n, 2).value = url
            self.sheet.cell(n, 8).value = rewards
            self.sheet.cell(n, 9).value = subtotal
            self.sheet.cell(n, 10).value = shipping
            self.sheet.cell(n, 11).value = duties
            self.sheet.cell(n, 12).value = order_discount
            self.sheet.cell(n, 13).value = total
            n += 1
            for ele in self.driver.find_elements_by_css_selector('.order-products-container .product-row-rd'):
                # ele.find_element_by_css_selector('share-button').get_attribute('url').split('=')[1].split('qty')[0]
                # product_price = ele.find_element_by_css_selector('.row > .status-column > p > strong').text
                # qty = ele.find_element_by_css_selector('.row > .status-column ~ div > p > strong').text
                product_number = ele.find_element_by_css_selector('span[name="AddToCart"]').get_attribute('data-part-number')
                product_dsk = ele.find_element_by_css_selector('.prod-desc').text
                product_price = ele.find_element_by_css_selector('div.col-buffer > span ~ span').text
                qty = ele.find_element_by_css_selector('div.col-buffer-sm > span ~ span').text
                try:
                    discount = ele.find_element_by_css_selector('span.discount-line-price').text.replace('-', '')
                except:
                    discount = None

                self.sheet.cell(n, 3).value = product_number
                self.sheet.cell(n, 4).value = product_dsk
                self.sheet.cell(n, 5).value = product_price
                self.sheet.cell(n, 6).value = qty
                self.sheet.cell(n, 7).value = discount
                n += 1
            n += 1
            self.book.save('new.xlsx')


# email = 'bh4708251@gmail.com'
# password = 'Assad098765@'
id_stop = input('- Enter Number')  # '953067002' 953391085
app = Main()
input('- Enter..')
app.get_product()
'''
https://github.com/skotz/cbl-js
'''
