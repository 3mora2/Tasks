import os
from requests_html import HTMLSession
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook

os.environ['WDM_LOG_LEVEL'] = '0'

driver = webdriver.Chrome(ChromeDriverManager().install())

# driver.get('https://madaresegypt.com/ar/Results/%D8%AD%D8%B6%D8%A7%D9%86%D8%A7%D8%AA/%D9%85%D8%AF%D9%8A%D9%86%D8%A9-%D8%A7%D9%84%D9%82%D8%A7%D9%87%D8%B1%D8%A9/%D8%AE%D8%A7%D8%B5%D8%A9?page=1')
# driver.find_elements_by_css_selector('div.block-heading > h4 > a')
book = load_workbook('hadana_link.xlsx')
sheet = book.active
num = 1
sheet[f'A{num}'].value = 'الاسم'
sheet[f'B{num}'].value = "رقم الهاتف"
sheet[f'C{num}'].value = "البريد الالكتروني"
sheet[f'D{num}'].value = "رابط الموقع"
sheet[f'E{num}'].value = "رابط الخريطه"
sheet[f'F{num}'].value = "رابط الفيسبوك"
sheet[f'G{num}'].value = "العنوان"
sheet[f'H{num}'].value = "معلومات عن "
sheet[f'I{num}'].value = "خصائص"
sheet[f'J{num}'].value = "الصور"
for num in range(2, sheet.max_row + 1):
    try:
        if num == 200:
            break
        url = sheet[f'Z{num}'].value
        if url is None:
            continue
        print(num)
        driver.get(url)
        name = driver.find_element_by_css_selector('.page-title > a:nth-child(1)').text
        add = driver.find_element_by_css_selector('.col_right > div.contact_info div:nth-child(1)').text.split('\n')
        add.pop(0)
        address = ' '.join(add)
        try:
            number = driver.find_element_by_css_selector('.col_right > div.contact_info > div > div:nth-child(2) ').text
        except:
            number = None
        try:
            mail = driver.find_element_by_css_selector('#middle > div > div > div > div > div:nth-child(5) > span').text
        except:
            mail = None
        try:
            face = driver.find_element_by_css_selector(' div.row.facebook > a').get_attribute('href')
        except:
            face = None
        try:
            web = driver.find_element_by_css_selector(' div.row.website > a').get_attribute('href')
        except:
            web = None
        try:
            Map = driver.find_element_by_css_selector('div.single-property h3:nth-child(2) > a').get_attribute('href')
        except:
            Map = None
        try:
            about = driver.find_element_by_css_selector(
                'div.single-property > div.Editor-body > div.blog-post-body > div.editor-item').text
        except:
            about = None
        try:
            Property = driver.find_element_by_css_selector('div.single-property > div.agent > div > div.col-md-8').text
        except:
            Property = None
        try:
            img = '\n '.join([im.get_attribute('src').replace('_thumb', '') for im in
                              driver.find_elements_by_css_selector('#grid > div > img')])
        except:
            img = None
        sheet[f'A{num}'].value = name
        sheet[f'B{num}'].value = number
        sheet[f'C{num}'].value = mail
        sheet[f'D{num}'].value = web
        sheet[f'E{num}'].value = Map
        sheet[f'F{num}'].value = face
        sheet[f'G{num}'].value = address
        sheet[f'H{num}'].value = about
        sheet[f'I{num}'].value = Property
        sheet[f'J{num}'].value = img
    except:
        pass
book.save('1030.xlsx')

# da['Z1'] = 'URL'
# address = driver.find_element_by_css_selector('#middle > div > div > div > div > div:nth-child(1)').text
# phone = driver.find_element_by_css_selector('#middle > div > div > div > div > div:nth-child(4) > div:nth-child(2) > a').text
# phone_tow = driver.find_element_by_css_selector('#middle > div > div > div > div > div:nth-child(4) > div:nth-child(2)').text

# i = 1
# n = 2
# while i != 76:
#     driver.get(
#         f'https://madaresegypt.com/ar/Results/%D8%AD%D8%B6%D8%A7%D9%86%D8%A7%D8%AA/%D9%85%D8%AF%D9%8A%D9%86%D8%A9-%D8%A7%D9%84%D9%82%D8%A7%D9%87%D8%B1%D8%A9/%D8%AE%D8%A7%D8%B5%D8%A9?page={i}')
#     for el in driver.find_elements_by_css_selector('div.block-heading > h4 > a'):
#         da[f'Z{n}'] = el.get_attribute("href")
#         n += 1
#     print(n)
#     i += 1
# data.save("hadana_link.xlsx")
