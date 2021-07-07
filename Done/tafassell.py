from selenium import webdriver
from openpyxl import load_workbook
from time import sleep
import os
from requests_html import HTMLSession,HTML
s = HTMLSession()
path = 'tassel_result.xlsx'
book = load_workbook(path)
sheet = book.active
i = 1
sheet[f'A{i}'].value = 'الاسم'
sheet[f'B{i}'].value = "السعر"
sheet[f'C{i}'].value = "السعر1"
sheet[f'D{i}'].value = "السعر2"
sheet[f'E{i}'].value = "القسم الرئسي"
sheet[f'F{i}'].value = "القسم الفرعي"
sheet[f'G{i}'].value = "الحالة"
sheet[f'H{i}'].value = "البائع"
sheet[f'I{i}'].value = "الوصف"
sheet[f'J{i}'].value = "المعلومات"
sheet[f'K{i}'].value = "الخيارات"
sheet[f'L{i}'].value = "الصور"

end_cell = sheet.max_row
i = 782 #782
links = [sheet[f'N{o}'].value for o in range(i, 783 + 1) if
         sheet[f'N{o}'].value is not None]
print(len(links))
driver = webdriver.Firefox()
for ur in links:
    driver.get(ur)
    try:
        r = HTML(html=driver.page_source)
        cat1 = r.find('body > div.site-wrapper > ul > li:nth-child(2)', first=True).text
        try:
            cat2 = r.find('body > div.site-wrapper > ul > li:nth-child(3)', first=True).text
        except:
            cat2 = None
        img = ' \n '.join([ele.attrs['src'] for ele in r.find('div.swiper.main-image div.swiper-container > div > div > img')])
        name = r.find('body > div.site-wrapper > h1 > span', first=True).text
        price = r.find('#product > div.product-price-group > div.price-wrapper > div > div', first=True).text
        price_after = r.find('#product > div.product-price-group > div.price-wrapper > li', first=True).text.split(':')[1]
        try:
            price_befor = r.find('#product > div.product-price-group > div.price-wrapper > div > div.product-price-old', first=True).text
        except:
            price_befor = None
        static = r.find('#product > div.product-price-group > div.product-stats > ul > li.product-stock > span', first=True).text
        des = r.find('#product > div.product-price-group > div.product-stats > ul ', first=True).text
        try:
            seller = r.find('#product > div.product-price-group > div.product-stats > div > a > span', first=True).text
        except:
            seller = None
        inf = r.find('div.tab-content .block-wrapper', first=True).text
        try:
            option = r.find('.product-options', first=True).text
        except:
            option = None
        sheet[f'A{i}'].value = name
        sheet[f'B{i}'].value = price
        sheet[f'C{i}'].value = price_after
        sheet[f'D{i}'].value = price_befor
        sheet[f'E{i}'].value = cat1 if cat1 != name else None
        sheet[f'F{i}'].value = cat2 if cat2 != name else None
        sheet[f'G{i}'].value = static
        sheet[f'H{i}'].value = seller
        sheet[f'I{i}'].value = des
        sheet[f'J{i}'].value = inf
        sheet[f'K{i}'].value = option
        sheet[f'L{i}'].value = img
        print(i, name, price)
    except Exception as e:
        print(e)
    i += 1
    book.save('tassel_result.xlsx')


#############
# try:
#     k = 2
#     book = Workbook()
#     sheet = book.active
#     links = []
#     while True:
#         sleep(5)
#         driver.execute_script("window.scrollTo(0, document.body.scrollHeight-1200);")
#         for u in driver.find_elements_by_css_selector(
#                 ' div.main-products.product-grid > div > div > div.caption > div.name > a'):
#             link = u.get_attribute('href')
#             if link not in links:
#                 links.append(link)
#                 print(len(links))
#         try:
#             driver.find_element_by_css_selector('div.ias-trigger-next a.btn').click()
#         except:
#             pass
# except Exception as r:
#     print(r)
#
# try:
#     k = 2
#     book = Workbook()
#     sheet = book.active
#     links = []
#     page = 0
#     while True:
#         driver.get(
#             f'https://www.extra.com/ar-sa/computer/laptops/c/3-303?q=%3Arelevance%3AinStock%3Atrue&pg={page}&pageSize=48')
#         page += 1
#         sleep(5)
#         if len(driver.find_elements_by_css_selector('div.row-flex.c_product-listing > div > div > a')) == 0:
#             break
#         for u in driver.find_elements_by_css_selector('div.row-flex.c_product-listing > div > div > a'):
#             link = u.get_attribute('href')
#             if link not in links:
#                 links.append(link)
#                 print(len(links))
#         sleep(3)
#         try:
#             driver.find_element_by_css_selector('.c_pagination-next a').click()
#         except:
#             pass
#     try:
#         for i in links:
#             sheet[f'N{k}'].value = i
#             k += 1
#         book.save('laptops.xlsx')
#     except Exception as e:
#         print(e)
# except Exception as r:
#     print(r)