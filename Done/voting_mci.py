from openpyxl import Workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from time import sleep

book = Workbook()
sheet = book.active

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get('https://voting.mci.gov.sa/Voters/Index?eid=402e1f56-1fb8-494b-8c24-d8eaf3f74eb5')
sleep(5)
driver.find_element_by_css_selector('select[name="votersTable_length"]').send_keys(500)
sleep(7)
num = 2
while True:
    all_text = driver.find_element_by_css_selector('#votersTable > tbody').text
    all_line = all_text.split('\n')
    print(len(all_line))
    for line in all_line:
        sheet.cell(num, 3).value = line.split(' ')[-1]
        sheet.cell(num, 2).value = line.split(' ')[-2]
        sheet.cell(num, 1).value = ' '.join(line.split(' ')[:-2])
        num += 1
    # for element in driver.find_elements_by_css_selector('#votersTable > tbody > tr'):
    #     FacilityName = element.find_element_by_css_selector('td.FacilityName').text
    #     CommercialRegistrationNo = element.find_element_by_css_selector('td.CommercialRegistrationNo').text
    #     COCNumber = element.find_element_by_css_selector('td.COCNumber').text
    #     print(num)
    #     sheet.cell(num, 1).value = FacilityName
    #     sheet.cell(num, 2).value = CommercialRegistrationNo
    #     sheet.cell(num, 3).value = COCNumber
    if 'disabled' not in driver.find_element_by_css_selector('#votersTable_next').get_attribute('class'):
        driver.find_element_by_css_selector('#votersTable_next').click()
        print('- Click')
        print(num)
        sleep(5)
    else:
        break


book.save('test.xlsx')
