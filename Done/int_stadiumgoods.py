import json
import os
import subprocess
from time import sleep
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from datetime import datetime
import re


class CollectPosts(object):
    def __init__(self):
        try:
            self.book = load_workbook('final_11.xlsx')
        except:
            print('cant find')
            self.book = Workbook()

        self.sheet = self.book.active
        self.old = [self.sheet.cell(i, 26).value for i in range(1, self.sheet.max_row + 1) if self.sheet.cell(i, 26).value]
        chrome_options = Options()
        self.user_data = os.path.expanduser("~") + r'\AppData\Local\Google\Chrome\User Data'
        chrome_options.add_argument("--user-data-dir={}".format(self.user_data))
        # chrome_options.add_argument('--profile-directory=Profile 4')
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
        self.driver.get('https://www.stadiumgoods.com/en-sa/shopping/air-force-1-low-07-white-on-white-16352608')

    def start(self, urls):
        num = self.sheet.max_row+2

        for v, url in enumerate(urls):
            print(v, num)
            if url in self.old:
                print('-found')
                continue
            self.driver.get(url)
            # sleep(6)
            WebDriverWait(self.driver, 20).until(
                (ec.visibility_of_element_located((By.CSS_SELECTOR, 'div.product-brand'))))
            brand = self.driver.find_element_by_css_selector('div.product-brand').text
            product_name = self.driver.find_element_by_css_selector('h1.product-name').text
            try:
                product_name_next = self.driver.find_element_by_css_selector('h1.product-name+p').text
            except:
                product_name_next = None
            cats = self.driver.find_elements_by_css_selector('[data-test="breadcrumbs-item"]')
            data = self.driver.execute_script("return document.querySelector('#root + script').text")

            json_data = json.loads(data.replace('window.__PRELOADED_STATE__ = ', ''))
            key = [key for key in json_data['entities']['products'].keys()][0]
            img = list(set(img['sources']['2048'] for img in json_data['entities']['products'][key]['images']))
            desk = json_data['entities']['products'][key]['description']
            sku = json_data['entities']['products'][key]['sku']

            # try:
            #     self.driver.find_element_by_css_selector('p > button.e1q51ql11').click()
            # except:
            #     pass
            # dsk = self.driver.find_element_by_css_selector('[data-test="productDescriptionContainer"]').text.replace('READ MORE', '').text.replace('READ LESS', '')
            WebDriverWait(self.driver, 20).until(
                (ec.element_to_be_clickable((By.CSS_SELECTOR, '[data-test="pdp-dropdown-button"]')))).click()

            # self.driver.find_element_by_css_selector('[data-test="pdp-dropdown-button"]').click()
            sleep(1)
            im = 0
            self.sheet.cell(num, 26).value = url
            for size in self.driver.find_elements_by_css_selector('.product-shop.css-gluvp2.em3pm8o3 > label'):
                try:
                    if 'UK' in size.find_element_by_css_selector('div.css-1962bhi.euap27q7').text:
                        continue
                except:
                    pass

                name = re.findall(r'\d{1,3}.\d{1,3}|\d{1,3}', size.find_element_by_css_selector('span.euap27q4').text)[0]
                if float(name.replace(',', '.')) > 13:
                    continue
                # print(name)
                # price = size.find_element_by_css_selector('input').get_attribute('data-price')
                price = size.find_element_by_css_selector('div > span').text
                self.sheet.cell(num, 1).value = product_name
                self.sheet.cell(num, 2).value = product_name_next
                self.sheet.cell(num, 3).value = brand
                self.sheet.cell(num, 4).value = sku
                self.sheet.cell(num, 5).value = name
                self.sheet.cell(num, 6).value = price
                self.sheet.cell(num, 7).value = desk

                c = 8
                for cat in cats[1:]:
                    self.sheet.cell(num, c).value = cat.text
                    c += 1

                # c = 13
                # for i in img:
                #     self.sheet.cell(num, c).value = i
                #     c += 1
                try:
                    if im < img.__len__():
                        self.sheet.cell(num, 13).value = img[im]
                        self.sheet.cell(num, 14).value = im+1
                        im += 1
                except:
                    pass
                num += 1

        self.book.save('final_11.xlsx')


# sheet = load_workbook(r"C:\Users\3mora\Downloads\ready to scraping.xlsx").active
# urls = [sheet.cell(i, 1).value for i in range(1, sheet.max_row+1) if sheet.cell(i, 1).value]
# print(urls.__len__())
# self = CollectPosts()
# input('- ')
# self.start(urls)
# book = load_workbook(r"C:\Users\3mora\Downloads\final_1 (1).xlsx")
# sheet = book.active
# code = None
#
# for i in range(2, sheet.max_row+1):
#     if code is None:
#         code = sheet.cell(i, 10).value
#         imgs = set([sheet.cell(i, c).value for c in range(14, sheet.max_column+1) if sheet.cell(i, c).value])
#         for v, img in enumerate(imgs):
#             sheet.cell(i+v, 13).value = img
#             sheet.cell(i+v, 12).value = v+1
#
#     if code == sheet.cell(i, 10).value:
#         continue
#
#     else:
#         code = sheet.cell(i, 10).value
#         imgs = set([sheet.cell(i, c).value for c in range(14, sheet.max_column+1) if sheet.cell(i, c).value])
#         for v, img in enumerate(imgs):
#             sheet.cell(i+v, 13).value = img
#             sheet.cell(i+v, 12).value = v+1
#     print(i)
#
# book.save('fo.xlsx')

# book = load_workbook(r"C:\Users\3mora\Downloads\toworkon.xlsx")
# sheet = book.active
# code = None
# for i in range(2, sheet.max_row+1):
#     if sheet.cell(i, 7).value:
#         code = sheet.cell(i, 7).value
#     else:
#         sheet.cell(i, 7).value = code
# book.save('for.xlsx')

# book = load_workbook(r"C:\Users\3mora\Downloads\edit.xlsx")
# sheet = book.active
#
# # مش مهمه اوي ممكن نحذفها الفكره اني بحجز مكان لمتغير من غير قيمة
# code = None
# for i in range(2, sheet.max_row+1):
#
#     # هنا بتشك لو قيمه العمود رقم 7 ليها قيمه ولا فضيه
#     if sheet.cell(i, 7).value:
#         # لو ليها قيمه بحطها في متغير علشان استخدمها في الباقي
#         code = sheet.cell(i, 7).value
#
#     # هنا لو المتغير الا هو كان اسم المنتج بيساوي الخليه رقم 10 حطلي قيمتها ب None  يعني خليها فاضيه
#     if code == sheet.cell(i, 10).value:
#         sheet.cell(i, 10).value = None
#
#     if code == sheet.cell(i, 11).value:
#         sheet.cell(i, 11).value = None
#
# book.save('for_edit_1.xlsx')
def return_(dict_: dict):

    li = [float(v) for v in dict_.values() if v]
    li.sort()
    if li.__len__() == 1:
        print('- only one')
        return li[0]
    elif li.__len__() <1:
        print('- No price')
    else:
        m = (li[0]+li[-1])//2
        print(m)
        return m
    return None


def change(sheet, val, list_):
    for i in list_.keys():
        if list_[i] is not None:
            sheet.cell(i, 5).value = list_[i]
        else:
            sheet.cell(i, 5).value = val


book = load_workbook(r"C:\Users\3mora\Downloads\edit.xlsx")
sheet = book.active

# مش مهمه اوي ممكن نحذفها الفكره اني بحجز مكان لمتغير من غير قيمة
code = None
list_ = dict()
for i in range(2, sheet.max_row+1):

    if code is None:
        code = sheet.cell(i, 2).value
        list_[i] = sheet.cell(i, 4).value

    if code == sheet.cell(i, 2).value:
        list_[i] = sheet.cell(i, 4).value

    if code != sheet.cell(i, 2).value:
        val = return_(list_)
        change(sheet, val, list_)
        code = sheet.cell(i, 2).value
        list_ = dict()
        list_[i] = sheet.cell(i, 4).value

book.save('edit_2.xlsx')

