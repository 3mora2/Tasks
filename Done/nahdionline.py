from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium import webdriver
from openpyxl import load_workbook
from time import sleep
import os

driver = webdriver.Chrome()
driver.implicitly_wait(4)
total = []
for file in os.listdir('file'):
    f = f'file//{file}'
    print(f)

    book = load_workbook(f)
    sheet = book.active
    print(sheet.max_row)
    i = 1
    sheet.cell(i, 1).value = 'Name'
    sheet.cell(i, 2).value = "الاسم"
    sheet.cell(i, 3).value = "البرند"
    sheet.cell(i, 4).value = "السعر"
    sheet.cell(i, 5).value = "السعر الجديد"
    sheet.cell(i, 6).value = "التقيم"
    sheet.cell(i, 7).value = "الوصف"
    sheet.cell(i, 8).value = "المميزات"
    sheet.cell(i, 9).value = "الصور"

    for i in range(2, sheet.max_row+1):
        url = sheet[f'W{i}'].value
        # if sheet[f'A{i}'].value is None:
        #     driver.get(url)
        #     driver.find_element_by_css_selector('#switcher-language').click()
        #     name_en = driver.find_element_by_css_selector('div.page-title-wrapper.product > h1 > span').text
        #     sheet.cell(i, 1).value = name_en
        # if sheet[f'G{i}'].value is None:
        #     driver.get(url)
        #     print(i)

            # sleep(7)
            # if sheet.cell(i, 6).value is None:
            #     try:
            #         rate = driver.find_element_by_css_selector('#ratings-summary > div.bv_avgRating_component_container.notranslate').text
            #         sheet.cell(i, 6).value = rate
            #     except:
            #         pass

            # photo = '\n'.join([f.get_attribute('src').split('?')[0] for f in driver.find_elements_by_css_selector(
            #                 'div.fotorama__stage__shaft.fotorama__grab > div.fotorama__loaded.fotorama__loaded--img > img.fotorama__img')])
            # if photo == '':
            #     photo = '\n'.join([f.get_attribute('src').split('?')[0] for f in driver.find_elements_by_css_selector(
            #                 'div.fotorama__stage > div.fotorama__stage__shaft img.fotorama__img')])
            # if photo != '':
            #     sheet.cell(i, 9).value = photo

            # if sheet.cell(i, 7).value is None:
            #     try:
            #         driver.find_element_by_css_selector('#description > span.read-more[data-collapsed="1"]').click()
            #     except:
            #         pass
            #
            #     desk = driver.find_element_by_css_selector('#description ').text.replace('وصف المنتج:', '').replace(
            #         'اقرأ المزيد', '').strip().replace('إخفاء', '')
            #     sheet.cell(i, 7).value = desk

            # if sheet.cell(i, 8).value is None:
            #     driver.find_element_by_css_selector('#tab-label-specifications-tab').click()
            #     sleep(1)
            #     specifications = driver.find_element_by_css_selector('#specifications-tab').text
            #     sheet.cell(i, 8).value = specifications

            # print(desk)
            # book.save('vitamins-supplements.xlsx')
        total.append(url)
        if url is None or sheet[f'B{i}'].value is not None:
            continue
        driver.get(url)
        sleep(5)
        try:
            name = driver.find_element_by_css_selector('div.page-title-wrapper.product > h1 > span').text
            try:
                brand = driver.find_element_by_css_selector('div.product-info-main > div.manufacturer > div.value').text
            except:
                brand = None
            try:
                price = driver.find_element_by_css_selector(
                    'div.product-info-main > div.product-info-price > div.price-box.price-final_price span[data-price-type="oldPrice"]').get_attribute(
                    'data-price-amount')
                new_price = driver.find_element_by_css_selector(
                'div.product-info-main > div.product-info-price > div.price-box.price-final_price span[data-price-type="finalPrice"]').get_attribute(
                'data-price-amount')
            except:
                new_price = None
                price = driver.find_element_by_css_selector(
                    'div.product-info-main > div.product-info-price > div.price-box.price-final_price span[data-price-type="finalPrice"]').get_attribute(
                    'data-price-amount')
            try:
                rate = driver.find_element_by_css_selector('#ratings-summary > div.bv_avgRating_component_container.notranslate').text
            except:
                rate = None

            photo = '\n'.join([f.get_attribute('src').split('?')[0] for f in driver.find_elements_by_css_selector(
                    'div.fotorama__stage__shaft.fotorama__grab > div.fotorama__loaded.fotorama__loaded--img > img.fotorama__img')])
            if photo == '':
                photo = '\n'.join([f.get_attribute('src').split('?')[0] for f in driver.find_elements_by_css_selector(
                    'div.fotorama__stage > div.fotorama__stage__shaft img.fotorama__img')])

            try:
                driver.find_element_by_css_selector('#description > span.read-more[data-collapsed="1"]').click()
            except:
                pass
            try:
                desk = driver.find_element_by_css_selector('#description ').text.replace('وصف المنتج:', '').replace('اقرأ المزيد', '').strip()
            except:
                desk = None
            # sleep(.5)
            try:
                driver.find_element_by_css_selector('#tab-label-specifications-tab').click()
            except:
                pass
            # sleep(1)
            try:
                specifications = driver.find_element_by_css_selector('#specifications-tab').text
            except:
                specifications = None

            # driver.get(url.replace('/ar/', '/en/'))
            driver.execute_script("window.scrollTo(0, 0)")
            sleep(.5)

            driver.find_element_by_css_selector('#switcher-language').click()
            try:
                name_en = WebDriverWait(driver, 15).until(EC.visibility_of_element_located((
                                By.CSS_SELECTOR, 'div.page-title-wrapper.product > h1 > span'))).text
            except:
                name_en = None
            # name_en = driver.find_element_by_css_selector('div.page-title-wrapper.product > h1 > span').text
            sheet.cell(i, 1).value = name_en
            sheet.cell(i, 2).value = name
            sheet.cell(i, 3).value = brand
            sheet.cell(i, 4).value = price
            sheet.cell(i, 5).value = new_price
            sheet.cell(i, 6).value = rate
            sheet.cell(i, 7).value = desk
            sheet.cell(i, 8).value = specifications
            sheet.cell(i, 9).value = photo
            print(len(total))
            print(f'{i}, name_en : {name_en}, name: {name}, brand : {brand}, price : {price}, new_price: {new_price}, rate: {rate}, {photo}')
        except Exception as e:
            print(e)
    book.save(f)

# lin = []
# linkts = []
# index = 0
# for el in driver.find_elements_by_css_selector('#ais-category-subtree a.ais-hierarchical-menu--link'):
#     linkts.append((el.text, index))
#     index += 1
#
# for title, index in linkts:
#
#     print(title)
#     links = []
#     # for page in range(1, 43):
#     # driver.get(f'{linkt}')
#     driver.find_elements_by_css_selector('#ais-category-subtree a.ais-hierarchical-menu--link')[index].click()
#     sleep(5)
#     while True:
#         sleep(5)
#         for ele in driver.find_elements_by_css_selector(' div.result-content > div.result-thumbnail > a[itemprop="url"]'):
#             if ele.get_attribute('href') not in lin:
#                 lin.append(ele.get_attribute('href'))
#                 if ele.get_attribute('href') not in links:
#                     links.append(ele.get_attribute('href'))
#         print(len(links))
#         try:
#             driver.find_element_by_css_selector(
#                 'li.ais-pagination--item.ais-pagination--item__next > a[aria-label="Next"]').click()
#         except:
#             break
#
#     from openpyxl import Workbook
#     wb = Workbook()
#     ws = wb.active
#
#     for i, item in enumerate(links):
#         print(i+2)
#         ws[f'W{i+2}'] = item
#     wb.save(f' العناية بالأم والطفل -{title.strip()}12.xlsx')

# https://www.nahdionline.com/ar/
# https://www.nahdionline.com/ar/vitamins-supplements?page=
# https://www.nahdionline.com/ar/healthy-food?page=2
# https://www.nahdionline.com/ar/medical-equipments?page=2


# https://www.nahdionline.com/ar/hair-care?page=2
# https://www.nahdionline.com/ar/skin-care?page=
#