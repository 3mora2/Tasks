#!/usr/bin/env python
# -*- coding: cp1256 -*-

# from win32com import client
# import win32gui


# def winEnumHandler(hwnd, ctx):
#     if win32gui.IsWindowVisible(hwnd):
#         title = win32gui.GetWindowText(hwnd)
#         cal = win32gui.GetClassName(hwnd)
#         cal = win32gui.GetLayeredWindowAttributes(hwnd)


# win32gui.EnumWindows(winEnumHandler, None)


# def Split_XSLX(file, out):
#     school_d = None
#     school_i = None
#     excel = client.Dispatch("Excel.Application")
#     excel.Visible = False
#     excel.DisplayAlerts = False
#     workbook = excel.Workbooks.Open(file)
#     num = workbook.Sheets.Count
#     for i in range(1, num + 1):
#         sheet = workbook.Worksheets(i)
#         ret = sheet.Range("AK13").Text
#         name = sheet.Range("AD19").Text
#         school_name = sheet.Range("AE6").Text
#         sheet.PageSetup.Orientation = 2
#         sheet.PageSetup.CenterHorizontally = True
#         sheet.PageSetup.PaperSize = 1
#         sheet.ExportAsFixedFormat(0, out + str(ret) + '.pdf')
#         print(i)
#
#     excel.Quit()

# if __name__ == '__main__':
# input_file = r'G:\programing\projects\drob\wael\school\1041591965\CS_SemesterNotificationBrief.xlsx'
#     out_f = 'G:\\programing\\projects\\drob\\wael\\school\\1041591965\\'
#     Split_XSLX(input_file, out_f)
#
# import openpyxl
# from reportlab.pdfgen import canvas
# from reportlab.pdfbase import pdfmetrics
# from reportlab.pdfbase.pdfmetrics import stringWidth
# from reportlab.pdfbase.ttfonts import TTFont
# import os
# from PyPDF2 import PdfFileReader, PdfFileMerger
#
# pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
# wb = openpyxl.load_workbook(input_file, read_only=True)
# sheet = wb.get_sheet_by_name("Sheet1")
#
# # print(sheet.cell(2,2).value)
#
# page_width = 2156
# page_height = 3050
# start = 200
# start_2 = 700
# spread = 80
# categories = ["Roll Number: ", "Name: ", "Gender: ", "Age: ", "Board: ", "City: ", "Email: "]
#
# university = "NED University Of Engineering And Technology"
#
#
# def create_data():
#     for i in range(2, 40):
#         std_id = sheet.cell(row=i, column=1).value
#         std_name = sheet.cell(row=i, column=2).value
#         std_gender = sheet.cell(row=i, column=3).value
#         std_age = sheet.cell(row=i, column=4).value
#         std_board = sheet.cell(row=i, column=5).value
#         std_city = sheet.cell(row=i, column=6).value
#         std_email = sheet.cell(row=i, column=8).value
#
#         data = [std_id, std_name, std_gender, std_age, std_board, std_city, std_email]
#
#         c = canvas.Canvas(str(std_name) + '.pdf')
#         c.setPageSize((page_width, page_height))
#         c.setFont('Arial', 80)
#         text_width = stringWidth(university, 'Arial', 80)
#         c.drawString((page_width - text_width) / 2, 2900, university)
#
#         y = 2500
#
#         for x in range(0, 7):
#             c.setFont('Arial', 40)
#             c.drawString(start, y, categories[x])
#             c.drawString(start_2, y, str(data[x]))
#             y -= spread
#
#         c.save()
#
#
# def merge_data():
#     files_dir = '.'
#     pdf_files = [f for f in os.listdir(files_dir) if f.endswith('.pdf')]
#     merger = PdfFileMerger()
#     for filename in pdf_files:
#         merger.append(PdfFileReader(os.path.join(files_dir, filename, 'rb')))
#     merger.write(os.path.join(files_dir, 'merged_data.pdf'))
#
#
# create_data()
# merge_data()
file = r'C:\Users\3mora\Downloads\CS_SemesterNotificationBrief.xlsx'
# file = r'C:\Users\3mora\Downloads\CS_SemesterNotificationBrief.docx'
#, encoding='cp1256')
from openpyxl import load_workbook
names = []
book = load_workbook(file, read_only=True)
sheets = book.worksheets
for num, sheet in enumerate(sheets):
    names.append(sheet['AD19'].value)
    print(num)
print(len(names))
# name = sheet.Range("AD19").Text
# import PyPDF2
#
# read_pdf = PyPDF2.PdfFileReader(file)
# number_of_pages = read_pdf.getNumPages()
# print(number_of_pages)
# for i in range(number_of_pages):
#     page = read_pdf.getPage(0)
#     pdf_writer = PyPDF2.PdfFileWriter()
#     pdf_writer.addPage(page)
#     with open(str(i)+'rotate_pages.pdf', 'wb') as fh:
#         pdf_writer.write(fh)

# from PyPDF2 import PdfFileReader, PdfFileWriter
#
# def split(path, name_of_split):
#     pdf = PdfFileReader(path)
#     for page in range(pdf.getNumPages()):
#         pdf_writer = PdfFileWriter()
#         pdf_writer.addPage(pdf.getPage(page))
#
#         output = f'{name_of_split}{page}.pdf'
#         with open(output, 'wb') as output_pdf:
#             pdf_writer.write(output_pdf)
#
# if __name__ == '__main__':
#     path = 'Jupyter_Notebook_An_Introduction.pdf'
#     split(file, 'jupyter_page')