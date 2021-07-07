from selenium import webdriver
from time import sleep
from openpyxl import Workbook, load_workbook
import os

import requests_html
f = []
s = requests_html.HTMLSession()
for file in os.listdir('text'):
    if file.endswith('.xlsx'):
        print(file)
        book = load_workbook('text\\' + file)
        sheet = book.active
        print(sheet.max_row)
        for i in range(2, sheet.max_row + 1):

            if sheet[f'I{i}'].value is not None and sheet[f'F{i}'].value is None:
                f.append((file, i))

            # if sheet[f'I{i}'].value is not None:
            #     if sheet[f'G{i}'].value is None or sheet[f'G{i}'].value == '':
            #         try:
            #             r = s.get(sheet[f'I{i}'].value + '?___store=en')
            #             name = r.html.find('div.product-name > h1', first=True).text
            #             sheet[f'G{i}'] = name
            #             print(i, name)
            #         except Exception as e:
            #             print(e)

            # try:
            #     r = s.get(sheet[f'I{i}'].value)# + '?___store=ar')
            #     try:
            #         if len([ele.attrs['href'] for ele in r.html.find('.slides > li > a')]) > 1:
            #             img = '\n '.join([ele.attrs['href'] for ele in r.html.find('.slides > li > a')])
            #             sheet[f'H{i}'] = img
            #     except Exception as e:
            #         print(e)
            #
            #     try:
            #         cat3 = r.html.find('.cat-tree', first=True).text.split('/')[-1].strip()
            #
            #         sheet[f'F{i}'] = cat3
            #     except Exception as e:
            #         cat3 = 'error'
            #         print(e)
            #     try:
            #         All = ' '.join(r.html.find('.cat-tree', first=True).text.split())
            #         sheet[f'J{i}'] = All
            #     except:
            #         pass
            #     print(i, cat3)
            # except Exception as e:
            #     print(e)
        book.save('text\\' + file)


# file = 'new.xlsx'
# book = load_workbook(file)
# sheet = book.active
#
# for i in range(2, sheet.max_row + 1):
#     if sheet[f'I{i}'].value is not None:
#         r = s.get(sheet[f'I{i}'].value + '?___store=en')
#         sleep(5)
#         sheet[f'G{i}'] = r.html.find('div.product-name > h1', first=True).text
#         print(i)


# book = Workbook()
# sheet = book.active
# sheet['A1'] = 'Names'
# sheet['B1'] = 'price'
# sheet['C1'] = 'price'
# sheet['D1'] = 'category1'
# sheet['E1'] = 'category2'
# sheet['F1'] = 'category3'
#
# sheet['G1'] = 'name_en'
# sheet['H1'] = 'img'
# sheet['I1'] = 'url'
#
# driver = webdriver.Chrome()
# driver.get('https://www.othaimmarkets.com/?___store=')
# driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
# links = []
# n = 2
# for element in driver.find_elements_by_css_selector('div.span2.product '):
#     if element.find_element_by_class_name('div.wrapper-hover h3.product-name > a').get_attribute('href') not in links:
#
#         links.append(element.find_element_by_css_selector('div.wrapper-hover h3.product-name > a').get_attribute('href'))
#         sheet[f'A{n}'] = element.find_element_by_css_selector('div.wrapper-hover h3.product-name > a').text
#         if len(element.find_elements_by_css_selector('div.wrapper-hover div.price-box span.price')) == 1:
#             price1 = element.find_element_by_css_selector('div.wrapper-hover div.price-box span.price').text
#             price2 = None
#         else:
#             price1 = element.find_elements_by_css_selector('div.wrapper-hover div.price-box span.price')[-1].text
#             price2 = element.find_elements_by_css_selector('div.wrapper-hover div.price-box span.price')[0].text
#         sheet[f'B{n}'] = price1
#         sheet[f'C{n}'] = price2
#         try:
#             sheet[f'D{n}'] = driver.find_element_by_css_selector('#content > div.row.top > div > div.braeed-crumms > div > a:nth-child(3)').text
#         except:
#             pass
#         try:
#             sheet[f'E{n}'] = driver.find_element_by_css_selector('#content > div.row.top > div > div.braeed-crumms > div > a:nth-child(5)').text
#         except:
#             pass
#         sheet[f'H{n}'] = element.find_element_by_css_selector(
#             'div.product-image-wrapper.nohover > a > img').get_attribute('src').replace('small_image/170x165', 'image')
#         sheet[f'I{n}'] = element.find_element_by_css_selector('div.wrapper-hover h3.product-name > a').get_attribute('href')
#         n += 1
#         print(len(links))
# book.save(file)


# cat = []
# for el in driver.find_elements_by_css_selector('#collapsed-menu > li > ul > li > label > a'):
#     cat.append((el.text, el.get_attribute('href')))
# cat = [('مشروبات ساخنة', 'http://www.othaimmarkets.com/grocery/hot-drinks.html'),
#        ('تصنيع الحلويات والمخبوزات', 'http://www.othaimmarkets.com/grocery/home-baking.html'),
#        ('معلبات', 'http://www.othaimmarkets.com/grocery/canned-foods.html'),
#        ('حليب', 'http://www.othaimmarkets.com/grocery/milk.html'),
#        ('أجبان وقشطة وزبدة', 'http://www.othaimmarkets.com/grocery/uht-cream.html'),
#        ('منتجات الحمية', 'http://www.othaimmarkets.com/grocery/dietary-products.html'),
#        ('صلصات ومعجون الطماطم', 'http://www.othaimmarkets.com/grocery/sauces-tomato-paste.html'),
#        ('زيوت وسمن', 'http://www.othaimmarkets.com/grocery/oils-ghee.html'),
#        ('معكرونة وشوربة', 'http://www.othaimmarkets.com/grocery/pasta-noodles-soup.html'),
#        ('الأرز والسكر', 'http://www.othaimmarkets.com/grocery/rice-sugar.html'),
#        ('مأكولات آسيوية', 'http://www.othaimmarkets.com/grocery/oriental-food.html'),
#        ('حبوب إفطار', 'http://www.othaimmarkets.com/grocery/cereal-food.html'),
#        ('أغذية إفطار', 'http://www.othaimmarkets.com/grocery/breakfast-food.html'),
#        ('الخبز والحلويات', 'http://www.othaimmarkets.com/fresh-and-bakery/bakery.html'),
#        ('دواجن وبيض', 'http://www.othaimmarkets.com/fresh-and-bakery/fresh-chicken-eggs.html'),
#        ('منتجات ألبان', 'http://www.othaimmarkets.com/fresh-and-bakery/dairy-products.html'),
#        ('فواكه وخضروات', 'http://www.othaimmarkets.com/fresh-and-bakery/fruit-vegetables.html'),
#        ('لحوم', 'http://www.othaimmarkets.com/fresh-and-bakery/fresh-meat.html'),
#        ('أجبان ولانشون وزيتون', 'http://www.othaimmarkets.com/fresh-and-bakery/delicatessens.html'),
#        ('آيس كريم وحلويات', 'http://www.othaimmarkets.com/frozen/ice-cream-cakes22.html'),
#        ('خضروات وفواكه', 'http://www.othaimmarkets.com/frozen/fruit-vegetables.html'),
#        ('معجنات للتحضير', 'http://www.othaimmarkets.com/frozen/pastries.html'),
#        ('مأكولات مسبقة التحضير', 'http://www.othaimmarkets.com/frozen/ready-to-eat.html'),
#        ('لحوم ودواجن', 'http://www.othaimmarkets.com/frozen/chicken-meat.html'),
#        ('أسماك وأطعمة بحرية', 'http://www.othaimmarkets.com/frozen/fish-seafood.html'),
#        ('حلويات وتسالي', 'http://www.othaimmarkets.com/beverage-confectionery/confectionery-snacks.html'),
#        ('مشروبات وعصائر', 'http://www.othaimmarkets.com/beverage-confectionery/drinks-juice.html'),
#        ('مستلزمات حلاقة', 'http://www.othaimmarkets.com/health-beauty/shaving-needs.html'),
#        ('عناية المرأة', 'http://www.othaimmarkets.com/health-beauty/woman-care.html'),
#        ('عناية الجسم والوجه', 'http://www.othaimmarkets.com/health-beauty/body-face-care.html'),
#        ('صابون الأيدي والجسم', 'http://www.othaimmarkets.com/health-beauty/hand-soap-shower-gel.html'),
#        ('مزيلات عرق', 'http://www.othaimmarkets.com/health-beauty/deodorant.html'),
#        ('عناية الأسنان', 'http://www.othaimmarkets.com/health-beauty/dental-care.html'),
#        ('عناية الشعر', 'http://www.othaimmarkets.com/health-beauty/hair-care.html'),
#        ('عناية الطفل', 'http://www.othaimmarkets.com/baby-needs-4/baby-food.html'),
#        ('أغذية الطفل', 'http://www.othaimmarkets.com/baby-needs-4/baby-food-1.html'),
#
#        ('أواني منزلية', 'http://www.othaimmarkets.com/home-care/kitchen-needs.html'), ('غسيل الملابس', 'http://www.othaimmarkets.com/home-care/household.html'), ('عناية الملابس', 'http://www.othaimmarkets.com/home-care/clothes-care.html'), ('غسيل وتنظيف الصحون', 'http://www.othaimmarkets.com/home-care/dishwashing.html'), ('مطهرات وملمعات ومعطرات', 'http://www.othaimmarkets.com/home-care/disinfectant-cleaner-freshener.html'), ('قصدير وحفظ وتغليف الماكؤلات', 'http://www.othaimmarkets.com/home-care/foils-cling-films.html'), ('منتجات بلاستيكية', 'http://www.othaimmarkets.com/home-care/plastic-products.html'), ('منتجات ورقية', 'http://www.othaimmarkets.com/home-care/paper-products.html'), ('طعام حيوانات', 'http://www.othaimmarkets.com/home-care/pet-food.html'), ('أدوات منزلية', 'http://www.othaimmarkets.com/home-care/household-4.html'), ('أجهزة عناية شخصية', 'http://www.othaimmarkets.com/electronics-appliances/personal-care-machines.html'), ('مستلزمات وعدد', 'http://www.othaimmarkets.com/electronics-appliances/tools-hardware.html'), ('مستلزمات رحلات', 'http://www.othaimmarkets.com/electronics-appliances/outdoor-camping.html'), ('أجهزة منزلية', 'http://www.othaimmarkets.com/electronics-appliances/home-appliances.html')]

# for name, url in cat:
#     book = Workbook()
#     sheet = book.active
#     driver.get(url)
#     sleep(3)
#     links = []
#     n = 2
#     while True:
#         driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#         sleep(4)
#
#         for element in driver.find_elements_by_css_selector('div.span2.product div.wrapper-hover h3.product-name > a'):
#             if element.get_attribute('href') not in links:
#
#                 links.append(element.get_attribute('href'))
#
#                 sheet[f'I{n}'] = element.get_attribute('href')
#                 n += 1
#                 print(len(links))
#
#     book.save(str(name)+'.xlsx')

# from openpyxl import Workbook, load_workbook
#
# import requests_html
# s = requests_html.HTMLSession()
# g = []
# for name, url in cat:
#     book = Workbook()
#     sheet = book.active
#     sheet['A1'] = 'Names'
#     sheet['B1'] = 'price'
#     sheet['C1'] = 'price'
#     sheet['D1'] = 'category1'
#     sheet['E1'] = 'category2'
#     sheet['F1'] = 'category3'
#     sheet['G1'] = 'name_en'
#     sheet['H1'] = 'img'
#     sheet['I1'] = 'url'
#     l = []
#     f = True
#     print(name)
#     n = 2
#     while True:
#         if f:
#             r = s.get(url + '?limit=30')
#             f = ''
#         elif f == '':
#             if len(r.html.find('.col-md-9 #toptoolbar .pages a.small_icon')) == 0:
#                 break
#             r = s.get(r.html.find('.col-md-9 #toptoolbar .pages a.small_icon')[0].attrs['href'])
#             f = False
#         else:
#             if len(r.html.find('.col-md-9 #toptoolbar .pages a.small_icon')) == 1:
#                 break
#             r = s.get(r.html.find('.col-md-9 #toptoolbar .pages a.small_icon')[1].attrs['href'])
#
#         for el in r.html.find('div.row > div.span2'):
#             l.append(el.find('div.wrapper-hover h3.product-name > a')[0].attrs['href'])
#             g.append(el.find('div.wrapper-hover h3.product-name > a')[0].attrs['href'])
#
#             sheet[f'A{n}'] = el.find('div.wrapper-hover h3.product-name > a')[0].text
#
#             if len(el.find('div.wrapper-hover div.price-box span.price')) == 1:
#                 price1 = el.find('div.wrapper-hover div.price-box span.price')[0].text
#                 price2 = None
#             else:
#                 price1 = el.find('div.wrapper-hover div.price-box span.price')[1].text
#                 price2 = el.find('div.wrapper-hover div.price-box span.price')[0].text
#
#             sheet[f'B{n}'] = price1
#             sheet[f'C{n}'] = price2
#             try:
#                 sheet[f'D{n}'] = r.html.find('#content > div.row.top > div > div.braeed-crumms > div > a:nth-child(3)')[0].text
#             except:
#                 pass
#             try:
#                 sheet[f'E{n}'] = r.html.find('#content > div.row.top > div > div.braeed-crumms > div > a:nth-child(5)')[0].text
#             except:
#                 pass
#             sheet[f'H{n}'] = el.find('div.product-image-wrapper > a > img')[0].attrs['src'].replace('small_image/170x165', 'image')
#             sheet[f'I{n}'] = el.find('div.wrapper-hover h3.product-name > a')[0].attrs['href']
#             n += 1
#
#         print(len(l))
#     print('********', r.html.find('.col-md-9 #toptoolbar .pull-left')[2].text.split(' ')[-1], '********')
#     book.save(str(name)+'.xlsx')
#
# print(len(g))

# 5138
