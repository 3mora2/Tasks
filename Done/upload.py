import os
import subprocess
from time import sleep
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, UnexpectedAlertPresentException
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
from selenium.webdriver.chrome.options import Options
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec


class Main:
    def __init__(self):
        # chrome_options = Options()
        # chrome_options.add_argument("--user-data-dir={}".format(
        #     r'C:\Users\3mora\AppData\Local\Google\Chrome\User Data'))
        self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())#, chrome_options=chrome_options)
        self.book = load_workbook('uploaded.xlsx')
        self.sheet = self.book.active

    def start(self):

        for i in range(2, self.sheet.max_row + 1):
            print(i)
            if self.sheet.cell(i, 4).value and not self.sheet.cell(i, 8).value:
                name = self.sheet.cell(i, 4).value
                path = rf'E:\New folder\3bdo\img bcg remover\{name}'
                if os.path.exists(path):
                    self.sheet.cell(i, 8).value = self.upload(path)

            if self.sheet.cell(i, 5).value and not self.sheet.cell(i, 9).value:
                name = self.sheet.cell(i, 5).value
                path = rf'E:\New folder\3bdo\img bcg remover\{name}'
                if os.path.exists(path):
                    self.sheet.cell(i, 9).value = self.upload(path)

            self.book.save('uploaded.xlsx')

    def upload(self, path):
        try:
            self.driver.get('https://www.linkpicture.com/en/index.php')
            sleep(4)
            WebDriverWait(self.driver, 60).until(
                (ec.presence_of_element_located((By.CSS_SELECTOR, '#mFile')))).send_keys(path)
            # self.driver.find_element_by_css_selector('#mFile').send_keys(path)
            sleep(1)
            WebDriverWait(self.driver, 60).until(
                (ec.element_to_be_clickable((By.CSS_SELECTOR, '[name="pcden"]')))).click()

            # self.driver.find_element_by_css_selector('[name="pcden"]').click()
            sleep(4)
            WebDriverWait(self.driver, 160).until(
                (ec.visibility_of_element_located((By.CSS_SELECTOR, '.form-control'))))
            url = self.driver.find_elements_by_css_selector('.form-control')[-1].get_attribute('value')
            print(url)
            return url
        except UnexpectedAlertPresentException as e:
            print('- error')
            return None
        except Exception as e:
            print(e)
            return None


self = Main()
self.start()
