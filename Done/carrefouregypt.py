from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import time
from openpyxl import Workbook
###############################################################
book = Workbook()
sheet = book.active
sheet['A1'] = 'Names'
sheet['B1'] = 'id'
sheet['C1'] = 'price'
sheet['D1'] = 'brand'
sheet['E1'] = 'category'
sheet['F1'] = 'time'
################################################################
driver=webdriver.Firefox()
driver.get('https://www.carrefouregypt.com/')
##################################################################
def xl(data,r):
    data=eval(data)
    sheet.cell(row=r, column=1).value =data['name']
    sheet.cell(row=r, column=2).value =data['id']
    sheet.cell(row=r, column=3).value =data['price']
    sheet.cell(row=r, column=4).value =data['brand']
    sheet.cell(row=r, column=5).value =data['category']
    sheet.cell(row=r, column=6).value =data['dimension26']
##################################################################
try:
    hov=WebDriverWait(driver,10).until(ec.visibility_of_element_located((By.CLASS_NAME, "css-a4hvbl")))
    hover = ActionChains(driver).move_to_element(hov)
    hover.perform()
    ###################################################################
    get_cl=WebDriverWait(driver,10).until(ec.visibility_of_all_elements_located((By.CLASS_NAME, "css-1an2mwo")))

    links=[]
    categ=[]

    if get_cl:
        for i in get_cl:
            categ.append(i.text)
            links.append(i.get_attribute('href'))
    ########################################################################
    r=2
    if len(links):
        for i in links:
            driver.get(i)
            #driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)
            try:
                num=int(driver.find_elements_by_class_name('plp-pagination__navpages')[-1].text)
                print(num)
                for o in range(0,num):
                    driver.get('{0}?&qsort=relevance&pg={1}'.format(i,o))
                    prod=driver.find_elements_by_class_name('js-gtmProdData')
                    for p in prod:
                        pro=p.get_attribute('data-gtm-prod-data')
                        xl(pro,r)
                        r+=1

            except:
                print('no')
                prod=driver.find_elements_by_class_name('js-gtmProdData')
                for p in prod:
                    pro=p.get_attribute('data-gtm-prod-data')
                    xl(pro, r)
                    r += 1
except Exception as e:
    print(e)
##########################################################################
book.save('data.xlsx')
driver.close()
