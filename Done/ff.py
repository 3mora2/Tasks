
from webdriver_manager.chrome import ChromeDriverManager

from selenium import webdriver
from time import sleep
from openpyxl import Workbook, load_workbook


class Main:
    def __init__(self):
        self.book = Workbook()
        self.sheet = self.book.active
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        l = ['http://www.140online.com/company/R2808/صحارى لتصنيع و تعبئه المشروبات/',
             'http://www.140online.com/company/C88745/كوكاكولا - اطلانتك اندستريز/',
             'http://www.140online.com/company/C523228/مان لتعبئه المياه/',
             'http://www.140online.com/company/C174630/السادات للتصنيع و التنميه - اكوا سيوا/',
             'http://www.140online.com/company/NW5773/المصرية للعبوات الزراعية/',
             'http://www.140online.com/company/NW5774/المصريه للخيوط الصناعيه و البلاستيك/',
             'http://www.140online.com/company/NW5770/سيوة للعبوات/',
             'http://www.140online.com/company/C37920/الدوليه المصريه العربيه للاستثمار/',
             'http://www.140online.com/company/C173820/الرواد/',
             'http://www.140online.com/company/C605886/الواحه لتعبئه المياه/',
             'http://www.140online.com/company/C87815/نستله للمياه مصر/',
             'http://www.140online.com/company/U2708200924766/سول ووتر/',
             'http://www.140online.com/company/C9469/اكواستون/',
             'http://www.140online.com/company/C624042/المنار للمياه المعدنيه/',
             'http://www.140online.com/company/NW3793/بامبينى/',
             'http://www.140online.com/company/C615709/بيور ووتر تكنولوجى لمعالجه المياه/',
             'http://www.140online.com/company/C187091/اكوا بيور/',
             'http://www.140online.com/company/N432/سونات - الوطنيه للصناعات الغذائيه/',
             'http://www.140online.com/company/C11808/اكوا دلتا - دلتا للمياه المعدنيه/',
             'http://www.140online.com/company/C10850/برجن للصناعه/',
             'http://www.140online.com/company/N1072/كراش - النيل للمياه الغازيه/',
             'http://www.140online.com/company/C20341/سيوه للتصنيع و التنميه/',
             'http://www.140online.com/company/R2432/الدويك لانظمة معالجة المياة/',
             'http://www.140online.com/company/NW19523/العجلان العربية للصناعة والاستثمار/',
             'http://www.140online.com/company/N678/العربية لمعالجة وتعبئة المياه - اكوا تاب /',
             'http://www.140online.com/company/N1350/المتحده للمياه المعدنيه - منيرال/',
             'http://www.140online.com/company/NW18773/بركة للمياه المعدنية/',
             'http://www.140online.com/company/NW28254/برومس جروب/',
             'http://www.140online.com/company/C613687/بولى بت للعبوات البلاستيكية/',
             'http://www.140online.com/company/NW6568/شركة واتر هاوس للمياة المعدنية/',
             'http://www.140online.com/company/N442/نستلة مصر/',
             'http://www.140online.com/company/NW18416/هيرو سبورتيف/',
             'http://www.140online.com/company/NW18417/واحة باريس للمياه الطبيعية - احدى شركات وزارة البترول/',
             'http://www.140online.com/company/NW27137/وادي الريان للخدمات البترولية والنقل/',
             'http://www.140online.com/company/C583122/ووتر سيتى تكنولوجى لمعالجة المياة وحمامات السباحة/',
             'http://www.140online.com/company/C57934/العالميه لتصنيع و تعبئه المياه الطبيعيه/',
             'http://www.140online.com/company/C410180/الماسه لمعالجه المياه/',
             'http://www.140online.com/company/C243950/منتجات مياه مصر - اكوا دلتا/',
             'http://www.140online.com/company/NW5771/مجموعة الأكوح جروب - المصرية للبلاستيك والكيماويات - بولي تك/',
             'http://www.140online.com/company/C174051/الرواء/', 'http://www.140online.com/company/C612744/اليكس جروب/',
             'http://www.140online.com/company/NW8990/الوطنية لتصنيع وتعبئة المياه الطبيعية - صافي/']
        i = 2
        for url in l:
            self.driver.get(url)
            name = self.driver.find_element_by_css_selector('#ctl09_lblCompanyName').text
            email = self.driver.find_element_by_css_selector('#ctl09_lnkEmail').get_attribute('href').split(':')[-1]
            link = self.driver.find_element_by_css_selector('#ctl09_lnkwebsite').get_attribute('href')
            clas = self.driver.find_element_by_css_selector('#ctl09_lblClass').text
            self.sheet.cell(i,1).value = name
            self.sheet.cell(i,2).value = email
            self.sheet.cell(i,3).value = link
            self.sheet.cell(i,4).value = clas

            self.sheet.cell(i,5).value = url
            i += 1

        self.book.save('140online.xlsx')
        i = 2
        for n in range(len(self.driver.find_elements_by_css_selector('div.span7 h3 a'))):
            self.driver.find_elements_by_css_selector('div.span7 h3 a')[n].click()
            name = self.driver.find_element_by_css_selector('#ctl09_lblCompanyName').text
            try:
                email = self.driver.find_element_by_css_selector('#ctl09_lnkEmail').get_attribute('href').split(':')[-1]
            except:
                email = None
            try:
                link = self.driver.find_element_by_css_selector('#ctl09_lnkwebsite').get_attribute('href')
            except:
                link = None
            clas = self.driver.find_element_by_css_selector('#ctl09_lblClass').text
            self.sheet.cell(i, 1).value = name
            self.sheet.cell(i, 2).value = email
            self.sheet.cell(i, 3).value = link
            self.sheet.cell(i, 4).value = clas
            self.sheet.cell(i, 5).value = url
            i += 1
            self.driver.get('http://www.140online.com/class/pages/584/%D9%85%D9%8A%D8%A7%D9%87%20%D9%85%D8%B9%D8%AF%D9%86%D9%8A%D9%87/2/')


if __name__ == '__main__':
    app = Main()


"""http://www.140online.com/class/pages/584/%D9%85%D9%8A%D8%A7%D9%87%20%D9%85%D8%B9%D8%AF%D9%86%D9%8A%D9%87/1/
https://www.yellowpages.com.eg/ar/category/%D9%85%D9%8A%D8%A7%D9%87-%D9%85%D8%B9%D8%A8%D8%A6%D8%A9-(%D9%85%D8%B9%D8%AF%D9%86%D9%8A%D8%A9)
https://eg.kompass.com/c/%D8%A7%D9%84%D9%88%D8%A7%D8%AD%D9%87-%D9%84%D8%AA%D9%86%D9%82%D9%8A%D9%87-%D9%88%D8%AA%D8%B9%D8%A8%D9%8A%D9%94%D9%87-%D8%A7%D9%84%D9%85%D9%8A%D8%A7%D9%87-%D8%A7%D9%84%D8%B7%D8%A8%D9%8A%D8%B9%D9%8A%D9%87/eg059960/"""