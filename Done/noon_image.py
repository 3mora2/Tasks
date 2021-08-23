from io import BytesIO
from time import sleep

from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from selenium import webdriver
import requests
from webdriver_manager.firefox import GeckoDriverManager
from requests_html import HTMLSession
import json

session = HTMLSession()


def headers():
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.5",
        "x-platform": "web",
        "x-cms": "v2",
        "x-mp": "noon",
        "Cache-Control": "no-cache, max-age=0, must-revalidate, no-store",
        # "x-Locale": "ar-eg",
        "x-content": "desktop",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin"
    }


class NOON:
    def __init__(self):
        # self.dict_pro = dict()
        # self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())
        self.book = Workbook()
        self.sheet = self.book.active
        i = 1
        self.sheet.cell(i, 1).value = 'SKU'
        self.sheet.cell(i, 3).value = 'Name'
        self.sheet.cell(i, 4).value = 'Price'
        self.sheet.cell(i, 5).value = 'Old Price'
        self.sheet.cell(i, 6).value = 'Brand'
        self.sheet.cell(i, 7).value = 'Rating Value'
        self.sheet.cell(i, 8).value = 'Rating Count'
        self.sheet.cell(i, 9).value = "URL"
        self.sheet.column_dimensions['B'].width = 25
        self.sheet.column_dimensions['A'].width = 12
        self.sheet.column_dimensions['C'].width = 40
        self.sheet.column_dimensions['D'].width = 8
        self.sheet.column_dimensions['E'].width = 8
        self.sheet.column_dimensions['F'].width = 20
        self.sheet.column_dimensions['G'].width = 11
        self.sheet.column_dimensions['H'].width = 11
        self.sheet.column_dimensions['I'].width = 60

        for column in range(1, 10):
            try:
                self.sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                 wrap_text=True)
            except:
                pass

    def start(self, link):
        dict_pro = dict()
        i = 2
        self.driver.get(link)
        try:
            self.driver.find_element_by_css_selector('.select-menu-btn-plp_display').click()
            sleep(1)
            self.driver.find_element_by_css_selector('li[data-value="150"]').click()
        except:
            pass
        sleep(2)
        while True:
            for element in self.driver.find_elements_by_css_selector('.productContainer'):
                code = element.find_element_by_css_selector('a').get_attribute('id').split('-')[-1]
                url = element.find_element_by_css_selector('a').get_attribute('href')
                # img = element.find_element_by_css_selector('a .grid img').get_attribute('src')
                name = element.find_element_by_css_selector('a .grid img').get_attribute('alt')
                try:
                    price = element.find_element_by_css_selector(
                        '[data-qa="product-name"]+div span.currency+strong').text
                except:
                    price = None
                try:
                    old_price = element.find_element_by_css_selector('[data-qa="product-name"]+div .oldPrice').text
                except:
                    old_price = None
                try:
                    ratingValue = element.find_element_by_css_selector('.ratingValue').text
                except:
                    ratingValue = None
                try:
                    ratingCount = element.find_element_by_css_selector('.ratingCount .count').text
                except:
                    ratingCount = None
                try:
                    element.find_element_by_css_selector('[alt="noon-express"]')
                    express = 'noon-express'
                except:
                    express = None
                d = session.get(f'https://www.noon.com/egypt-ar/{code}/p?')
                img = d.html.find('[property="og:image"]', first=True).attrs['content'].replace('tr:n-t_120/', '')
                res = session.get(img)
                image_file = BytesIO(res.content)
                img = Image(image_file)
                img.width = 200
                img.height = 270
                self.sheet.row_dimensions[i].height = 206
                self.sheet.add_image(img, f'B{i}')
                self.sheet.cell(i, 1).value = code
                self.sheet.cell(i, 3).value = name
                self.sheet.cell(i, 4).value = price
                self.sheet.cell(i, 5).value = old_price
                self.sheet.cell(i, 6).value = express
                self.sheet.cell(i, 7).value = ratingValue
                self.sheet.cell(i, 8).value = ratingCount
                self.sheet.cell(i, 9).value = url
                for column in range(i, 10):
                    try:
                        self.sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                         wrap_text=True)
                    except:
                        pass
                dict_pro[code] = {
                    'n': i,
                    'name': name,
                    'price': price,
                    'old_price': old_price,
                    'express': express,
                    'ratingValue': ratingValue,
                    'ratingCount': ratingCount,
                    'url': url,
                    'img': None
                }
                print(i)
                i += 1
            try:
                if 'false' in self.driver.find_element_by_css_selector('a[aria-label="Next page"]').get_attribute(
                        'aria-disabled'):
                    self.driver.find_element_by_css_selector('a[aria-label="Next page"]').click()
                    sleep(3)
                else:
                    break
            except:
                break
        self.book.save(link.split('/')[4] + '.xlsx')

    def start_h(self, link, file):
        list_n = []
        i = 2
        if 'limit' in link:
            n = int(link.split('limit=')[-1].split('&')[0])
            link = link.replace('limit=' + str(n), 'limit=150')
        else:
            if '?' in link:
                link += '&limit=150'
            else:
                link += '?limit=150'
        while True:
            print(link)
            r = session.get(link, headers=headers())
            if not r.ok:
                sleep(60)
                continue

            if 'page' in link:
                n = int(link.split('page=')[-1].split('&')[0])
                link = link.replace('page=' + str(n), 'page=' + str(n + 1))
            else:
                if '?' in link:
                    link += '&page=2'
                else:
                    link += '?page=2'

            data = json.loads(r.html.find('#__NEXT_DATA__', first=True).text)
            if len(data['props']['pageProps']['catalog']['hits']) == 0:
                break

            for element in data['props']['pageProps']['catalog']['hits']:
                sku = element.get('sku')
                if sku not in list_n:
                    list_n.append(sku)
                else:
                    print('- found')
                    continue
                brand = element.get('brand')
                name = element['name']
                price = element.get('price')
                sale_price = element['sale_price']
                try:
                    product_rating_value = element['product_rating']['value']
                    product_rating_count = element['product_rating']['count']
                except:
                    product_rating_count = None
                    product_rating_value = None

                image_key = element['image_key']
                url = element['url']
                img = f'https://z.nooncdn.com/products/{image_key}.jpg'
                try:
                    res = session.get(img, headers=headers())
                    image_file = BytesIO(res.content)
                    img = Image(image_file)

                    img.width = 200
                    img.height = 270

                    self.sheet.row_dimensions[i].height = 206
                    self.sheet.add_image(img, f'B{i}')
                except:
                    print('- error', img)
                self.sheet.cell(i, 1).value = sku
                self.sheet.cell(i, 3).value = name
                self.sheet.cell(i, 4).value = price if not sale_price else sale_price
                self.sheet.cell(i, 5).value = sale_price if not sale_price else price
                self.sheet.cell(i, 6).value = brand
                self.sheet.cell(i, 7).value = product_rating_value
                self.sheet.cell(i, 8).value = product_rating_count
                self.sheet.cell(i, 9).value = url
                print(i)
                i += 1

        self.book.save(str(file) + '.xlsx')


self = NOON()

self.start_h(input('- Enter Url: '), input('- Enter Filename: '))
# https://www.noon.com/egypt-ar/beauty-and-health/beauty/fragrance/eau-de-parfum?limit=150
# pyinstaller --add-data c:\users\3mora\anaconda3\envs\python3.6\lib\site-packages\pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info
