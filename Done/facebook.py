from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from requests_html import HTML
from openpyxl import Workbook
from time import sleep
import warnings

warnings.filterwarnings("ignore")


class Facebook:
    def __init__(self, data):
        ##############################
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet['C1'] = 'profile pic'
        self.sheet['A1'] = 'name'
        self.sheet['B1'] = 'comment'
        ##############################
        # number of post which you want
        mail = data[0]
        pas = data[1]
        self.link = data[3]
        self.num = data[4]
        self.ilike = data[5]
        self.ishare = data[6]
        self.icomment = data[7]
        self.idata = data[8]
        self.get_comment = data[9]

        chrome_options = Options()
        # disable notifications
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-notifications")
        # make chrom hideen
        # chrome_options.add_argument("--headless")
        # open chrom
        self.driver = webdriver.Chrome('G:\\programing\\projects\\automation\\chromedriver.exe',
                                       chrome_options=chrome_options)
        try:
            # link of ligin
            self.driver.get("https://facebook.com/login/")
        except:
            # if return error it will refresh
            self.driver.refresh()
        sleep(2)
        self.driver.refresh()
        # to make sture you need login
        if self.driver.current_url != 'https://www.facebook.com/':
            self.driver.find_element_by_id('email').send_keys(mail)
            sleep(2)
            self.driver.find_element_by_id('pass').send_keys(pas)
            sleep(2)
            self.driver.find_element_by_id('loginbutton').click()
            sleep(2)
        # start
        self.like_commont_share()
        # after it finsh it will clise
        self.driver.close()

    def like_commont_share(self):
        # open link page
        self.driver.get(self.link)
        # to select all posts
        all_posts = self.driver.find_elements_by_class_name('_427x')
        self.Make_like_comment(all_posts)
        print('tern')

    def Make_like_comment(self, post):
        p = 1
        for i in post:
            # to select part of like and commont and share
            if self.num > 0:
                bar1 = WebDriverWait(i, 10).until(ec.visibility_of_element_located((By.CLASS_NAME, "_3vuz")))
                # i.find_element_by_class_name('_3vuz')
                sleep(1)
                ########################################################
                # to select botton share
                try:
                    if self.ishare == 1:
                        share_ = WebDriverWait(i, 10).until(ec.visibility_of_element_located((By.CLASS_NAME, "_2nj7")))
                        a = share_.location
                        a['y'] = a['y'] - 100
                        self.driver.execute_script(f"window.scrollTo({a['x']}, {a['y']})")
                        share_.click()
                        sleep(5)
                        box_share = self.driver.find_elements_by_class_name('_54ng')[-1]
                        # WebDriverWait(self.driver,10).until(ec.visibility_of_all_elements_located((By.CLASS_NAME, "_54ng")))
                        sleep(1)
                        box_Item = box_share.find_elements_by_class_name('__MenuItem')
                        print(len(box_Item))
                        box_Item[0].find_element_by_tag_name('a').click()
                        print('make share')
                        sleep(4)
                except Exception as e:
                    print(e)
                #######################################################
                try:
                    if self.ilike == 1:
                        # select like botton
                        hov = WebDriverWait(bar1, 10).until(ec.visibility_of_element_located((By.CLASS_NAME, "_3l2t")))
                        # bar1.find_element_by_class_name('_3l2t')
                        # to make sure it is'n liked
                        if hov.get_attribute('aria-pressed') != 'true':
                            sleep(2)
                            # to make love
                            '''hover = ActionChains(self.driver).move_to_element(hov)
                            hover.perform()
                            sleep(2)
                            lik=self.driver.find_elements_by_class_name('_iuw')
                            lik[len(lik)-7].click()
                            print('make love')'''
                            hov.click()
                            try:
                                # some time apper notfication it will close
                                self.driver.find_element_by_xpath(
                                    '//*[@id="facebook"]/body/div[16]/div[2]/div[1]/div[1]/div/div/div/div/div[1]/div').click()
                            except:
                                pass
                        sleep(2)
                except:
                    pass
                ###############################################
                '''try:
                    if self.icomment!='0' or self.icomment!= 0:
                        #select comment botton
                        comment_c =WebDriverWait(bar1,10).until(ec.visibility_of_element_located((By.CLASS_NAME, "_42ft")))
                        comment_c.click()
                        #bar1.find_element_by_class_name('_42ft').click()
                        
                        sleep(3)
                        # where it will write
                        h =WebDriverWait(i,10).until(ec.visibility_of_element_located((By.CLASS_NAME, "notranslate")))
                        #i.find_element_by_class_name('notranslate')
                        #text comment
                        h.send_keys(self.icomment)
                        #to put enter and post
                        h.send_keys(Keys.ENTER)
                        print('comment')
                except:
                    pass'''
                ##############################################_2nj7
                # to get information about post
                try:
                    if self.idata == 1:
                        title_post = i.find_elements_by_class_name('fcg')[0].text
                        time_post = i.find_elements_by_class_name('fcg')[2].text
                        content_post = i.find_element_by_class_name('userContent').text
                        bar_likes_comment = i.find_element_by_class_name('_3vum')
                        number_likes = bar_likes_comment.find_elements_by_class_name('_81hb')[0].text
                        number_share = bar_likes_comment.find_elements_by_class_name('_42ft')[1].text
                        number_comment = bar_likes_comment.find_elements_by_class_name('_42ft')[0].text
                        print('title_post :', title_post)
                        print('time_post :', time_post)
                        print('content_post :', content_post)
                        print('number_likes :', number_likes)
                        print('number_comment', number_comment)
                        print('number_share', number_share)
                except:
                    pass

                try:
                    if self.get_comment == 1:
                        i.find_element_by_class_name('_7794').find_element_by_class_name('_4ssp').click()
                        sleep(4)
                        # html=HTML(html=self.driver.page_source)
                        box_d = i.find_element_by_class_name('_3w53')
                        comments_ = box_d.find_elements_by_class_name('_72vr')
                        # box=html.find('div._3w53')
                        # print(len(box))
                        # comments=bb.find('div._72vr')
                        for bb in comments_:
                            name_ = bb.find_element_by_class_name('_6qw4').text
                            url_ = bb.find_element_by_class_name('_6qw4').get_attribute('href')
                            comment_ = bb.find_elements_by_tag_name('span')[2].text

                            # name=bb.find('a._6qw4',first=True).text
                            # url=bb.find('a._6qw4',first=True).attrs['href']
                            # comment=bb.find('span')[2].text
                            self.save_data(p=p, data=[name_, comment_, url_])
                            print(name, comment, url)
                            p += 1
                except:
                    pass
                self.num -= 1
                self.book.save('Facebook.xlsx')
            else:
                return False

    def save_data(self, p, data):
        if True:
            self.sheet.cell(row=p + 1, column=1).value = data[0]
            self.sheet.cell(row=p + 1, column=2).value = data[1]
            self.sheet.cell(row=p + 1, column=3).value = data[2]


'''def DataBase():
    import sqlite3
    link1='https://www.facebook.com/pg/OrangeEgyptOfficial/posts/?ref=page_internal'
    link='https://www.facebook.com/pg/Vodafone.Egypt/posts/?ref=page_internal'
    db = sqlite3.connect('database.db')
    db1=db.cursor()
    db1.execute('CREATE TABLE IF NOT EXISTS TEST(mail TEXT,password TEXT,time TEXT,link TEXT,numberpost INT,like INT,shara INT,comment TEXT,getdpost INT,get_comment INT)')
    db.row_factory = sqlite3.Row
    #db1.execute('INSERT INTO TEST(time,link,numberpost,like,shara,comment,getdpost,get_comment ) VALUES(?,?,?,?,?,?,?,?)',('5-50',link,4,1,1,'vodafon',1,1))            
    
    db1.execute('select mail,password,time,link,numberpost,like,shara,comment,getdpost,get_comment from TEST')
    p=db1.fetchall()
    if len(p)==0:
        db1.execute('INSERT INTO TEST(mail,password,time,link,numberpost,like,shara,comment,getdpost,get_comment ) VALUES(?,?,?,?,?,?,?,?,?,?)',('moduhondu@gmail.com','ammar2020','5-50',link1,4,1,1,'OrangeEgyptOfficial',1,1))
        db.commit()
        db1.execute('select mail,password,time,link,numberpost,like,shara,comment,getdpost,get_comment from TEST')
        p=db1.fetchall()
    lists=[p[0][0],p[0][1],p[0][2],p[0][3],p[0][4],p[0][5],p[0][6],p[0][7],p[0][8],p[0][9]]
    db1.close()
    db.close()
    return lists'''


def Time(tim):
    import time
    tt = time.localtime()
    h1 = int(time.strftime("%H"))
    m1 = int(time.strftime("%M"))
    t = tim.split('-')
    h2 = int(t[0])
    m2 = int(t[1])
    h = h2 - h1
    m = m2 - m1
    if m < 0:
        m = m * (-1)
    if h < 0:
        h = h * (-1)
    s = h * 60 * 60 + m * 69
    return s


from openpyxl import load_workbook

path = "facebook_book.xlsx"
book = load_workbook(path)
sheet = book.active

email = sheet.cell(row=2, column=1).value
password = sheet.cell(row=2, column=2).value
time = sheet.cell(row=2, column=3).value
link = sheet.cell(row=2, column=4).value
number = sheet.cell(row=2, column=5).value
like = sheet.cell(row=2, column=6).value
share = sheet.cell(row=2, column=7).value
comment = sheet.cell(row=2, column=8).value
gpost = sheet.cell(row=2, column=9).value
gcomment = sheet.cell(row=2, column=10).value
data = [email, password, time, link, number, like, share, comment, gpost, gcomment]
print(number)
# data=DataBase()
S = Time(data[2])
print(f'it will start after {S} s')
# sleep(S)
print(data)
Facebook(data)
