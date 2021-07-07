from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from time import sleep
from datetime import datetime
import os


class SouCode:

    def __init__(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.implicitly_wait(5)
        self.login(self.driver)

    @staticmethod
    def login(driver):
        try:
            driver.get('https://core.noon.partners/en-sa/sales')
            sleep(5)
            if 'login.noon' in driver.current_url:
                driver.find_element_by_name('email').send_keys(email)
                driver.find_element_by_name('password').send_keys(password)
                driver.find_element_by_css_selector('#formContainer > button').click()
                sleep(5)
                while True:
                    if 'login.noon' not in driver.current_url:
                        driver.get('https://core.noon.partners/en-sa/sales')
                        break

        except Exception as error:
            print(error)
            print('- Login Error')

    def Get_URL(self):
        URL = []
        url = self.driver.current_url
        if all_type is False:
            url = url.replace('status=all', 'status=returned') if 'status=' in url else url + '&status=returned'
        url = url.replace('limits=20', 'limits=100') if 'limits=' in url else url+'&limits=100'
        print(url)
        self.driver.get(url)
        sleep(5)
        q = 0
        while True:
            for element in self.driver.find_elements_by_css_selector('table > tbody > tr'):

                if not all_type:
                    un = element.find_element_by_xpath('td /a[contains(@href,"/sales/")]').text.strip()
                    if un in l:
                        print('- Found')
                        continue

                URL.append(element.find_element_by_xpath('td /a[contains(@href,"/sales/")]').get_attribute('href'))
                q += 1
            print(q)
            try:
                if 'disabled' not in self.driver.find_element_by_css_selector('.next').get_attribute('class'):
                    self.driver.execute_script("document.querySelector('li.next a').click();")
                    sleep(2)
                else:
                    break
            except:
                break
        return URL

    def Get_Products(self):
        file_save = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")+' - product.xlsx'
        data = self.Get_URL()
        print(len(data))
        book = Workbook()
        sheet = book.active
        sheet.cell(1, 1).value = 'SKU'
        sheet.cell(1, 2).value = 'Price'
        sheet.cell(1, 3).value = 'New Price'
        sheet.cell(1, 5).value = 'Static'
        try:
            sheet.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except:
            pass

        cont = 2
        for url in data:
            print(url)
            u = True
            while u:
                self.driver.get(url)
                sleep(2)
                try:
                    price = WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(
                        (By.CSS_SELECTOR, 'div.spacer ~ div.value')))).text.replace('SAR', '').strip()
                    sleep(1)
                    u = False
                except:
                    sleep(90)
                    continue

                try:
                    SKU = self.driver.find_element_by_css_selector('div:nth-child(7) > div > div.attrVal').text.strip()
                    unc = self.driver.find_element_by_css_selector('div:nth-child(3) > div > div.attrVal').text
                    static = self.driver.find_element_by_css_selector('div.solid').text
                    if unc.strip() in l and not all_type:
                        print('- Found')
                        cont -= 1
                        continue
                    try:
                        new_price = float(price) + vol
                    except:
                        new_price = None
                    sheet.cell(cont, 1).value = SKU
                    sheet.cell(cont, 2).value = price
                    sheet.cell(cont, 3).value = new_price
                    sheet.cell(cont, 5).value = static
                    sheet.cell(cont, 6).value = url
                    if all_type is False:
                        sheet.cell(cont, 7).value = unc

                    print(f'{cont - 1} - {SKU} - Price : {price}')
                    sleep(1)
                except Exception as e:
                    sheet.cell(cont, 6).value = url
                    print(e)
            cont += 1
            book.save(file_save)

        self.driver.quit()


if __name__ == "__main__":
    email = input('- Enter your email : ')
    password = input('- Enter your password : ')
    vol = float(input('- Enter value (+) : '))
    type_se = input('- Enter (1 ----> All) Or (2 ----> Returned) : ')
    l = []

    if type_se == 2 or type_se.strip() == '2':
        all_type = False
        for file in os.listdir():
            if file.endswith('product.xlsx'):
                try:
                    s = load_workbook(file)
                    ss = s.active
                    for i in range(2, ss.max_row+1):
                        if ss.cell(i, 7).value is not None:
                            l.append(ss.cell(i, 7).value.strip())
                except:
                    pass

    else:
        all_type = True
    print(len(l))
    bot = SouCode()
    input('- Enter : ')
    bot.Get_Products()

# bekj.119@gmail.com
