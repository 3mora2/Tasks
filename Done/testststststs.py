from openpyxl import load_workbook

book = load_workbook('dddd.xlsx')
sheet = book.active

name = [sheet.cell(i, 2).value for i in range(1, 19280 + 1)]
print(len(name))  # 1170
d = e = nn = 0
for n in name:
    if n == '' or n == None:
        print(n)
        continue

    n = n.strip()
    n = n.replace('  ', ' ')
    if len(n) < 50:
        pp = n
        d += 1
    else:
        if n.find(',') > -1:
            if n.find(',') <= 50:

                '''if n.find(',',40) <= 50:
                    pp = n[:n.find(',',40)]
                elif n.find(',',30) <= 50:
                    pp = n[:n.find(',', 30)]
                elif n.find(',',20) <= 50:
                    pp = n[:n.find(',', 20)]
                elif n.find(',') <= 50:
                    pp = n[:n.find(',')]'''
                e += 1
            else:
                nn += 1
        else:
            nn += 1
print('ready', d, 'need', e, 'cont', nn)
