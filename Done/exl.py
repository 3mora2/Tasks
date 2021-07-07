from openpyxl import load_workbook

file_name = 'Book1.xlsx'
book = load_workbook(file_name)
sheet = book.active
max_row = sheet.max_row
max_column = sheet.max_column
# cells = []
# for item in sheet.merged_cell_ranges:
#     cells += item.bottom
#     text = item.start_cell.value
#     new = []
#     if text is None:
#         continue
#     txt = text.split(' ')
#     for t in txt:
#         if t.isalpha():
#             t = t[::-1]
#         elif "(" in t:
#             t = f"({t.replace('(', '').replace(')', '')[::-1]})"
#         elif "?" in t:
#             t = t[::-1]
#         new.append(t)
#     value = ' '.join(new[::-1])
#     print(value)
#     item.start_cell.value = value.replace('?', '')
#
# item.bottom = row, column
for row in range(1, max_row+1):
    for column in range(1, max_column + 1):
        # if (row, column) in cells:
        #     continue
        text = sheet.cell(row=row, column=column).value
        if text is None:
            continue
        new = []
        txt = text.split(' ')
        if len(txt) == 1:
            if text.find('\n') > -1:
                txt = text.split('\n')
                for t in txt:
                    t = t[::-1]
                    new.append(t)
                value = '\n'.join(new[::-1])
                sheet.cell(row=row, column=column).value = value.replace('?', '')
                continue
        for t in txt:
            if t.isalpha():
                t = t[::-1]
            elif "(" in t:
                t = f"({t.replace('(', '').replace(')', '')[::-1]})"
            elif "?" in t:
                t = t[::-1]
            new.append(t)
        value = ' '.join(new[::-1])
        sheet.cell(row=row, column=column).value = value.replace('?', '')

book.save(file_name)
