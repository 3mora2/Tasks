import openpyxl
from openpyxl.styles import Font, Color, colors
from bs4 import BeautifulSoup
import os
import urllib.request

import urllib.parse


class AppURLopener(urllib.request.FancyURLopener):
    version = "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.69 Safari/537.36"


urllib._urlopener = AppURLopener()


def Download_pics(pic_url, pic_name, rowa):
    print("[*] Downloading '" + pic_url + "<><><><><>")
    try:
        try:
            pic_name = pic_name.replace('\\', '/')
            pic_url = '/'.join(pic_url.split('/')[:-1]+[urllib.request.quote(pic_url.split('/')[-1])])
            urllib._urlopener.retrieve(pic_url, pic_name)
            print(" Done.")

        except Exception as e:
            print(e)
            f = urllib.request.urlopen(pic_url)
            htmlSource = f.read()
            soup = BeautifulSoup(htmlSource, 'html.parser')
            metaTag = soup.find_all('meta', {'property': 'og:image'})
            imgURL = metaTag[0]['content']
            fileName = pic_name
            urllib._urlopener.retrieve(imgURL, fileName)
            print(" ERROR2!")
            sh.cell(row=rowa, column=2).font = Font(color=colors.BLUE, bold=True)
            wb.save(newpath)

    except Exception as e:
        print(e)
        print(" ERROR!")
        sh.cell(row=rowa, column=2).font = Font(color=colors.BLUE, bold=True)
        wb.save(newpath)


if __name__ == '__main__':

    path = input("enter the file direction: ")
    newpath = input("the new file dir: ")
    while not os.path.exists(path):
        os.system('cls' if os.name == 'nt' else 'clear')
        path = input("[!] wrong path!!...enter the excel file path : ")

    wb = openpyxl.load_workbook(path)
    sh = wb.active
    first = input("put the first row number : ")
    last = input("and here the last row number : ")
    for s in range(int(first), int(last)):
        pic_url = sh.cell(s, 1).value
        pic_name = sh.cell(s, 2).value
        # pic_path = r"."
        pic_path = r"D:/picdown1"
        if os.name == 'nt':
            if pic_path[len(pic_path) - 1] != "\\":
                pic_path += "\\"
        else:
            if pic_path[len(pic_path) - 1] != "/":
                pic_path += "/"

        Download_pics(pic_url, pic_path + pic_name, s)
