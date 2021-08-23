from selenium import webdriver
from openpyxl import load_workbook, Workbook
from webdriver_manager.firefox import GeckoDriverManager
from pyurls import links
driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())
wb = load_workbook('qtr.xlsx')
sh = wb.active
# links = list()

for i in range(7, sh.max_row+1):
    url = sh.cell(i, 1).value
    cat = sh.cell(i, 2).value
    if url:
        driver.get(url)
        while True:
            for elem in driver.find_elements_by_css_selector(".site > h2 > a"):
                if url in links:
                    print('- found')
                    continue
                links.append((elem.get_attribute('href'), cat))
            print(links.__len__(), cat)
            try:
                driver.find_element_by_css_selector('a[title="Next"]').click()
            except:
                break

wb = Workbook()
sh = wb.active
row = 2
for link, cat in links:
    print(row)
    # sh.cell(row, column=1).value = name
    # sh.cell(row, column=2).value = phone
    sh.cell(row, column=3).value = cat
    sh.cell(row, column=4).value = link
    row += 1

wb.save('new.xlsx')
#https://www.qtr.company/section-110.html