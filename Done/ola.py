from requests_html import HTMLSession

s = HTMLSession()
from openpyxl import Workbook, load_workbook
# pip install requests-html openpyxl
book = load_workbook('ID.xlsx')
sheet = book.active
n = sheet.max_row + 1
print(n)  # 9

governorates = {'01': 'Cairo',
                '02': 'Alexandria',
                '03': 'Port Said',
                '04': 'Suez',
                '11': 'Damietta',
                '12': 'Dakahlia',
                '13': 'Ash Sharqia',
                '14': 'Kaliobeya',
                '15': 'Kafr El - Sheikh',
                '16': 'Gharbia',
                '17': 'Monoufia',
                '18': 'El Beheira',
                '19': 'Ismailia',
                '21': 'Giza',
                '22': 'Beni Suef',
                '23': 'Fayoum',
                '24': 'El Menia',
                '25': 'Assiut',
                '26': 'Sohag',
                '27': 'Qena',
                '28': 'Aswan',
                '29': 'Luxor',
                '31': 'Red Sea',
                '32': 'New Valley',
                '33': 'Matrouh',
                '34': 'North Sinai',
                '35': 'South Sinai',
                '88': 'Foreign'}
'''
book = Workbook()
sheet = book.active
# 1615619 1615969 1590517
# while True:
#     try:
#         r = s.post("https://www.azhar.eg/DesktopModules/NatigaSecWebService/WebService1.asmx/GetNatiga3ItemWithStringNational", headers={
#                                     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
#                                     "Accept": "application/json, text/javascript, */*; q=0.01",
#                                     "Accept-Language": "en-US,en;q=0.5",
#                                     "Content-Type": "application/json; charset=utf-8",
#                                     "X-Requested-With": "XMLHttpRequest",
#                                     "Sec-Fetch-Dest": "empty",
#                                     "Sec-Fetch-Mode": "cors",
#                                     "Sec-Fetch-Site": "same-origin",
#                                     'Referer': 'https://www.azhar.eg/seatsno.html'
#                                 },
#                    data="{'ItemId':30402081204085,'SeatNo':130101}"
#                    )
#         print('- DOne')
#         break
#     except:
#         pass
# import requests
#
# s = requests.session()
n = 2
# for i in range(1659169, 1659372):
for i in range(1009169, 1590608+1)[::-1]:
    r = s.post("http://natega.dostor.org/Home/Result", headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Content-Type": "application/x-www-form-urlencoded",
        "Alt-Used": "natega.dostor.org:443",
        "Upgrade-Insecure-Requests": "1"
        
    },
               data=f"seating_no={i}"
               )
    try:
        # sheet.cell(n, 1).value = i
        # sheet.cell(n, 2).value = r.html.find('#tab-list-active + span')[0].text
        # sheet.cell(n, 3).value = r.html.find('li>h1')[2].text
        # sheet.cell(n, 4).value = r.html.find('.resultItem > #tab + span')[0].text
        # sheet.cell(n, 5).value = r.html.find('.resultItem > #tab + span')[1].text
        # sheet.cell(n, 6).value = r.html.find('.resultItem > #tab + span')[2].text
        # sheet.cell(n, 7).value = r.html.find('.resultItem > #tab + span')[3].text
        # sheet.cell(n, 8).value = r.html.find('.resultItem > #tab + span')[4].text
        # n += 1
        print(r.html.find('#tab-list-active + span')[0].text, [i.text for i in r.html.find('.resultItem > #tab + span')], i)
        # print(r.html.find('.resultItem > #tab + span')[0].text, r.html.find('.resultItem > #tab + span')[2].text, r.html.find('li>h1')[2].text,
        #       r.html.find('#tab-list-active + span')[0].text, i)
    except:
        pass
'''
'''
30402081204085 
1  2 2 9  4 9  28  99999 

10 >>1
'''
other_s = 628
for y in [3, 4]:  # 3
    for m in range(1, 13):  # 4, 5
        if m in [1, 3, 5, 7, 8, 10, 12]:
            day_range = range(1, 32)
        elif m in [4, 6, 9, 11]:
            day_range = range(1, 31)
        elif m == 2:
            day_range = range(30) if y % 4 == 0 else range(29)
        else:
            day_range = range(32)

        for d in day_range:  # 6,7
            for c in governorates.keys():  # 8,9
                if other_s:
                    others = range(other_s, 1000)
                    other_s = None
                else:
                    others = range(0, 1000)
                for other in others:  # 10,11,12 # 3 03 01 01 03 445 8 6
                    for g in range(1, 10):  # 13
                        for check in range(1, 10):  # 14
                            number = f'3{y:02d}{m:02d}{d:02d}{c}{other:03d}{g}{check}'
                            print(number, number.__len__(), n)
                            s.get('https://www.azhar.eg/seatsno.html')
                            r = s.post(
                                "https://www.azhar.eg/DesktopModules/NatigaSecWebService/WebService1.asmx/GetQueryResult1",
                                headers={
                                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
                                    "Accept": "application/json, text/javascript, */*; q=0.01",
                                    "Accept-Language": "en-US,en;q=0.5",
                                    "Content-Type": "application/json; charset=utf-8",
                                    "X-Requested-With": "XMLHttpRequest",
                                    "Sec-Fetch-Dest": "empty",
                                    "Sec-Fetch-Mode": "cors",
                                    "Sec-Fetch-Site": "same-origin",
                                    'Referer': 'https://www.azhar.eg/seatsno.html'
                                },
                                data="{'ItemId':'" + number + "'}",
                            )
                            try:

                                name = r.json()['d']['Name']
                                id_ = r.json()['d']['Info1']
                                number_id = r.json()['d']['Info2']
                                print(number_id)
                                address = [r.json()['d'][f'Info{i}'] for i in range(3, 12)]
                                sheet.cell(n, 1).value = name
                                sheet.cell(n, 2).value = id_
                                sheet.cell(n, 3).value = number_id
                                sheet.cell(n, 4).value = address[0]
                                sheet.cell(n, 5).value = address[1]
                                sheet.cell(n, 6).value = address[2]
                                sheet.cell(n, 7).value = address[3]
                                sheet.cell(n, 8).value = address[4]
                                sheet.cell(n, 9).value = address[5]
                                sheet.cell(n, 10).value = address[6]
                                sheet.cell(n, 11).value = address[7]
                                sheet.cell(n, 12).value = address[8]
                                book.save('ID.xlsx')
                                n += 1
                                break
                            except:
                                pass
# book.save('test.xlsx')
book.save('ID.xlsx')
# 2267999




from requests_html import HTMLSession
s = HTMLSession()
from openpyxl import Workbook, load_workbook


def get_natiga(file_name):
    book = load_workbook(file_name)
    sheet = book.active
    for i in range(2, sheet.max_row+1):
        id_ = sheet.cell(i, 2).value
        si = sheet.cell(i, 3).value
        if id_ and si and not sheet.cell(i, 14).value:
            r = s.post(
                "https://www.azhar.eg/DesktopModules/NatigaSecWebService/WebService1.asmx/GetNatiga3ItemWithStringNational",
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "Accept-Language": "en-US,en;q=0.5",
                    "Content-Type": "application/json; charset=utf-8",
                    "X-Requested-With": "XMLHttpRequest",
                    "Sec-Fetch-Dest": "empty",
                    "Sec-Fetch-Mode": "cors",
                    "Sec-Fetch-Site": "same-origin",
                    'Referer': 'https://www.azhar.eg/seatsno.html'
                },
                data="{'ItemId':"+f"{id_},'SeatNo':{si}"+"}"
                )

            sheet.cell(i, 14).value = r.json()['d']['Name']
            sheet.cell(i, 15).value = r.json()['d']['Info2']
            sheet.cell(i, 16).value = r.json()['d']['Info3']
            sheet.cell(i, 17).value = r.json()['d']['Info4']
            sheet.cell(i, 18).value = r.json()['d']['Info5']
            sheet.cell(i, 19).value = r.json()['d']['Info6']
            sheet.cell(i, 20).value = r.json()['d']['Info8']
            print(i)

    book.save(file_name)
