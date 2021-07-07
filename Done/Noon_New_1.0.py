import os
import threading
import traceback
import uuid
from datetime import datetime
from time import sleep
from random import randrange
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import smtplib
from email.message import EmailMessage


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class NOON:
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-sa/catalog'
    URL_CAT_ADD = 'https://catalog.noon.partners/en-sa/noon-catalog'
    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'tr:nth-child(1) > td > a')
    FIRST_ELEMENT_SELECTOR_ADD = (By.CSS_SELECTOR, 'a[href*="/en-sa/noon-catalog/preview/"]')

    PRICE_MAX_SELECTOR = (
        By.XPATH,
        "//div[contains(@class, 'inputInfo') and contains(text(), 'Noon SKU Price')] / *[@class='priceRange']")
    PRICE_INPUT_SELECTOR = (By.NAME, 'price')

    QUANTITY_VALUE_SELECTOR = (By.CSS_SELECTOR, 'div.WhStockNet > div > div.WhItemValue')
    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')

    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="offerActive"]\').checked == false){document.querySelector(\'input[id="offerActive"]\').click()};'

    # SAVE_CHANGE_SELECTOR = (By.CSS_SELECTOR, 'div.transparent ~ div.solid')
    # SAVE_CHANGE_SELECTOR = (By.CSS_SELECTOR, 'div.tabsWrapper ~ div  div.solid')
    SAVE_CHANGE_SELECTOR = (By.XPATH, '//*[contains(text(),"Save Changes")]')
    SUBMIT_SELECTOR = (
        # By.XPATH, '//div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')
        # SUBMIT_SELECTOR = (
        By.XPATH,
        '//div[calss="btnWrapper"]/div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')
    #############################################
    URL_SEARCH = 'https://www.noon.com/saudi-en/'
    RATE_SELECTOR = (By.CSS_SELECTOR, 'div.detail_percentage')
    BRAND_SELECTOR = (By.CSS_SELECTOR, 'a > div.MCFyV')
    PRICE_SCRIPT = "return parseFloat(document.querySelector('.priceNow').textContent.replace('SAR',''));"
    SELLER_TEXT_SELECTOR = (By.CSS_SELECTOR, 'a.storeLink')
    EXPRESS_SELECTOR = (By.CSS_SELECTOR, 'div.estimator_right > div > img')
    ALL_OFFER_SELECTOR = (By.CSS_SELECTOR, '.allOffers')
    S_PRICE_SELECTOR = (By.CSS_SELECTOR, 'div.regular > strong')
    S_SELLER_SELECTOR = (By.CSS_SELECTOR, '.soldby > strong')
    S_EXPRESS_SELECTOR = (By.CSS_SELECTOR, 'div[direction="row"] > div > img')
    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################
    SKU_TEXT_SELECTOR = (By.CSS_SELECTOR, 'div.tabsWrapper > div')
    SKU_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[name="partner_sku"]')
    SKU_BUTTON_SELECTOR = (By.CSS_SELECTOR, '.solid')
    QUANTITY_SITE_SELECTOR = (By.CSS_SELECTOR, 'div.placeholder')
    QUANTITY_SITE_CHOSE_SELECTOR = (By.CSS_SELECTOR, 'div.active > div > div.countryLabel')
    BARCODE_INSIDE_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="text"][placeholder="Enter Barcode"]')
    BUTTON_ADD_BARCODE_INSIDE_SELECTOR = (By.CSS_SELECTOR, 'div.eachInput > div > div.solid')

    def __init__(self, file_name):
        self.file_save = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + file_name

        try:
            self.book = load_workbook(file_name)
            self.sheet = self.book.active
        except Exception as err:
            print(err)
            exit()
        self.load_file()
        try:
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
        except:
            print(traceback.print_exc())
            self.driver = webdriver.Chrome()
        self.login()

    def load_file(self):
        try:
            self.sheet[f'A1'] = 'Noon_SKU'
            self.sheet[f'B1'] = 'Price'
            self.sheet[f'C1'] = 'Quantity'
            self.sheet[f'D1'] = 'Seller Name'
            self.sheet[f'E1'] = 'Current Price'
            self.sheet[f'F1'] = 'Express'
            self.sheet[f'G1'] = 'Earn'
            self.sheet[f'H1'] = 'New price'
            self.sheet[f'I1'] = 'Status'
            self.sheet[f'J1'] = 'Brand'
            self.sheet[f'K1'] = 'Account'
            self.sheet[f'L1'] = 'Minima Price'
            self.sheet[f'M1'] = 'Change Value -'
            self.sheet[f'N1'] = 'Exception'
            self.sheet[f'O1'] = 'Quantity'
            self.sheet[f'P1'] = 'Note 1'
            self.sheet[f'Q1'] = 'Note 2'
            self.sheet.column_dimensions['A'].width = 14
            self.sheet.column_dimensions['B'].width = 15
            self.sheet.column_dimensions['C'].width = 15
            self.sheet.column_dimensions['D'].width = 15
            self.sheet.column_dimensions['E'].width = 10
            self.sheet.column_dimensions['F'].width = 14
            self.sheet.column_dimensions['G'].width = 10
            self.sheet.column_dimensions['H'].width = 11
            self.sheet.column_dimensions['I'].width = 18
            self.sheet.column_dimensions['J'].width = 30
            self.sheet.column_dimensions['K'].width = 20
            self.sheet.column_dimensions['L'].width = 16
            self.sheet.column_dimensions['M'].width = 16
            self.sheet.column_dimensions['N'].width = 16
            self.sheet.column_dimensions['O'].width = 13
            self.sheet.column_dimensions['P'].width = 85
            self.sheet.column_dimensions['Q'].width = 50
        except:
            pass

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    @staticmethod
    def price_count(price, express, earn):
        if express:
            p = (price * NoonCoinsExpress / 100)
            if p > 6:
                p = 6
            elif p < 2:
                p = 2
        else:
            p = (price * NoonCoinsNonExpr / 100)
            if p > 10:
                p = 10
            elif p < 3:
                p = 3

        min_pr = price + (price * 9 / 100) + p + (price * float(earn) / 100)
        return min_pr

    @staticmethod
    def price_eran(price, your_price, express):
        if express:
            p = (your_price * NoonCoinsExpress / 100)
            if p > 6:
                p = 6
            elif p < 2:
                p = 2
        else:
            p = (your_price * NoonCoinsNonExpr / 100)
            if p > 10:
                p = 10
            elif p < 3:
                p = 3

        earn = price - (your_price + (your_price * 9 / 100) + p)
        return earn

    def add_qu(self, q):
        sleep(1)
        try:
            WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
        except:
            try:
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_element_located(self.QUANTITY_SITE_SELECTOR))).click()
                sleep(1)
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_all_elements_located(self.QUANTITY_SITE_CHOSE_SELECTOR)))[0].click()
                sleep(1)
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
                sleep(1)
            except:
                print(traceback.print_exc())
                print('- Cant Change quantity')

    def search_bar(self, code, url, add=False):
        try:
            self.driver.get(url)
            sleep(1)
            try:
                WebDriverWait(self.driver, 10).until(
                    (ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))).send_keys(
                    code, Keys.ENTER)
            except:
                print('- Not found Bar')
                return False

            sleep(3)
            if add:
                print('- add')
                first = self.FIRST_ELEMENT_SELECTOR_ADD
                try:
                    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(first)).click()
                    sleep(1)
                    return True
                except:
                    return False
            else:
                try:
                    for i in range(3):
                        sleep(1)
                        if code in WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(
                                (By.CSS_SELECTOR, 'tbody > tr:nth-child(1) > td:nth-child(5)'))).text or \
                                code in WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(
                            (By.CSS_SELECTOR, 'tbody > tr:nth-child(1) > td:nth-child(4)'))).text or \
                                code in WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(
                            (By.CSS_SELECTOR, 'tbody > tr:nth-child(1) > td:nth-child(6)'))).text:
                            print('- found')
                            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(
                                (By.CSS_SELECTOR, 'tbody > tr:nth-child(1) > td > a'))).click()
                            return True
                    print('- Not found')
                    return False
                except:
                    return 'ADD'
        except:
            print(traceback.print_exc())
            return False

    def change_price(self, price):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a',
                                                                                         Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a',
                                                                                         Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(str(price))
            sleep(1)
        except:
            print(traceback.print_exc())
            print('- Cant Change Price')

    def change_quantity(self, i):
        sleep(1)
        try:
            content = WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_VALUE_SELECTOR))).text
            print(content)
            if int(content) < int(content_inp):
                if int(content) == 0:
                    self.sheet[f'Q{i}'] = 'Warning: Quantity is ' + str(content)
                else:
                    WebDriverWait(self.driver, 6).until(
                        (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(content_new))
                    print('Change Quantity to ', content_new)

                    self.sheet[f'Q{i}'] = 'Warning: Quantity is ' + str(content)
                    self.sheet[f'O{i}'] = 'Change Quantity to ' + str(content_new)
                try:
                    self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                    self.sheet[f'O{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                except:
                    pass
        except:
            try:
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_element_located(self.QUANTITY_SITE_SELECTOR))).click()
                sleep(1)
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_all_elements_located(self.QUANTITY_SITE_CHOSE_SELECTOR)))[0].click()
                sleep(1)
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(content_new))
                sleep(1)
            except:
                self.sheet[f'Q{i}'] = 'Cant Change quantity'
                self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                self.sheet[f'O{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                print(traceback.print_exc())
                print('- Cant Change quantity')

    def add_active(self):
        sleep(1)
        try:
            self.driver.execute_script(self.ACTIVE_SCRIPT)
        except:
            print(traceback.print_exc())
            print('Cant active')

    def add_bar(self, sk):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.BARCODE_INSIDE_INPUT_SELECTOR)
                                                 ).send_keys(Keys.CONTROL, 'a', Keys.BACKSPACE)
            WebDriverWait(self.driver, 10).until(
                ec.visibility_of_element_located(self.BARCODE_INSIDE_INPUT_SELECTOR)).send_keys(sk)
            sleep(1)
            WebDriverWait(self.driver, 10).until(
                ec.visibility_of_element_located(self.BUTTON_ADD_BARCODE_INSIDE_SELECTOR)).click()
            sleep(1)
        except:
            print(traceback.print_exc())
            print('- Cant Change Barcode')

    def add_serial(self):
        sleep(1)
        try:
            serial = \
                WebDriverWait(self.driver, 5).until(
                    ec.visibility_of_element_located(self.SKU_TEXT_SELECTOR)).text.split('\n')[-1].replace('N',
                                                                                                           '').replace(
                    'A', '') + str(randrange(9990, 999999))

            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_INPUT_SELECTOR)).send_keys(
                serial)
            sleep(3)
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_BUTTON_SELECTOR)).click()
            return serial
        except:
            print('- Exist, Continue')
            return False

    def Search(self):
        Note = None
        static = None
        self.sheet['K2'] = str(email)
        self.sheet['K4'] = str(main_seller)
        self.sheet['K3'] = str(password)
        end_cell = self.sheet.max_row
        exception_seller = [self.sheet[f'N{i}'].value for i in range(2, end_cell + 1) if
                            self.sheet[f'N{i}'].value is not None]

        group1 = [self.sheet[f'R{i}'].value for i in range(2, end_cell + 1) if self.sheet[f'R{i}'].value is not None]
        group2 = [self.sheet[f'S{i}'].value for i in range(2, end_cell + 1) if self.sheet[f'S{i}'].value is not None]
        group3 = [self.sheet[f'T{i}'].value for i in range(2, end_cell + 1) if self.sheet[f'T{i}'].value is not None]
        group4 = [self.sheet[f'U{i}'].value for i in range(2, end_cell + 1) if self.sheet[f'U{i}'].value is not None]
        group5 = [self.sheet[f'V{i}'].value for i in range(2, end_cell + 1) if self.sheet[f'V{i}'].value is not None]

        brand_exception = [self.sheet[f'W{i}'].value for i in range(2, end_cell) if
                           self.sheet[f'V{i}'].value is not None]
        brand_group = {}
        for i in range(2, end_cell + 1):
            if self.sheet[f'X{i}'].value is not None and self.sheet[f'Y{i}'].value is not None:
                brand_group[self.sheet[f'X{i}'].value] = self.sheet[f'Y{i}'].value

        thr = Thread()
        thr.start()
        for i in range(2, end_cell + 1):
            bar = self.sheet[f'A{i}'].value

            if not bar:
                continue
            print(bar)
            try:
                your_price = float(self.sheet[f'B{i}'].value)
            except:
                continue
            try:
                earn = float(self.sheet[f'C{i}'].value)
            except:
                try:
                    earn = float(self.sheet[f'C{i}'].value.replace('%', ''))
                except:
                    earn = earn_all

            url = f'{self.URL_SEARCH}{bar}/p/'
            try:
                self.driver.get(url)
            except:
                print(traceback.print_exc())
                continue

            try:
                price = float(self.driver.execute_script(self.PRICE_SCRIPT))

                seller = WebDriverWait(self.driver, 3).until(
                    (ec.visibility_of_element_located(self.SELLER_TEXT_SELECTOR))).text
                try:
                    express = WebDriverWait(self.driver, 3).until(
                        (ec.visibility_of_element_located(self.EXPRESS_SELECTOR))).get_attribute('alt')
                    if 'express' in express:
                        express_static = True
                    else:
                        express_static = False
                except:
                    express = None
                    express_static = False

                try:
                    brand = WebDriverWait(self.driver, 4).until(
                        (ec.visibility_of_element_located(self.BRAND_SELECTOR))).text
                    self.sheet[f'J{i}'] = brand
                except:
                    brand = None
                try:
                    if brand in brand_exception:
                        self.sheet[f'D{i}'] = seller
                        self.sheet[f'E{i}'] = price

                        try:
                            self.sheet[f'I{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
                            if 'express' in express:
                                self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
                        except:
                            pass

                        self.sheet[f'F{i}'] = express
                        self.sheet[f'I{i}'] = 'Exception Brand'
                        print(f'- {i} - {bar} > seller : {seller}, price : {price}, static : Exception Brand')
                        continue
                except:
                    pass

                if seller in group1:
                    average = float(group1[0])
                elif seller in group2:
                    average = float(group2[0])
                elif seller in group3:
                    average = float(group3[0])
                elif seller in group4:
                    average = float(group4[0])
                elif seller in group5:
                    average = float(group5[0])

                elif brand in brand_group.keys():
                    average = float(brand_group[brand])
                else:
                    average = gaverage

                if main_seller in seller:
                    try:
                        Offer_number = len(WebDriverWait(self.driver, 1).until(
                            (ec.visibility_of_all_elements_located(self.ALL_OFFER_SELECTOR))))
                    except:
                        Offer_number = 0

                    if Offer_number == 0:
                        static = 'main_seller'
                        new_price = None
                        yearn = self.price_eran(price=price, your_price=your_price, express=express_static)
                        self.sheet[f'G{i}'] = yearn
                        if yearn < 0:
                            new_price = self.price_count(your_price, express, earn)
                            self.sheet[f'L{i}'].value = new_price
                        print(f'- You are Main, Earn = {yearn}')
                    else:
                        min_pr = self.price_count(your_price, express_static, earn)
                        WebDriverWait(self.driver, 2).until(
                            (ec.visibility_of_element_located(self.ALL_OFFER_SELECTOR))).click()
                        sleep(3)
                        prices = [float(p.text) for p in
                                  self.driver.find_elements_by_css_selector('.dPvBIm [data-qa="productPrice"]')[1:]]
                        prices.sort()
                        seller_price = prices[0]

                        new_price = float(seller_price) - average
                        try:
                            self.sheet[f'M{i}'] = average
                        except:
                            pass
                        if new_price < price:
                            if price < min_pr:
                                new_price = min_pr
                                static = 'Minimum price'
                            else:
                                static = 'main_seller'
                                new_price = None

                        elif new_price < min_pr:
                            new_price = min_pr
                            static = 'Minimum price'

                elif seller in exception_seller:
                    static = 'Exception'
                    print('Exception')
                    new_price = None
                else:
                    try:
                        express_static = self.driver.execute_script('''
                        for (r of document.querySelectorAll('.dPvBIm > div')) {
                            name = r.querySelector('.offerSname').textContent
                            express = r.querySelector('.emGBiz > img').attributes["alt"].textContent
                            if ('هداف' == name) {
                            if ('fulfilment_express' == express) {
                            return true; 
                        }
                        }
                        }
                        return false;
                        ''')
                    except:
                        print(traceback.print_exc())

                    min_pr = self.price_count(your_price, express_static, earn)
                    self.sheet[f'L{i}'].value = min_pr
                    if price < min_pr:
                        new_price = min_pr
                        static = 'Minimum price'
                    else:
                        new_price = float(price) - average
                        if new_price < min_pr:
                            new_price = min_pr
                            static = 'Minimum price'
                        else:
                            try:
                                self.sheet[f'M{i}'] = average
                            except:
                                pass

            except Exception as e:
                print(e)
                seller = price = express = new_price = None
                self.sheet[f'P{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                static = "Not found"

            if new_price is not None:
                res = self.search_bar(code=bar, url=self.URL_CAT)
                if res == 'ADD':
                    res = self.search_bar(code=bar, url=self.URL_CAT_ADD, add=True)
                    if not res:
                        continue

                    serial = self.add_serial()
                    if serial:
                        self.change_price(new_price)
                        sleep(1)
                        self.add_active()
                        sleep(1)
                        self.add_qu(content_new)
                        sleep(1)
                        self.add_bar(serial)
                        sleep(1)
                        try:
                            WebDriverWait(self.driver, 10).until(
                                (ec.element_to_be_clickable(self.SAVE_CHANGE_SELECTOR))).click()
                            sleep(1)
                            WebDriverWait(self.driver, 10).until(
                                (ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                            sleep(1)
                        except:
                            print(traceback.print_exc())
                            print('Cant Save')

                elif res:
                    try:
                        pp = WebDriverWait(self.driver, 3).until((ec.visibility_of_element_located(
                            self.PRICE_MAX_SELECTOR))).text.replace('SAR', '').replace(',', '').strip().split('-')
                        p_n = float(pp[0].strip())
                        p_m = float(pp[1].strip())

                        print(f'- Max noon : {p_m}, Minimum Noon : {p_n}, Your New Price {new_price}')

                        # if new price small than noon mine, new price = noon min price
                        if float(new_price) < float(p_n):
                            new_price = p_n
                            print(f'- Edit To Noon Min {new_price}')
                            static = 'Noon Minimum price'

                        # if new price > noon max price, new price = noon max price
                        elif float(new_price) > float(p_m):
                            new_price = p_m
                            print(f'- Edit To Noon Max {new_price}')
                            static = 'Noon Maximum price'

                    except:
                        print(traceback.print_exc())

                    self.change_price(new_price)
                    sleep(1)
                    # self.add_active()
                    self.change_quantity(i)
                    sleep(2)
                    try:
                        WebDriverWait(self.driver, 10).until(
                            (ec.element_to_be_clickable(self.SAVE_CHANGE_SELECTOR))).click()
                        sleep(1)
                        WebDriverWait(self.driver, 10).until(
                            (ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                        sleep(1)
                    except:
                        self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                        self.sheet[f'Q{i}'] = 'Cant Save'
                        print('Cant Save')
                        print(traceback.print_exc())

            for column in range(1, 18):
                try:
                    self.sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                     wrap_text=True)
                except:
                    pass
            self.sheet[f'D{i}'] = seller
            self.sheet[f'E{i}'] = price
            try:
                if 'express' in express:
                    self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
            except:
                pass

            self.sheet[f'F{i}'] = express
            self.sheet[f'H{i}'] = new_price
            if static == 'Minimum price':
                self.sheet[f'H{i}'].fill = PatternFill(fill_type='solid', fgColor='FF9999')
            self.sheet[f'I{i}'] = static
            self.sheet[f'P{i}'] = Note

            self.book.save(self.file_save)
            print(f'- {i} - {bar} > seller : {seller}, price : {price}, new price : {new_price}')
            if thr.p == 'pause' or thr.p == 'p':
                start = input('(Type start or s and press Enter)>>>')
                while True:
                    if start == 'start' or start == 's':
                        thr = Thread()
                        thr.start()
                        break
                    elif start == 'exit' or start == 'e':
                        self.driver.quit()
                        exit()
                        break
                    else:
                        start = input()
            elif thr.p == 'exit' or thr.p == 'e':
                self.driver.quit()
                break

        self.driver.quit()

        if SENDER_EMAIL:
            print('- send excel')
            send_mail_with_excel(Recipient_email1, topic1, subject1, self.file_save)
            send_mail_with_excel(Recipient_email2, topic2, subject2, self.file_save)


def add_mac():
    try:
        os.remove(str(uuid.getnode()))
    except:
        pass
    try:
        with open(str(uuid.getnode()), 'w+') as fil:
            fil.write(str(uuid.getnode()))
        os.popen('attrib +h ' + str(uuid.getnode()))
        return True
    except:
        print('delete old file')
        return False


def check_mac():
    try:
        with open(str(uuid.getnode()), 'r') as fil:
            text = fil.readlines()
        if str(uuid.getnode()) in text:
            return True
        else:
            print('ref')
            return False
    except FileNotFoundError:
        print('ref')
        return False


def send_mail_with_excel(recipient_email, subject, content, excel_file):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = SENDER_EMAIL
    msg['To'] = recipient_email
    msg.set_content(content)

    with open(excel_file, 'rb') as f:
        file_data = f.read()
    msg.add_attachment(file_data, maintype="application", subtype="xlsx", filename=excel_file)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(SENDER_EMAIL, APP_PASSWORD)
        smtp.send_message(msg)


if __name__ == "__main__":
    file = None
    pr = check_mac()
    while True:
        try:
            file = input('- Enter file name : ').strip()
            load_workbook(file).close()
            break
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)
        if file == 'scrt*python*add*mac*address':
            pr = add_mac()

    try:
        with open('data.json') as f:
            text = f.read()
        text = json.loads(text)
        NoonCoinsExpress = float(text['NoonCoinsExpress'])
        NoonCoinsNonExpr = float(text['NoonCoinsNonExpr'])
        SENDER_EMAIL = text['SENDER_EMAIL']
        APP_PASSWORD = text['APP_PASSWORD']
        Recipient_email1 = text['Recipient_email1']
        Recipient_email2 = text['Recipient_email2']
        topic1 = text['topic1']
        topic2 = text['topic2']
        subject1 = text['subject1']
        subject2 = text['subject2']
    except Exception as e:
        print(e)
        print('- Cant Found Json file')
        NoonCoinsExpress = 6
        NoonCoinsNonExpr = 10
        SENDER_EMAIL = None
        APP_PASSWORD = None
        Recipient_email1 = None
        Recipient_email2 = None

    main_seller = input('- Main seller : ')
    email = input('- Enter your email : ')
    password = input('- Enter your password : ')
    # main_seller = 'Be beautiful'  # input('- Main seller : ')
    # email = 'bekj.119@gmail.com'  # input('- Enter your email : ')
    # password = 'bekj.119'  # input('- Enter your password : ')
    gaverage = float(input('- Enter value (-) '))
    content_inp = int(input('- Enter minimum quantity : '))
    content_new = int(input('- Enter new quantity : '))
    # content_inp = 1  # int(input('- Enter minimum quantity : '))
    # content_new = 5  # int(input('- Enter new quantity : '))
    earn_all = float(input('- Enter Earn : '))
    if pr:
        while True:
            my_bot = NOON(file_name=file)
            my_bot.Search()
'''
for (r of document.querySelectorAll('.dPvBIm > div')) {
	name = r.querySelector('.offerSname').textContent
    express = r.querySelector('.emGBiz > img').attributes["alt"].textContent
	if ('اهداف' == name) {
    if ('fulfilment_express' == express) {
 	console.log(`${name}: ${express}`); 
}
}
}
'''
