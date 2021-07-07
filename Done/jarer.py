from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from requests_html import HTMLSession
s = HTMLSession()
# book = Workbook()
i = 1
path = 'printer-inks-toners.xlsx'
book = load_workbook(path)
sheet = book.active
sheet[f'A{i}'].value = 'Title'
sheet[f'B{i}'].value = 'Bender'
sheet[f'C{i}'].value = 'Price'
sheet[f'D{i}'].value = 'Category1'
sheet[f'E{i}'].value = 'Category2'
sheet[f'F{i}'].value = 'Category3'
sheet[f'G{i}'].value = 'Category4'
sheet[f'H{i}'].value = 'Small des'
sheet[f'I{i}'].value = 'Desk'
sheet[f'J{i}'].value = 'Title'
sheet[f'K{i}'].value = 'price'
sheet[f'L{i}'].value = 'category1'
sheet[f'M{i}'].value = 'category2'
sheet[f'N{i}'].value = 'category3'
sheet[f'O{i}'].value = 'category4'
sheet[f'P{i}'].value = 'desk1'
sheet[f'Q{i}'].value = 'desk2'
sheet[f'R{i}'].value = 'img'


i = 212

end_cell = sheet.max_row
links = [sheet[f'S{o}'].value for o in range(i, end_cell+1) if
         sheet[f'S{o}'].value is not None]
print(len(links))
# for ur in links:
#     r = s.get(ur)
#     sleep(3)
#     title = r.html.find(' div.product-title.product-title--main > h1.product-title__title')[0].text
#     try:
#         bender = r.html.find(
#             ' div.product-title.product-title--main > div.product-title__brand')[0].text
#     except:
#         bender = None
#     price = r.html.find('div.price')[0].text if r.html.find('div.price')[0].text != '' else r.html.find('div.price')[1].text
#     print(price)
#     cat1 = r.html.find('div.main-container > div > div > ul > li > a')[1].text
#     try:
#         daman = r.html.find('#card-warranty')[0].text
#     except:
#         daman = None
#     try:
#         cat2 = r.html.find('div.main-container > div > div > ul > li > a')[2].text
#     except:
#         cat2 = None
#     try:
#         cat3 = r.html.find('div.main-container > div > div > ul > li > a')[3].text
#     except:
#         cat3 = None
#     try:
#         cat4 = r.html.find('div.main-container > div > div > ul > li > a')[4].text
#     except:
#         cat4 = None
#
#     img = ' ,\n'.join([ele.attrs['data-src'] for ele in r.html.find('#gallery-id  div > div  img')])
#     small_des = r.html.find(' div.product-title.product-title--main > p.product-title__info')[0].text.replace('عرض المزيد','')
#     des = r.html.find('#card-specifications > div > table > tbody')[0].text
#     ############################################################################################
#     r = s.get(ur.replace('www.jarir.com/', 'www.jarir.com/sa-en/'))
#     sleep(3)
#     title_en = r.html.find(' div.product-title.product-title--main > h1.product-title__title')[0].text
#     price_en = r.html.find('div.price')[0].text if r.html.find('div.price')[0].text != '' else r.html.find('div.price')[1].text
#     cat1_en = r.html.find('div.main-container > div > div > ul > li > a')[1].text
#     try:
#         cat2_en = r.html.find('div.main-container > div > div > ul > li > a')[2].text
#     except:
#         cat2_en = None
#     try:
#         cat3_en = r.html.find('div.main-container > div > div > ul > li > a')[3].text
#     except:
#         cat3_en = None
#     try:
#         cat4_en = r.html.find('div.main-container > div > div > ul > li > a')[4].text
#     except:
#         cat4_en = None
#     try:
#         daman_en = r.html.find('#card-warranty')[0].text
#     except:
#         daman_en = None
#     small_des_en = r.html.find(' div.product-title.product-title--main > p.product-title__info')[0].text.replace('Show More','')
#     try:
#         des_en = r.html.find('#card-specifications > div > table > tbody')[0].text
#     except:
#         des_en = None
#     sheet[f'A{i}'].value = title
#     sheet[f'B{i}'].value = bender
#     sheet[f'C{i}'].value = price.replace('.', '')
#     sheet[f'D{i}'].value = cat1
#     sheet[f'E{i}'].value = cat2
#     sheet[f'F{i}'].value = cat3
#     sheet[f'G{i}'].value = cat4
#     sheet[f'H{i}'].value = small_des
#     sheet[f'I{i}'].value = des
#     sheet[f'J{i}'].value = title_en
#     sheet[f'K{i}'].value = price_en.replace('.', '')
#     sheet[f'L{i}'].value = cat1_en
#     sheet[f'M{i}'].value = cat2_en
#     sheet[f'N{i}'].value = cat3_en
#     sheet[f'O{i}'].value = cat4_en
#     sheet[f'P{i}'].value = small_des_en
#     sheet[f'Q{i}'].value = des_en
#     sheet[f'R{i}'].value = img
#     sheet[f'T{i}'].value = daman
#     sheet[f'U{i}'].value = daman_en
#     i += 1
#     book.save(path)
#     print(i, title_en,)

driver = webdriver.Chrome()
# driver.get('https://www.jarir.com/office-supplies/paper.html')
# driver.get('https://www.jarir.com/office-supplies/pens-and-markers/roco-pens-345434.html')
for ur in links:

    driver.get(ur)
    sleep(2)
    title = driver.find_element_by_css_selector(' div.product-title.product-title--main > h1.product-title__title').text
    try:
        bender = driver.find_element_by_css_selector(
            ' div.product-title.product-title--main > div.product-title__brand').text
    except:
        bender = None
    try:
        cat1 = driver.find_elements_by_css_selector('div.main-container > div > div > ul > li > a')[1].text
    except:
        cat1 = None
    try:
        cat2 = driver.find_elements_by_css_selector('div.main-container > div > div > ul > li > a')[2].text
    except:
        cat2 = None
    try:
        cat3 = driver.find_elements_by_css_selector('div.main-container > div > div > ul > li > a')[3].text
    except:
        cat3 = None
    try:
        cat4 = driver.find_elements_by_css_selector('div.main-container > div > div > ul > li > a')[4].text
    except:
        cat4 = None
    sleep(1)
    img = '\n'.join([ele.get_attribute('src') for ele in driver.find_elements_by_css_selector(
        '#gallery-id > div > div > div.slick-slide > div > div > div > img') if ele.get_attribute('src') is not None])
    small_des = driver.find_element_by_css_selector(' div.product-title.product-title--main > p.product-title__info').text.split('\n')[0]
    sleep(1)
    try:
        driver.find_element_by_css_selector('#card-specifications > a.card__show.card__show--more.link.link--icon').click()
        des = driver.find_element_by_css_selector('#card-specifications > div > table > tbody').text.replace('\x1a', '')
    except:
        des = None
    try:
        daman = driver.find_elements_by_css_selector('#card-warranty')[0].text
    except:
        daman = None
    if driver.find_elements_by_css_selector('div.price')[0].text != '':
        price = driver.find_elements_by_css_selector('div.price')[0].text
    elif driver.find_elements_by_css_selector('div.price')[1].text != '':
        price = driver.find_elements_by_css_selector('div.price')[1].text
    else:
        price = driver.find_elements_by_css_selector('div.price')[2].text
    print(price)
    ############################################################################################
    driver.get(ur.replace('www.jarir.com/', 'www.jarir.com/sa-en/'))
    sleep(2)
    title_en = driver.find_element_by_css_selector(' div.product-title.product-title--main > h1.product-title__title').text
    sleep(1)

    # price_en = driver.find_elements_by_css_selector('div.price')[0].text if driver.find_elements_by_css_selector('div.price')[0].text != '' else driver.find_elements_by_css_selector('div.price')[1].text
    try:
        bender_en = driver.find_element_by_css_selector(
            ' div.product-title.product-title--main > div.product-title__brand').text
    except:
        bender_en = None
    try:
        cat1_en = driver.find_elements_by_css_selector('div.main-container > div > div > ul > li > a')[1].text
    except:
        cat1_en = None
    try:
        cat2_en = driver.find_elements_by_css_selector('div.main-container > div > div > ul > li > a')[2].text
    except:
        cat2_en = None
    try:
        cat3_en = driver.find_elements_by_css_selector('div.main-container > div > div > ul > li > a')[3].text
    except:
        cat3_en = None
    try:
        cat4_en = driver.find_elements_by_css_selector('div.main-container > div > div > ul > li > a')[4].text
    except:
        cat4_en = None
    small_des_en = driver.find_element_by_css_selector(' div.product-title.product-title--main > p.product-title__info').text.split('\n')[0]
    sleep(1)
    try:
        driver.find_element_by_css_selector('#card-specifications > a.card__show.card__show--more.link.link--icon').click()
        des_en = driver.find_element_by_css_selector('#card-specifications > div > table > tbody').text.replace('\x1a', '')
    except:
        des_en = None
    try:
        daman_en = driver.find_elements_by_css_selector('#card-warranty')[0].text
    except:
        daman_en = None
    if driver.find_elements_by_css_selector('div.price')[0].text != '':
        price_en = driver.find_elements_by_css_selector('div.price')[0].text
    elif driver.find_elements_by_css_selector('div.price')[1].text != '':
        price_en = driver.find_elements_by_css_selector('div.price')[1].text
    else:
        price_en = driver.find_elements_by_css_selector('div.price')[2].text
    sheet[f'A{i}'].value = title
    sheet[f'B{i}'].value = bender
    sheet[f'C{i}'].value = price
    sheet[f'D{i}'].value = cat1
    sheet[f'E{i}'].value = cat2
    sheet[f'F{i}'].value = cat3
    sheet[f'G{i}'].value = cat4
    sheet[f'H{i}'].value = small_des
    sheet[f'I{i}'].value = des
    sheet[f'J{i}'].value = title_en
    sheet[f'K{i}'].value = price_en.replace('.', '')
    sheet[f'L{i}'].value = cat1_en
    sheet[f'M{i}'].value = cat2_en
    sheet[f'N{i}'].value = cat3_en
    sheet[f'O{i}'].value = cat4_en
    sheet[f'P{i}'].value = small_des_en
    sheet[f'Q{i}'].value = des_en
    sheet[f'R{i}'].value = img
    sheet[f'T{i}'].value = daman
    sheet[f'U{i}'].value = daman_en
    sheet[f'W{i}'].value = bender_en
    print(i, title_en)
    i += 1
    book.save(path)


# number_sanf = [san.find_element_by_css_selector('td.table__item').text for san in driver.find_elements_by_css_selector('#card-specifications > div > table > tbody > tr') if 'رقم الصنف' in san.text][0]
# number_p = [san.find_element_by_css_selector('td.table__item').text for san in driver.find_elements_by_css_selector('#card-specifications > div > table > tbody > tr') if 'رقم المنتج' in san.text][0]
# color = [san.find_element_by_css_selector('td.table__item').text for san in driver.find_elements_by_css_selector('#card-specifications > div > table > tbody > tr') if 'لون' in san.text][0]
# size = [san.find_element_by_css_selector('td.table__item').text for san in driver.find_elements_by_css_selector('#card-specifications > div > table > tbody > tr') if 'مقاس' in san.text][0]
'''
try:
    for el in driver.find_elements_by_css_selector('ul > li > div > h3 > a'):
        u = el.get_attribute('href')
        if u not in links:
            links.append(u)
        print(len(links))
except:
    pass    
'''


'''
try:
    k = 2
    book = Workbook()
    sheet = book.active
    links = []
    while True:
        sleep(2)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        sleep(1)
        
        for el in driver.find_elements_by_css_selector('ul > li > div > h3 > a'):
            u = el.get_attribute('href')
            if u not in links:
                links.append(u)
                print(len(links))
        driver.find_elements_by_css_selector('.amscroll-page.loading')[-1].click()
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    
except Exception as e:
    print(e)
try:
    for i in links:
        sheet[f'S{k}'].value = i
        k += 1
    book.save('smart-home-devices.xlsx')
except Exception as e:
    print(e)

'''