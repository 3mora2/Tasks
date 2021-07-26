from openpyxl import load_workbook
from io import BytesIO
from openpyxl.drawing.image import Image
from requests_html import HTMLSession

session = HTMLSession()


def headers(location):
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.5",
        "x-platform": "web",
        "x-cms": "v2",
        "x-mp": "noon",
        "Cache-Control": "no-cache, max-age=0, must-revalidate, no-store",
        "x-Locale": location,
        "x-content": "desktop",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin"
    }


while True:
    try:
        path = input('- Enter file name : ').strip()
        book = load_workbook(path)
        break
    except FileNotFoundError:
        print('- No such file or directory')
    except Exception as e:
        print(e)

sheet = book.active
sheet.column_dimensions['C'].width = 24
dict_data = dict()
dict_error = dict()
for i in range(2, sheet.max_row + 1):
    if sheet.cell(i, 1).value is not None:
        dict_data[sheet.cell(i, 1).value] = None

print('- Get URLs')
for loc in ["ar-eg", "ar-ae", "ar-sa"]:
    print(loc)
    nl = []
    for s in dict_data.keys():
        if dict_data[s] is not None:
            continue

        nl.append(s)
        if len(nl) == 50:
            sku = ','.join(nl)
            d = session.get(f"https://www.noon.com/_svc/catalog/api/search?sku={sku}",
                            headers=headers(loc))
            if d.ok:
                for v, element in enumerate(d.json()['hits']):
                    dict_data[element['sku']] = f'https://z.nooncdn.com/products/{element["image_key"]}.jpg'

            nl = []

for s in dict_data.keys():
    if dict_data[s] is not None:
        continue
    d = session.get(f'https://www.noon.com/egypt-ar/{s}/p?')
    try:
        dict_data[s] = d.html.find('[property="og:image"]', first=True).attrs['content'].replace('tr:n-t_120/', '')
        print(f'- {s} Wait.....', end='\r')
    except:
        pass

for i in range(2, sheet.max_row + 1):
    if sheet.cell(i, 1).value is None:
        continue
    main_img = dict_data[sheet.cell(i, 1).value]
    print(i, main_img, end='\r')
    if main_img:
        res = session.get(main_img)
        image_file = BytesIO(res.content)
        img = Image(image_file)

        img.width = 200
        img.height = 270

        sheet.row_dimensions[i].height = 206
        sheet.add_image(img, f'C{i}')

book.save(path.split('.')[0] + '_IMG.xlsx')
# pyinstaller --add-data C:/Users/3mora/AppData/Local/Programs/Python/Python39/Lib/site-packages/pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info

# for i in range(2, sheet.max_row+1):
#     if sheet.cell(i, 1).value is None:
#         continue
#     for loc in ["ar-ae", "ar-eg", "ar-sa"]:
#         d = s.get("https://www.noon.com/_svc/catalog/api/u/search?q=" + sheet.cell(i, 1).value,
#                   headers=headers(loc))
#         if d.ok:
#             try:
#                 element = d.json()['hits'][0]
#             except:
#                 continue
#
#             main_img = f'https://z.nooncdn.com/products/{element["image_key"]}.jpg'
#             print(i, main_img)
#             res = s.get(main_img)
#             image_file = BytesIO(res.content)
#             img = Image(image_file)
#             img.width = 117
#             img.height = 95
#             sheet.row_dimensions[i].height = 72
#             sheet.add_image(img, f'C{i}')
#             break
#
# book.save(path.split('.')[0] + '_IMG.xlsx')


# n = 2
# l = 2
# dict_data = dict()
# dict_error = dict()
# while True:
#
#     sku = ','.join([sheet.cell(i, 1).value for i in range(n, n + 50) if sheet.cell(i, 1).value is not None])
#     d = s.get("https://www.noon.com/_svc/catalog/api/search?sku=" + sku,
#               headers=headers)
#     if d.ok:
#         n += 50
#         for v, element in enumerate(d.json()['hits']):
#             main_img = f'https://z.nooncdn.com/products/{element["image_key"]}.jpg'
#             print(l, main_img)
#             res = s.get(main_img)
#             image_file = BytesIO(res.content)
#             img = Image(image_file)
#             img.width = 117
#             img.height = 95
#             sheet.row_dimensions[l].height = 72
#             sheet.add_image(img, f'C{l}')
#             l += 1
#     if n - sheet.max_row + 1 > 0:
#         break

# path = r"C:\Users\3mora\Downloads\pics.xlsx"
# book = load_workbook(path)
# sheet = book.active
# for i in range(2, sheet.max_row + 1):
#     if sheet.cell(i, 1).value is None:
#         continue
#     print(i, sheet.cell(i, 1).value)
#     d = session.get(f'https://www.noon.com/egypt-ar/{sheet.cell(i, 1).value}/p')
#     # headers=headers("ar-eg"))
#     name = d.html.find('h1', first=True).text
#     dsk = d.html.find('[name="TabArea"]~div', first=True).text
#     sheet.cell(i, 5).value = name
#     sheet.cell(i, 6).value = dsk
#
# book.save(path.split('.')[0] + 'des.xlsx')
# 'N29342295A'
