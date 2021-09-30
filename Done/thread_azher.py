# pip install requests-html openpyxl
from time import sleep

from requests_html import HTMLSession
from threading import Thread

s = HTMLSession()

from openpyxl import Workbook, load_workbook
full_book = load_workbook('full.xlsx')
full_sheet = full_book.active


def start_(brvx, others):
    count = 0
    file_name = f'{brvx}-ID.xlsx'
    try:
        book = load_workbook(file_name)
    except:
        book = Workbook()
    sheet = book.active
    n = sheet.max_row + 1
    for other in others:  # 10,11,12 # 3 03 01 01 01 628 2 2
        for g in range(1, 10):  # 13
            for check in range(1, 10):  # 14
                count += 1
                number = f'{brvx}{other:03d}{g}{check}'
                s.get('https://www.azhar.eg/seatsno.html')
                r = s.post(
                    "https://www.azhar.eg/DesktopModules/NatigaSecWebService/WebService1.asmx/GetQueryResult1",
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
                        "Accept": "application/json, text/javascript, */*; q=0.01",
                        "Accept-Language": "en-US,en;q=0.5",
                        "Content-Type": "application/json; charset=utf-8",
                        "X-Requested-With": "XMLHttpRequest",
                        "Sec-Fetch-Dest": "empty",
                        "Sec-Fetch-Mode": "cors",
                        "Sec-Fetch-Site": "same-origin",
                        'Referer': 'https://www.azhar.eg/seatsno.html'
                    },
                    data="{'ItemId':'" + number + "'}",
                )
                try:

                    name = r.json()['d']['Name']
                    id_ = r.json()['d']['Info1']
                    number_id = r.json()['d']['Info2']
                    print(count, brvx, number, number.__len__(), n, number_id)
                    address = [r.json()['d'][f'Info{i}'] for i in range(3, 12)]
                    sheet.cell(n, 1).value = name
                    sheet.cell(n, 2).value = id_
                    sheet.cell(n, 3).value = number_id
                    sheet.cell(n, 4).value = address[0]
                    sheet.cell(n, 5).value = address[1]
                    sheet.cell(n, 6).value = address[2]
                    sheet.cell(n, 7).value = address[3]
                    sheet.cell(n, 8).value = address[4]
                    sheet.cell(n, 9).value = address[5]
                    sheet.cell(n, 10).value = address[6]
                    sheet.cell(n, 11).value = address[7]
                    sheet.cell(n, 12).value = address[8]
                    book.save(file_name)
                    n += 1
                    break
                except:
                    pass
            print(count, brvx, n)
            print('*'*100)


threads = []
for i in range(1, full_sheet.max_row+1):
    brvx = full_sheet.cell(i, 1).value
    if brvx:
        others = range(0, 1000)

        process = Thread(target=start_, args=(brvx, others,))
        process.start()
        threads.append(process)
        print('start', i)
        sleep(3)

