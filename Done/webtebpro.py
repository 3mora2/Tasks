from openpyxl import load_workbook, Workbook
from requests_html import HTMLSession
from time import sleep
# https://www.webtebpro.com/5bb45936bfaf560bd0adb78f_%D8%AF%D9%83%D8%AA%D9%88%D8%B1-%D8%B3%D9%85%D9%8A%D8%B1-%D9%85%D8%AD%D9%85%D8%AF-%D8%B2%D8%A7%D9%87%D8%AF
book = load_workbook('webtebpro.xlsx')
sheet = book.active

s = HTMLSession()
# links = []
# for i in range(1, 558):
#     r = s.get(f'https://www.webtebpro.com/doctors/search?pagenumber={i}')
#     for ele in r.html.find('#endless > div > div > a'):
#         links.append(ele.attrs['href'])
#         sheet.cell(i, 25).value = ele.attrs['href']
#         print(len(links))
# book.save('webtebpro.xlsx')
# ii = 1
# sheet.cell(ii, 1).value = 'الجنس'
# sheet.cell(ii, 2).value = 'الأسم'
# sheet.cell(ii, 3).value = 'التخصص'
# sheet.cell(ii, 4).value = 'الموقع'
# sheet.cell(ii, 5).value = 'سيرة ذاتية'
# sheet.cell(ii, 6).value = 'الموقع الإلكتروني'
# sheet.cell(ii, 7).value = 'ملفات التواصل الاجتماعي'
# sheet.cell(ii, 8).value = 'واتس'
# sheet.cell(ii, 9).value = 'رقم الهاتف'
# sheet.cell(ii, 10).value = 'البريد الإلكتروني'
# sheet.cell(ii, 11).value = 'العيادات'
# sheet.cell(ii, 12).value = 'الموقع'
# sheet.cell(ii, 13).value = 'رقم هاتف العيادة'
# sheet.cell(ii, 14).value = 'رقم الفاكس'
# sheet.cell(ii, 15).value = 'ايام وساعات العمل'
# sheet.cell(ii, 16).value = 'المنشورات'
# sheet.cell(ii, 17).value = 'الخدمات الطبية'
# sheet.cell(ii, 18).value = 'الزمالة'
# sheet.cell(ii, 19).value = 'التعليم'
# sheet.cell(ii, 20).value = 'البورد'
# sheet.cell(ii, 21).value = 'الجوائز'
# sheet.cell(ii, 22).value = 'ملاحظات'
# sheet.cell(ii, 23).value = 'اللغات التي أتحدثها'
# sheet.cell(ii, 24).value = 'التأمين الصحي المقبول'
# sheet.cell(ii, 26).value = 'الصورة'
for ii in range(2, sheet.max_row+1):
    # if ii == 300:
    #     break
    if sheet.cell(ii, 2).value is not None:
        continue
    url = sheet.cell(ii, 25).value
    print(ii)
    try:
        r = s.get(url)
        prefix = r.html.find('.pro-profile-prefix', first=True).text
        name = r.html.find('.pro-profile-name', first=True).text
        print(ii, '-', name)
        img = r.html.find(' div.pro-profile-img > img', first=True).attrs['src']
        speciality = '\n'.join([s.text for s in r.html.find('.pro-profile-speciality')])
        place = r.html.find('#About p.doctor-txt', first=True).text
        try:
            cv = r.html.find('#About > div:nth-child(3) > p:nth-child(2)', first=True).text
        except:
            cv = None
        try:
            web = r.html.find('#About > div:nth-child(4) > p.doctor-txt > a', first=True).attrs['href']
        except:
            web = None
        try:
            info_social = '\n'.join([f"{inf.attrs['data-track-label']} : {inf.attrs['href']} " for inf in r.html.find('#About div.doctor-card-info-social a')])
        except:
            info_social = None
        # facebook = r.html.find('#About div.doctor-card-info-social a.facebook', first=True).attrs['href']
        try:
            whats_app = r.html.find('#ContactInfo > div.contact.doctor-card-info-number > div > a.whatsapp', first=True).attrs['href']
        except:
            whats_app = None
        try:
            phone_public = r.html.find('#ContactInfo > div.contact.doctor-card-info-number > div > a.dark-blue.mobile-number',
                                       first=True).text
        except:
            phone_public = None
        try:
            mail = r.html.find('#ContactInfo > div:nth-child(3) > a[data-track-label="email"]', first=True).text
        except:
            mail = None
        try:
            title_pro = r.html.find('#pro-location > p.location-title.section-title', first=True).text
        except:
            title_pro = None
        try:
            pro_location = r.html.find('#pro-location > div.address > p.doctor-txt', first=True).text
        except:
            pro_location = None
        try:
            pro_number = r.html.find('#pro-location div.doctor-card-info-number', first=True).text
        except:
            pro_number = None
        try:
            fax = r.html.find('#pro-location > div:nth-child(5) div.doctor-card-info-number', first=True).text
        except:
            fax = None
        try:
            day_hours = r.html.find('#pro-location > div.working-hours > div > ul', first=True).text
        except:
            day_hours = None
        # day = r.html.find('#pro-location > div.working-hours > div > ul > li > span.doctor-txt.days', first=True).text
        # hours = r.html.find('#pro-location > div.working-hours > div > ul > li > span.doctor-txt.hours', first=True).text
        try:
            posts = '\n'.join([po.attrs['href'] for po in r.html.find('#Publication a')])
        except:
            posts = None
        try:
            add_doc = r.html.find('#AdditionalInfo > div.boards.tags', first=True).text.replace('الخدمات الطبية', '').strip()
        except:
            add_doc = None
        try:
            fellowship = r.html.find('#AdditionalInfo > div.fellowship > p.doctor-txt', first=True).text
        except:
            fellowship = None
        try:
            education = r.html.find('#AdditionalInfo > div.education > p.doctor-txt', first=True).text
        except:
            education = None
        try:
            board = r.html.find('#AdditionalInfo > div:nth-child(5) > p:nth-child(2)', first=True).text
        except:
            board = None
        try:
            win = r.html.find('#AdditionalInfo > div:nth-child(6) > p:nth-child(2)', first=True).text
        except:
            win = None
        try:
            note = r.html.find('#AdditionalInfo > div:nth-child(7) > p:nth-child(2)', first=True).text
        except:
            try:
                note = r.html.find('#AdditionalInfo > div:nth-child(7) > p:nth-child(2)', first=True).text
            except:
                note = None
        try:
            languages = r.html.find('#AdditionalInfo > div.languages > ul', first=True).text
        except:
            languages = None
        try:
            pro_insurance = r.html.find('#pro-insurance > ul', first=True).text
        except:
            pro_insurance = None

        sheet.cell(ii, 1).value = prefix
        sheet.cell(ii, 2).value = name
        sheet.cell(ii, 3).value = speciality
        sheet.cell(ii, 4).value = place
        sheet.cell(ii, 5).value = cv
        sheet.cell(ii, 6).value = web
        sheet.cell(ii, 7).value = info_social
        sheet.cell(ii, 8).value = whats_app
        sheet.cell(ii, 9).value = phone_public
        sheet.cell(ii, 10).value = mail
        sheet.cell(ii, 11).value = title_pro
        sheet.cell(ii, 12).value = pro_location
        sheet.cell(ii, 13).value = pro_number
        sheet.cell(ii, 14).value = fax
        sheet.cell(ii, 15).value = day_hours
        sheet.cell(ii, 16).value = posts
        sheet.cell(ii, 17).value = add_doc
        sheet.cell(ii, 18).value = fellowship
        sheet.cell(ii, 19).value = education
        sheet.cell(ii, 20).value = board
        sheet.cell(ii, 21).value = win
        sheet.cell(ii, 22).value = note
        sheet.cell(ii, 23).value = languages
        sheet.cell(ii, 24).value = pro_insurance
        sheet.cell(ii, 26).value = img
    except:
        sleep(10)

book.save('webtebpro.xlsx')