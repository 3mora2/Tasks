from selenium import webdriver
from openpyxl import Workbook
from openpyxl.styles import Alignment
from time import sleep

from selenium.webdriver.chrome.webdriver import WebDriver


class Main_app:
    driver: WebDriver

    def __init__(self):
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            #####################################
            self.ws.column_dimensions['A'].width = 13
            self.ws.column_dimensions['B'].width = 23
            self.ws.column_dimensions['C'].width = 21
            self.ws.column_dimensions['D'].width = 31
            self.ws.column_dimensions['E'].width = 26
            self.ws.column_dimensions['F'].width = 25
            self.ws.column_dimensions['G'].width = 19
            self.ws.column_dimensions['H'].width = 18
            self.ws.column_dimensions['I'].width = 14
            ######################################
            self.start()
        except Exception as e:
            print(e)

    def start(self):
        u = True
        first = True
        adds = 1
        try:
            self.driver = webdriver.Chrome()
            self.driver.get(
                'https://www.moj.gov.sa/ar/opendata/bi/birealestate/Dashboards/200_kpiTown/201_Monthly/kpi201_04_G.aspx')
            input('When you finish, Press Enter')
            while u:
                if first is True:
                    elements = self.driver.find_elements_by_css_selector('div > table.ms-bigrid-table > tbody > tr')
                    first = False
                else:
                    elements = self.driver.find_elements_by_css_selector('div > table.ms-bigrid-table > tbody > tr')[1:]
                for row, element in enumerate(elements):
                    list_row = []
                    for column, ele in enumerate(element.find_elements_by_css_selector('td')):
                        if ele.get_attribute('title') != '':
                            data = ele.get_attribute('title')
                            list_row.append(ele.get_attribute('title'))
                        else:
                            list_row.append(ele.text)
                            data = ele.text
                        self.ws.cell(row=row + adds, column=column + 1).value = data
                        try:
                            self.ws.cell(row=row + adds, column=column + 1).alignment = Alignment(horizontal='center',
                                                                                                  vertical='center',
                                                                                                  wrap_text=True)
                        except:
                            pass
                    print(row + adds, list_row)
                self.wb.save('Table.xlsx')
                adds = row + adds + 1
                try:
                    if self.driver.find_element_by_css_selector('#olap_toolbar_ti_nppi').get_attribute('onclick') is None:
                        u = False
                    else:
                        self.driver.find_element_by_css_selector('#olap_toolbar_ti_nppi').click()
                        sleep(10)
                except:
                    u = False
            self.driver.quit()
            self.wb.save('Table.xlsx')
        except Exception as e:
            print(e)


if __name__ == '__main__':
    m = Main_app()
