import os

import requests

from openpyxl import Workbook, load_workbook
from datetime import datetime
from time import sleep
import requests
import json
import random
from random import choice

headers_and = {
    'User-Agent': 'Instagram 76.0.0.15.395 Android (24/7.0; 640dpi; 1440x2560; samsung; SM-G930F; herolte; samsungexynos8890; en_US; 138226743)',
    'Connection': 'close', 'Accept': '*/*', 'Accept-Language': 'en-US', 'Accept-Encoding': 'gzip, deflate',
    'X-IG-Capabilities': '3brTvw==', 'X-IG-Connection-Type': 'WIFI', 'X-IG-Connection-Speed': '4932kbps',
    'X-IG-App-ID': '567067343352427', 'X-IG-Bandwidth-Speed-KBPS': '-1.000', 'X-IG-Bandwidth-TotalBytes-B': '0',
    'X-IG-Bandwidth-TotalTime-MS': '0', 'X-FB-HTTP-Engine': 'Liger'}


class Insta:
    def __init__(self):
        print('''- please wait while starting........''')
        self.session1 = self.login(username='201028946519', password='ammar2020')
        sleep(1)
        self.session2 = self.login(username='ebrahemelmorsy22001@gmail.com', password='ammar2020')
        sleep(1)
        self.session3 = self.login(username='ebrahemelmorsy22002@gmail.com', password='ammar2020')

        # self.book = Workbook()
        # self.sheet = self.book.active
        # self.cur = 2

    def login(self, username, password):
        session = requests.Session()
        headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
                   'Accept-Language': 'en-GB,en;q=0.9,en-US;q=0.8,ml;q=0.7',
                   'User-Agent': 'Mozilla/S.0 (X11; Linux x86_64) AppleWebKit/S37.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36'}
        session.headers.update(headers)
        session.get('https://www.instagram.com/web/__mid/')
        csrf_token = session.cookies.get_dict()['csrftoken']
        session.headers.update({'X-CSRFToken': csrf_token})
        enc_password = '#PWD_INSTAGRAM_BROWSER:0:{}:{}'.format(int(datetime.now().timestamp()), password)
        resp = session.post('https://www.instagram.com/accounts/login/ajax/',
                            data={'enc_password': enc_password, 'username': username},
                            allow_redirects=True)  # ,proxies=proxyDict)
        print(resp.status_code)
        ###################################################################################################
        try:
            resp_json = resp.json()
        except Exception as e:
            print("Login error: JSON decode fail, {} - {}.".format(resp.status_code, resp.reason))
        try:
            if resp_json.get('checkpoint_url'):
                print("Login: Checkpoint required. Point your browser to "
                      "https://www.instagram.com{} - "
                      "follow the instructions, then retry.".format(resp_json.get('checkpoint_url')))
            if resp_json['status'] != 'ok':
                if 'message' in resp_json:
                    print("Login error: \"{}\" status, message \"{}\".".format(resp_json['status'],
                                                                               resp_json['message']))
                else:
                    print("Login error: \"{}\" status.".format(resp_json['status']))
            if 'authenticated' not in resp_json:
                if 'message' in resp_json:
                    print("Login error: Unexpected response, \"{}\".".format(resp_json['message']))
                else:
                    print("Login error: Unexpected response, this might indicate a blocked IP.")
            if not resp_json['authenticated']:
                if resp_json['user']:
                    print('Login error: Wrong password.')
                else:
                    print('Login error: User {} does not exist.'.format(username))
        except:
            pass
        return session

    def get_likes(self):
        while True:
            dict_l = dict()
            url = input('- put url  ( example:-https://www.instagram.com/p/code/ ) : ')  # chhc
            if url != '':
                break
            else:
                pass
        if url.find('instagram.com/p/') > -1:
            code = url.split('/p/')[1].split('/')[0]
        else:
            print('- wrong url')
            return False

        u = True

        print('- getting Likes ')
        resp = self.get_like(code=code)
        num = resp['data']['shortcode_media']['edge_liked_by']['count']
        print('- Number of Likes {}'.format(num))
        end_cursor = ''
        while u:
            if resp != 'errors':
                u = resp['data']['shortcode_media']['edge_liked_by']['page_info']['has_next_page']
                end_cursor = resp['data']['shortcode_media']['edge_liked_by']['page_info']['end_cursor']
                for i in resp['data']['shortcode_media']['edge_liked_by']['edges']:
                    user = i['node']['username']
                    full_name = i['node']['full_name']
                    id = i['node']['id']
                    profile_pic_url = i['node']['profile_pic_url']
                    pr = f'ID: {id} ,username: {user} ,Full Name: {full_name}'
                    dict_l[id] = {'name': full_name, 'user': user, 'img': profile_pic_url}
                    self.print_percent_done(len(dict_l), num, data=pr)

                sleep(2)
                resp = self.get_like(code=code, end_cursor=end_cursor)
            else:
                print('- please wait...')
                sleep(15)
                resp = self.get_like(code=code, end_cursor=end_cursor)
        print('- Done')
        winner = dict_l[random.choice(list(dict_l.keys()))]
        print(f'- Winner Is (name: {winner["name"]}, username: {winner["user"]}, img: {winner["img"]} )', end='\r')

    def get_like(self, code, end_cursor=None):
        if end_cursor is None:
            url = 'https://www.instagram.com/graphql/query/?query_id=17864450716183058&{}'.format(
                'variables={"shortcode":"%s","first":50}') % code
            str_html = self.session.get(url)
        else:
            url = 'https://www.instagram.com/graphql/query/?query_id=17864450716183058&{}'.format(
                'variables={"shortcode":"%s","first":50,"after":"%s"}') % (code, end_cursor)
            str_html = self.session.get(url)

        if str_html.ok:
            resp = json.loads(str_html.text)
            return resp
        else:
            print(str_html.status_code)
            return 'errors'

    @staticmethod
    def print_percent_done(index, total, data='', bar_len=50):
        percent_done = (index - 1) / total * 100
        percent_done = round(percent_done, 1)

        done = round(percent_done / (100 / bar_len))
        togo = bar_len - done

        done_str = '█' * int(done)
        togo_str = '░' * int(togo)

        print(f'- ⏳ {total}\\{index - 1} - [{done_str}{togo_str}] {data} ({percent_done} %)', end='\r')

    def extract_all_posts(self, username):
        r = self.session3.get('https://www.instagram.com/{0}/?__a=1'.format(username))
        user_id = r.json()['graphql']['user']['id']
        #######################################
        end_cursor = ''
        next_page = True

        while next_page:
            session = choice([self.session3, self.session1, self.session2])
            r = session.get('https://www.instagram.com/graphql/query/',
                                  params={'query_id': '17880160963012870', 'id': user_id, 'first': 12*4,
                                          'after': end_cursor})
            o = r.json()
            graphql = o['data']
            for edge in graphql['user']['edge_owner_to_timeline_media']['edges']:
                shortcode = edge['node']['shortcode']
                url1 = 'https://www.instagram.com/p/{0}/'.format(shortcode)
                self.sheet.cell(self.cur, 5).value = url1
                print(self.cur)
                self.cur += 1
                # self.Download(url1)

            end_cursor = graphql['user']['edge_owner_to_timeline_media']['page_info']['end_cursor']
            next_page = graphql['user']['edge_owner_to_timeline_media']['page_info']['has_next_page']
            # sleep(5)
        self.book.save('final_posts.xlsx')

    def open(self):
        self.book = load_workbook('final_posts_1.xlsx')
        self.sheet = self.book.active
        for i in range(2, self.sheet.max_row+1):
            url = self.sheet.cell(i, 5).value
            if url and not self.sheet.cell(i, 2).value:
                self.Download(url, i)

        self.book.save('final_posts_1.xlsx')

    def Download(self, url, i):
        img = []
        vi = []
        url = str(url) + '?__a=1'
        print(url)
        session = choice([self.session3, self.session1])#, self.session2])
        session.headers.update(headers_and)
        self.r = re = session.get(url)  # B_XP3L_pjjl   B_B4QNUhOUp CAfbugFAw5o
        o = re.json()
        dsk = '\n'.join([i['node']['text'] for i in o['graphql']['shortcode_media']['edge_media_to_caption']['edges']])
        if o['graphql']['shortcode_media']['__typename'] == 'GraphSidecar':
            for edge in o['graphql']['shortcode_media']['edge_sidecar_to_children']['edges']:
                if edge['node']['is_video']:
                    pass
                    # video_url = edge['node']['video_url']
                    # vi.append(video_url)
                else:
                    display_url = edge['node']['display_url']
                    img.append(display_url)
        # elif o['graphql']['shortcode_media']['__typename'] == 'GraphVideo':
        #     video_url = o['graphql']['shortcode_media']['video_url']
        #     vi.append(video_url)

        elif o['graphql']['shortcode_media']['__typename'] == 'GraphImage':
            display_url = o['graphql']['shortcode_media']['display_url']
            img.append(display_url)
        print(i)
        self.sheet.cell(i, 1).value = dsk
        self.sheet.cell(i, 2).value = ','.join(img)
        sleep(4)
        # self.sheet.cell(i, 3).value = ','.join(vi)


# if __name__ == '__main__':
    # inst = Insta()
    # inst.open()
    # inst.extract_all_posts('camelcorners')
    # inst.get_likes()
# https://www.instagram.com/camelcorners/
# import requests_html
# s = requests_html.HTMLSession()
# n = 0
# l = []
# book = load_workbook('final_posts_1.xlsx')
# sheet = book.active
# book_new = Workbook()
# sheet_new = book_new.active
# for i in range(2, sheet.max_row+1):
#     url = sheet.cell(i, 2).value
#     if url:# and sheet.cell(i, 1).value:3686 2584 4541
#         code = sheet.cell(i, 5).value.split('/')[4]
#         for n, u in enumerate(url.split(',')):
#             if n == 2:
#                 break
#             if u in l:
#                 print('- found')#1045
#                 continue
#
#             l.append(u)
#             print(l.__len__())
#             if os.path.isfile(f'img/{code}-{i}-{n}.jpg'):
#                 continue
#             with open(f'img/{code}-{i}-{n}.jpg', 'wb') as f:
#                 print('add')
#                 r = s.get(u, headers=headers_and)
#                 f.write(r.content)
# dict_ = dict()
# for file in os.listdir(r'E:\New folder\3bdo\img bcg remover'):
#     f = file.split('.')[0].split('-')
#     c = int(f[-1])
#     r = int(f[-2])
#     code = '-'.join(f[:-2])
#     if not sheet_new.cell(r, 1).value == sheet.cell(r, 1).value:
#         for i in range(1, sheet.max_column+1):
#             sheet_new.cell(r, i).value = sheet.cell(r, i).value
#     sheet_new.cell(r, 20+c).value = file
#     print(r)
#
# book_new.save('test.xlsx')
# book = load_workbook('test.xlsx')
# sheet = book.active
# nn = 0
# for i in range(2, sheet.max_row+1):
#     if sheet.cell(i, 4).value:#and not sheet.cell(i, 5).value:
#         # print(sheet.cell(i, 4).value)
#         try:
#             if os.path.exists(rf'E:\New folder\3bdo\img bcg remover\{sheet.cell(i, 4).value.replace("-0.png","-1.png")}'):
#                 sheet.cell(i, 5).value = sheet.cell(i, 4).value.replace("-0.png","-1.png")
#             else:
#                 sheet.cell(i, 5).value = None
#             # with open(rf'E:\New folder\img\{sheet.cell(i, 3).value.replace("-0.png","-1.jpg")}', 'rb') as fr:
#             #     with open(rf'im\{sheet.cell(i, 3).value.replace("-0.png","-1.jpg")}', 'wb') as fw:
#             #         fw.write(fr.read())
#             nn += 1
#         except:
#             continue
#
#     # print(nn)
# book.save('test.xlsx')