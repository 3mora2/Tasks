# from selenium import webdriver
from time import sleep
from openpyxl import load_workbook,Workbook
from io import BytesIO
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
# from requests_html import HTMLSession, HTML
import os
# s = HTMLSession()
# driver = webdriver.Firefox()
for file in os.listdir('noon'):
    if file != 'mob - Copy.xlsx':
        continue
    path = 'noon\\'+file
    print(path)
    book = load_workbook(path)
    sheet = book.active
    sheet.column_dimensions['B'].width = 17
    end = sheet.max_row
    i = 2
    print(end)
    for n in range(i, end + 1):
        if sheet[f'I{n}'].value is not None and sheet[f'C{n}'].value is None:
            print(n)
            sheet.delete_rows(n, 1)
    print(sheet.max_row)
    book.save('1.xlsx')
    # print(len(links))
    # links = [(sheet[f'I{n}'].value, n) for n in range(i, end + 1) if sheet[f'I{n}'].value is not None and sheet[f'C{n}'].value is None]
    # for link, i in links:
    #     try:
    #         link = link.split('?o=')[0]
    #         driver.get(link)
    #
    #         # r = s.get(link)
    #         html = HTML(html=driver.page_source)
    #         try:
    #             cat = html.find('div.breadcrumbContainer > div > div > span > button')[-1].text
    #         except:
    #             cat = None
    #         name = html.find('div.coreWrapper > div > h1', first=True).text
    #         try:
    #             brand = html.find('div.coreWrapper > div > a.brand', first=True).text
    #         except:
    #             brand = None
    #         price = html.find('div.priceRow > div > p > span > span > span.sellingPrice > span > span.value',
    #                           first=True).text
    #         main_img = \
    #             html.find('div.imageGalleryWrapper > div > div.standardSlider > div > div > div > div > div > div > img',
    #                       first=True).attrs['src']
    #         photo = '\n'.join([img.attrs['src'] for img in html.find(
    #             'div.imageGalleryWrapper > div > div.standardSlider > div > div > div > div > div > div > img')])
    #         # print(photo)
    #         driver.get(link.replace('saudi-ar', 'saudi-en'))
    #         html = HTML(html=driver.page_source)
    #         # r = s.get(link.replace('saudi-ar', 'saudi-en'))
    #         name_en = html.find('div.coreWrapper > div > h1', first=True).text
    #         try:
    #             brand_en = html.find('div.coreWrapper > div > a.brand', first=True).text
    #         except:
    #             brand_en = None
    #         try:
    #             if main_img is not None:
    #                 res = s.get(main_img)
    #                 image_file = BytesIO(res.content)
    #                 img = Image(image_file)
    #                 img.width = 117
    #                 img.height = 95
    #                 sheet.row_dimensions[i].height = 72
    #                 sheet.add_image(img, f'B{i}')
    #         except Exception as er:
    #             print(er)
    #         sheet[f'A{i}'].value = cat
    #         sheet[f'C{i}'].value = name_en
    #         sheet[f'D{i}'].value = name
    #         sheet[f'E{i}'].value = brand
    #         sheet[f'F{i}'].value = brand_en
    #         sheet[f'G{i}'].value = price
    #         sheet[f'H{i}'].value = photo
    #         print(i, name, main_img)
    #         book.save(path)
    #     except:
    #         pass
    #     # i += 1
    # book.save(path)

#
# driver = webdriver.Chrome()
# driver.get('https://www.noon.com/saudi-ar/electronics-and-mobiles/mobiles-and-accessories/accessories-16176/chargers-17982?sort[by]=new_arrivals&sort[dir]=desc&limit=100')
# input('jjj')
# try:
#     book = Workbook()
#     sheet = book.active
#     i = 1
#     sheet[f'A{i}'].value = 'القسم'
#     sheet[f'B{i}'].value = 'الصور'
#     sheet[f'C{i}'].value = 'name'
#     sheet[f'D{i}'].value = 'الاسم'
#     sheet[f'E{i}'].value = 'البرند'
#     sheet[f'F{i}'].value = 'brand'
#     sheet[f'G{i}'].value = 'السعر'
#
#     l = []
#     while True:
#         for element in driver.find_elements_by_css_selector(
#                 '.productListContainer > div > div.productListWrapper > div.productList.gridView > div > div > a.product'):
#             if element.get_attribute('href') not in l:
#                 l.append(element.get_attribute('href'))
#         print(len(l))
#         # if driver.find_element_by_css_selector('.nextLink').get_attribute('aria-disabled') != 'true':
#         #     driver.find_element_by_css_selector('.nextLink').click()
#         #     sleep(9)
#         # else:
#         #     break
#         break
#     n = 2
#     print('save')
#     for e in l:
#         sheet[f'I{n}'].value = e
#         n += 1
#     book.save('power-banks.xlsx')
#
# except Exception as e:
#     print(e)

"""
www.noon.com/saudi-ar/electronics-and-mobiles/mobiles-and-accessories/mobiles-20905?page=4
قائمة بالجوالات الحديثه وجديده وليست مجددة خلال سنتين من مختلف الشركات فقط

www.noon.com/saudi-ar/electronics-and-mobiles/mobiles-and-accessories/accessories-16176
هذا رابط اكسسوارات الموبيلات كله ويندرج تحته بعض الاقسام ولكن المسميات في الموقع مختلفه عن بعض المسميات اللتي ذكرتها فبرجاء تحديد الاسماء مطابقة للموقع كما في الصورة المرفقة .

www.noon.com/saudi-ar/mobiles

قائمة اكسوارات الجوال

www.noon.com/saudi-ar/mobiles

قائمة اكسوارات الجوال
شواحن - شواحن متنقله- كيابل- اغطية جوال- حامي الشاشه- سماعات
———-
انواع الجوالات

www.noon.com/saudi-ar/electronics-and-mobiles/mobiles-and-accessories/mobiles-20905?page=4
قائمة بالجوالات الحديثه وجديده وليست مجددة خلال سنتين من مختلف الشركات فقط

"""
