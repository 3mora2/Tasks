from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from time import sleep
from openpyxl import load_workbook, Workbook

# wb = load_workbook('thanwy.xlsx')
wb = Workbook()
ws = wb.active
num = 1

driver = webdriver.Chrome()
for i in range(582200, 582200+2000):
    try:
        sleep(2)
        driver.get('https://g12.emis.gov.eg/')
        sleep(2)
        WebDriverWait(driver, 5).until(ec.visibility_of_element_located((By.ID, "SeatingNo"))).send_keys(i)
        WebDriverWait(driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".btn.btn-success.text-center"))).click()
        print(WebDriverWait(driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,
                                                                               "#Result > div > div.row.no-gutter.mb-5 > div.col-lg-4 > div > table > tbody > tr:nth-child(1) > td"))).text,
              driver.find_element_by_css_selector('#Result > div > div.row.no-gutter.mb-5 > div.col-lg-4 > div > table > tbody > tr:nth-child(2) > td').text,
              driver.find_element_by_css_selector('#Result > div > div.row.no-gutter.mb-5 > div.col-lg-4 > div > table > tbody > tr:nth-child(3) > td').text,
              driver.find_element_by_css_selector(
                  '#Result > div > div.row.no-gutter.mb-5 > div.col-lg-4 > div > table > tbody > tr:nth-child(4) > td').text,
              driver.find_element_by_css_selector(
                  '#Result > div > div.row.no-gutter.mb-5 > div.col-lg-4 > div > table > tbody > tr:nth-child(5) > td').text,
              driver.find_element_by_css_selector(
                  '#Result > div > div.row.no-gutter.mb-5 > div.col-lg-8 > div:nth-child(1) > div > div > div > table > tbody > tr:nth-child(8) > th:nth-child(2)').text,
              )

        # cat = WebDriverWait(driver, 5).until(
        #     ec.visibility_of_element_located((By.CSS_SELECTOR, "#pills-tab > li:nth-child(3) > h1"))).text
        # cat = WebDriverWait(driver, 5).until(
        #      ec.visibility_of_element_located((By.CSS_SELECTOR, "#Result > div > div.row.no-gutter.mb-5 > div.col-lg-4 > div > table > tbody > tr:nth-child(2) > td"))).text
        # col = WebDriverWait(driver, 5).until(
        #      ec.visibility_of_element_located((By.CSS_SELECTOR, "#Result > div > div.row.no-gutter.mb-5 > div.col-lg-4 > div > table > tbody > tr:nth-child(3) > td"))).text
        # # WebDriverWait(driver, 5).until(
        # #     ec.visibility_of_element_located((By.CSS_SELECTOR, "#pills-tab > li:nth-child(2) > h1"))).text
        # name = WebDriverWait(driver, 5).until(ec.visibility_of_element_located(
        #     (By.CSS_SELECTOR, "#pills-tab > li:nth-child(1) > span:nth-child(2)"))).text
        # scoll = WebDriverWait(driver, 5).until(ec.visibility_of_element_located(
        #     (By.CSS_SELECTOR, "#pills-tab > li:nth-child(2) > span:nth-child(2)"))).text
        # stat = WebDriverWait(driver, 5).until(ec.visibility_of_element_located(
        #     (By.CSS_SELECTOR, "#pills-tab > li:nth-child(5) > span:nth-child(2)"))).text
        # far = WebDriverWait(driver, 5).until(
        #     ec.visibility_of_element_located(
        #         (By.CSS_SELECTOR, "#pills-tab > li:nth-child(7) > span:nth-child(2)"))).text
        # print(f'{name}-{cat}{col} {stat}  {scoll}  {far}')
        #
        # ws[f'A{num}'] = i
        # ws[f'B{num}'] = name
        # ws[f'C{num}'] = cat
        # ws[f'D{num}'] = col
        # ws[f'E{num}'] = stat
        # ws[f'F{num}'] = scoll
        # ws[f'G{num}'] = far
        # num += 1
        # wb.save('thanwy2.xlsx')
    except Exception as e:
        print(e)
