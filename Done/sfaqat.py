import os
import sys
import warnings
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from time import sleep
from requests_html import HTML
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from webdriver_manager.firefox import GeckoDriverManager

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class Main:
    def __init__(self):
        self.book = load_workbook('sfaqat.xlsx')
        self.sheet = self.book.active
        try:
            self.driver = webdriver.Firefox()
        except:
            self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())

    def wait_load(self):
        sleep(1)
        while True:
            try:
                if 'block' not in self.driver.find_element_by_css_selector('.loading-mask[data-role="loader"]').get_attribute('style'):
                    break
            except:
                break
        sleep(1)

    def wait_load_p(self):
        sleep(2)
        while True:
            try:
                if not self.safe_find_element_by(By.CSS_SELECTOR, '.loading-mask[data-role="loader"]'):
                    break
            except:
                break
        sleep(1)

    def safe_find_element_by(self, by, elem):
        try:
            return self.driver.find_element(by, elem)
        except NoSuchElementException:
            return None

    def start(self):
        products = []
        url = 'https://www.sfaqat.com/ar/products.html?'
        while True:
            self.driver.get(url)
            sleep(1)
            self.wait_load()
            sleep(1)
            for element in self.driver.find_elements_by_css_selector('#product-list-container li > div > a.product'):
                if element.get_attribute('href') in products:
                    print('- found')
                else:
                    products.append(element.get_attribute('href'))
                print(products.__len__())
            try:
                url = self.driver.find_element_by_css_selector('a.next').get_attribute('href')
            except:
                break

        n = 2
        for url in products:
            self.sheet.cell(n, 6).value = url
            n += 1

        self.book.save('sfaqat.xlsx')

    def start_get(self):
        for i in range(2, self.sheet.max_row+1):
            url = self.sheet.cell(i, 6).value
            if url and not self.sheet.cell(i, 1).value:
                self.driver.get(url)
                self.wait_load_p()
                try:
                    name = self.driver.find_element_by_css_selector('meta[name="title"]').get_attribute('content')
                    description = self.driver.find_element_by_css_selector('meta[name="description"]').get_attribute('content')
                    price = self.driver.find_element_by_css_selector('span.price').text
                    sku = self.driver.find_element_by_css_selector('[itemprop="sku"]').text
                    state = self.driver.find_element_by_css_selector('.product-info-stock-sku').text
                    img = ',\n'.join([im.get_attribute('href') for im in self.driver.find_elements_by_css_selector('[data-gallery-role="stage-shaft"] > div') if im.get_attribute('href')])
                except:
                    continue
                self.sheet.cell(i, 1).value = name
                self.sheet.cell(i, 2).value = price
                self.sheet.cell(i, 3).value = sku
                self.sheet.cell(i, 4).value = state
                self.sheet.cell(i, 5).value = description
                self.sheet.cell(i, 7).value = img
                print(i, sku)

        self.book.save('sfaqat.xlsx')


if __name__ == '__main__':
    self = Main()
    self.start_get()

