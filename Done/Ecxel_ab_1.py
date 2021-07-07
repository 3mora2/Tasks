from openpyxl import load_workbook, Workbook

sub_cat = dict()

book = load_workbook(r'C:\Users\3mora\Downloads\jarir-com.xlsx')
sheet = book.active

new_book = Workbook()
new_sheet = new_book.active

# Add Headers
for c in range(1, sheet.max_column+1):
    new_sheet.cell(1, c).value = sheet.cell(1, c).value

count_new = 2
for i in range(2, sheet.max_row+1):
    sub_title = sheet[f'C{i}'].value
    if sub_title is None or sub_title == '':
        for c in range(1, sheet.max_column + 1):
            new_sheet.cell(count_new, c).value = sheet.cell(i, c).value
        count_new += 1
    else:
        if sub_title in sub_cat.keys():
            if int(sub_cat[sub_title]) < 3:
                for c in range(1, sheet.max_column + 1):
                    new_sheet.cell(count_new, c).value = sheet.cell(i, c).value
                count_new += 1
            sub_cat[sub_title] = int(sub_cat[sub_title])+1
        else:
            sub_cat[sub_title] = 1
            for c in range(1, sheet.max_column + 1):
                new_sheet.cell(count_new, c).value = sheet.cell(i, c).value
            count_new += 1
    print(i)
new_book.save(r'C:\Users\3mora\Downloads\jarir-com_new_no_none.xlsx')
