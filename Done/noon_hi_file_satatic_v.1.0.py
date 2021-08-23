import threading
import traceback
from datetime import datetime
from time import sleep
import os
import sys
import warnings

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from requests_html import HTMLSession
from openpyxl import load_workbook, Workbook

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 's':
                break


class Main:
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-sa/catalog'

    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'tr:nth-child(1) > td > a')

    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')
    DEACTIVATE_SCRIPT = 'if (document.querySelector(\'input[id="offerActive"]\').checked == true){document.querySelector(\'input[id="offerActive"]\').click()};'
    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="offerActive"]\').checked == false){document.querySelector(\'input[id="offerActive"]\').click()};'

    SAVE_CHANGE_SELECTOR = (By.XPATH, '//*[contains(text(),"Save Changes")]')
    SUBMIT_SELECTOR = (By.XPATH,
                       '//div[calss="btnWrapper"]/div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')

    def __init__(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.session = HTMLSession()

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def add_new(self, code):
        try:
            self.driver.get(self.URL_CAT)
            sleep(2)
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))
                                                 ).send_keys(code, Keys.ENTER)
            sleep(3)
            WebDriverWait(self.driver, 10).until(
                ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
            sleep(3)
            return True
        except:
            if 'login.noon.partners' in self.driver.current_url:
                self.login()
            try:
                self.driver.get(self.URL_CAT)
                sleep(2)
                WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))
                                                     ).send_keys(code, Keys.ENTER)
                sleep(3)
                WebDriverWait(self.driver, 10).until(
                    ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
                sleep(3)
                return True
            except:
                sleep(30)
                print(traceback.print_exc())
                return False

    def add_qu(self, q):
        sleep(1)
        try:
            WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
            return True
        except:
            print(traceback.print_exc())
            print('- Cant Change quantity')
            return False

    def update_cookies(self):
        self.driver.refresh()
        sleep(1)
        cookies = self.driver.get_cookies()
        for cookie in cookies:
            self.session.cookies.set(str(cookie['name']), str(cookie['value']))

    def get_noon_product(self):
        dict_noon = dict()
        headers = {
            "accept": "application/json, text/plain, */*",
            "accept-language": "en-US,en;q=0.9",
            "cache-control": "no-cache, no-store",
            "content-type": "application/json",
            "sec-ch-ua": "\" Not;A Brand\";v=\"99\", \"Google Chrome\";v=\"91\", \"Chromium\";v=\"91\"",
            "sec-ch-ua-mobile": "?0",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "x-locale": "en-sa",
            "x-platform": "web"
        }
        body = ['{"productList":{"page":', ',"perPage":100}}']
        page = 1
        self.update_cookies()
        while True:
            try:
                r = self.session.post('https://catalog.noon.partners/_svc/mp-partner-catalog/catalog/product/list',
                                      headers=headers, data=f'{page}'.join(body))
                if r.ok:
                    sleep(4)
                    page += 1
                    try:
                        products = r.json()['products']
                        if len(products) == 0:
                            break
                        for product in products:
                            static = product['psku']['isActive']
                            code = product['psku']['partnerSku']
                            express = True if product['fbnStock'] else False
                            dict_noon[code] = {'static': static, 'express': express}
                        print(len(dict_noon))
                    except Exception as e:
                        print(e)
                        sleep(60)
                        self.update_cookies()

                else:
                    print(r.status_code)
                    sleep(60)
                    self.update_cookies()

            except Exception as e:
                print(e)
                sleep(60)
                self.update_cookies()

        return dict_noon

    def noon_file(self, data_noon, data_ih, story):
        thr = Thread()
        thr.start()
        file = 'data-' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + '.xlsx'
        book = Workbook()
        sheet = book.active
        n = 2
        for code in data_noon.keys():
            if thr.p == 's':
                break
            if data_noon[code]['express']:
                continue

            if code not in data_ih.keys():
                print('- Code Not in Data ih ')
                s = HTMLSession()
                cookie = {'name': 'iher-pref1',
                          'value': 'storeid=0&sccode=SA&lan=en-US&scurcode=SAR&wp=2&lchg=1&ifv=1&accsave=0'}
                s.cookies.set(cookie['name'], cookie['value'])
                try:
                    r = s.get('https://sa.iherb.com/search?kw=' + code)
                    static = r.html.find('#stock-status')[0].text
                    if 'متوفر حاليا' in static or 'In Stock' in static:
                        data_ih[code] = 'InStock'
                    else:
                        data_ih[code] = 'OutStock'
                    print(code, data_ih[code])
                except:
                    pass

            if data_noon[code]['static'] == 0 and data_ih.get(code, ) == 'InStock':
                action = self.ACTIVE_SCRIPT
                static = True
            elif data_noon[code]['static'] == 1 and data_ih.get(code, ) == 'OutStock':
                if code in story and story[code] > 0:
                    action = self.ACTIVE_SCRIPT
                    static = True
                else:
                    action = self.DEACTIVATE_SCRIPT
                    static = False

            else:
                continue
            print(n, code, data_ih.get(code, ), data_noon[code]['static'], static)
            sheet.cell(n, 1).value = code
            sheet.cell(n, 2).value = data_noon[code]['static']
            sheet.cell(n, 3).value = data_ih.get(code, )
            sheet.cell(n, 4).value = static

            result_add = self.add_new(code)
            if result_add:
                if static:
                    result_q = self.add_qu('30')
                    sleep(1)
                    if not result_q:
                        sheet.cell(n, 5).value = 'error with q'

                try:
                    self.driver.execute_script(action)
                    sleep(4)
                    WebDriverWait(self.driver, 10).until(
                        (ec.element_to_be_clickable(self.SAVE_CHANGE_SELECTOR))).click()
                    sleep(1)
                    WebDriverWait(self.driver, 10).until(
                        (ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                    sleep(1)
                except:
                    sheet.cell(n, 5).value = 'error save'
                sleep(4)
            else:
                sheet.cell(n, 5).value = 'error'
            n += 1
        book.save(file)
        self.driver.quit()


def test_request():
    new_edit = dict()
    cont = 0
    s = HTMLSession()
    cookie = {'name': 'iher-pref1', 'value': 'storeid=0&sccode=SA&lan=en-US&scurcode=SAR&wp=2&lchg=1&ifv=1&accsave=0'}
    s.cookies.set(cookie['name'], cookie['value'])
    r = s.get('https://sa.iherb.com/')
    if r.ok:
        urls = ['https://sa.iherb.com' + element.attrs['href'] + '?noi=192' for element in
                r.html.find('.category') if '/c/' in element.attrs['href']]

        for url in urls:
            r_ = s.get(url)
            next_pag = True
            while next_pag:
                for ele in r_.html.find('div.product-inner.product-inner-wide'):
                    code = ele.find('div.absolute-link-wrapper > a', first=True).attrs['data-part-number']
                    try:
                        static = ele.find('meta+link')[0].attrs['href']
                    except:
                        static = ''

                    if 'InStock' in static:
                        new_edit[code] = 'InStock'
                    elif static == '':
                        new_edit[code] = ''
                    else:
                        new_edit[code] = 'OutStock'

                    cont += 1
                print(f'- {cont}', end='\r')
                try:
                    url = 'https://sa.iherb.com' + r_.html.find('.pagination-next')[0].attrs['href']
                    r_ = s.get(url)
                    if not r_.ok:
                        sleep(30)
                        r_ = s.get(url)
                except:
                    break

        for elem in new_edit.keys():
            if new_edit[elem] == '':
                r = s.get('https://sa.iherb.com/search?kw=' + elem)
                static = r.html.find('#stock-status')[0].text
                if 'متوفر حاليا' in static or 'In Stock' in static:
                    new_edit[elem] = 'InStock'
                else:
                    new_edit[elem] = 'OutStock'
    else:
        print(r.status_code)
    book = Workbook()
    sheet = book.active
    for n, elem in enumerate(new_edit.keys()):
        sheet.cell(n + 1, 1).value = elem
        sheet.cell(n + 1, 2).value = new_edit[elem]
    book.save('ih.xlsx')
    return new_edit


def open_test(path):
    dict_data = dict()
    sheet = load_workbook(path).active
    for i in range(1, sheet.max_row + 1):
        code = sheet.cell(i, 1).value
        static = sheet.cell(i, 2).value
        if code:
            dict_data[code] = static

    return dict_data


def read_file():
    data_ = dict()
    sheet = load_workbook(file_name).active
    for i in range(2, sheet.max_row + 1):
        code = sheet.cell(i, 1).value
        value = sheet.cell(i, 2).value
        if code and value:
            data_[code] = value

    return data_


if __name__ == "__main__":
    while True:
        try:
            file_name = input('- Enter file name : ')
            if os.path.isfile(file_name) and file_name.endswith('.xlsx'):
                break
            print('- No such file or directory')
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)

    email = input('- Enter your email : ')
    password = input('- Enter your password : ')
    while True:
        storage = read_file()
        data = test_request()
        print('- Get Data from Hi Done')
        self = Main()
        self.login()
        noon_data = self.get_noon_product()
        print('- Get Data from Noon Done')
        self.noon_file(noon_data, data, storage)

# pyinstaller --add-data C:/Users/3mora/AppData/Local/Programs/Python/Python39/Lib/site-packages/pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info
'''
bekj.119@gmail.com

perfect203070@gmail.com
Mm0987654
'''