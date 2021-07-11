import os
from time import sleep
from openpyxl import Workbook
from selenium import webdriver

from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import requests
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities


class CollectPosts(object):
    def __init__(self):
        caps = DesiredCapabilities().CHROME
        caps["pageLoadStrategy"] = "none"
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('blink-settings=imagesEnabled=false')
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), desired_capabilities=caps,
                                       chrome_options=chrome_options)

    def safe_find_element_by(self, by, elem):
        try:
            return self.driver.find_element(by, elem)
        except NoSuchElementException:
            return None

    def search(self, text):
        self.current_cell = 1
        self.book = Workbook()
        self.sheet = self.book.active

        self.sheet.cell(self.current_cell, 1).value = 'Desk'
        self.sheet.cell(self.current_cell, 2).value = 'IMG name'
        self.sheet.cell(self.current_cell, 3).value = 'IMG Url'
        self.sheet.cell(self.current_cell, 4).value = 'Url'
        self.current_cell += 1
        main_path = 'skeletal dysplasia'
        cat_path = main_path + '/' + text
        urls = []
        self.driver.get(f'https://radiopaedia.org/search?lang=us&modality=X-ray&q={text}&scope=cases')
        sleep(10)
        while True:
            for element in self.driver.find_elements_by_css_selector('a.search-result.search-result-case'):
                if element.get_attribute('href') not in urls:
                    urls.append(element.get_attribute('href'))
            print(len(urls))
            try:
                if 'disabled' in self.driver.find_element_by_css_selector('a.next_page').get_attribute('class'):
                    break
                else:
                    _ = self.driver.find_element_by_css_selector('a.next_page').location_once_scrolled_into_view
                    self.driver.find_element_by_css_selector('a.next_page').click()
                    sleep(6)
            except:
                break
        if not os.path.exists(main_path):
            os.mkdir(main_path)
        if not os.path.exists(cat_path):
            os.mkdir(cat_path)

        for url in urls:
            img_urls = []
            self.driver.get(url)
            sleep(4)
            title = self.driver.find_element_by_css_selector('h1.header-title').text
            try:
                dsk = self.driver.find_element_by_css_selector('.case-study > div.sub-section.study-findings').text
            except:
                dsk = None
            try:
                img_url = set()
                img_req = 'https://radiopaedia.org'+self.driver.find_element_by_css_selector('.user-generated-content > div.case-study').get_attribute('data-study-stacks-url')
                r = requests.get(img_req)
                for data in r.json():
                    for img in data['images']:
                        img_url.add(img['fullscreen_filename'])

            except:
                print('- old')
                img_url = set(element.get_attribute('src').replace('_small.', '_jumbo.') for element in
                              self.driver.find_elements_by_css_selector('#case-images img')
                              if element.get_attribute('src') and not element.get_attribute('src').endswith('.gif'))

            name_page = self.driver.current_url.split("?")[0].split('/')[-1]
            img_path = cat_path + "/" + name_page
            if len(img_url) > 0:
                if not os.path.exists(img_path):
                    os.mkdir(img_path)
                for num, img in enumerate(img_url):
                    end = img.split('.')[-1]
                    name = f'{img_path}/{name_page}-{str(num)}.'
                    r = requests.get(img)
                    try:
                        with open(name+end, 'wb') as f:
                            f.write(r.content)
                    except:
                        name = f'{img_path}/{"-".join(name_page.split("-")[-4:])}-{str(num)}.'
                        with open(name+end, 'wb') as f:
                            f.write(r.content)

                    if dsk:
                        with open(name + 'txt', 'w') as f:
                            f.write(dsk)

                    img_urls.append(f'{name_page}-{str(num)}.{end}')

            self.sheet.cell(self.current_cell, 1).value = dsk
            self.sheet.cell(self.current_cell, 2).value = '\n'.join(img_urls)
            self.sheet.cell(self.current_cell, 3).value = '\n'.join(img_url)
            self.sheet.cell(self.current_cell, 4).value = url
            print(self.current_cell)
            self.current_cell += 1

        self.book.save(cat_path + '/' + text + '.xlsx')


if __name__ == '__main__':
    while True:
        try:
            init = input('- Enter Text : ')
            if init != "":
                break
        except:
            pass

    self = CollectPosts()
    for txt in init.split(','):
        self.search(txt)
'''
Developmental dysplasia of the hip, talipes deformity, cleidocranial dysplasia, osteogenesis imperfecta,osteopetrosis,fibrous dysplasia,osteopathia striata,multiple epiphyseal dysplasia,diaphyseal aclasis,achondroplasia,Mucopolysaccharidoses,sprengel's shoulder,madelung's deformity
'''