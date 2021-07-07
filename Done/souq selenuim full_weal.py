import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import pyautogui
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup



wb = Workbook()
ws = wb.active
ws['A1'] = 'Name'
ws['B1'] = 'Price'
ws['C1'] = 'Img'
ws['D1'] = 'Desc link'
ws['E1'] = 'full Desc'
ws['F1'] = 'some desc'
ws['G1'] = 'all desc'
ws['H1'] = 'Code1'
ws['I1'] = 'Code2'
ws['J1'] = 'Code3'
ws['K1'] = 'Code4'
ws['L1'] = 'Pic'



#ws['B1'] = datetime.datetime.now()
x=2
y=2
z=2
w=2
s=2
p=2
t=2
m=2
n=2
i=2
l=2
o=2
q=2
j=2
url='https://deals.souq.com/sa-en/beverages/t/1015?ref=nav'



driver = webdriver.Chrome()
driver.maximize_window() 
driver.get(url)
time.sleep(5)

'''for i in range(20):
    pyautogui.scroll(-3000)
    time.sleep(3)   '''


names=driver.find_elements_by_class_name('itemTitle')
#prices=driver.find_elements_by_class_name('price')
#images=driver.find_elements_by_tag_name('img')
Desc_links = driver.find_elements_by_class_name('img-link')

for name in names:
    driver.find_elements_by_class_name('itemTitle')

    ws['A{}'.format(x)]=name.text
    x+=1

    print(name.text)
'''
for price in prices:
    driver.find_elements_by_class_name('price')

    ws['B{}'.format(y)]=price.text
    y+=1
    
    print(price.text)

for image in images:
    driver.find_elements_by_tag_name('img')

    ws['C{}'.format(z)]=image.get_attribute('data-src')
    z+=1
    print(image.get_attribute('data-src'))

 
drivera = webdriver.Chrome()
for Desc_link in Desc_links:

    ws['D{}'.format(w)]=Desc_link.get_attribute('href')
    w+=1
    print(Desc_link.get_attribute('href'))

    dls=Desc_link.get_attribute('href')
    
    drivera.get(dls)
    l=''

    soup = BeautifulSoup(drivera.page_source, 'html.parser')
    
    descs = soup.find(id='description-short')
    ws['E{}'.format(s)]=descs.text
    s+=1
    print(descs.text)
    
    fulls = soup.find(id='specs-short')
    ws['F{}'.format(t)]=fulls.text
    t+=1
    print(fulls.text)
    
    descsall = soup.find(id='specs-full')
    ws['G{}'.format(m)]=descsall.text
    m+=1
    print(descsall.text)

    code1 = drivera.find_element_by_xpath('//*[@id="specs-short"]/dl/dd[1]') # or [1,2,3]
    ws['H{}'.format(n)]=code1.text
    n+=1
    print(code1.text)

    code2 = drivera.find_element_by_xpath('//*[@id="specs-short"]/dl/dd[2]') # or [1,2,3]
    ws['I{}'.format(i)]=code2.text
    i+=1
    print(code2.text)

    code3 = drivera.find_element_by_xpath('//*[@id="specs-short"]/dl/dd[3]') # or [1,2,3]
    ws['J{}'.format(o)]=code3.text
    o+=1
    print(code3.text)

    code4 = drivera.find_element_by_xpath('//*[@id="specs-short"]/dl/dd[4]') # or [1,2,3]
    ws['K{}'.format(q)]=code4.text
    q+=1
    print(code4.text)

    descs_pics = soup.find('div', {'class' : 'slick-track'})
    for descs_pic in descs_pics.find_all('img'):
        l=descs_pic.attrs['src']+'\n'+l

        print(descs_pic.attrs['src'])
  
    ws['L{}'.format(p)]=l
    p+=1'''
    
        

    #drivera.quit()    
        


        

     

wb.save("KSA Beverages.xlsx")    
    
