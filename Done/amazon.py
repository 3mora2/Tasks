import subprocess
import traceback
from time import sleep
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


class Main:
    URL_LOGIN = 'https://sellercentral.amazon.sa/signin?ref_=scsa_soa_wp_signin_n&ld=SCSAWPDirect'
    # URL_ADD = 'https://sellercentral.amazon.sa/productsearch/ref=xx_catadd_dnav_userperms'
    URL_ADD = 'https://sellercentral.amazon.sa/product-search/search?q='

    def __init__(self):
        # chrome_options = Options()
        # try:
        #     subprocess.Popen(
        #         '"C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222')
        # except:
        #     subprocess.Popen(
        #         '"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222')
        # chrome_options.add_argument("--disable-infobars")
        # chrome_options.add_argument("--disable-notifications")
        # chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        # self.driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), chrome_options=chrome_options)
        self.driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())
        self.driver.get(self.URL_ADD)
        self.book = load_workbook(filename)
        self.sheet = self.book.active

    def check_if_exist(self):
        if len(self.driver.find_elements_by_css_selector('kat-icon[name="error"]')):
            return False

        return True

    def check_load(self):
        if len(self.driver.find_elements_by_css_selector('kat-box')):
            return True

        return False

    @staticmethod
    def check_available(first_box):
        if len(first_box.find_elements_by_css_selector('kat-dropdown')):
            return True

        return False

    def start(self):
        for index in range(2, self.sheet.max_row+1):
            exist = True
            price = self.sheet.cell(index, 2).value
            quantity = self.sheet.cell(index, 3).value
            barcode = self.sheet.cell(index, 8).value
            if barcode:
                try:
                    self.driver.get(f'{self.URL_ADD}{barcode}')
                    sleep(2)
                    n_try = 0
                    while True:
                        if self.check_load():
                            break

                        exist = self.check_if_exist()
                        if not exist:
                            break

                        n_try += 1
                        sleep(1)
                        if n_try >= 10:
                            exist = False
                            break

                    if exist:
                        first_box = self.driver.find_element_by_css_selector('#search-result kat-box:first-child')
                        static = first_box.find_element_by_css_selector('kat-button').text
                        if self.check_available(first_box):
                            first_box.find_element_by_css_selector('kat-dropdown').click()
                            sleep(1)
                            first_box.find_element_by_css_selector('.select-options').click()
                            sleep(1)
                            first_box.find_element_by_css_selector('span > a').click()
                            sleep(1)
                            if len(self.driver.window_handles) > 1:
                                self.driver.close()
                                sleep(1)
                                self.driver.switch_to.window(self.driver.window_handles[0])
                                sleep(2)
                                if self.driver.find_element_by_css_selector('kat-toggle#advanced-view-switch').get_attribute('checked') != 'true':
                                    self.driver.find_element_by_css_selector('kat-toggle#advanced-view-switch').click()
                                    sleep(2)

                                self.driver.find_element_by_css_selector('kat-input#item_sku').send_keys(str(barcode))
                                sleep(1)
                                self.driver.find_element_by_css_selector('kat-input#standard_price'
                                                                         ).send_keys(str(price))
                                sleep(1)
                                self.driver.find_element_by_css_selector('kat-input#quantity').send_keys(str(quantity))
                                sleep(1)
                                self.driver.find_element_by_css_selector('kat-input#fulfillment_latency').send_keys(latency)
                                sleep(1)
                                _ = self.driver.find_element_by_css_selector(
                                    'kat-button#EditSaveAction').location_once_scrolled_into_view

                                WebDriverWait(self.driver, 5).until((ec.element_to_be_clickable(
                                    (By.CSS_SELECTOR, 'kat-button#EditSaveAction')))).click()
                                sleep(2)
                                static = None
                            else:
                                try:
                                    self.sheet.cell(index, 10).fill = PatternFill(fill_type='solid', fgColor='86E507')
                                except:
                                    pass
                                static = 'Some thing go error'
                        else:
                            try:
                                self.sheet.cell(index, 10).fill = PatternFill(fill_type='solid', fgColor='86E507')
                            except:
                                pass
                            message = self.driver.execute_script(
                                "return document.querySelector('#search-result kat-box .limitations-box').textContent")
                            self.sheet.cell(index, 11).value = message

                    else:
                        static = 'غير موجود'
                        try:
                            self.sheet.cell(index, 10).fill = PatternFill(fill_type='solid', fgColor='86E507')
                        except:
                            pass

                    self.sheet.cell(index, 10).value = static
                    print(f'- {index} - {barcode} : {price} : {quantity} : {static}')
                except:
                    print(traceback.print_exc())

        self.book.save(filename)


if __name__ == '__main__':
    while True:
        try:
            filename = input('- Enter file name : ').strip()
            load_workbook(filename).close()
            break
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)
    latency = input('- Enter latency number(8): ')
    self = Main()
    input('- Enter .....')
    self.start()

# Assad098765
# bh2030098765@gmail.com
'''
perfect203070@gmail.com
Assad102030

'''