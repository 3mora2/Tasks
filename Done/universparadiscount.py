from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from selenium import webdriver
from time import sleep
from openpyxl.drawing.image import Image
from io import BytesIO
from requests_html import HTMLSession, HTML

session = HTMLSession()


def print_percent_done(index, total, bar_len=50):
    percent_done = (index-1)/total*100
    percent_done = round(percent_done, 1)

    done = round(percent_done/(100/bar_len))
    togo = bar_len-done

    done_str = '█'*int(done)
    togo_str = '░'*int(togo)

    print(f'- ⏳ {total}\\{index - 1} - [{done_str}{togo_str}] ({percent_done} %)', end='\r')


class Main:
    def __init__(self):
        i = 1
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.cell(i, 1).value = 'IMG'
        self.sheet.column_dimensions['A'].width = 11
        self.sheet.cell(i, 2).value = 'Name'
        self.sheet.cell(i, 3).value = 'Marque'
        self.sheet.cell(i, 4).value = 'Price'
        self.sheet.cell(i, 5).value = 'Description'
        self.sheet.cell(i, 6).value = 'Marque IMG'
        self.sheet.cell(i, 7).value = 'IMG URL'
        self.sheet.cell(i, 8).value = 'URL'

    def Start(self):
        self.driver.get(url_main)
        # self.driver.get('https://universparadiscount.ma/4-visage#/page-130')
        links = list()
        while True:
            sleep(1)
            for element in self.driver.find_elements_by_css_selector('li a.product-name'):
                links.append(element.get_attribute('href'))
                print(f'- Number Of Links : {len(links)}', end='\r')

            if 'disabled' not in self.driver.find_element_by_css_selector('#pagination_next_bottom'
                                                                          ).get_attribute('class'):
                self.driver.find_element_by_css_selector('#pagination_next_bottom > a').click()
                sleep(2)
            else:
                break
        i = 2
        len_l = len(links)
        for link in links:
            self.driver.get(link)
            mark_name = eval(self.driver.find_element_by_css_selector('.fab-perso').get_attribute('innerText')).strip().upper()
            # name = self.driver.find_element_by_css_selector('h1[itemprop="name"]').text
            name = self.driver.find_element_by_css_selector('h1[itemprop="name"]').text.replace(mark_name, '').replace(mark_name.replace(' ', ''), '').strip()
            price = self.driver.find_element_by_css_selector('.content_prices span.price').get_attribute('content')
            mark_img = self.driver.find_element_by_css_selector('#product_manufacturer .product_img_manufacturer'
                                                                ).get_attribute('src')
            desk = self.driver.find_element_by_css_selector('#tab_product_page .tab-content').text
            img_url = '\n'.join([i.get_attribute('src').replace('-cart_default', '') for i in
                             self.driver.find_elements_by_css_selector(
                                 '.product-section > .pb-left-column #views_block img')])
            first_img = img_url.split('\n')[0]
            if first_img != '':
                try:
                    img_data = session.get(first_img)
                    image_file = BytesIO(img_data.content)
                    img = Image(image_file)
                    img.width = 90
                    img.height = 75

                    self.sheet.row_dimensions[i].height = 56
                    self.sheet.add_image(img, 'A{}'.format(i))
                except:
                    pass
            self.sheet.cell(i, 2).value = name
            self.sheet.cell(i, 3).value = mark_name
            self.sheet.cell(i, 4).value = price
            self.sheet.cell(i, 5).value = desk
            self.sheet.cell(i, 6).value = mark_img
            self.sheet.cell(i, 7).value = img_url
            self.sheet.cell(i, 8).value = link
            print_percent_done(i, len_l)
            i += 1

        self.book.save('new_sheet.xlsx')


url_main = input('- Enter URL : ')

app = Main()
app.Start()
self = app
