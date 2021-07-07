from warnings import filterwarnings

filterwarnings("ignore")
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from openpyxl import load_workbook
from random import randint
from time import sleep

import threading


class Thread(threading.Thread):
    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        self.p = None
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class souq_code:
    def __init__(self, file, rang, en=None):
        self.rang = rang.split('-')
        self.file = file
        self.Bar_Code()
        ###########################
        chrome_options = Options()
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-notifications")
        self.driver = webdriver.Chrome('chromedriver.exe',
                                       chrome_options=chrome_options)
        self.driver.maximize_window()
        if en == 'ar':
            self.url = 'https://egypt.souq.com//eg-ar/'
        else:
            self.url = 'https://egypt.souq.com//eg-en/'
        self.driver.get(self.url)
        sleep(randint(4, 9))

    def Bar_Code(self):
        try:
            path = self.file
            self.book = load_workbook(path)
            self.sheet = self.book.active
            end_cell = self.sheet.max_row
            try:
                if len(self.rang) == 2:
                    self.frst = int(self.rang[0]) if self.rang[0].isnumeric() and int(self.rang[0]) > 1 and int(
                        self.rang[0]) < end_cell else 2
                    self.scnd = int(self.rang[1]) if self.rang[1].isnumeric() and int(self.rang[1]) > 1 and int(
                        self.rang[1]) < end_cell else end_cell
                else:
                    self.frst = 2
                    self.scnd = end_cell
            except:
                self.frst = 2
                self.scnd = end_cell
            self.bar_code = [self.sheet.cell(i, 1).value for i in range(self.frst, self.scnd + 1)]
        except Exception as e:
            print(e)
            exit()

    def Search(self):
        self.index = 1
        self.sheet[f'D{self.index}'] = 'Item Title'
        self.sheet[f'J{self.index}'] = 'New Price'
        self.sheet[f'K{self.index}'] = 'Product Rate'
        self.sheet[f'L{self.index}'] = 'Seller Name'
        self.sheet[f'M{self.index}'] = 'Seller Rate'
        self.sheet[f'N{self.index}'] = 'Seller URL'
        self.sheet[f'O{self.index}'] = 'Fulfilled'
        self.sheet.column_dimensions['A'].width = 15
        self.sheet.column_dimensions['B'].width = 15
        self.sheet.column_dimensions['D'].width = 170
        self.sheet.column_dimensions['J'].width = 13
        self.sheet.column_dimensions['K'].width = 15
        self.sheet.column_dimensions['L'].width = 15
        self.sheet.column_dimensions['M'].width = 20
        self.sheet.column_dimensions['N'].width = 46
        self.sheet.column_dimensions['O'].width = 18

        self.index = self.frst
        ####################
        self.thr = Thread()
        self.thr.start()
        for bar in self.bar_code:
            ##########################################
            if self.thr.p == 'pause' or self.thr.p == 'p':
                start = input('(put start)>>>')
                while True:

                    if start == 'start' or start == 's':
                        self.thr = Thread()
                        self.thr.start()
                        break
                    elif start == 'exit' or start == 'e':
                        exit()
                    else:
                        start = input()
            elif self.thr.p == 'exit' or self.thr.p == 'e':
                exit()
            ##########################################
            self.bar = bar
            self.driver.get(f'{self.url}{bar}/s/')
            sleep(2)
            try:
                _ = self.driver.find_element_by_class_name('header-product-fulfilled').get_attribute('src')
                self.fulfilled = 'Fulfilled by SOUQ'
            except:
                self.fulfilled = None
            try:
                _ = self.driver.find_element_by_class_name('zero-results').text
                self.name_ = self.price_ = self.usrname_ = self.product_rata_ = self.seller_rate_ = self.url_usernam_ = static = 'Not Found'
                print(f'{self.index} - {self.bar} is : {static} ')
                self.save_data()
                continue
            except:
                pass
            try:
                _ = self.driver.find_element_by_class_name('notice').text
                self.name_ = self.driver.find_element_by_class_name('product-title').find_element_by_tag_name('h6').text
                try:
                    self.product_rata_ = self.driver.find_element_by_class_name('avg').text
                except:
                    self.product_rata_ = 'Not Available'
                self.price_ = self.usrname_ = self.seller_rate_ = self.url_usernam_ = static = 'Not Available'
                print(f'{self.index} - {self.bar} is : {static} ')
                self.save_data()
                continue

            except:
                pass
            try:
                static = 'OK'
                self.name_ = self.driver.find_element_by_class_name('product-title').find_element_by_tag_name('h1').text

                try:
                    pri = self.driver.find_element_by_class_name('price-container').find_element_by_class_name('price')
                    v = pri.find_element_by_tag_name('span').text
                    self.price_ = pri.text.replace(v, '').strip()
                except:
                    self.price_ = None
                try:
                    seller = self.driver.find_element_by_class_name('unit-seller-link').find_element_by_tag_name('a')
                    self.usrname_ = seller.text
                    self.url_usernam_ = seller.get_attribute('href')
                except:
                    self.usrname_ = None
                    self.url_usernam_ = None
                try:
                    self.seller_rate_ = self.driver.find_element_by_class_name('seller-rating').text
                except:
                    self.seller_rate_ = None
                try:
                    self.product_rata_ = self.driver.find_element_by_class_name('avg').text
                except:
                    self.product_rata_ = None
                print(f'{self.index} - {self.bar} is : {static}   ')
                self.save_data()
                continue
            except Exception as e:
                print(e)

        self.book.close()
        self.driver.quit()

    def save_data(self):
        self.sheet[f'D{self.index}'] = self.name_
        self.sheet[f'J{self.index}'] = self.price_
        self.sheet[f'K{self.index}'] = self.product_rata_
        self.sheet[f'L{self.index}'] = self.usrname_
        self.sheet[f'M{self.index}'] = self.seller_rate_
        self.sheet[f'N{self.index}'] = self.url_usernam_
        self.sheet[f'O{self.index}'] = self.fulfilled
        self.index += 1
        self.book.save(self.file)


file = str(input('enter name of file Eaxmple(project.xlsx) : '))
lang = str(input('enter ar for Arabic or en for English : '))
rang = str(input('entar range example(2-50) : '))

if __name__ == "__main__":
    my_bot = souq_code(en=lang, file=file, rang=rang)
    my_bot.Search()
