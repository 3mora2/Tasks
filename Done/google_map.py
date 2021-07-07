import warnings

warnings.filterwarnings("ignore")

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium import webdriver

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

from requests_html import HTMLSession
from io import BytesIO
from time import sleep
import subprocess

session = HTMLSession()


class Main:
    def __init__(self):
        while True:
            self.sear = input('Search :')
            if self.sear != '':
                break
        self.search = self.sear.replace(' ', '+')
        self.book = Workbook()
        self.sheet = self.book.active
        try:
            try:
                subprocess.Popen(
                    '"C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222')
            except:
                subprocess.Popen(
                    '"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222')
            chrome_options = Options()
            chrome_options.add_argument("--disable-infobars")
            chrome_options.add_argument("--disable-notifications")
            chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
            self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
        except:
            self.driver = webdriver.Chrome()
        ##############################################################################################################
        self.driver.maximize_window()

    def open(self):
        self.sheet['A1'] = 'Image'
        self.sheet['B1'] = 'Name'
        self.sheet['C1'] = 'Type'
        self.sheet['D1'] = 'Rate'
        self.sheet['E1'] = 'Address'
        self.sheet['F1'] = 'Plus Code'
        self.sheet['G1'] = 'Phone'
        self.sheet['H1'] = 'Website'
        self.sheet['I1'] = "Note"
        self.sheet['J1'] = "URL"
        self.sheet.column_dimensions['A'].width = 13
        self.sheet.column_dimensions['B'].width = 50
        self.sheet.column_dimensions['C'].width = 30
        self.sheet.column_dimensions['D'].width = 10
        self.sheet.column_dimensions['E'].width = 100
        self.sheet.column_dimensions['F'].width = 70
        self.sheet.column_dimensions['G'].width = 25
        self.sheet.column_dimensions['H'].width = 30
        self.sheet.column_dimensions['I'].width = 50
        self.sheet.column_dimensions['J'].width = 120
        for column in range(1, 11):
            try:
                self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except:
                pass

        self.driver.get(f'https://www.google.com/maps/search/{self.search}')
        print('Program is loading,please wait!')
        sleep(5)
        input('- Enter...')
        num = 2
        while True:
            sleep(1)
            num_pl = len(self.driver.find_elements_by_class_name('section-result-content'))
            if num_pl == 0:
                break
            for i in range(0, num_pl):
                address = None
                plus_code = None
                phone_number = None
                website = None
                sleep(1)
                try:
                    w = self.driver.find_elements_by_class_name('section-result-content')[
                        i].find_elements_by_css_selector(
                        'div:nth-child(1) > a.section-result-action.section-result-action-wide')
                    try:
                        if len(w) != 0:
                            website = w[0].get_attribute('href')
                    except:
                        pass
                    self.driver.find_elements_by_class_name('section-result-content')[i].click()
                    sleep(3)
                    try:
                        main_photo = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((
                            By.CSS_SELECTOR,
                            'div.section-hero-header-image-hero-container.collapsible-hero-image > button > img'))).get_attribute(
                            'src')
                    except:
                        main_photo = None

                    try:
                        name = WebDriverWait(self.driver, 4).until(
                            EC.visibility_of_element_located((By.XPATH, '//h1/span[1]'))).text

                    except:
                        name = None

                    try:
                        typ = WebDriverWait(self.driver, 4).until(EC.visibility_of_element_located(
                            (By.CSS_SELECTOR, 'span.section-rating-term > span:nth-child(1) > button'))).text
                    except:
                        typ = None

                    try:
                        rat = WebDriverWait(self.driver, 4).until(
                            EC.visibility_of_element_located((By.CLASS_NAME, 'section-star-display'))).text
                    except:
                        try:
                            rat = self.driver.find_element_by_css_selector(
                                '#pane > div > div.widget-pane-content.scrollable-y > div > div > div.jqnFjrOWMVU__root > div > div.jqnFjrOWMVU__right > div.gm2-display-2')
                        except:
                            rat = None

                    try:
                        note = WebDriverWait(self.driver, 4).until(EC.visibility_of_element_located(
                            (By.CSS_SELECTOR, 'div.section-editorial-attribute-container'))).text
                    except:
                        note = None

                    try:
                        _ = self.driver.find_elements_by_class_name('ugiz4pqJLAG__content')[
                            -1].location_once_scrolled_into_view
                    except:
                        pass

                    try:
                        for element in self.driver.find_elements_by_xpath(
                                '//*[@id="pane"]/div/div[1]/div/div/div/button'):
                            if element.get_attribute('data-tooltip') is not None:
                                if 'العنوان' in element.get_attribute(
                                        'data-tooltip') or 'address' in element.get_attribute('data-tooltip'):
                                    address = element.text
                                    break
                    except:
                        pass

                    try:
                        for element in self.driver.find_elements_by_xpath(
                                '//*[@id="pane"]/div/div[1]/div/div/div/button'):
                            if element.get_attribute('data-tooltip') is not None:
                                if 'الموقع المفتوح' in element.get_attribute(
                                        'data-tooltip') or 'plus code' in element.get_attribute('data-tooltip'):
                                    plus_code = element.text
                                    break
                    except:
                        pass

                    sleep(1)
                    try:
                        for element in self.driver.find_elements_by_xpath(
                                '//*[@id="pane"]/div/div[1]/div/div/div/button'):
                            if element.get_attribute('data-tooltip') is not None:
                                if 'هاتف' in element.get_attribute(
                                        'data-tooltip') or 'phone number' in element.get_attribute('data-tooltip'):
                                    phone_number = element.text
                                    break
                    except:
                        pass

                    try:
                        if website is None:
                            for element in self.driver.find_elements_by_xpath(
                                    '//*[@id="pane"]/div/div[1]/div/div/div/button'):
                                if element.get_attribute('data-tooltip') is not None:
                                    if 'الموقع الإلكتروني' in element.get_attribute(
                                            'data-tooltip') or 'website' in element.get_attribute('data-tooltip'):
                                        website = element.text
                                        break
                    except:
                        pass

                    try:
                        if main_photo is not None:
                            res = session.get(main_photo)
                            image_file = BytesIO(res.content)
                            img = Image(image_file)
                            img.width = 90
                            img.height = 75
                            self.sheet.row_dimensions[num].height = 56
                            self.sheet.add_image(img, f'A{num}')
                    except:
                        pass
                    self.sheet[f'B{num}'] = name
                    self.sheet[f'C{num}'] = typ
                    self.sheet[f'D{num}'] = rat
                    self.sheet[f'E{num}'] = address
                    self.sheet[f'F{num}'] = plus_code
                    self.sheet[f'G{num}'] = phone_number
                    self.sheet[f'H{num}'] = website
                    self.sheet[f'I{num}'] = note
                    try:
                        self.sheet[f'J{num}'] = self.driver.current_url
                    except:
                        pass
                    try:
                        self.book.save(f'{self.sear}.xlsx')
                    except:
                        pass
                    for column in range(2, 11):
                        try:
                            self.sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                               wrap_text=True)
                            self.sheet.cell(num + 1, column).alignment = Alignment(horizontal='center',
                                                                                   vertical='center', wrap_text=True)
                        except:
                            pass
                    print(f'{num - 1} - {name}')
                    num += 1
                    try:
                        WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                            (By.CLASS_NAME, 'section-back-to-list-button'))).click()
                    except:
                        try:
                            WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                                (By.CLASS_NAME, 'section-back-to-list-button'))).click()
                        except:
                            pass
                    sleep(2)
                except:
                    pass
            try:
                self.driver.find_elements_by_class_name('n7lv7yjyC35__button')[1].click()
            except:
                try:
                    self.driver.find_elements_by_class_name('n7lv7yjyC35__button')[1].click()
                except:
                    break
            sleep(1)
        self.driver.quit()


if __name__ == '__main__':
    app = Main()
    app.open()
