from requests_html import HTMLSession
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 'Title'
ws['B1'] = 'Location'
ws['C1'] = 'Subtitle'
ws['D1'] = 'price'
ws['E1'] = 'ref'
ws['F1'] = 'Country'
ws['G1'] = 'Area'
ws['H1'] = 'Classe'
ws['I1'] = 'Rooms'
ws['J1'] = 'Bathrooms'
ws['K1'] = 'Description'
ws['L1'] = 'URL'

ws['M1'] ='Img_url'
#####################################################################################
ur='http://www.horusimmo.com'
session = HTMLSession()
num=2
page=0
while page !=30:
    respons = session.get(f'http://www.horusimmo.com/v02/acheter-bien-maroc/search/start/{page}/type/8')
    page+=5
    html=respons.html
    h=[ur+i.find('a',first=True).attrs['href'] for i in html.find('.jePropImage')]

    for urll in h:
        respons = session.get(urll)
        html=respons.html
        try:
            title=html.xpath('//*[@id="jePropertyDetailed"]/h1/span[1]',first=True).text
        except:
            title=None
        ########################################################
        try:
            location=html.xpath('//*[@id="jePropertyDetailed"]/h1/span[2]',first=True).text
        except:
            location=None
        ########################################################
        try:
            subtitle=html.xpath('//*[@id="jePropertyDetailed"]/div[1]/h2',first=True).text
        except:
            subtitle=None
        ########################################################
        try:
            price=html.xpath('//*[@id="jePropertyDetailed"]/div[1]/div[1]/span[1]',first=True).text
        except:
            price=None
        ########################################################
        try:
            ref=html.xpath('//*[@id="jePropertyDetailed"]/div[2]/span/span',first=True).text.split(':')[-1].strip()
        except:
            ref=None
        ########################################################
        try:
            country=[i.xpath('//p/text()')[-1].strip() for i in html.find('.detailed_line') if 'Pays' in i.text][0]
        except:
            country=None
        ########################################################
        try:
            area=[i.xpath('//p/text()')[-1].strip() for i in html.find('.detailed_line') if 'Surface' in i.text][0]
        except:
            area=None
        ########################################################
        try:
            classe=[i.text.splitlines()[1] for i in html.find('.detailed_line') if 'Classe' in i.text][0]
        except:
            classe=None
        ########################################################
        try:
            rooms=[i.xpath('//p/text()')[-1].strip() for i in html.find('.detailed_line') if 'Chambres' in i.text][0]
        except:
            rooms=None
        ########################################################
        try:
            bathrooms=[i.xpath('//p/text()')[-1].strip() for i in html.find('.detailed_line') if 'Salle de Bain' in i.text][0]
        except:
            bathrooms=None

        #############################################
        try:
            o=html.find('.jeDescription',first=True).text
            description=o[o.find('\n')+1:o.rfind('\n')]
        except:
            description=None
        #############################################
        try:
            pic=''
            img_url=[ur+i.attrs['src'] for i in html.find('#pikame',first=True).xpath('//img')]
            for img in img_url:
                pic=pic+img+'\n'
        except:
            img_url=None
        #############################################
        ws[f'A{num}']=title
        ws[f'B{num}']=location
        ws[f'C{num}']=subtitle
        ws[f'D{num}']=price
        ws[f'E{num}']=ref
        ws[f'F{num}']=country
        ws[f'G{num}']=area
        ws[f'H{num}']=classe
        ws[f'I{num}']=rooms
        ws[f'J{num}']=bathrooms
        ws[f'K{num}']=description
        ws[f'L{num}']=urll
        ws[f'M{num}']=pic

        num+=1

        '''print('title : ',title,'location : ',location,'subtitle : ',subtitle,'price : ',price,'ref : ',
              ref,'country : ',country,'area : ',area,'classe : ',classe,'rooms : ',rooms,'bathrooms : ',bathrooms,'description  : ',description,'img_url : ',img_url)
        print('\n \n')'''

wb.save("Horusimmo-Acheter Villa.xlsx")
