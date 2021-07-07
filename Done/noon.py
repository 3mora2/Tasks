from warnings import filterwarnings

filterwarnings("ignore")
from selenium import webdriver
from openpyxl import load_workbook
from time import sleep
from requests_html import HTML
import threading


class Thread(threading.Thread):
    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        self.p = None
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class SouCode:
    def __init__(self, file, rang):
        self.bar_code_re = []
        self.Rang = rang.split('-')
        self.file = file
        self.url = 'https://www.noon.com/egypt-en/'

    def Search(self):
        print('Extracting Data From Excel........')
        Rang = self.Rang
        try:
            self.book = load_workbook(file)
            self.sheet = self.book.active
            end_cell = self.sheet.max_row
            try:
                if len(Rang) == 2:
                    self.first = int(Rang[0]) if Rang[0].isnumeric() and int(Rang[0]) > 1 and int(
                        Rang[0]) < end_cell else 2
                    scad = int(Rang[1]) if Rang[1].isnumeric() and int(Rang[1]) > 1 and int(
                        Rang[1]) < end_cell else end_cell
                else:
                    self.first = 2
                    scad = end_cell
            except:
                self.first = 2
                scad = end_cell
            self.bar_code = [self.sheet[f'F{i}'].value for i in range(self.first, scad + 1)]
        except Exception as e:
            print(e)
            exit()
        self.index = self.first
        self.sheet[f'AA1'] = 'New Price'
        self.sheet[f'AB1'] = 'Seller Name'
        self.sheet[f'AC1'] = 'Always In Stock'
        self.sheet[f'AD1'] = 'Ships On Time'
        self.sheet[f'AE1'] = 'Express Noon'
        self.sheet.column_dimensions['AA'].width = 10
        self.sheet.column_dimensions['AB'].width = 20
        self.sheet.column_dimensions['AC'].width = 14
        self.sheet.column_dimensions['AD'].width = 14
        self.sheet.column_dimensions['AE'].width = 13
        ####################
        self.thr = Thread()
        self.thr.start()
        print('Getting Data..........')
        self.driver = webdriver.Chrome()
        for bar in self.bar_code:
            self.driver.get(f'{self.url}{bar}/p/')
            # sleep(1)
            r = HTML(html=self.driver.page_source)
            try:
                price = r.find('span.sellingPrice span.value', first=True).text
            except:
                price = 'Not Found'
            try:
                seller = r.find('p.sellerName a', first=True).text
            except:
                seller = 'Not Found'
            try:
                express = r.find('div.bottomRow div.container img', first=True).attrs['alt']
            except:
                express = 'Not Found'
            try:
                rata1 = r.find('div.container > div> p:nth-child(3)')[0].text
            except:
                rata1 = 'Not Found'
            try:
                rata2 = r.find('div.container > div> p:nth-child(3)')[1].text
            except:
                rata2 = 'Not Found'

            print(self.index, '-', bar, '>', price)
            self.sheet[f'AA{self.index}'] = price
            self.sheet[f'AB{self.index}'] = seller
            self.sheet[f'AC{self.index}'] = rata1
            self.sheet[f'AD{self.index}'] = rata2
            self.sheet[f'AE{self.index}'] = express
            self.index += 1
            self.book.save(self.file)
            ##########################################
            if self.thr.p == 'pause' or self.thr.p == 'p':
                start = input('(put start)>>>')
                while True:
                    if start == 'start' or start == 's':
                        self.thr = Thread()
                        self.thr.start()
                        break
                    elif start == 'exit' or start == 'e':
                        exit()
                    else:
                        start = input()
            elif self.thr.p == 'exit' or self.thr.p == 'e':
                exit()
            ##########################################
        self.book.close()
        self.driver.quit()

    def None_re(self):
        print('Please Wait While Detect None Value')
        try:
            self.book = load_workbook(file)
            self.sheet = self.book.active
        except Exception as e:
            print(e)
            exit()
        for i in range(2, self.sheet.max_row):
            if self.sheet[f'AA{i}'].value == 'Not Found' or \
                    self.sheet[f'AB{i}'].value == 'Not Found' or self.sheet[f'AC{i}'].value == 'Not Found':
                if self.sheet[f'F{i}'].value not in self.bar_code_re:
                    self.bar_code_re.append((self.sheet[f'F{i}'].value, i))
        print('Number of Non Found value is : ', len(self.bar_code_re))
        if len(self.bar_code_re) != 0:
            self.driver = webdriver.Chrome()
            for elem in self.bar_code_re:
                self.index = elem[1]
                self.driver.get(f'{self.url}{elem[0]}/p/')
                sleep(3)
                r = HTML(html=self.driver.page_source)
                try:
                    price = r.find('span.sellingPrice span.value', first=True).text
                except:
                    price = None
                try:
                    seller = r.find('p.sellerName a', first=True).text
                except:
                    seller = None
                try:
                    express = r.find('div.bottomRow div.container img', first=True).attrs['alt']
                except:
                    express = None
                print(self.index - 1, price, seller, express)
                self.sheet[f'AA{self.index}'] = price
                self.sheet[f'AB{self.index}'] = seller
                self.sheet[f'AC{self.index}'] = express
                self.book.save(self.file)

            self.driver.quit()


if __name__ == "__main__":
    file = str(input('Enter the name of file, For example( P.update-2.xlsx ) : ')).strip()
    re_find = str(input('Enter 1 to extract Data, 2 to extract Non Found value : ')).strip()
    if re_find == '1':
        rang = str(input('Enter range, For example(2-50) : '))
        my_bot = SouCode(file=file, rang=rang)
        my_bot.Search()
    elif re_find == '2':
        rang = ''
        my_bot = SouCode(file=file, rang=rang)
        my_bot.None_re()
    else:
        print('Wrong Choice')
        rang = ''
        exit()
