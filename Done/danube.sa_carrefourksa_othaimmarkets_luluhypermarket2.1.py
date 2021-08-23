import subprocess
from time import sleep

from openpyxl.styles import Alignment
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from openpyxl import Workbook
from requests_html import HTML, HTMLSession
import json
from datetime import datetime
import os
import sys
import warnings

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")

session = HTMLSession()


class Carrefour:
    def __init__(self):
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.cell(row=1, column=1).value = 'Name'
        self.sheet.cell(row=1, column=2).value = 'Brand'
        self.sheet.cell(row=1, column=3).value = 'Category'
        self.sheet.cell(row=1, column=4).value = 'Price'
        self.sheet.cell(row=1, column=5).value = 'Original Price'
        self.sheet.cell(row=1, column=6).value = 'Supplier'
        self.sheet.cell(row=1, column=7).value = 'URL'
        self.sheet.column_dimensions['A'].width = 60
        self.sheet.column_dimensions['B'].width = 20
        self.sheet.column_dimensions['C'].width = 15
        self.sheet.column_dimensions['D'].width = 6
        self.sheet.column_dimensions['E'].width = 10
        self.sheet.column_dimensions['F'].width = 9
        self.sheet.column_dimensions['G'].width = 170
        for column in range(1, 8):
            try:
                self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                 wrap_text=True)
            except:
                pass

    def start(self, url, size):

        file_save = 'Carrefour-' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + '.xlsx'

        r = session.get(url)
        data = r.html.find('#__NEXT_DATA__', first=True).text
        data = json.loads(data)
        category_id = data['props']['initialProps']['pageProps']['query']['categoryId']
        area_code = data['props']['initialProps']['pageProps']['appConfig']['areaCode']
        token = data['props']['initialProps']['pageProps']['appConfig']['accessToken']
        num = 0
        page = 0
        list_element = []
        print(category_id)
        while True:
            r = session.get(
                f"https://www.carrefourksa.com/api/v5/zones/302/search/categories/{category_id}?filter=&sortBy=relevance&"
                f"currentPage={page}&pageSize={size}&"
                f"areaCode=Granada%20-%20Riyadh&lang=ar&expressPos=302&displayCurr=SAR"
                # "&foodPos=302&nonFoodPos=302&latitude=24.7136&longitude="
                ,
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0",
                    "Accept": "*/*",
                    "Accept-Language": "en-US,en;q=0.5",
                    "env": "prod",
                    "appId": "Reactweb",
                    "token": token,
                    "credentials": "include",
                    "storeId": "mafsau",
                    "Sec-Fetch-Dest": "empty",
                    "Sec-Fetch-Mode": "cors",
                    "Sec-Fetch-Site": "same-origin",
                    "userid": ""
                })
            page += 1
            if len(r.json()['products']) == 0:
                break
            for element in r.json()['products']:
                if element in list_element:
                    print('- found')
                name = element['name']
                brand = element['brand']['name']
                category = element['category'][-1]['name']
                price = element['price']['price']
                try:
                    discount = element['price']['discount']['price']
                except:
                    discount = None
                supplier = element['supplier']
                product_url = 'https://www.carrefourksa.com' + element['links']['productUrl']['href']
                list_element.append(element)
                self.sheet.cell(row=num + 2, column=1).value = name
                self.sheet.cell(row=num + 2, column=2).value = brand
                self.sheet.cell(row=num + 2, column=3).value = category
                self.sheet.cell(row=num + 2, column=4).value = discount
                self.sheet.cell(row=num + 2, column=5).value = price
                self.sheet.cell(row=num + 2, column=6).value = supplier
                self.sheet.cell(row=num + 2, column=7).value = product_url
                for column in range(1, 8):
                    try:
                        self.sheet.cell(num + 2, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                               wrap_text=True)
                    except:
                        pass
                print(num, name)
                num += 1

        self.book.save(file_save)


class OthaimMarkets:
    def __init__(self):
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.cell(row=1, column=1).value = 'Name'
        self.sheet.cell(row=1, column=2).value = 'Category'
        self.sheet.cell(row=1, column=3).value = 'Price'
        self.sheet.cell(row=1, column=4).value = 'Old Price'
        self.sheet.cell(row=1, column=5).value = 'URL'
        self.sheet.column_dimensions['A'].width = 60
        self.sheet.column_dimensions['B'].width = 15
        self.sheet.column_dimensions['C'].width = 6
        self.sheet.column_dimensions['D'].width = 10
        self.sheet.column_dimensions['E'].width = 80
        for column in range(1, 6):
            try:
                self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                 wrap_text=True)
            except:
                pass
        self.driver = webdriver.Chrome(ChromeDriverManager().install())

    def start(self, url):
        file_save = 'OthaimMarkets-' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + '.xlsx'
        list_u = []
        num = 2
        self.driver.get(url)
        old_product = 0
        while True:
            if 'block' in self.driver.find_element_by_css_selector('#preloader .loader').get_attribute('style'):
                sleep(3)
                continue
            sleep(2)
            if old_product == len(self.driver.find_elements_by_css_selector('.product')):
                break
            old_product = len(self.driver.find_elements_by_css_selector('.product'))
            print('- Total', old_product)
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            sleep(6)

        cat = self.driver.find_element_by_css_selector('.last-breadcrumb').text
        for element in self.driver.find_elements_by_css_selector('.product'):
            name = element.find_element_by_css_selector('h3 a').text
            url = element.find_element_by_css_selector('h3 a').get_attribute('href')
            try:
                price = element.find_element_by_css_selector('.special-price .price').text.replace('ريال', '')
                old_price = element.find_element_by_css_selector('.old-price .price').text
            except:
                price = element.find_element_by_css_selector('.price').text.replace('ريال', '')
                old_price = None

            if url in list_u:
                print('- found')
                continue

            list_u.append(url)
            self.sheet.cell(row=num, column=1).value = name
            self.sheet.cell(row=num, column=2).value = cat
            self.sheet.cell(row=num, column=3).value = price
            self.sheet.cell(row=num, column=4).value = old_price
            self.sheet.cell(row=num, column=5).value = url
            for column in range(1, 6):
                try:
                    self.sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                       wrap_text=True)
                except:
                    pass
            print(num, name)
            num += 1

        self.book.save(file_save)


class LuluhyperMarket:
    def __init__(self):
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-gpu')
        # chrome_options.add_argument('blink-settings=imagesEnabled=false')
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
        self.driver.maximize_window()
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.cell(row=1, column=1).value = 'Name'
        self.sheet.cell(row=1, column=2).value = 'Brand'
        self.sheet.cell(row=1, column=3).value = 'Category'
        self.sheet.cell(row=1, column=4).value = 'Price'
        self.sheet.cell(row=1, column=5).value = 'Original Price'
        self.sheet.cell(row=1, column=6).value = 'INFO'
        self.sheet.cell(row=1, column=7).value = 'URL'
        self.sheet.column_dimensions['A'].width = 60
        self.sheet.column_dimensions['B'].width = 10
        self.sheet.column_dimensions['C'].width = 17
        self.sheet.column_dimensions['D'].width = 10
        self.sheet.column_dimensions['E'].width = 12
        self.sheet.column_dimensions['F'].width = 12
        self.sheet.column_dimensions['G'].width = 80
        for column in range(1, 7):
            try:
                self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                 wrap_text=True)
            except:
                pass

    def start(self, url, s):
        file_save = 'LuluhyperMarket-' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + '.xlsx'
        list_u = []
        self.driver.get(url)
        sleep(3)
        n = 0
        while True:
            # for element in self.driver.find_elements_by_css_selector('.product-tile-main a.js-gtm-product-link'):
            for element in self.driver.find_elements_by_css_selector('.product-img > a.js-gtm-product-link'):
                if element.get_attribute('href') in list_u:
                    continue
                list_u.append(element.get_attribute('href'))

            print('total:', len(list_u))
            if len(self.driver.find_elements_by_css_selector('#loaderonplp')) > 0:
                if 'd-none' not in self.driver.find_element_by_css_selector('#loaderonplp').get_attribute('class'):
                    sleep(1)
                    continue
                sleep(1)
                if n == len(self.driver.find_elements_by_css_selector('.product-img > a.js-gtm-product-link')):
                    break
                n = len(self.driver.find_elements_by_css_selector('.product-img > a.js-gtm-product-link'))
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                sleep(2)
            else:
                try:
                    self.driver.get(self.driver.find_element_by_css_selector('a[rel="next"]').get_attribute('href'))
                    sleep(1)
                except Exception as e:
                    print(e)
                    break

        for num, link in enumerate(list_u):
            try:
                self.driver.get(link)
                sleep(1)
                title = self.driver.find_element_by_css_selector('h1').text
                try:
                    brand = self.driver.find_element_by_css_selector('[property="product:brand"]').get_attribute('content')
                except:
                    brand = None
                try:
                    cat = self.driver.find_elements_by_css_selector('nav>ol>li>a')[-1].text
                except:
                    cat = None
                try:
                    old_price = self.driver.find_element_by_css_selector('.price-tag.detail .off').text.replace('AED','').strip()
                except:
                    old_price = None
                try:
                    # price = self.driver.find_elements_by_css_selector('.prod-price > ul > li > span')[-1].text
                    price = self.driver.find_element_by_css_selector('[property="product:price:amount"]').get_attribute('content')
                except:
                    price = None
                try:
                    inf = self.driver.find_element_by_css_selector('#home').text
                except:
                    inf = None
                self.sheet.cell(row=num + 2, column=1).value = title
                self.sheet.cell(row=num + 2, column=2).value = brand
                self.sheet.cell(row=num + 2, column=3).value = cat
                self.sheet.cell(row=num + 2, column=4).value = price
                self.sheet.cell(row=num + 2, column=5).value = old_price
                self.sheet.cell(row=num + 2, column=6).value = inf
                self.sheet.cell(row=num + 2, column=7).value = link
                for column in range(1, 8):
                    try:
                        self.sheet.cell(num + 2, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                               wrap_text=True)
                    except:
                        pass
                print(num, title)
            except:
                pass
        self.book.save(file_save)


class Danube:
    def __init__(self):
        self.user_data = os.path.expanduser("~") + r'\AppData\Local\Google\Chrome\User Data'
        try:
            pros = subprocess.Popen(
                '"C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"'
                rf' --user-data-dir="{self.user_data}" '
                ' --remote-debugging-port=9222')
        except:
            pros = subprocess.Popen(
                '"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"'
                rf' --user-data-dir="{self.user_data}"'
                '  --remote-debugging-port=9222')
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.cell(row=1, column=1).value = 'Name'
        self.sheet.cell(row=1, column=2).value = 'Category'
        self.sheet.cell(row=1, column=3).value = 'Price'
        self.sheet.cell(row=1, column=4).value = 'Original Price'
        self.sheet.cell(row=1, column=5).value = 'URL'
        self.sheet.column_dimensions['A'].width = 60
        self.sheet.column_dimensions['B'].width = 15
        self.sheet.column_dimensions['C'].width = 10
        self.sheet.column_dimensions['D'].width = 15
        self.sheet.column_dimensions['E'].width = 100
        for column in range(1, 7):
            try:
                self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                 wrap_text=True)
            except:
                pass

    def start(self, url):
        file_save = 'Danube-' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + '.xlsx'
        list_u = []
        num = 0
        self.driver.get(url)
        sleep(3)
        input('- enter: ')
        while True:
            cat = self.driver.find_element_by_css_selector('h1').text
            for element in self.driver.find_elements_by_css_selector('.product-box'):
                url = element.find_element_by_css_selector('a').get_attribute('href')
                if url in list_u:
                    continue
                title = element.find_element_by_css_selector('.product-box__name').text
                price = element.find_element_by_css_selector('.product-price__current-price').text
                try:
                    old_price = element.find_element_by_css_selector('.product-price__original-price__span').text
                except:
                    old_price = None
                self.sheet.cell(row=num + 2, column=1).value = title
                self.sheet.cell(row=num + 2, column=2).value = cat
                self.sheet.cell(row=num + 2, column=3).value = price if old_price else None
                self.sheet.cell(row=num + 2, column=4).value = old_price if old_price else price
                self.sheet.cell(row=num + 2, column=5).value = url

                for column in range(1, 7):
                    try:
                        self.sheet.cell(num + 2, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                               wrap_text=True)
                    except:
                        pass
                print(num + 2, title)
                num += 1
                list_u.append(url)

            try:
                self.driver.execute_script("document.querySelector('a[aria-label=\"Next\"]').click()")
                sleep(5)
            except:
                break

        self.book.save(file_save)


if __name__ == '__main__':
    while True:
        init = input('- 1 >>> OthaimMarkets\n'
                     '- 2 >>> Carrefour \n'
                     '- 3 >>> LuluhyperMarket \n'
                     '- 4 >>> Danube \n'
                     '- warning.... to use 4 you must close chrome if has open '
                     ': '
                     )
        if init == '1':
            urls = input('- url: ')  # https://www.othaimmarkets.com/baby-needs-4.html
            oself = OthaimMarkets()
            oself.start(urls)

        elif init == '2':
            cself = Carrefour()
            urls = input('- url: ')
            try:
                s = int(input('- enter number: '))
            except:
                s = 60
            cself.start(urls, s)
        elif init == '3':
            lself = LuluhyperMarket()
            url = input(
                '- url: ')  # 'https://www.luluhypermarket.com/en-sa/department-store-mobiles-mobiles-tablets/c/HY00214728'
            try:
                s = int(input('- enter second number: '))
            except:
                s = 30
            lself.start(url, s)
        elif init == '4':
            dself = Danube()
            url = input('- url: ')  # 'https://danube.sa/en?dismiss_popover=true'
            dself.start(url)

# pyinstaller --add-data C:/Users/3mora/AppData/Local/Programs/Python/Python39/Lib/site-packages/pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info
# pyinstaller --add-data c:\users\3mora\anaconda3\envs\autoenv\lib\site-packages\pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info
'https://www.luluhypermarket.com/en-saBreakfast%20&%20Spreads/c/HY00214912'