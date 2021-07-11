from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from tkinter import filedialog
from datetime import datetime
from io import BytesIO
from time import sleep
import urllib3

urllib3.disable_warnings()
import urllib, os, requests, json, re


# kwefii 6:47--------9:49   3500 43
# kevindavid
# hasanalhalaby
# azizbasha1 6:26
headers_and = {
    'User-Agent': 'Instagram 76.0.0.15.395 Android (24/7.0; 640dpi; 1440x2560; samsung; SM-G930F; herolte; samsungexynos8890; en_US; 138226743)',
    'Connection': 'close', 'Accept': '*/*', 'Accept-Language': 'en-US', 'Accept-Encoding': 'gzip, deflate',
    'X-IG-Capabilities': '3brTvw==', 'X-IG-Connection-Type': 'WIFI', 'X-IG-Connection-Speed': '4932kbps',
    'X-IG-App-ID': '567067343352427', 'X-IG-Bandwidth-Speed-KBPS': '-1.000', 'X-IG-Bandwidth-TotalBytes-B': '0',
    'X-IG-Bandwidth-TotalTime-MS': '0', 'X-FB-HTTP-Engine': 'Liger'}

class Insta:
    def __init__(self):
        print('''
----- plze wait while starting ------''')
        self.images_path = None
        self.videos_path = None
        self.numm = []
        ######################################################################
        #######################################################################
        self.session1 = self.login(username='201028946519', password='ammar2020')
        sleep(1)
        self.session2 = self.login(username='ebrahemelmorsy22001@gmail.com', password='ammar2020')
        sleep(1)
        # self.session3 = self.login(username='ebrahemelmorsy22002@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session4 = self.login(username='pywpywpy', password='pYpY2222')
        # sleep(1)
        # self.session5 = self.login(username='pywpywpywpy', password='pYpY2222')
        # sleep(1)
        # self.session6 = self.login(username='moduhondu@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session7 = self.login(username='uhiotyui6@gmail.com', password='ammar2020')
        # sleep(1)

        # self.session8 = self.login(username='uiuyhjyy@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session9 = self.login(username='jvbjyhg@gmail.com', password='ammar2020')
        # sleep(1)
        # self.session10 = self.login(username='shakermouty@gmail.com', password='ammar2020')
        # sleep(1)
    #     (username='3mora534', password='ammar2200')

    #####################################################################
    #####################################################################
    def login(self, username, password):
        session = requests.Session()
        headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
                   'Accept-Language': 'en-GB,en;q=0.9,en-US;q=0.8,ml;q=0.7',
                   'User-Agent': 'Mozilla/S.0 (X11; Linux x86_64) AppleWebKit/S37.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36'}
        session.headers.update(headers)
        session.get('https://www.instagram.com/web/__mid/')
        csrf_token = session.cookies.get_dict()['csrftoken']
        session.headers.update({'X-CSRFToken': csrf_token})
        enc_password = '#PWD_INSTAGRAM_BROWSER:0:{}:{}'.format(int(datetime.now().timestamp()), password)
        resp = session.post('https://www.instagram.com/accounts/login/ajax/',
                            data={'enc_password': enc_password, 'username': username},
                            allow_redirects=True)  # ,proxies=proxyDict)
        print(resp.status_code)
        ###################################################################################################
        try:
            resp_json = resp.json()
        except Exception as e:
            print("Login error: JSON decode fail, {} - {}.".format(resp.status_code, resp.reason))
        try:
            if resp_json.get('checkpoint_url'):
                print("Login: Checkpoint required. Point your browser to "
                      "https://www.instagram.com{} - "
                      "follow the instructions, then retry.".format(resp_json.get('checkpoint_url')))
            if resp_json['status'] != 'ok':
                if 'message' in resp_json:
                    print("Login error: \"{}\" status, message \"{}\".".format(resp_json['status'],
                                                                               resp_json['message']))
                else:
                    print("Login error: \"{}\" status.".format(resp_json['status']))
            if 'authenticated' not in resp_json:
                if 'message' in resp_json:
                    print("Login error: Unexpected response, \"{}\".".format(resp_json['message']))
                else:
                    print("Login error: Unexpected response, this might indicate a blocked IP.")
            if not resp_json['authenticated']:
                if resp_json['user']:
                    print('Login error: Wrong password.')
                else:
                    print('Login error: User {} does not exist.'.format(username))
        except:
            pass
        return session

    ########################################################################################################################
    ########################################################################################################################
    def sessions(self, b):
        if b == 1:
            session = self.session1
        elif b == 2:
            session = self.session2
        # elif b == 3:
        #     session = self.session3
        # elif b == 4:
        #     session = self.session10
        # elif b == 5:
        #     session = self.session1
        # elif b == 6:
        #     session = self.session6
        # elif b == 7:
        #     session = self.session7
        # elif b == 8:
        #     session = self.session8
        # elif b == 9:
        #     session = self.session9
        else:
            session = self.session10
        return session

    ########################################################################################################################
    ########################################################################################################################
    def get_followers(self):
        try:
            while True:
                url = input(' put username  ( example:- chhc  ) : ')  # chhc
                if url != '':
                    break
                else:
                    pass
            if url.find('instagram.com/') > -1:
                username = url.split('com/')[1].split('/')[0]
            else:
                username = url
            roots = input(' if you want excel with img put 1 or excel with url of img put 2 ')
            if roots == '1':
                self.roots = 1
            else:
                self.roots = 2
            now1 = datetime.now()
            start_time = now1.strftime("%H:%M:%S")
            '''try:
                path = "Insta_Followers.xlsx"
                self.book = load_workbook(path)
                self.sheet = self.book.active
                self.end_cell = self.sheet.max_row
                self.root=1
            except:#48
                self.book = Workbook()
                self.sheet = self.book.active
                self.sheet['B1'] = 'user name'
                self.sheet.column_dimensions['B'].width = 21
                self.sheet['D1'] = 'id'
                self.sheet.column_dimensions['D'].width = 13
                self.sheet['C1'] = 'full name'
                self.sheet.column_dimensions['C'].width = 34
                self.sheet['A1'] = 'profile pic'
                #self.sheet.column_dimensions['A'].width = 60
                self.sheet.column_dimensions['A'].width = 13
                self.root=2
                self.end_cell=None'''
            print('getting followers')
            p = 1
            b = 1
            u = True
            end_cursor = None
            #####################################################################
            strHtml = self.session1.get("https://www.instagram.com/%s" % username).text
            strjson = strHtml[strHtml.index('window._sharedData =') + 20:]
            strjson = strjson[0:strjson.index(';</script>')]
            jObject = json.loads(strjson)
            user_id = jObject['entry_data']['ProfilePage'][0]['graphql']['user']['id']
            print(user_id)
            resp = self.gf(user_id)
            #####################################################################
        #     while u:
        #         if resp != 'erorr':
        #             u = resp['data']['user']['edge_followed_by']['page_info']['has_next_page']
        #             end_cursor = resp['data']['user']['edge_followed_by']['page_info']['end_cursor']
        #             for i in resp['data']['user']['edge_followed_by']['edges']:
        #                 user = i['node']['username']
        #                 '''id=i['node']['id']
        #                 full_name=i['node']['full_name']
        #                 profile_pic_url=i['node']['profile_pic_url']'''
        #                 print('{0}:{1}'.format(p, user))
        #                 self.user_info(user, b)
        #                 #################################################################################
        #                 # self.save_data(p=p,root=self.root,end_cell=self.end_cell,data=[user,id,full_name,profile_pic_url])
        #                 ##################################################################################
        #
        #                 p += 1
        #                 if b == None:
        #                     b = 1
        #                 elif b < 10:
        #                     b += 1
        #                 elif b == 10:
        #                     b = 1
        #
        #             resp = self.gf(user_id=user_id, end_cursor=end_cursor, b=b)
        #         else:
        #             print('*********** plze wait **************')
        #             sleep(25)
        #             if b == None:
        #                 b = 1
        #             elif b < 10:
        #                 b += 1
        #             elif b == 10:
        #                 b = 1
        #             resp = self.gf(user_id=user_id, end_cursor=end_cursor, b=b)
        #         # self.book.save('Insta_Followers.xlsx')

            while u:
                try:
                    if b is None:
                        b = 1
                    elif b < 10:
                        b += 1
                    elif b == 10:
                        b = 1

                    if resp != 'error':
                        if len(resp['data']['user']['edge_followed_by']['edges']) != 0:
                            u = resp['data']['user']['edge_followed_by']['page_info']['has_next_page']
                            end_cursor = resp['data']['user']['edge_followed_by']['page_info']['end_cursor']
                            for i in resp['data']['user']['edge_followed_by']['edges']:
                                user = i['node']['username']
                                Id = i['node']['id']
                                # self.user_info(Id, b)
                                print(f'{p} : {user}')
                                p += 1
                        else:
                            print('- Please Wait....')
                            sleep(5)

                        sleep(1.5)
                        resp = self.gf(user_id=user_id, end_cursor=end_cursor, b=b)
                    else:
                        print('- Error, Please Wait....')
                        sleep(40)
                        if b is None:
                            b = 1
                        elif b < 10:
                            b += 1
                        elif b == 10:
                            b = 1
                        resp = self.gf(user_id=user_id, end_cursor=end_cursor, b=b)
                except Exception as e:
                    print(e)

                except:
                    pass
        except Exception as e:
            print(e)
        # self.book.save('Insta_Followers.xlsx')
        # self.book.close()
        now = datetime.now()
        end_time = now.strftime("%H:%M:%S")
        print(start_time, end_time)
        print('--Done--')
        print('Number of Followers {0}'.format(resp['data']['user']['edge_followed_by']['count']))
        print(self.numm)

    #####################################################################
    def gf(self, user_id, end_cursor=None, b=None):
        session = self.session2
        if end_cursor is None:
            strHtml = session.get(
                f"https://www.instagram.com/graphql/query/?query_id=17851374694183129&id={user_id}&first=100")
        else:
            strHtml = session.get(
                f"https://www.instagram.com/graphql/query/?query_id=17851374694183129&id={user_id}&first=100&after={end_cursor}")
        if strHtml.ok:
            resp = json.loads(strHtml.text)
            return resp
        else:
            print(strHtml.status_code)
            return 'error'

    ########################################################################################################################
    ########################################################################################################################
    def get_following(self):
        now1 = datetime.now()
        start_time = now1.strftime("%H:%M:%S")
        try:
            while True:
                url = input('put username  ( example:- chhc  ) : ')  # chhc
                if url != '':
                    break
                else:
                    pass
            if url.find('instagram.com/') > -1:
                username = url.split('com/')[1].split('/')[0]
            else:
                username = url
            roots = input(' if you want excel with img put 1 or excel with url of img put 2 ')
            if roots == '1':
                self.roots = 1
            else:
                self.roots = 2
            path = "Insta_Following.xlsx"
            self.book = load_workbook(path)
            self.sheet = self.book.active
            self.end_cell = self.sheet.max_row
            self.root = 1
        except:
            self.book = Workbook()
            self.sheet = self.book.active
            self.sheet['A1'] = 'profile pic'
            self.sheet.column_dimensions['A'].width = 13
            self.sheet['B1'] = 'user name'
            self.sheet.column_dimensions['B'].width = 13
            self.sheet['C1'] = 'full name'
            self.sheet.column_dimensions['C'].width = 34
            self.sheet['D1'] = 'id'
            # self.sheet.column_dimensions['D'].width = 60
            self.sheet.column_dimensions['D'].width = 21
            self.root = 2
            self.end_cell = None
        print(' getting following ')
        p = 1
        b = None
        u = True
        end_cursor = None
        #####################################################################
        strHtml = self.session1.get("https://www.instagram.com/%s" % username).text
        strjson = strHtml[strHtml.index('window._sharedData =') + 20:]
        strjson = strjson[0:strjson.index(';</script>')]
        jObject = json.loads(strjson)
        user_id = jObject['entry_data']['ProfilePage'][0]['graphql']['user']['id']
        resp = self.gfing(user_id)
        print('Number of Following {}'.format(resp['data']['user']['edge_follow']['count']))
        #####################################################################
        # while u:
        #     if resp != 'erorr':
        #         u = resp['data']['user']['edge_follow']['page_info']['has_next_page']
        #         end_cursor = resp['data']['user']['edge_follow']['page_info']['end_cursor']
        #         for i in resp['data']['user']['edge_follow']['edges']:
        #             user = i['node']['username']
        #             id = i['node']['id']
        #             full_name = i['node']['full_name']
        #             profile_pic_url = i['node']['profile_pic_url']
        #             print('{0}:{1}'.format(p, user))
        #             # self.user_info(user)
        #             #################################################################################
        #             # self.save_data(p=p, root=self.root, end_cell=self.end_cell,data=[user, id, full_name, profile_pic_url])
        #             ##################################################################################
        #             p += 1
        #         if b == None:
        #             b = 1
        #         elif b < 10:
        #             b += 1
        #         elif b == 10:
        #             b = 1
        #         resp = self.gfing(user_id=user_id, end_cursor=end_cursor, b=b)
        #     else:
        #         print('*********** plze wait **************')
        #         sleep(15)
        #         if b == None:
        #             b = 1
        #         elif b < 10:
        #             b += 1
        #         elif b == 10:
        #             b = 1
        #         resp = self.gfing(user_id=user_id, end_cursor=end_cursor, b=b)
            # self.book.save('Insta_Following.xlsx')
        while u:
            try:
                if b is None:
                    b = 1
                elif b < 10:
                    b += 1
                elif b == 10:
                    b = 1

                if resp != 'error':
                    if len(resp['data']['user']['edge_follow']['edges']) != 0:
                        u = resp['data']['user']['edge_follow']['page_info']['has_next_page']
                        end_cursor = resp['data']['user']['edge_follow']['page_info']['end_cursor']
                        for i in resp['data']['user']['edge_follow']['edges']:
                            user = i['node']['username']
                            self.user_info(user, b)
                            print(f'{p} : {user}')
                            p += 1
                    else:
                        print('- Please Wait....')
                        sleep(5)

                    sleep(1.5)
                    resp = self.gfing(user_id=user_id, end_cursor=end_cursor, b=b)
                else:
                    print('- Error, Please Wait....')
                    sleep(40)
                    if b is None:
                        b = 1
                    elif b < 10:
                        b += 1
                    elif b == 10:
                        b = 1
                    resp = self.gfing(user_id=user_id, end_cursor=end_cursor, b=b)
            except Exception as e:
                print(e)

            except:
                pass
        self.book.close()
        print('--Done--')
        now = datetime.now()
        end_time = now.strftime("%H:%M:%S")
        print(start_time, end_time)
        print(self.numm)

    #####################################################################
    def gfing(self, user_id, end_cursor=None, b=None):
        session = self.sessions(b)
        #########################
        if end_cursor == None:
            strHtml = session.get(
                'https://www.instagram.com/graphql/query/?query_id=17874545323001329&id={}&first=100'.format(user_id))
        else:
            strHtml = session.get(
                'https://www.instagram.com/graphql/query/?query_id=17874545323001329&id={}&first=100&after={}'.format(
                    user_id, end_cursor))
        #########################
        if strHtml.ok:
            resp = json.loads(strHtml.text)
            return resp
        else:
            print(strHtml.status_code)
            return 'erorr'

    #####################################################################
    #####################################################################
    def save_data(self, p, root, end_cell=None, data=[]):
        if root == 1:
            self.sheet.cell(row=(end_cell + p), column=2).value = data[0]
            self.sheet.cell(row=(end_cell + p), column=4).value = data[1]
            self.sheet.cell(row=(end_cell + p), column=3).value = data[2]
            if self.roots != 1:
                self.sheet.cell(row=(end_cell + p), column=1).value = data[3]
            #############################################################
            else:
                http = urllib3.PoolManager()
                r = http.request('GET', data[3])
                image_file = BytesIO(r.data)
                img = Image(image_file)
                img.width = 90
                img.height = 75

                self.sheet.row_dimensions[end_cell + p].height = 56
                self.sheet.add_image(img, 'A{}'.format(end_cell + p))

        else:
            self.sheet.cell(row=p + 1, column=2).value = data[0]
            self.sheet.cell(row=p + 1, column=4).value = data[1]
            self.sheet.cell(row=p + 1, column=3).value = data[2]
            if self.roots != 1:
                self.sheet.cell(row=p + 1, column=1).value = data[3]
            else:
                http = urllib3.PoolManager()
                r = http.request('GET', data[3])
                image_file = BytesIO(r.data)
                img = Image(image_file)
                img.width = 90
                img.height = 75
                self.sheet.row_dimensions[p + 1].height = 56
                self.sheet.add_image(img, 'A{}'.format(p + 1))

    ########################################################################################################################
    ########################################################################################################################
    def user_info(self, user_id, b=None):

        try:
            session = self.sessions(b)
            # sleep(2.5)
            r = session.get(f'https://i.instagram.com/api/v1/users/{user_id}/info/', headers=headers_and).json()
            if r.get('status')=='ok':
                public_email = r.get('user').get('public_email')
                public_phone_number = r.get('user').get('public_phone_number')
                contact_phone_number = r.get('user').get('contact_phone_number')
                bio = r.get('user').get('biography')
                full_name = r.get('user').get('full_name')
                username = r.get('user').get('username')
                userid = r.get('user').get('pk')
                is_business = r.get('user').get('is_business')
                if not public_email:
                    if re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', bio):
                        email = re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', bio)
                    elif re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', full_name):
                        email = re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', full_name)
                    else:
                        email = []
                    print(email)
                if not public_phone_number and not contact_phone_number:
                    if re.findall(r'\d{8,15}', bio):
                        number = re.findall(r'\d{8,15}', bio)
                    elif re.findall(r'\d{8,15}', full_name):
                        number = re.findall(r'\d{8,15}', full_name)
                    else:
                        number = []
                    print(number)

                print(f'username:{username} is_business_account:{is_business} email:{public_email}  number : {public_phone_number}, {contact_phone_number}')
        except Exception as e:
            print(e)

    def user_info_2(self, user_name, b=None):
        try:
            session = self.sessions(b)
            # sleep(2.5)
            strHtml = session.get(f"https://www.instagram.com/{user_name}/?__a=1").text
            resp = json.loads(strHtml)
            user = resp['graphql']['user']
            bio = user['biography']
            is_business_account = user['is_business_account']
            full_name = user['full_name']
            username = user['username']
            '''
            external_url=user['external_url']
            followed_by=user['edge_followed_by']['count']
            follow=user['edge_follow' ]['count']
            user_id=user['id']
            is_business_account=user['is_business_account']
            is_joined_recently=user['is_joined_recently']
            business_category_name=user['business_category_name']
            category_id=user['category_id']
            overall_category_name=user['overall_category_name']
            
            connected_fb_page=user['connected_fb_page']
            is_private=user['is_private' ]
            posts=user['edge_owner_to_timeline_media']['count']

            has_channel=user['has_channel']
            country_block=user['country_block']
            followed_by_viewer=user['followed_by_viewer']
            follows_viewer=user['follows_viewer']
            has_ar_effects=user['has_ar_effects']
            has_blocked_viewer=user['has_blocked_viewer']
            highlight_reel_count=user['highlight_reel_count']
            has_requested_viewer=user['has_requested_viewer']
            user_id=user['id']
            
            is_joined_recently=user['is_joined_recently']
            business_category_name=user['business_category_name']
            category_id=user['category_id']
            overall_category_name=user['overall_category_name']
            category_enum=user['category_enum']
            is_verified=user['is_verified' ]
            mutual_followed_by=user['edge_mutual_followed_by']['count']
            profile_pic_url_hd=user['profile_pic_url_hd' ]
            requested_by_viewer=user['requested_by_viewer']
            felix_video_timeline=user['edge_felix_video_timeline']['count']

            saved_media=user['edge_saved_media']['count']
            media_collections=user['edge_media_collections']['count']'''
            if re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', bio):
                email = re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', bio)
            elif re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', full_name):
                email = re.findall(r'[\w\.-]+@[\w-]+\.[\w-]+', full_name)
            else:
                email = []

            self.numm += email

            if re.findall(r'\d{8,15}', bio):
                number = re.findall(r'\d{8,15}', bio)
            elif re.findall(r'\d{8,15}', full_name):
                number = re.findall(r'\d{8,15}', full_name)
            else:
                number = []
            self.numm += number

            print(f'username:{username} is_business_account:{is_business_account} email:{email}  number : {number}')
            print(len(self.numm))
            print(self.numm)
        except Exception as e:
            sleep(5)
            print(e)

    ########################################################################################################################
    ########################################################################################################################
    def get_comments(self):
        while True:
            url = input('put url  ( example:-https://www.instagram.com/p/code/ ) : ')  # chhc
            if url != '':
                break
            else:
                pass
        if url.find('instagram.com/p/') > -1:
            code = url.split('/p/')[1].split('/')[0]
        else:
            print('wrong url')
            return False
        roots = input(' if you want excel with img put 1 or excel with url of img put 2 ')
        if roots == '1':
            self.roots = 1
        else:
            self.roots = 2
        try:
            path = "Insta_Comments.xlsx"
            self.book = load_workbook(path)
            self.sheet = self.book.active
            end_cell = self.sheet.max_row
            root = 1
        except:  # 48
            self.book = Workbook()
            self.sheet = self.book.active
            self.sheet['B1'] = 'user name'
            self.sheet.column_dimensions['B'].width = 21
            self.sheet['D1'] = 'id'
            self.sheet.column_dimensions['D'].width = 13
            self.sheet['C1'] = 'comment'
            self.sheet.column_dimensions['C'].width = 34
            self.sheet['A1'] = 'profile pic'
            # self.sheet.column_dimensions['A'].width = 60
            self.sheet.column_dimensions['A'].width = 13
            root = 2
            end_cell = None
        u = True
        now1 = datetime.now()
        start_time = now1.strftime("%H:%M:%S")
        print(' getting comments ')
        resp = self.gcomment(code=code)
        print('Number of Comments {}'.format(resp['data']['shortcode_media']['edge_media_to_comment']['count']))
        p = 1
        while u:
            if resp != 'erorr':
                u = resp['data']['shortcode_media']['edge_media_to_comment']['page_info']['has_next_page']
                end_cursor = resp['data']['shortcode_media']['edge_media_to_comment']['page_info']['end_cursor']
                for i in resp['data']['shortcode_media']['edge_media_to_comment']['edges']:
                    text = i['node']['text']
                    user = i['node']['owner']['username']
                    id = i['node']['owner']['id']
                    profile_pic_url = i['node']['owner']['profile_pic_url']
                    print(p, '{0}:{1}'.format(user, text))
                    #################################################################################
                    # self.save_data(p=p, root=self.root, end_cell=self.end_cell,data=[user, id, full_name, profile_pic_url])
                    self.save_comments(root=root, p=p, end_cell=end_cell, data=[user, id, text, profile_pic_url])
                    ##################################################################################
                    p += 1
                resp = self.gcomment(code=code, end_cursor=end_cursor)
            else:
                print('*********** plze wait **************')
                sleep(15)
                resp = self.gcomment(code=code, end_cursor=end_cursor)
            self.book.save('Insta_Comments.xlsx')
        self.book.close()
        now = datetime.now()
        end_time = now.strftime("%H:%M:%S")
        print('done')
        print(start_time, end_time)

    #####################################################################
    def gcomment(self, code, end_cursor=None):
        #########################
        if end_cursor == None:
            url = 'https://www.instagram.com/graphql/query/?query_hash=33ba35852cb50da46f5b5e889df7d159&{}'.format(
                'variables={"shortcode":"%s","first":50}') % (code)
            strHtml = requests.get(url)
        else:
            url = 'https://www.instagram.com/graphql/query/?query_hash=33ba35852cb50da46f5b5e889df7d159&{}'.format(
                'variables={"shortcode":"%s","first":50,"after":"%s"}') % (code, end_cursor)
            strHtml = requests.get(url)
        #########################
        if strHtml.ok:
            resp = json.loads(strHtml.text)
            return resp
        else:
            print(strHtml.status_code)
            return 'erorr'

    #####################################################################
    def save_comments(self, root, p, data, end_cell):
        if root == 1:
            self.sheet.cell(row=(end_cell + p), column=2).value = data[0]
            self.sheet.cell(row=(end_cell + p), column=4).value = data[1]
            self.sheet.cell(row=(end_cell + p), column=3).value = data[2]
            if self.roots != 1:
                self.sheet.cell(row=(end_cell + p), column=1).value = data[3]
            #############################################################
            else:
                http = urllib3.PoolManager()
                r = http.request('GET', data[3])
                image_file = BytesIO(r.data)
                img = Image(image_file)
                img.width = 90
                img.height = 75

                self.sheet.row_dimensions[end_cell + p].height = 56
                self.sheet.add_image(img, 'A{}'.format(end_cell + p))

        else:
            self.sheet.cell(row=p + 1, column=2).value = data[0]
            self.sheet.cell(row=p + 1, column=4).value = data[1]
            self.sheet.cell(row=p + 1, column=3).value = data[2]
            if self.roots != 1:
                self.sheet.cell(row=p + 1, column=1).value = data[3]
            else:
                http = urllib3.PoolManager()
                r = http.request('GET', data[3])
                image_file = BytesIO(r.data)
                img = Image(image_file)
                img.width = 90
                img.height = 75
                self.sheet.row_dimensions[p + 1].height = 56
                self.sheet.add_image(img, 'A{}'.format(p + 1))

    ########################################################################################################################
    ########################################################################################################################
    def get_likes(self):
        while True:
            url = input('put url  ( example:-https://www.instagram.com/p/code/ ) : ')  # chhc
            if url != '':
                break
            else:
                pass
        if url.find('instagram.com/p/') > -1:
            code = url.split('/p/')[1].split('/')[0]
        else:
            print('wrong url')
            return False
        roots = input(' if you want excel with img put 1 or excel with url of img put 2 ')
        if roots == '1':
            self.roots = 1
        else:
            self.roots = 2
        try:
            path = "Insta_Likes.xlsx"
            self.book = load_workbook(path)
            self.sheet = self.book.active
            end_cell = self.sheet.max_row
            root = 1
        except:  # 48
            self.book = Workbook()
            self.sheet = self.book.active
            self.sheet['B1'] = 'user name'
            self.sheet.column_dimensions['B'].width = 21
            self.sheet['D1'] = 'id'
            self.sheet.column_dimensions['D'].width = 13
            self.sheet['C1'] = 'Full name'
            self.sheet.column_dimensions['C'].width = 34
            self.sheet['A1'] = 'profile pic'
            # self.sheet.column_dimensions['A'].width = 60
            self.sheet.column_dimensions['A'].width = 13
            root = 2
            end_cell = None
        u = True
        now1 = datetime.now()
        start_time = now1.strftime("%H:%M:%S")
        print(' getting Likes ')
        resp = self.glikes(code=code)
        print('Number of Likes {}'.format(resp['data']['shortcode_media']['edge_liked_by']['count']))
        p = 1
        while u:
            if resp != 'erorr':
                u = resp['data']['shortcode_media']['edge_liked_by']['page_info']['has_next_page']
                end_cursor = resp['data']['shortcode_media']['edge_liked_by']['page_info']['end_cursor']
                for i in resp['data']['shortcode_media']['edge_liked_by']['edges']:
                    user = i['node']['username']
                    full_name = i['node']['full_name']
                    id = i['node']['id']
                    profile_pic_url = i['node']['profile_pic_url']
                    print('{0}:{1}'.format(p, user))
                    #################################################################################
                    self.save_data(p=p, root=root, end_cell=end_cell, data=[user, id, full_name, profile_pic_url])
                    ##################################################################################
                    p += 1
                sleep(2)
                resp = self.glikes(code=code, end_cursor=end_cursor)
            else:
                print('*********** plze wait **************')
                sleep(15)
                resp = self.glikes(code=code, end_cursor=end_cursor)
            self.book.save('Insta_Likes.xlsx')
        self.book.close()
        now = datetime.now()
        end_time = now.strftime("%H:%M:%S")
        print('done')
        print(start_time, end_time)

    #####################################################################
    def glikes(self, code, end_cursor=None):
        #########################
        if end_cursor == None:
            url = 'https://www.instagram.com/graphql/query/?query_id=17864450716183058&{}'.format(
                'variables={"shortcode":"%s","first":50}') % (code)
            strHtml = self.session2.get(url)
        else:
            url = 'https://www.instagram.com/graphql/query/?query_id=17864450716183058&{}'.format(
                'variables={"shortcode":"%s","first":50,"after":"%s"}') % (code, end_cursor)
            strHtml = self.session1.get(url)
        #########################
        if strHtml.ok:
            resp = json.loads(strHtml.text)
            return resp
        else:
            print(strHtml.status_code)
            return 'erorr'

    ########################################################################################################################
    ########################################################################################################################
    def Download(self, url):

        if url.find('instagram.com/p/') > -1:
            if url.startswith('https') == False:
                url = 'https://' + str(url)
            if url.endswith('/') == False:
                url = url + '/'
            url = str(url) + '?__a=1'
            re = requests.get(url)  # B_XP3L_pjjl   B_B4QNUhOUp CAfbugFAw5o
            o = json.loads(re.text)
            if self.images_path == None:
                Folder_saved = filedialog.askdirectory(title="Select the folder where you want to save: ")
                Folder_saved = Folder_saved.replace('/', '\\')
                self.images_path = str(Folder_saved) + "//Images"
                self.videos_path = str(Folder_saved) + "//Videos"
                if not os.path.exists(self.images_path):
                    os.makedirs(self.images_path)
                if not os.path.exists(self.videos_path):
                    os.makedirs(self.videos_path)
            if o['graphql']['shortcode_media']['__typename'] == 'GraphSidecar':
                self.sidecar_downloader(o)
            elif o['graphql']['shortcode_media']['__typename'] == 'GraphVideo':
                self.download_vido(o)
            elif o['graphql']['shortcode_media']['__typename'] == 'GraphImage':
                self.image_downloader(o)
        else:
            print('  ***** wrong url ******')

    #####################################################################
    #####################################################################
    def image_downloader(self, o):
        img = o['graphql']['shortcode_media']
        display_url = img['display_url']
        file_name = img['taken_at_timestamp']
        download_path = self.images_path + '\\' + str(file_name) + '.jpg'
        if not os.path.exists(download_path):
            print('Downloading ' + str(file_name) + '.jpg...........')
            urllib.request.urlretrieve(display_url, download_path)
            print(str(file_name) + '.jpg Downloaded')
        else:
            print(str(file_name) + '.jpg has been downloaded before')

    #####################################################################
    def download_vido(self, o):
        try:
            video = o['graphql']['shortcode_media']
            title = video['taken_at_timestamp']
            video_url = video['video_url']
            video_view_count = video['video_view_count']
            video_duration = video['video_duration']
            download_path = self.videos_path + '\\' + str(title) + '.mp4'
            if not os.path.exists(download_path):
                print('Downloading ' + str(title) + '.mp4...........')
                resource = urllib.request.urlopen(video_url)
                output = open(download_path, "wb")
                output.write(resource.read())
                output.close()
                print('File Dowenload in :' + str(self.videos_path))
            else:
                print('.mp4 has been downloaded before')
        except Exception as e:
            print(e)

    #####################################################################
    def sidecar_downloader(self, o):
        try:
            num = 1
            for edge in o['graphql']['shortcode_media']['edge_sidecar_to_children']['edges']:
                if edge['node']['is_video']:
                    video_url = edge['node']['video_url']
                    file_name = o['graphql']['shortcode_media']['taken_at_timestamp']
                    download_path = self.videos_path + '\\' + str(file_name) + '_' + str(num) + '.mp4'
                    if not os.path.exists(download_path):
                        print('Downloading ' + str(file_name) + '_' + str(num) + '.mp4...........')
                        urllib.request.urlretrieve(video_url, download_path)
                        print(str(download_path) + ' ...........Downloaded ......')
                    else:
                        print(str(file_name) + '_' + str(num) + '.mp4 has been downloaded before')
                else:
                    display_url = edge['node']['display_url']
                    file_name = o['graphql']['shortcode_media']['taken_at_timestamp']
                    download_path = self.images_path + '\\' + str(file_name) + '_' + str(num) + '.jpg'
                    if not os.path.exists(download_path):
                        print('Downloading ' + str(file_name) + '_' + str(num) + '.jpg...........')
                        urllib.request.urlretrieve(display_url, download_path)
                        print(str(file_name) + '_' + str(num) + '.jpg Downloaded')
                    else:
                        print(str(file_name) + '_' + str(num) + '.jpg has been downloaded before')
                num += 1
        except Exception as e:
            print(e)

    #####################################################################
    #####################################################################
    def Download_all(self):
        while True:
            url = input('put username  ( example: ammar  ) : ')  # chhc
            if url != '':
                break
            else:
                pass
        if url.find('instagram.com/') > -1:
            username = url.split('com/')[1].split('/')[0]
        else:
            username = url
        inp = input('''if you want only video put v
        if you want only photo put p
           Or put a for all 
           ''')
        if inp == 'v' or inp == 'V':
            out = 1
        elif inp == 'p' or inp == 'P':
            out = 2
        else:
            out = 3
        now1 = datetime.now()
        start_time = now1.strftime("%H:%M:%S")
        #####################################
        r = requests.get('https://www.instagram.com/{0}/?__a=1'.format(username))
        user_id = r.json()['graphql']['user']['id']
        #######################################
        end_cursor = ''
        next_page = True
        Folder_saved = filedialog.askdirectory(title="Select the folder where you want to save: ")
        Folder_saved = Folder_saved.replace('/', '\\')
        self.images_path = str(Folder_saved) + "//Images"
        self.videos_path = str(Folder_saved) + "//Videos"
        if not os.path.exists(self.images_path):
            os.makedirs(self.images_path)
        if not os.path.exists(self.videos_path):
            os.makedirs(self.videos_path)
        while next_page:
            r = requests.get('https://www.instagram.com/graphql/query/',
                             params={'query_id': '17880160963012870', 'id': user_id, 'first': 12, 'after': end_cursor})
            o = r.json()
            graphql = o['data']
            for edge in graphql['user']['edge_owner_to_timeline_media']['edges']:
                __typename = edge['node']['__typename']
                shortcode = edge['node']['shortcode']
                url1 = 'https://www.instagram.com/p/{0}/'.format(shortcode)
                print(url1)
                if __typename == 'GraphImage':
                    if out == 1:
                        continue
                    self.Download(url1)
                elif __typename == 'GraphVideo':
                    if out == 2:
                        continue
                    self.Download(url1)
                elif __typename == 'GraphSidecar':
                    if out == 2 or out == 1:
                        continue
                    self.Download(url1)

            end_cursor = graphql['user']['edge_owner_to_timeline_media']['page_info']['end_cursor']
            next_page = graphql['user']['edge_owner_to_timeline_media']['page_info']['has_next_page']
            sleep(5)


############################# End ###################################
#####################################################################
def main():
    print('--------------------- welcom ---------------------')


#     while True:
#         Input = str(input('''--------------------------------------------------
#  This script to extract information from instagram
#  1-To Get Followers of someone
#     put Followers or followers or 1
#  2-To Get Following of someone
#     put Following or following or 2
#  3-To Get Comments of post
#     put Comment or comment or 3
#  4-To Get likes  of post
#     put Like or like or 4
#  5-To Dowenload one media
#     put Dowenload or dowenload or 5
#  6-To Dowenload  All Media(video and photo)
#     put Media or media or 6
#  7-To Exite
#     put close or c or 7
# --------------------------------------------------
#  >>>  '''))
#         if Input == 'Followers' or Input == 'followers' or Input == '1':
#             inst.get_followers()
#         elif Input == 'Following' or Input == 'following' or Input == '2':
#             inst.get_following()
#         elif Input == 'Comment' or Input == 'comment' or Input == '3':
#             inst.get_comments()
#         elif Input == 'Like' or Input == 'like' or Input == '4':
#             inst.get_likes()
#         elif Input == 'Dowenload' or Input == 'dowenload' or Input == '5':
#             while True:
#                 url = input('put url to download ( example:-https://www.instagram.com/p/code/ ) : ')  # chhc
#                 if url != '':
#                     break
#                 else:
#                     pass
#             inst.Download(url)
#         elif Input == 'Media' or Input == 'media' or Input == '6':
#             inst.Download_all()
#         elif Input == 'close' or Input == 'c' or Input == '7':
#             exit()

    # inst.get_followers()
    # inst.get_comments()
    # inst.get_likes()
    # inst.Download()
    # img "https://www.instagram.com/p/CAYP9-zJjVA/?__a=1"
    # side "https://www.instagram.com/p/CAfbugFAw5o/?__a=1"
    # videi "http://www.instagram.com/p/B_XP3L_pjjl/?__a=1"


if __name__ == '__main__':  # 10:48-->
    inst = Insta()
    inst.get_likes()
