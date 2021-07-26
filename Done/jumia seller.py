import json
from time import sleep

from openpyxl import load_workbook, Workbook
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

from requests_html import HTMLSession

book = Workbook()
sheet = book.active
n = 1
sheet.cell(n, 1).value = 'SKU'
sheet.cell(n, 3).value = 'Price'
sheet.cell(n, 4).value = 'Name'
sheet.cell(n, 5).value = 'Name en'
sheet.cell(n, 6).value = 'Old Price'
sheet.cell(n, 7).value = 'Brand'
sheet.cell(n, 8).value = 'Categories'
sheet.cell(n, 9).value = 'Rating'
sheet.cell(n, 10).value = 'Total Ratings'
sheet.cell(n, 11).value = 'URL'
sheet.column_dimensions['B'].width = 25
sheet.column_dimensions['A'].width = 25
sheet.column_dimensions['C'].width = 7
sheet.column_dimensions['D'].width = 50
sheet.column_dimensions['E'].width = 50
sheet.column_dimensions['F'].width = 9
sheet.column_dimensions['G'].width = 30
sheet.column_dimensions['H'].width = 50
sheet.column_dimensions['I'].width = 7
sheet.column_dimensions['J'].width = 13
sheet.column_dimensions['K'].width = 60
for column in range(1, 12):
    try:
        sheet.cell(n, column).alignment = Alignment(horizontal='center', vertical='center',
                                                    wrap_text=True)
    except:
        pass

session = HTMLSession()
main_url = url = input('- Enter URL: ')  # https://www.jumia.com.eg/ar/jg-gzpuluz/
# url = "https://www.jumia.com.eg/ar/saudi-gorub-marketplace/"

n = 2
while True:
    r = session.get(url,
                    headers={
                        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
                        "accept-language": "ar,en-US;q=0.9,en;q=0.8",
                        "cache-control": "no-cache",
                        "pragma": "no-cache",
                        "sec-ch-ua": "\"Chromium\";v=\"92\", \" Not A;Brand\";v=\"99\", \"Google Chrome\";v=\"92\"",
                        "sec-ch-ua-mobile": "?0",
                        "sec-fetch-dest": "document",
                        "sec-fetch-mode": "navigate",
                        "sec-fetch-site": "cross-site",
                        "sec-fetch-user": "?1",
                        "upgrade-insecure-requests": "1"
                    }
                    )
    sleep(3)
    data = json.loads(
        r.html.find('script+div+script+script', first=True).text.replace('window.__STORE__=', '').replace(';', ''))
    for element in data['products']:
        try:
            sku = element['sku']
            displayName = element.get('displayName')
            brand = element.get('brand')
            categories = element.get('categories')
            prices = element.get('prices').get('rawPrice')
            oldPrice = element.get('prices').get('oldPrice')
            rating = element.get('rating').get('average')
            totalRatings = element.get('rating').get('totalRatings')
            main_img = element.get('image').replace('300x300', '500x500')
            u = 'https://www.jumia.com.eg' + element.get('url')
            word = ''
            for w in element['url'].split('/')[-1].split('-')[:-1]:
                f = True
                for c in w:
                    if f:
                        c = c.upper()
                        f = False
                    word += c
                word += ' '
            name_en = word
            if main_img:
                res = session.get(main_img)
                image_file = BytesIO(res.content)
                img = Image(image_file)
                img.width = 200
                img.height = 270
                sheet.row_dimensions[n].height = 206
                sheet.add_image(img, f'B{n}')
            sheet.cell(n, 1).value = sku
            sheet.cell(n, 3).value = prices
            sheet.cell(n, 4).value = displayName
            sheet.cell(n, 5).value = name_en
            sheet.cell(n, 6).value = oldPrice
            sheet.cell(n, 7).value = brand
            sheet.cell(n, 8).value = categories
            sheet.cell(n, 9).value = rating
            sheet.cell(n, 10).value = totalRatings
            sheet.cell(n, 11).value = u
            for column in range(1, 12):
                try:
                    sheet.cell(n, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrap_text=True)
                except:
                    pass
            print(n)
            n += 1
        except:
            pass
    try:
        url = 'https://www.jumia.com.eg' + r.html.find("a.pg[aria-label='الصفحة التالية']", first=True).attrs['href']
    except:
        break
try:
    book.save(main_url.split('/')[4] + '.xlsx')
except:
    book.save('final.xlsx')

# pyinstaller --add-data c:\users\3mora\anaconda3\envs\python6\lib\site-packages\pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info