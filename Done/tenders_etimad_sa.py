from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill
from time import sleep
import uuid, os
import traceback
driver = webdriver.Chrome(ChromeDriverManager().install())
book = load_workbook('AllVisitorAnnouncements.xlsx')
sheet = book.active
i = 1
sheet.cell(i, 1).value = "إسم المنافسة"
sheet.cell(i, 2).value = "رقم المنافسة"
sheet.cell(i, 3).value = "الرقم المرجعي"
sheet.cell(i, 4).value = "الغرض "
sheet.cell(i, 5).value = "قيمة وثائق المنافسة"
sheet.cell(i, 6).value = "مكان فتح العرض"
sheet.cell(i, 7).value = "حالة المنافسة"
sheet.cell(i, 8).value= "نوع المنافسة"
sheet.cell(i, 9).value = "الجهة الحكوميه"
sheet.cell(i, 10).value = "الوقت المتبقى"
sheet.cell(i, 11).value = "طريقة تقديم العروض"
sheet.cell(i, 12).value = "مطلوب ضمان الإبتدائي"
sheet.cell(i, 13).value = "عنوان الضمان الإبتدائى"

sheet.cell(i, 14).value = "اخر موعد لإستلام الاستفسارات"
sheet.cell(i, 15).value = "آخر موعد لتقديم العروض"
sheet.cell(i, 16).value = "تاريخ فتح العروض"
sheet.cell(i, 17).value = "تاريخ فحص العروض"

sheet.cell(i, 18).value = "مكان التنفيذ"
sheet.cell(i, 19).value = "المناطق"
sheet.cell(i, 20).value = "التفاصيل"
sheet.cell(i, 21).value = "نشاط المنافسة"
sheet.cell(i, 22).value = "الإنشاء"
sheet.cell(i, 23).value = "أعمال الصيانة والتشغيل"
sheet.cell(i, 24).value = "الرابط"
sheet.cell(i, 25).value = "رابط الطباعة"
for i in range(2, sheet.max_row+1):
    try:
        print(i)
        url = sheet.cell(i, 24).value
        if sheet.cell(i, 1).value is not None:
            continue

        driver.get(url)
        sleep(4)
        name = driver.find_element_by_xpath('// div[div[contains(text(), "اسم الإعلان")]] / div[contains(@class,"etd-item-info")]').text
        number_m = driver.find_element_by_xpath('// div[div[contains(text(), "رقم المنافسة")]] / div[contains(@class,"etd-item-info")]').text
        number_return = driver.find_element_by_xpath('// div[div[contains(text(), "الرقم المرجعي")]] / div[contains(@class,"etd-item-info")]').text
        driver.find_element_by_css_selector('i.readMore').click()
        sleep(.5)
        prpose = driver.find_element_by_css_selector('#purposeSpan').text.replace(' ...عرض الأقل...', '')
        # prpose = driver.execute_script('return document.querySelector("#purposeSpan").textContent').replace(' ...عرض الأقل...', '')
        value_price = driver.find_element_by_xpath('// div[div[contains(text(), "قيمة وثائق المنافسة")]] / *[contains(@class,"etd-item-info")]').text
        place = driver.find_element_by_xpath('// div[div[contains(text(), "مكان فتح العرض")]] / *[contains(@class,"etd-item-info")]').text
        state = driver.find_element_by_xpath('// div[div[contains(text(), "حالة المنافسة")]] / *[contains(@class,"etd-item-info")]').text
        typee= driver.find_element_by_xpath('// div[div[contains(text(), "نوع المنافسة")]] / *[contains(@class,"etd-item-info")]').text
        gvar = driver.find_element_by_xpath('// div[div[contains(text(), "اسم الجهة")]] / *[contains(@class,"etd-item-info")]').text
        rem_time = driver.find_element_by_xpath('// div[div[contains(text(), "الوقت المتبقى")]] / *[contains(@class,"etd-item-info")]').text
        try:
            way_intro = driver.find_element_by_xpath('// div[div[contains(text(), "طريقة تقديم العروض")]] / *[contains(@class,"etd-item-info")]').text
        except:
            way_intro = None
        pry_dman = driver.find_element_by_xpath('// div[div[contains(text(), "مطلوب ضمان الإبتدائي")]] / *[contains(@class,"etd-item-info")]').text
        try:
            add_pry_d = driver.find_element_by_xpath('// div[div[contains(text(), "عنوان الضمان الإبتدائى")]] / *[contains(@class,"etd-item-info")]').text
        except:
            add_pry_d = None
        driver.find_element_by_css_selector('#tenderDatesTab').click()
        sleep(3)
        end_time_res = driver.find_element_by_xpath('// div[div[contains(text(), "اخر موعد لإستلام الاستفسارات")]] / *[contains(@class,"etd-item-info")]').text
        end_time_intro = driver.find_element_by_xpath('// div[div[contains(text(), "آخر موعد لتقديم العروض")]] / *[contains(@class,"etd-item-info")]').text
        open_time = driver.find_element_by_xpath('// div[div[contains(text(), "تاريخ فتح العروض")]] / *[contains(@class,"etd-item-info")]').text
        check_time = driver.find_element_by_xpath('// div[div[contains(text(), "تاريخ فحص العروض")]] / *[contains(@class,"etd-item-info")]').text
        driver.find_element_by_css_selector('#relationStepTab').click()
        done_place = WebDriverWait(driver, 10).until((ec.visibility_of_element_located((By.XPATH, '// div[div[contains(text(), "مكان التنفيذ")]] / *[contains(@class,"etd-item-info")]')))).text
        # done_place = driver.find_element_by_xpath('// div[div[contains(text(), "مكان التنفيذ")]] / *[contains(@class,"etd-item-info")]').text
        try:
            places = driver.find_element_by_xpath('// div[div[contains(text(), "المناطق")]] / *[contains(@class,"etd-item-info")]').text
        except:
            places = None
        detel = driver.find_element_by_xpath('// div[div[contains(text(), "التفاصيل")]] / *[contains(@class,"etd-item-info")]').text
        activate = driver.find_element_by_xpath('// div[div[contains(text(), "نشاط المنافسة")]] / *[contains(@class,"etd-item-info")]').text
        work_d = driver.find_element_by_xpath('// div[div[contains(text(), "الإنشاء")]] / *[contains(@class,"etd-item-info")]').text
        work_dd = driver.find_element_by_xpath('// div[div[contains(text(), "أعمال الصيانة والتشغيل")]] / *[contains(@class,"etd-item-info")]').text
        print_url = driver.find_element_by_css_selector('div.dropdown > a').get_attribute('href')

        sheet.cell(i, 1).value = name
        sheet.cell(i, 2).value = number_m
        sheet.cell(i, 3).value = number_return
        sheet.cell(i, 4).value = prpose
        sheet.cell(i, 5).value = value_price
        sheet.cell(i, 6).value = place
        sheet.cell(i, 7).value = state
        sheet.cell(i, 8).value = typee
        sheet.cell(i, 9).value = gvar
        sheet.cell(i, 10).value = rem_time
        sheet.cell(i, 11).value = way_intro
        sheet.cell(i, 12).value = pry_dman
        sheet.cell(i, 13).value = add_pry_d

        sheet.cell(i, 14).value = end_time_res
        sheet.cell(i, 15).value = end_time_intro
        sheet.cell(i, 16).value = open_time
        sheet.cell(i, 17).value = check_time

        sheet.cell(i, 18).value = done_place
        sheet.cell(i, 19).value = places
        sheet.cell(i, 20).value = detel
        sheet.cell(i, 21).value = activate
        sheet.cell(i, 22).value = work_d
        sheet.cell(i, 23).value = work_dd
        sheet.cell(i, 25).value = print_url
        print(name)
    except:
        print(traceback.print_exc())
book.save('AllVisitorAnnouncements.xlsx')

"""
book = load_workbook('QualificationsForVisitor.xlsx')
sheet = book.active
i = 1
sheet.cell(i, 1).value = "اسم دعوة التأهيل"
sheet.cell(i, 2).value = "الرقم المرجعي"
sheet.cell(i, 3).value = "الجهة الحكوميه"
sheet.cell(i, 4).value = "اسم لجنة الفحص"
sheet.cell(i, 5).value = "اسم اللجنة الفنية"
sheet.cell(i, 6).value = "الوقت المتبقى"
sheet.cell(i, 7).value = "اخر موعد لإستلام الاستفسارات"
sheet.cell(i, 8).value = "آخر موعد لتقديم وثائق التأهيل"
sheet.cell(i, 9).value = "تاريخ ووقت تقييم وثائق التأهيل"
sheet.cell(i, 10).value = "مكان التنفيذ"
sheet.cell(i, 11).value = "المناطق"
sheet.cell(i, 12).value = "نشاط"
sheet.cell(i, 13).value = "الإنشاء"
sheet.cell(i, 14).value = "أعمال الصيانة والتشغيل"

for i in range(2, sheet.max_row+1):
    try:
        print(i)
        url = sheet.cell(i, 24).value
        if sheet.cell(i, 1).value is not None:
            continue

        driver.get(url)
        sleep(4)
        name = driver.find_element_by_xpath('// div[div[contains(text(), "اسم دعوة التأهيل")]] / div[contains(@class,"etd-item-info")]').text
        number_return = driver.find_element_by_xpath('// div[div[contains(text(), "الرقم المرجعي")]] / div[contains(@class,"etd-item-info")]').text
        gvar = driver.find_element_by_xpath('// div[div[contains(text(), "الجهة الحكومية")]] / *[contains(@class,"etd-item-info")]').text
        check_name = driver.find_element_by_xpath('// div[div[contains(text(), "اسم لجنة الفحص")]] / *[contains(@class,"etd-item-info")]').text
        name_work = driver.find_element_by_xpath('// div[div[contains(text(), "اسم اللجنة الفنية")]] / *[contains(@class,"etd-item-info")]').text
        rem_time = driver.find_element_by_xpath('// div[div[contains(text(), "الوقت المتبقى")]] / *[contains(@class,"etd-item-info")]').text

        end_time_res = driver.find_element_by_xpath('// div[div[contains(text(), "اخر موعد لإستلام الاستفسارات")]] / *[contains(@class,"etd-item-info")]').text
        end_time_intro = driver.find_element_by_xpath('// div[div[contains(text(), "آخر موعد لتقديم  وثائق التأهيل")]] / *[contains(@class,"etd-item-info")]').text
        open_time = driver.find_element_by_xpath('// div[div[contains(text(), "تاريخ ووقت تقييم وثائق التأهيل")]] / *[contains(@class,"etd-item-info")]').text

        done_place = WebDriverWait(driver, 10).until((ec.visibility_of_element_located((By.XPATH, '// div[div[contains(text(), "مكان التنفيذ")]] / *[contains(@class,"etd-item-info")]')))).text
        try:
            places = driver.find_element_by_xpath('// div[div[contains(text(), "المناطق")]] / *[contains(@class,"etd-item-info")]').text
        except:
            places = None
        activate = driver.find_element_by_xpath('// div[div[contains(text(), "النشاط")]] / *[contains(@class,"etd-item-info")]').text
        work_d = driver.find_element_by_xpath('// div[div[contains(text(), "الإنشاء")]] / *[contains(@class,"etd-item-info")]').text
        work_dd = driver.find_element_by_xpath('// div[div[contains(text(), "أعمال الصيانة والتشغيل")]] / *[contains(@class,"etd-item-info")]').text

        sheet.cell(i, 1).value = name
        sheet.cell(i, 2).value = number_return
        sheet.cell(i, 3).value = gvar
        sheet.cell(i, 4).value = check_name
        sheet.cell(i, 5).value = name_work
        sheet.cell(i, 6).value = rem_time
        sheet.cell(i, 7).value = end_time_res
        sheet.cell(i, 8).value = end_time_intro
        sheet.cell(i, 9).value = open_time
        sheet.cell(i, 10).value = done_place
        sheet.cell(i, 11).value = places
        sheet.cell(i, 12).value = activate
        sheet.cell(i, 13).value = work_d
        sheet.cell(i, 14).value = work_dd
        print(name)
    except:
        print(traceback.print_exc())
        break
book.save('QualificationsForVisitor.xlsx')
"""
# book = Workbook()
# sheet = book.active
# driver.get('https://tenders.etimad.sa/Announcement/AllVisitorAnnouncements')
# input('- enter')
# urls = []
# while True:
#     for element in driver.find_elements_by_css_selector('div.border-bottom  div > div:nth-child(3) > h3 > a'):
#         urls.append(element.get_attribute('href'))
#         print(len(urls))
#     try:
#         if driver.find_element_by_css_selector('button[aria-label="Next"]').get_attribute('disabled') != 'true':
#             _ = driver.find_element_by_css_selector('button[aria-label="Next"]').location_once_scrolled_into_view
#             _ = driver.find_element_by_css_selector('button[aria-label="Next"]').location_once_scrolled_into_view
#             sleep(3)
#             driver.find_element_by_css_selector('button[aria-label="Next"]').click()
#         else:
#             break
#         sleep(3)
#     except:
#         print('cant')
#         break
# i = 2
# urls = list(set(urls))
# for url in urls:
#     sheet.cell(i, 24).value = url
#     i += 1
# book.save('AllVisitorAnnouncements.xlsx')
'''
book = load_workbook('AllTendersForVisitor.xlsx')
sheet = book.active
i = 1
sheet.cell(i, 1).value = "إسم المنافسة"
sheet.cell(i, 2).value = "رقم المنافسة"
sheet.cell(i, 3).value = "الرقم المرجعي"
sheet.cell(i, 4).value = "الغرض "
sheet.cell(i, 5).value = "قيمة وثائق المنافسة"
sheet.cell(i, 6).value = "مكان فتح العرض"
sheet.cell(i, 7).value = "حالة المنافسة"
sheet.cell(i, 8).value= "نوع المنافسة"
sheet.cell(i, 9).value = "الجهة الحكوميه"
sheet.cell(i, 10).value = "الوقت المتبقى"
sheet.cell(i, 11).value = "طريقة تقديم العروض"
sheet.cell(i, 12).value = "مطلوب ضمان الإبتدائي"
sheet.cell(i, 13).value = "عنوان الضمان الإبتدائى"

sheet.cell(i, 14).value = "اخر موعد لإستلام الاستفسارات"
sheet.cell(i, 15).value = "آخر موعد لتقديم العروض"
sheet.cell(i, 16).value = "تاريخ فتح العروض"
sheet.cell(i, 17).value = "تاريخ فحص العروض"

sheet.cell(i, 18).value = "مكان التنفيذ"
sheet.cell(i, 19).value = "المناطق"
sheet.cell(i, 20).value = "التفاصيل"
sheet.cell(i, 21).value = "نشاط المنافسة"
sheet.cell(i, 22).value = "الإنشاء"
sheet.cell(i, 23).value = "أعمال الصيانة والتشغيل"
sheet.cell(i, 24).value = "الرابط"
sheet.cell(i, 25).value = "رابط الطباعة"
for i in range(2, sheet.max_row+1):
    try:
        print(i)
        url = sheet.cell(i, 24).value
        if sheet.cell(i, 1).value is not None:
            continue

        driver.get(url)
        sleep(4)
        name = driver.find_element_by_xpath('// div[div[contains(text(), "إسم المنافسة")]] / div[contains(@class,"etd-item-info")]').text
        number_m = driver.find_element_by_xpath('// div[div[contains(text(), "رقم المنافسة")]] / div[contains(@class,"etd-item-info")]').text
        number_return = driver.find_element_by_xpath('// div[div[contains(text(), "الرقم المرجعي")]] / div[contains(@class,"etd-item-info")]').text
        driver.find_element_by_css_selector('i.readMore').click()
        sleep(.5)
        prpose = driver.find_element_by_css_selector('#purposeSpan').text.replace(' ...عرض الأقل...', '')
        # prpose = driver.execute_script('return document.querySelector("#purposeSpan").textContent').replace(' ...عرض الأقل...', '')
        value_price = driver.find_element_by_xpath('// div[div[contains(text(), "قيمة وثائق المنافسة")]] / *[contains(@class,"etd-item-info")]').text
        place = driver.find_element_by_xpath('// div[div[contains(text(), "مكان فتح العرض")]] / *[contains(@class,"etd-item-info")]').text
        state = driver.find_element_by_xpath('// div[div[contains(text(), "حالة المنافسة")]] / *[contains(@class,"etd-item-info")]').text
        typee= driver.find_element_by_xpath('// div[div[contains(text(), "نوع المنافسة")]] / *[contains(@class,"etd-item-info")]').text
        gvar = driver.find_element_by_xpath('// div[div[contains(text(), "الجهة الحكوميه")]] / *[contains(@class,"etd-item-info")]').text
        rem_time = driver.find_element_by_xpath('// div[div[contains(text(), "الوقت المتبقى")]] / *[contains(@class,"etd-item-info")]').text
        try:
            way_intro = driver.find_element_by_xpath('// div[div[contains(text(), "طريقة تقديم العروض")]] / *[contains(@class,"etd-item-info")]').text
        except:
            way_intro = None
        pry_dman = driver.find_element_by_xpath('// div[div[contains(text(), "مطلوب ضمان الإبتدائي")]] / *[contains(@class,"etd-item-info")]').text
        try:
            add_pry_d = driver.find_element_by_xpath('// div[div[contains(text(), "عنوان الضمان الإبتدائى")]] / *[contains(@class,"etd-item-info")]').text
        except:
            add_pry_d = None
        driver.find_element_by_css_selector('#tenderDatesTab').click()
        sleep(3)
        end_time_res = driver.find_element_by_xpath('// div[div[contains(text(), "اخر موعد لإستلام الاستفسارات")]] / *[contains(@class,"etd-item-info")]').text
        end_time_intro = driver.find_element_by_xpath('// div[div[contains(text(), "آخر موعد لتقديم العروض")]] / *[contains(@class,"etd-item-info")]').text
        open_time = driver.find_element_by_xpath('// div[div[contains(text(), "تاريخ فتح العروض")]] / *[contains(@class,"etd-item-info")]').text
        check_time = driver.find_element_by_xpath('// div[div[contains(text(), "تاريخ فحص العروض")]] / *[contains(@class,"etd-item-info")]').text
        driver.find_element_by_css_selector('#relationStepTab').click()
        done_place = WebDriverWait(driver, 10).until((ec.visibility_of_element_located((By.XPATH, '// div[div[contains(text(), "مكان التنفيذ")]] / *[contains(@class,"etd-item-info")]')))).text
        # done_place = driver.find_element_by_xpath('// div[div[contains(text(), "مكان التنفيذ")]] / *[contains(@class,"etd-item-info")]').text
        try:
            places = driver.find_element_by_xpath('// div[div[contains(text(), "المناطق")]] / *[contains(@class,"etd-item-info")]').text
        except:
            places = None
        detel = driver.find_element_by_xpath('// div[div[contains(text(), "التفاصيل")]] / *[contains(@class,"etd-item-info")]').text
        activate = driver.find_element_by_xpath('// div[div[contains(text(), "نشاط المنافسة")]] / *[contains(@class,"etd-item-info")]').text
        work_d = driver.find_element_by_xpath('// div[div[contains(text(), "الإنشاء")]] / *[contains(@class,"etd-item-info")]').text
        work_dd = driver.find_element_by_xpath('// div[div[contains(text(), "أعمال الصيانة والتشغيل")]] / *[contains(@class,"etd-item-info")]').text
        print_url = driver.find_element_by_css_selector('div.dropdown > a').get_attribute('href')

        sheet.cell(i, 1).value = name
        sheet.cell(i, 2).value = number_m
        sheet.cell(i, 3).value = number_return
        sheet.cell(i, 4).value = prpose
        sheet.cell(i, 5).value = value_price
        sheet.cell(i, 6).value = place
        sheet.cell(i, 7).value = state
        sheet.cell(i, 8).value = typee
        sheet.cell(i, 9).value = gvar
        sheet.cell(i, 10).value = rem_time
        sheet.cell(i, 11).value = way_intro
        sheet.cell(i, 12).value = pry_dman
        sheet.cell(i, 13).value = add_pry_d

        sheet.cell(i, 14).value = end_time_res
        sheet.cell(i, 15).value = end_time_intro
        sheet.cell(i, 16).value = open_time
        sheet.cell(i, 17).value = check_time

        sheet.cell(i, 18).value = done_place
        sheet.cell(i, 19).value = places
        sheet.cell(i, 20).value = detel
        sheet.cell(i, 21).value = activate
        sheet.cell(i, 22).value = work_d
        sheet.cell(i, 23).value = work_dd
        sheet.cell(i, 25).value = print_url
        print(name)
    except:
        print(traceback.print_exc())
book.save('AllTendersForVisitor.xlsx')
'''

# i = 1
# book = Workbook()
# sheet = book.active
# sheet.cell(i, 1).value = "العنوان"
# sheet.cell(i, 2).value = "الشراء"
# sheet.cell(i, 3).value = "السعر"
# sheet.cell(i, 4).value = "الوصف"
# sheet.cell(i, 5).value = "النشاط الأساسي"
# sheet.cell(i, 6).value = "الرقم المرجعي"
# sheet.cell(i, 7).value = "تاريخ النشر "
# sheet.cell(i, 8).value = "الوقت المتبقي"
# sheet.cell(i, 9).value = "اخر موعد لإستلام الاستفسارات"
# sheet.cell(i, 10).value = "آخر موعد لتقديم العروض"
# sheet.cell(i, 11).value = "آخر موعد لتقديم العروض"
# sheet.cell(i, 12).value = "تاريخ ووقت فتح العروض"
# sheet.cell(i, 13).value = "تاريخ ووقت فتح العروض"
# sheet.cell(i, 14).value = "المنافسه"
# i = 2
# for element in driver.find_elements_by_css_selector('#cardsresult > div.row.justify-content-center > div'):
#     date = element.find_element_by_css_selector('div.col-12.border-bottom > div > div:nth-child(1) > span').text
#     shr = element.find_element_by_css_selector('.badge.badge-primary').text
#     title = element.find_element_by_css_selector('div.col-12.border-bottom > div > div:nth-child(3) > h3 > a').text
#     url = element.find_element_by_css_selector(
#         'div.col-12.border-bottom > div > div:nth-child(3) > h3 > a').get_attribute('href')
#     des = \
#     element.find_element_by_css_selector('div.col-12.border-bottom > div > div:nth-child(4) > p').text.split('\n')[0]
#     main_ex = element.find_element_by_css_selector('div:nth-child(1) > div > div > div.col-12.pt-2 > span').text
#     end_time = element.find_element_by_css_selector(
#         'div:nth-child(2) > div > div.text-center.text-chart-indicator').text
#     return_namber = element.find_element_by_css_selector(
#         'div:nth-child(3) > div > div > div:nth-child(1) > span').text
#     end_date_re = element.find_element_by_css_selector('div:nth-child(3) > div > div > div:nth-child(2) > span').text
#     end_date_sh = element.find_element_by_css_selector(
#         'div:nth-child(3) > div > div > div:nth-child(3) > span.ml-3').text
#     end_time_sh = element.find_element_by_css_selector(
#         'div:nth-child(3) > div > div > div:nth-child(3) > span:nth-child(3)').text
#     try:
#         end_open = element.find_element_by_css_selector(
#             'div:nth-child(3) > div > div > div:nth-child(4) > span.ml-3').text
#     except:
#         end_open = None
#     try:
#         end_open_time = element.find_element_by_css_selector(
#             'div:nth-child(3) > div > div > div:nth-child(4) > span:nth-child(3)').text
#     except:
#         end_open_time = None
#     price = element.find_element_by_css_selector('div:nth-child(4) > div > div').text
#     sheet.cell(i, 1).value = title
#     sheet.cell(i, 2).value = shr
#     sheet.cell(i, 3).value = price
#     sheet.cell(i, 4).value = des
#     sheet.cell(i, 5).value = main_ex
#     sheet.cell(i, 6).value = return_namber
#     sheet.cell(i, 7).value = date
#     sheet.cell(i, 8).value = end_time
#     sheet.cell(i, 9).value = end_date_re
#     sheet.cell(i, 10).value = end_date_sh
#     sheet.cell(i, 11).value = end_time_sh
#     sheet.cell(i, 12).value = end_open
#     sheet.cell(i, 13).value = end_open_time
#     sheet.cell(i, 14).value = url
#     i += 1
# book.save('text.xlsx')
'''
https://tenders.etimad.sa/Tender/AllTendersForVisitor
https://tenders.etimad.sa/Qualification/QualificationsForVisitor

https://tenders.etimad.sa/Announcement/AllVisitorAnnouncements
'''
