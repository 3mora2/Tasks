from playwright import sync_playwright
from openpyxl import load_workbook, Workbook
from time import sleep
book = Workbook()
sheet = book.active
playwright = sync_playwright().start()
driver = playwright.chromium.launch(headless=False)
page = driver.newPage()
page.goto('https://www.webtebpro.com/doctors/search?pagenumber=20')


# page.goto('https://sa.iherb.com/c/beauty?noi=192')
# input('ttttttttttttt')
# file = 'beauty.xlsx'
# book = load_workbook(file)
# sheet = book.active
# for i in range(2, sheet.max_row+1):
#     link = sheet[f'X{i}'].value
#     if link is not None and sheet[f'A{i}'].value is None:
#         try:
#             print(i)
#             if i == 110:
#                 break
#             page.goto(link)
#             sleep(1)
#             try:
#                 UPC_Code = \
#                 [li.innerText().split(':')[-1].strip() for li in page.querySelectorAll('#product-specs-list > li') if
#                  'UPC' in li.innerText()][0]
#             except:
#                 sleep(5)
#                 page.reload()
#                 UPC_Code = \
#                     [li.innerText().split(':')[-1].strip() for li in
#                      page.querySelectorAll('#product-specs-list > li') if
#                      'UPC' in li.innerText()][0]
#
#             Product_Code = page.querySelector('#modelProperties').getAttribute('data-part-number')
#             product_id = page.querySelector('#modelProperties').getAttribute('data-product-id')
#             title = page.querySelector('#modelProperties').getAttribute('data-product-name')
#             price = page.querySelector('#modelProperties').getAttribute('data-list-price')
#             discounted_price = page.querySelector('#modelProperties').getAttribute('data-discounted-price')
#             state = page.querySelector('#stock-status > strong').innerText()
#             brand = page.querySelectorAll('#brand > a > span > bdi')[-1].innerText()
#
#             try:
#                 page.querySelector('#product-reviews div.top-rating-wrapper > div').scrollIntoViewIfNeeded()
#                 rate = page.querySelector('#product-reviews div.top-rating-wrapper > div').innerText()
#             except:
#                 rate = None
#
#             cat1 = page.querySelector('#breadCrumbs > a:nth-child(6)').innerText()
#             cat2 = page.querySelectorAll('#breadCrumbs > a.last')[1].innerText()
#             try:
#                 cat3 = page.querySelectorAll('#breadCrumbs > a.last')[2].innerText()
#             except:
#                 cat3 = None
#
#             shipping_weight_kg = page.querySelector('#modelProperties').getAttribute('data-shipping-weight-kg')
#             actual_weight_kg = page.querySelector('#modelProperties').getAttribute('data-actual-weight-kg')
#             shipping_weight_lb = page.querySelector('#modelProperties').getAttribute('data-shipping-weight-lb')
#             actual_weight_lb = page.querySelector('#modelProperties').getAttribute('data-actual-weight-lb')
#             dimensions_in = page.querySelector('#modelProperties').getAttribute('data-dimensions-in')
#             dimensions_cm = page.querySelector('#modelProperties').getAttribute('data-dimensions-cm')
#
#             try:
#                 Expiration_Date = \
#                 [li.innerText().split(':?')[-1].strip() for li in page.querySelectorAll('#product-specs-list > li')
#                  if 'Expiration' in li.innerText()][0]
#             except:
#                 Expiration_Date = None
#             try:
#                 quantity = [li.innerText().split(':')[-1].strip() for li in page.querySelectorAll('#product-specs-list > li') if 'Quantity' in li.innerText()][0]
#             except:
#                 quantity = None
#             try:
#                 des = page.querySelector('div[itemprop="description"]').innerText()
#             except:
#                 des = None
#             all_des = page.querySelector('section > div.inner-content').innerText()
#             photo = [elem.getAttribute('data-large-img') for elem in
#                      page.querySelectorAll('#product-image > div.thumbnail-container > img')]
#             brandCrumbs = page.querySelectorAll('#breadCrumbs > a.last')[0].innerText()
#             sheet.cell(i, 1).value = UPC_Code
#             sheet.cell(i, 2).value = Product_Code
#             sheet.cell(i, 3).value = product_id
#             sheet.cell(i, 4).value = title
#             sheet.cell(i, 5).value = price
#             sheet.cell(i, 6).value = discounted_price
#             sheet.cell(i, 7).value = state
#             sheet.cell(i, 8).value = brand
#             sheet.cell(i, 9).value = rate
#             sheet.cell(i, 10).value = cat1
#             sheet.cell(i, 11).value = cat2
#             sheet.cell(i, 12).value = cat3
#             sheet.cell(i, 13).value = shipping_weight_kg
#             sheet.cell(i, 14).value = actual_weight_kg
#             sheet.cell(i, 15).value = shipping_weight_lb
#             sheet.cell(i, 16).value = actual_weight_lb
#             sheet.cell(i, 17).value = dimensions_in
#             sheet.cell(i, 18).value = dimensions_cm
#             sheet.cell(i, 19).value = Expiration_Date
#             sheet.cell(i, 20).value = quantity
#             sheet.cell(i, 21).value = brandCrumbs
#             sheet.cell(i, 22).value = des
#             sheet.cell(i, 23).value = all_des
#             for co in range(len(photo)):
#                 sheet.cell(i, co+25).value = photo[co]
#             book.save('text1play.xlsx')
#         except Exception as e:
#             print(e)
# page.expect_console_message('document.evaluate("//*[@id="product-specs-list"]/li[1]/following-sibling::text()", document, null, XPathResult.STRING_TYPE, null ).stringValue;"')