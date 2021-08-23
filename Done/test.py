from selenium import webdriver
from openpyxl import load_workbook, Workbook
from webdriver_manager.chrome import ChromeDriverManager

driver = webdriver.Chrome(ChromeDriverManager().install())
wb = load_workbook("D:/work.xlsx")
ws = wb.active

book = Workbook()
sheet = book.active
row = 2
for a in range(1, ws.max_row+1):
    url = ws.cell(row=a, column=1).value
    if not url:
        continue
    # for url in ['https://www.autopartsco.com.au/efs-elite-lift-kit-mitsubishi-pajero-nm-nw-05-2000',
    #             'https://www.autopartsco.com.au/hpd-upgraded-front-mount-intercooler-mitsubishi-pa~17252',
    #             'https://www.autopartsco.com.au/tough-dog-40mm-lift-kit-mitsubishi-pajero-nh-nl-5',
    #             'https://www.autopartsco.com.au/fuel-manager-diesel-pre-filter-kit-mitsubishi-paje'
    #             ]:
    next_row = row+1
    driver.get(url)
    category = driver.find_element_by_xpath('//*[@id="main-content"]/div/div/nav/ol/li[2]/a/span').text
    name = driver.find_element_by_xpath('//*[@id="_jstl__header_r"]/div/div[1]/h1').text
    price = driver.execute_script("return document.querySelector('.productprice.productpricetext').textContent;").strip()
    # price = driver.find_element_by_css_selector('.productprice.productpricetext').text
    desc = driver.find_element_by_xpath('//*[@id="description"]').get_attribute("outerHTML")
    desc2 = driver.find_element_by_xpath('//*[@id="description"]').get_attribute("innerHTML")

    try:
        sku = driver.execute_script(
                "return document.evaluate('//tr[td/strong[text()=\"SKU\"]]/td[2]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.innerText;")
    except:
        sku = None

    sheet.cell(row, 1).value = url
    sheet.cell(row, 3).value = name
    sheet.cell(row, 4).value = category
    sheet.cell(row, 6).value = desc2

    column = 8
    for option in driver.find_elements_by_css_selector('#extra-options-table > tbody > tr')[1:]:
        row_ = row
        name_option = option.find_element_by_css_selector('td > label').text
        sheet.cell(row_, column).value = name_option
        column += 1
        for list_option in option.find_elements_by_css_selector('td > select > option'):
            sheet.cell(row_, column).value = list_option.text
            sheet.cell(row_, 2).value = name
            sheet.cell(row_, 5).value = sku
            sheet.cell(row_, 7).value = price
            row_ += 1

        next_row = row_ if row_ > next_row else next_row
        column += 1
        if column == 14:
            break

    photo = [img.get_attribute('src').replace('_thumb/', '/') for img in driver.find_elements_by_css_selector('#_jstl__images_r img') if '/thumbL/' not in img.get_attribute('src')]
    column = 14
    for index, img in enumerate(photo):
        row_ = row + index
        sheet.cell(row_, column).value = img
        sheet.cell(row_, column+1).value = index
        sheet.cell(row_, 2).value = name
        sheet.cell(row_, 5).value = sku
        sheet.cell(row_, 7).value = price
        next_row = row_+1 if row_ > next_row else next_row

    row = next_row if next_row > row else row+1

book.save('text.xlsx')