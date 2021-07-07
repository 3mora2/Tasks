from time import sleep
from requests_html import HTMLSession, HTML
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from selenium import webdriver
driver = webdriver.Firefox()
s = HTMLSession()
book = load_workbook('101.xlsx')
sheet = book.active
# num = 1
# sheet[f'A{num}'] = 'الاسم'
# sheet[f'B{num}'] = 'السعر بعد'
# sheet[f'C{num}'] = 'السعر قبل'
# sheet[f'D{num}'] = 'الشركة'
# sheet[f'E{num}'] = 'النوع'
# sheet[f'F{num}'] = 'تفاصيل'
# sheet[f'G{num}'] = 'Name'
# sheet[f'H{num}'] = 'New Price'
# sheet[f'I{num}'] = 'Old Price'
# sheet[f'J{num}'] = 'Brand'
# sheet[f'K{num}'] = 'Product Code'
# sheet[f'L{num}'] = 'Description'
# sheet[f'M{num}'] = 'Images'
# sheet[f'N{num}'] = 'URL'
# for column in range(2, 14):
#     try:
#         sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center')  # , wrap_text=True)
#     except:
#         pass

for num in range(2, sheet.max_row+1):
    if sheet[f'A{num}'].value is not None:
        continue

    print(num)
    url = sheet[f'N{num}'].value
    driver.get(url)

    # r = s.get(url)
    # sleep(4)
    # try:
    # name = r.html.find('.title-product-detail')[0].text
    # except:
    r = HTML(html=driver.page_source)
    name = r.find('div.product-info__title.for-desktop > h2')[0].text
    # try:
    #     price1 = r.html.find('div.row > div > div > div.product-price > ins')[0].text
    # except:
    price1 = r.find('div.price-box.product-info__price > span.price-box__new')[0].text
    # try:
    #     price2 = r.html.find('div.row > div > div > div.product-price > del')[0].text
    # except:
    try:
        price2 = r.find('div.price-box.product-info__price > span.price-box__old')[0].text
    except:
        price2 = None
    # try:
    #     mark = r.html.find('div.row > div:nth-child(2) > div > div.product-add-form > a')[0].text
    # except:
    try:
        mark = r.find('div.product-info__description > a')[0].text
    except:
        mark = None
    # try:
    #     model = r.html.find('div.product-add-form > br > p.product-code')[0].text.replace('الموديل:', '').strip()
    # except:
    try:
        model = r.find('#product-model')[0].text.split(':')[-1].strip()
    except:
        model = None
    # try:
    #     desk = r.html.find('#home')[0].text
    # except:
    desk = r.find('#Tab1')[0].text

    img = '\n '.join([im.attrs['data-zoom-image'] for im in r.find('#smallGallery li a')])
    if img == '':
        try:
            img = r.find('.product-zoom')[0].attrs['data-zoom-image']
        except:
            img = r.find('.product-zoom')[0].attrs['src']
    # img = '\n '.join([im.attrs['src'] for im in r.html.find('ul > li > a > img')])
    driver.get(url.replace('/ar/', '/en/'))
    # r = s.get(url.replace('/ar/', '/en/'))
    r = HTML(html=driver.page_source)
    name_en = r.find('div.product-info__title.for-desktop > h2')[0].text
    price1_en = r.find('div.price-box.product-info__price > span.price-box__new')[0].text
    try:
        price2_en = r.find('div.price-box.product-info__price > span.price-box__old')[0].text
    except:
        price2_en = None
    try:
        mark_en = r.find('div.product-info__description > a')[0].text
    except:
        mark_en = None
    try:
        model_en = r.find('#product-model')[0].text.split(':')[-1].strip()
    except:
        model_en = None
    desk_en = r.find('#Tab1')[0].text

    # name_en = r.html.find('.title-product-detail')[0].text
    # price1_en = r.html.find('div.row > div > div > div.product-price > ins')[0].text
    # try:
    #     price2_en = r.html.find('div.row > div > div > div.product-price > del')[0].text
    # except:
    #     price2_en = None
    # try:
    #     mark_en = r.html.find('div.row > div:nth-child(2) > div > div.product-add-form > a')[0].text
    # except:
    #     mark_en = None
    # try:
    #     model_en = r.html.find('div.product-add-form > br > p.product-code')[0].text.replace('Model:', '').strip()
    # except:
    #     model_en = None
    # desk_en = r.html.find('#home')[0].text
    sheet[f'A{num}'] = name
    sheet[f'B{num}'] = price1
    sheet[f'C{num}'] = price2
    sheet[f'D{num}'] = mark
    sheet[f'E{num}'] = model
    sheet[f'F{num}'] = desk
    sheet[f'G{num}'] = name_en
    sheet[f'H{num}'] = price1_en
    sheet[f'I{num}'] = price2_en
    sheet[f'J{num}'] = mark_en
    sheet[f'K{num}'] = model_en
    sheet[f'L{num}'] = desk_en
    sheet[f'M{num}'] = img
    for column in range(2, 14):
        try:
            sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center')  # , wrap_text=True)
        except:
            pass
    if num == 1200:
        break


book.save('101.xlsx')

# book = Workbook()
# sheet = book.active
# n = 1
# num = 2
# product = []
# while True:
#     # r = s.get(f'https://aristostationery.com/ar/category/%D8%A7%D8%B1%D8%AA-102/?page={n}')
#     r = s.get(f'https://aristostationery.com/ar/category/أدوات-مدرسية-و-مكتبية-101/?page={n}')
#     for e in r.html.find('div.product-preview div.product-preview__info div.product-preview__info__title > h2 > a'):
#         if e.attrs['href'] not in product:
#             product.append(e.attrs['href'])
#         sheet[f'N{num}'] = e.attrs['href']
#         num += 1
#
#     print(num)
#     if len(r.html.find('a.pagination__next')) == 0:
#         break
#     n += 1
# book.save('101.xlsx')
