from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from time import sleep
from openpyxl import load_workbook

# wb = Workbook()
wb = load_workbook('3tor.xlsx')
ws = wb.active

# https://www.goldenscent.com/fragrances.html?action=cat&id=3
# driver = webdriver.Chrome()
# driver = webdriver.Firefox()
'''
driver.get("https://www.goldenscent.com/fragrances.html?action=cat&id=3")
sleep(7)
# 3tor
driver.find_element_by_css_selector('div > div > div:nth-child(4) > div > div > ul > li:nth-child(1)').click()
sleep(7)
# more
links = []
while True:
    try:
        for element in driver.find_elements_by_css_selector('ol > li> div > a'):
            if element.get_attribute('href') not in links:
                links.append(element.get_attribute('href'))
        print(len(links))
        driver.find_element_by_css_selector('.ais-InfiniteHits-loadMore').click()
        sleep(3)
    except Exception as e:
        print(e)
        break
# product

'''
"""
ws['A1'] = 'Title'
ws['B1'] = 'Type'
ws['C1'] = 'date'
ws['D1'] = 'aval'
ws['E1'] = 'rate'
ws['F1'] = 'market'
ws['G1'] = 'Gender'
ws['H1'] = 'Product_Type'
ws['I1'] = 'Character'
ws['J1'] = 'Fragrance'
ws['K1'] = 'Occasion'
ws['L1'] = 'Concentration'
ws['M1'] = 'keys'
ws['N1'] = 'size1'
ws['O1'] = 'Price1'
ws['P1'] = 'size2'
ws['Q1'] = 'price2'
ws['R1'] = 'size3'
ws['S1'] = 'price3'
ws['T1'] = 'Ingredients'
ws['U1'] = 'des'
ws['V1'] = 'img'

"""
# num = 1
"""
ws[f'X{num}'] = 'name'
ws[f'Y{num}'] = 'Type'
ws[f'Z{num}'] = 'date'
ws[f'AA{num}'] = 'aval'
ws[f'AB{num}'] = 'rat'
ws[f'AC{num}'] = 'mak'
ws[f'AD{num}'] = 'Gender'
ws[f'AE{num}'] = 'Product_Type'
ws[f'AF{num}'] = 'Character'
ws[f'AG{num}'] = 'Fragrance'
ws[f'AH{num}'] = 'Occasion'
ws[f'AI{num}'] = 'Concentration'
ws[f'AJ{num}'] = 'kiys'
ws[f'AK{num}'] = 'size1'
ws[f'AL{num}'] = 'price1'
ws[f'AM{num}'] = 'size2'
ws[f'AN{num}'] = 'price2'
ws[f'AO{num}'] = 'size3'
ws[f'AP{num}'] = 'price3'
ws[f'AQ{num}'] = 'Ingredients'
ws[f'AR{num}'] = 'des'
"""
# num = 2148
# # int(input('enter number'))
# # 1573
#
# links = [ws[f'W{i}'].value for i in range(num, 2320)]  # ws.max_row + 1)]
# print(len(links))
#
# for url in links:
#     driver.get(url.replace('goldenscent.com/', 'goldenscent.com/en/'))
#     wb.save('3tor.xlsx')
#     # sleep(4)
#     try:
#         name = WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "h1.fw300"))).text
#         # name = driver.find_element_by_css_selector('h1.fw300').text
#     except:
#         driver.refresh()
#         name = WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "h1.fw300"))).text
#         # name = driver.find_element_by_css_selector('h1.fw300').text
#
#     Type = driver.find_element_by_css_selector('h2.fw400').text
#     aval = driver.find_element_by_css_selector('div.stock').text
#     try:
#         price1 = driver.find_element_by_css_selector('div.col-xs-6.fw400 span:nth-child(1)').text
#     except:
#         price1 = None
#     try:
#         size1 = driver.find_elements_by_css_selector('.type-perfume')[0].text
#     except:
#         size1 = None
#     """try:
#         if len(driver.find_elements_by_css_selector('.type-perfume')) == 2:
#             i = driver.find_elements_by_css_selector('.type-perfume')
#
#             i[0].click()
#
#             size1 = i[0].text
#             price1 = driver.find_element_by_css_selector('div.col-xs-6.fw400 span:nth-child(1)').text
#             i[1].click()
#             size2 = i[1].text
#             price2 = driver.find_element_by_css_selector('div.col-xs-6.fw400 span:nth-child(1)').text
#             price3 = None
#             size3 = None
#         elif len(driver.find_elements_by_css_selector('.type-perfume')) == 3:
#             i = driver.find_elements_by_css_selector('.type-perfume')
#             i[0].click()
#             size1 = i[0].text
#             price1 = driver.find_element_by_css_selector('div.col-xs-6.fw400 span:nth-child(1)').text
#             i[1].click()
#             size2 = i[1].text
#             price2 = driver.find_element_by_css_selector('div.col-xs-6.fw400 span:nth-child(1)').text
#             i[2].click()
#             size3 = i[2].text
#             price3 = driver.find_element_by_css_selector('div.col-xs-6.fw400 span:nth-child(1)').text
#         elif len(driver.find_elements_by_css_selector('.type-perfume')) == 1:
#             i = driver.find_elements_by_css_selector('.type-perfume')
#             i[0].click()
#             size1 = i[0].text
#             price1 = driver.find_element_by_css_selector('div.col-xs-6.fw400 span:nth-child(1)').text
#             price2 = price3 = size3 = size2 = None
#         else:
#             price1 = price2 = price3 = size3 = size2 = size1 = None
#         print(price1, price2, price3, size1, size2, size3)
#     except Exception as e:
#         price1 = price2 = price3 = size3 = size2 = size1 = None
#         print(e)"""
#
#     # img = ' ,\n'.join([el.get_attribute('src') for el in
#     #                    driver.find_elements_by_css_selector('img.product-image.inline-flex.pointer') if
#     #                    el.get_attribute('src') is not None])
#
#     # driver.find_element_by_css_selector('.product-description').click()
#     # sleep(2)
#     des = WebDriverWait(driver, 5).until(
#         ec.visibility_of_element_located((By.CSS_SELECTOR, '#product-description > p'))).text
#     # driver.find_element_by_css_selector('#product-description > p').text
#     driver.find_element_by_css_selector('.product-information').click()
#     sleep(1)
#
#     try:
#         mak = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                'الماركة' in elem.text or 'Brand' in elem.text][
#             0]
#     except:
#         mak = None
#     try:
#         Gender = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                   driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                   'الجنس' in elem.text or 'Gender' in elem.text][0]
#     except:
#         Gender = None
#     try:
#         Product_Type = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                         driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                         'نوع المنتج' in elem.text or 'Product Type' in elem.text][0]
#     except:
#         Product_Type = None
#     try:
#         Character = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                      driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                      'شخصية عطرك' in elem.text or 'Character' in elem.text][0]
#     except:
#         Character = None
#     try:
#         Fragrance = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                      driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                      'العائلة العطرية' in elem.text or 'Fragrance Family' in elem.text][0]
#     except:
#         Fragrance = None
#     try:
#         Occasion = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                     driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                     'مناسب' in elem.text or 'Occasion' in elem.text][0]
#     except:
#         Occasion = None
#     try:
#         Ingredients = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                        driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                        'المكونات' in elem.text or 'Ingredients' in elem.text][0]
#     except:
#         Ingredients = None
#     try:
#         Concentration = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                          driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                          'نسبة التركيز' in elem.text or 'Concentration' in elem.text][0]
#     except:
#         Concentration = None
#     try:
#         date = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                 driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                 'سنة الإصدار' in elem.text or 'Year of Launch' in elem.text][0]
#     except:
#         date = None
#     try:
#         kiys = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                 driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                 'مفتاح البحث' in elem.text or 'Search Key' in elem.text][0]
#     except:
#         kiys = None
#     try:
#         if size1 is None:
#             size1 = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                      driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                      'الحجم' in elem.text or 'Size' in elem.text][0]
#     except:
#         size1 = None
#
#     """try:
#         driver.find_element_by_css_selector('.product-reviews').click()
#         sleep(4)
#         rat = driver.find_element_by_css_selector('.average').text
#     except:
#         rat = None"""
#     """try:
#         mak = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                driver.find_elements_by_css_selector('#product-information > div > ul > li') if 'الماركة' in elem.text][
#             0]
#     except:
#         mak = None
#     try:
#         Gender = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                   driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                   'Gender' in elem.text][0]
#     except:
#         Gender = None
#     try:
#         Product_Type = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                         driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                         'Product Type' in elem.text][0]
#     except:
#         Product_Type = None
#     try:
#         Character = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                      driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                      'Character' in elem.text][0]
#     except:
#         Character = None
#     try:
#         Fragrance = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                      driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                      'Fragrance Family' in elem.text][0]
#     except:
#         Fragrance = None
#     try:
#         Occasion = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                     driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                     'Occasion' in elem.text][0]
#     except:
#         Occasion = None
#     try:
#         Ingredients = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                        driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                        'Ingredients' in elem.text][0]
#     except:
#         Ingredients = None
#     try:
#         Concentration = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
#                          driver.find_elements_by_css_selector('#product-information > div > ul > li') if
#                          'Concentration' in elem.text][0]
#     except:
#         Concentration = None
#
#     """
#     """ws[f'A{num}'] = name
#     ws[f'B{num}'] = Type
#     ws[f'C{num}'] = date
#     ws[f'D{num}'] = aval
#     ws[f'E{num}'] = rat
#     ws[f'F{num}'] = mak
#     ws[f'G{num}'] = Gender
#     ws[f'H{num}'] = Product_Type
#     ws[f'I{num}'] = Character
#     ws[f'J{num}'] = Fragrance
#     ws[f'K{num}'] = Occasion
#     ws[f'L{num}'] = Concentration
#     ws[f'M{num}'] = kiys
#     # 40
#     ws[f'N{num}'] = size1
#     ws[f'O{num}'] = price1
#     ws[f'P{num}'] = size2
#     ws[f'Q{num}'] = price2
#     ws[f'R{num}'] = size3
#     ws[f'S{num}'] = price3
#
#     ws[f'T{num}'] = Ingredients
#     ws[f'U{num}'] = des
#     ws[f'V{num}'] = img
#     """
#     ws[f'X{num}'] = name
#     ws[f'Y{num}'] = Type
#     ws[f'Z{num}'] = date
#     ws[f'AA{num}'] = aval
#     # ws[f'AB{num}'] = rat
#     ws[f'AC{num}'] = mak
#     ws[f'AD{num}'] = Gender
#     ws[f'AE{num}'] = Product_Type
#     ws[f'AF{num}'] = Character
#     ws[f'AG{num}'] = Fragrance
#     ws[f'AH{num}'] = Occasion
#     ws[f'AI{num}'] = Concentration
#     ws[f'AJ{num}'] = kiys
#     # 40
#     ws[f'AK{num}'] = size1
#     ws[f'AL{num}'] = price1
#     # ws[f'AM{num}'] = size2
#     # ws[f'AN{num}'] = price2
#     # ws[f'AO{num}'] = size3
#     # ws[f'AP{num}'] = price3
#
#     ws[f'AQ{num}'] = Ingredients
#     ws[f'AR{num}'] = des
#     # ws[f'AS{num}'] = img
#     num += 1
#     print(price1, size1)
#     print(num - 1)
#
# wb.save('3tor.xlsx')

'''
لوحة التحكم
على الرابط
http://valleyroz.com/admin/
اليوزر 
Naif_admin
الباسوورد
abc@123
'''
# https://assets.goldenscent.com/catalog/product/cache/2/small_image/750x750/9df78eab33525d08d6e5fb8d27136e95/6/6/663350065558_air-val_air-val_paw_patrol_edt_30_ml_30_multiplex_international_llc_cross_dock_1.jpg ,
# https://assets.goldenscent.com/catalog/product/6/6/663350065558_air-val_air-val_paw_patrol_edt_30_ml_30_multiplex_international_llc_cross_dock_2.jpg
# .replace('https://assets.goldenscent.com/catalog/product/','').replace('/','-')
# try:
#     import requests
#     from openpyxl import load_workbook
#     wb = load_workbook('text_3tor.xlsx')
#     ws = wb.active
#     links = [ws[f'V{i}'].value for i in range(2, ws.max_row + 1)]
#     n = 2
#     for ur in links:
#         co = ur.split(' ,\n')
#         l = []
#         for url in co:
#             name_img = url.replace('https://assets.goldenscent.com/catalog/product/', '').replace('/', '-')
#             if name_img not in l:
#                 l.append(name_img)
#                 r = requests.get(url)
#                 with open('photo//'+name_img, 'wb') as f:
#                     f.write(r.content)
#
#         ws[f'AS{n}'].value = ','.join(l)
#         n += 1
#     wb.save('text_3tor.xlsx')
# except Exception as e:
#     print(e)


try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as ec
    from time import sleep
    from openpyxl import load_workbook

    wb = load_workbook('3tor203.xlsx')
    ws = wb.active
    l = []
    for i in range(2, ws.max_row + 1):
        if ws[f'E{i}'].value is None:
            l.append((i, ws[f'W{i}'].value))
            print(i)
    print(len(l))
    driver = webdriver.Firefox()
    for num, ulr in l:
        try:
            driver.get(ulr)  # .replace('goldenscent.com/', 'goldenscent.com/en/'))
            sleep(4)
            try:
                name = WebDriverWait(driver, 10).until(
                    ec.visibility_of_element_located((By.CSS_SELECTOR, "h1.fw300"))).text
                # name = driver.find_element_by_css_selector('h1.fw300').text
            except:
                driver.refresh()
                name = WebDriverWait(driver, 10).until(
                    ec.visibility_of_element_located((By.CSS_SELECTOR, "h1.fw300"))).text
                # name = driver.find_element_by_css_selector('h1.fw300').text

            Type = driver.find_element_by_css_selector('h2.fw400').text
            aval = driver.find_element_by_css_selector('div.stock').text
            try:
                price1 = driver.find_element_by_css_selector('div.col-xs-6.fw400 span:nth-child(1)').text
            except:
                price1 = None
            try:
                size1 = driver.find_elements_by_css_selector('.type-perfume')[0].text
            except:
                size1 = None

            des = WebDriverWait(driver, 5).until(
                ec.visibility_of_element_located((By.CSS_SELECTOR, '#product-description > p'))).text
            # driver.find_element_by_css_selector('#product-description > p').text
            driver.find_element_by_css_selector('.product-information').click()
            sleep(1)

            try:
                mak = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                       driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                       'الماركة' in elem.text or 'Brand' in elem.text][
                    0]
            except:
                mak = None
            try:
                Gender = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                          driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                          'الجنس' in elem.text or 'Gender' in elem.text][0]
            except:
                Gender = None
            try:
                Product_Type = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                                driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                                'نوع المنتج' in elem.text or 'Product Type' in elem.text][0]
            except:
                Product_Type = None
            try:
                Character = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                             driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                             'شخصية عطرك' in elem.text or 'Character' in elem.text][0]
            except:
                Character = None
            try:
                Fragrance = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                             driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                             'العائلة العطرية' in elem.text or 'Fragrance Family' in elem.text][0]
            except:
                Fragrance = None
            try:
                Occasion = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                            driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                            'مناسب' in elem.text or 'Occasion' in elem.text][0]
            except:
                Occasion = None
            try:
                Ingredients = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                               driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                               'المكونات' in elem.text or 'Ingredients' in elem.text][0]
            except:
                Ingredients = None
            try:
                Concentration = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                                 driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                                 'نسبة التركيز' in elem.text or 'Concentration' in elem.text][0]
            except:
                Concentration = None
            try:
                date = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                        driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                        'سنة الإصدار' in elem.text or 'Year of Launch' in elem.text][0]
            except:
                date = None
            try:
                kiys = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                        driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                        'مفتاح البحث' in elem.text or 'Search Key' in elem.text][0]
            except:
                kiys = None
            try:
                if size1 is None:
                    size1 = [elem.find_element_by_css_selector('div:nth-child(2)').text for elem in
                             driver.find_elements_by_css_selector('#product-information > div > ul > li') if
                             'الحجم' in elem.text or 'Size' in elem.text][0]
            except:
                size1 = None
            try:
                driver.find_element_by_css_selector('.product-reviews').click()
                sleep(7)
                rat = driver.find_element_by_css_selector('.average').text
            except:
                rat = None
            ws[f'A{num}'] = name
            ws[f'B{num}'] = Type
            ws[f'C{num}'] = date
            ws[f'D{num}'] = aval
            ws[f'E{num}'] = rat
            ws[f'F{num}'] = mak
            ws[f'G{num}'] = Gender
            ws[f'H{num}'] = Product_Type
            ws[f'I{num}'] = Character
            ws[f'J{num}'] = Fragrance
            ws[f'K{num}'] = Occasion
            ws[f'L{num}'] = Concentration
            ws[f'M{num}'] = kiys
            ws[f'N{num}'] = size1
            ws[f'O{num}'] = price1
            ws[f'T{num}'] = Ingredients
            ws[f'U{num}'] = des

            print(num, price1, rat)
            wb.save('3tor203.xlsx')
        except Exception as e:
            print(e, num, ulr)
except:
    pass
