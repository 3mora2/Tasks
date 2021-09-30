import os
from time import sleep
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from requests_html import HTMLSession
from webdriver_manager.firefox import GeckoDriverManager


class CollectPosts:
    def __init__(self):
        self.links = []
        self.book = load_workbook('dub-final_4.xlsx')
        # self.book = Workbook()
        self.sheet = self.book.active
        # chrome_options = Options()
        # chrome_options.add_argument('--no-sandbox')
        # chrome_options.add_argument('--disable-gpu')
        # chrome_options.add_argument('blink-settings=imagesEnabled=false')
        # chrome_options.add_argument('--disable-plugins')
        # chrome_options.add_argument('--disable-popup-blocking')
        #
        # self.user_data = os.path.expanduser("~") + r'\AppData\Local\Google\Chrome\User Data\Test'
        # chrome_options.add_argument("--user-data-dir={}".format(self.user_data))
        # self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)

    def start(self):
        self.driver.get('https://www.dnb.com/business-directory/company-search.html?term=industries&page=1')
        sleep(4)
        while True:
            sleep(10)
            for element in self.driver.find_elements_by_css_selector('.row.search_result .primary_name>a'):
                if element.get_attribute('href') not in self.links:
                    self.links.append(element.get_attribute('href'))
            print(len(self.links))
            if len(self.links) > 1200:
                break
            self.driver.find_element_by_css_selector('.next>a').click()
            sleep(2)
            while True:
                if not self.driver.find_elements_by_css_selector('.amp-loader'):
                    break
        for n, link in enumerate(self.links):
            self.sheet.cell(n + 2, 28).value = link
        self.book.save('dub-final.xlsx')

    def get_(self):
        for i in range(2, self.sheet.max_row + 1):
            print(i)
            url = self.sheet.cell(i, 28).value
            if url and not self.sheet.cell(i, 1).value:
                self.driver.delete_all_cookies()
                self.driver.get(url)

                # sleep(2)
                try:
                    role = self.driver.find_element_by_css_selector('.company-role-comp.role').text
                except:
                    role = None
                try:
                    type_ = self.driver.find_element_by_css_selector('.company-type-comp.type').text
                except:
                    type_ = None
                try:
                    name = self.driver.find_element_by_css_selector('.profile-title h2.profile_header_title').text
                except:
                    name = None
                try:
                    company_link = self.driver.find_element_by_css_selector('#hero-company-link').get_attribute('href')
                except:
                    company_link = None
                try:
                    location = self.driver.find_element_by_css_selector('.location').text
                except:
                    location = None
                try:
                    phone = self.driver.find_element_by_css_selector('.profile-phone-element').text
                except:
                    phone = None
                try:
                    company_summary = self.driver.find_element_by_css_selector('.company_summary').text
                except:
                    company_summary = None
                try:
                    name_profile = self.driver.find_elements_by_css_selector('div.profile_see_more_col>span')[1].text
                except:
                    name_profile = None
                try:
                    tags = '\n'.join([el.get_attribute('innerText') for el in
                                      self.driver.find_elements_by_css_selector('.profile-industry-item')])
                except:
                    tags = None
                self.sheet.cell(i, 1).value = name
                self.sheet.cell(i, 2).value = role
                self.sheet.cell(i, 3).value = type_
                self.sheet.cell(i, 4).value = phone
                self.sheet.cell(i, 5).value = location
                self.sheet.cell(i, 6).value = company_link
                self.sheet.cell(i, 7).value = name_profile
                self.sheet.cell(i, 8).value = tags
                self.sheet.cell(i, 9).value = company_summary
                if i % 10 == 0:
                    print('save')
                    self.book.save('dub-final_2.xlsx')

    def get_2(self):
        try:
            for i in range(2, self.sheet.max_row + 1):
                url = self.sheet.cell(i, 10).value
                if url and not self.sheet.cell(i, 11).value:
                    # self.driver.delete_all_cookies()
                    # self.driver.get(url)
                    # num = self.driver.find_element_by_css_selector('.rev_title_number').text
                    s = HTMLSession()
                    r = s.get(url)
                    try:
                        num = r.html.find('.rev_title_number', first=True).text
                        self.sheet.cell(i, 11).value = num
                        print(i, num)
                    except Exception as e:
                        print(e, r.ok)
                        pass
                    if i % 10 == 0:
                        print('save')
                        self.book.save('dub-final_4.xlsx')
        except Exception as e:
            print(e)
            self.r = r
            self.book.save('dub-final_4.xlsx')


self = CollectPosts()
self.get_2()
'https://www.dnb.com/business-directory/company-search.html?term=industries&page=15'
