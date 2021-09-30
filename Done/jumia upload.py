# Source Generated with Decompyle++
# File: jumia_upload.pyc (Python 3.8)

from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as ec
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
import traceback

class Upload:
    driver: WebDriver = 'Upload'
    URL_LOGIN = 'https://sellercenter.jumia.com.eg/user/auth/login'
    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.NAME, 'submit')
    URL_CREATE = 'https://sellercenter.jumia.com.eg/product/create/from-shop/sku/'
    
    def __init__(self):
        self.book = load_workbook(file_name)
        self.sheet = self.book.active
        chrome_options = Options()
        chrome_options.add_argument('--disable-infobars')
        chrome_options.add_argument('--disable-notifications')
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options, **('options',))
        self.driver.maximize_window()
        self.login()

    
    def login(self):
        
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR)).send_keys(email)
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR)).send_keys(password)
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR)).click()
            if 'user/auth/login' not in self.driver.current_url:
                print('- Please Wait 30s...')
                sleep(30)
                self.alert1()
                self.alert()
            
        finally:
            pass
        print(traceback.print_exc())
        print('- Login Error')
        return None


    
    def alert(self):
        
        try:
            self.driver.find_element_by_css_selector('div.walkme-action-close').click()
        finally:
            pass
        return None


    
    def alert1(self):
        
        try:
            self.driver.find_element_by_css_selector('button.wm-visual-design-button').click()
        finally:
            pass
        return None


    
    def upload(self):
        for i in range(2, self.sheet.max_row - 4):
            
            try:
                bar = self.sheet.cell(i, 1).value
                price = self.sheet.cell(i, 2).value
                quantity = self.sheet.cell(i, 3).value
                if bar is None:
                    pass
            finally:
                continue
                print(bar)
                self.driver.get(str(self.URL_CREATE) + str(bar))
                sleep(2)
                
                try:
                    self.driver.switch_to_alert().accept()
                finally:
                    pass
                print('- Not alert')
                self.alert()
                self.alert1()
                sleep(1)
                self.driver.find_element_by_css_selector('input.sku_ean[required="true"]').send_keys(bar)
                sleep(1)
                if quantity:
                    self.driver.find_element_by_css_selector('input[name="variations[0][quantity]"][type="number"]').send_keys(quantity)

                sleep(1)
                if price:
                    self.driver.find_element_by_css_selector('input[name="variations[0][price]"]').send_keys(price)
                
                try:
                    self.driver.find_element_by_css_selector('#product-form-submit-btn').click()
                finally:
                    pass
                print("- Can' Save")
                continue
                print(traceback.print_exc())
                continue
                continue
                self.driver.quit()
                return None




# WARNING: Decompyle incomplete
