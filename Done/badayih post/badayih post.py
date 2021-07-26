from selenium.webdriver.firefox.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
import random
from webdriver_manager.firefox import GeckoDriverManager
from time import sleep
from PySide2.QtCore import QThread, Signal
from PySide2.QtWidgets import QMainWindow, QApplication, QFileDialog, QCheckBox, QRadioButton, QListWidgetItem
from selenium import webdriver
from selenium.webdriver.common.by import By
import os
import sys
import warnings
from main import Ui_MainWindow
from openpyxl import load_workbook, Workbook

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.Handel_Button()
        self.Main = Main()
        self.Main.finish.connect(self.addlist)
        self.Main.end.connect(self.end)

    def Handel_Button(self):
        self.ui.pushButton.clicked.connect(self.Open_File_Test)
        self.ui.pushButton_2.clicked.connect(self.start)
        self.ui.pushButton_3.clicked.connect(self.open)
        path = os.path.expanduser("~") + '\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles'
        for file in os.listdir(path):
            prof = file
            box = QRadioButton(prof)
            item = QListWidgetItem()
            self.ui.listWidget_2.addItem(item)
            self.ui.listWidget_2.setItemWidget(item, box)

    def Open_File_Test(self):
        save_file = QFileDialog.getOpenFileName(self, "اختار", '', '*.xlsx')[0]
        if save_file != '':
            sheet = load_workbook(save_file).active
            self.Main.sheet = sheet
            self.ui.pushButton.setDisabled(True)
            # with open(save_file, 'r', encoding="utf-8") as f:
            #     self.Main.data = f.read()

    def addlist(self):
        print('- add')
        self.Main.driver.execute_script(
            'document.querySelector(\'[href="#category-all"]\').click()')
        sleep(1)
        for i, element in enumerate(self.Main.driver.find_elements_by_css_selector('#categorychecklist > li> label')):
            box = QCheckBox(element.text.strip())
            item = QListWidgetItem()
            self.ui.listWidget.addItem(item)
            self.ui.listWidget.setItemWidget(item, box)

    def start(self):
        cb_list = [self.ui.listWidget.itemWidget(self.ui.listWidget.item(i))
                   for i in range(self.ui.listWidget.count())]
        chooses = []
        for cb in cb_list:
            if cb.isChecked():
                chooses.append(cb.text())
        print(chooses)
        self.Main.cat = chooses
        self.Main.start()
        self.ui.listWidget.setDisabled(True)
        self.ui.pushButton.setDisabled(True)
        self.ui.pushButton_2.setDisabled(True)
        self.ui.pushButton_3.setDisabled(True)

    def open(self):
        path = os.path.expanduser("~") + '\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles'
        cb_list = [self.ui.listWidget_2.itemWidget(self.ui.listWidget_2.item(i))
                   for i in range(self.ui.listWidget_2.count())]

        for cb in cb_list:
            if cb.isChecked():
                self.Main.prof = path + '\\' + cb.text()

        self.Main.open()
        self.ui.pushButton_3.setDisabled(True)

    def end(self):
        self.ui.listWidget.setDisabled(False)
        self.ui.pushButton.setDisabled(False)
        self.ui.pushButton_2.setDisabled(False)
        self.ui.pushButton_3.setDisabled(False)


class Main(QThread):
    driver: WebDriver
    finish = Signal()
    end = Signal()

    def __init__(self):
        super().__init__()
        try:
            self.old_book = load_workbook('old.xlsx')
            self.old_sheet = self.old_book.active
        except:
            self.old_book = Workbook()
            self.old_sheet = self.old_book.active

        self.old = [self.old_sheet.cell(i, 1).value.strip() for i in range(1, self.old_sheet.max_row + 1) if
                    self.old_sheet.cell(i, 1).value]

        self.current = self.old_sheet.max_row + 2
        # self.f = open('Done.txt', 'a+', encoding="utf-8")
        # self.old = self.f.read()
        # path = os.path.expanduser("~") + '\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles'
        # for file in os.listdir(path):
        #     if file.endswith('release'):
        #         self.prof = path + '\\' + file

    def open(self):
        print('- open browser')
        user_data = self.prof
        fp = webdriver.FirefoxProfile(user_data)
        options = webdriver.FirefoxOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-setuid-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        try:
            self.driver = webdriver.Firefox(firefox_profile=fp, options=options)
        except:
            self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install(), firefox_profile=fp,
                                            options=options)
        self.driver.get('https://badayih.com/wp-admin/post-new.php?wp-post-new-reload=true')
        self.finish.emit()
        print('- end')

    def run(self):
        # for post in self.data.split('*******************************************'):
        for i in range(2, self.sheet.max_row + 1):
            print(i)
            post = self.sheet.cell(i, 1).value
            if post:
                post = post.strip()
                if post in self.old:
                    print('- found')
                    continue
                if post.replace('*', '').strip() == '':
                    print('- ****')
                    continue
                if post or post.replace('\n', '').replace(' ', '') != '':
                    self.post(post)
                    self.old_sheet.cell(self.current, 1).value = post
                    self.current += 1
                    self.old_book.save('old.xlsx')
                    ss = random.choice(range(60 * 10, 60 * 40))
                    print('Wait:', ss)
                    sleep(ss)
        self.end.emit()

    def post(self, post):
        print(post)
        self.driver.get('https://badayih.com/wp-admin/post-new.php?wp-post-new-reload=true')
        sleep(6)
        WebDriverWait(self.driver, 20).until(
            (ec.presence_of_element_located((By.CSS_SELECTOR, 'iframe')))).send_keys(' ')
        WebDriverWait(self.driver, 20).until(
            (ec.presence_of_element_located((By.CSS_SELECTOR, 'iframe')))).send_keys(post)
        sleep(3)
        # self.driver.find_element_by_css_selector('iframe').send_keys('gg')
        title = post.split('\n')[0]
        if len(title) < 10:
            title = post[:50]
        self.driver.find_element_by_css_selector('#title').send_keys(title)
        sleep(3)
        self.driver.find_element_by_css_selector('#new-tag-post_tag').send_keys(title)
        sleep(3)
        self.driver.find_element_by_css_selector('#new-tag-post_tag + input').click()
        sleep(3)
        if not self.driver.execute_script(
                'return document.querySelectorAll(\'input[name="publish_now_accounts[]"]\')[0].checked'):
            self.driver.find_elements_by_css_selector('input[name="publish_now_accounts[]"]')[0].click()
            sleep(3)

        if not self.driver.execute_script(
                'return document.querySelectorAll(\'input[name="publish_now_accounts[]"]\')[1].checked'):
            self.driver.find_elements_by_css_selector('input[name="publish_now_accounts[]"]')[1].click()
            sleep(3)
        # for element in self.driver.find_elements_by_css_selector('#categorychecklist-pop > li> label'):
        #     if "سوق الاثاث - بدايه الخبر" == element.text.strip():
        #         element.find_element_by_tag_name('input').click()
        self.driver.execute_script(
            'document.querySelector(\'[href="#category-all"]\').click()')

        sleep(3)

        for element in self.driver.find_elements_by_css_selector('#categorychecklist > li> label'):
            if element.text.strip() in self.cat:
                if element.find_element_by_tag_name('input').is_selected():
                    pass
                try:
                    element.find_element_by_tag_name('input').click()
                except:
                    element.find_element_by_tag_name('input').click()
                sleep(3)
            else:
                if element.find_element_by_tag_name('input').is_selected():
                    element.find_element_by_tag_name('input').click()

        self.driver.find_element_by_css_selector(
            '.aioseo-input.add-focus-keyphrase-metabox-input > input.medium').send_keys(title)
        sleep(3)
        while True:
            try:
                self.driver.find_element_by_css_selector('#add-focus-keyphrase').click()
                break
            except:
                pass

        sleep(3)

        self.driver.find_element_by_css_selector('#publish').click()
        sleep(5)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec_()

'''
https://badayih.com/wp-admin/post-new.php?wp-post-new-reload=true
https://badayih.com/wp-admin/
user

qSVTZzzM#o9gi*q6cJxza02I
pyside2-uic mainwindow.ui -o mainwindow.py

'''
