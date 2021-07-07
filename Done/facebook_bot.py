import subprocess
from time import sleep
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime


class CollectPosts(object):
    def __init__(self):
        self.file_save = 'face.xlsx'
        self.is_group = None
        self.current_cell_likes = 2
        self.data_likes = set()
        self.book = Workbook()
        self.sheet = self.book.active
        # fp = webdriver.FirefoxProfile(
        #     r'C:\Users\3mora\AppData\Roaming\Mozilla\Firefox\Profiles\63z4wsqh.default-release')
        # self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install(), firefox_profile=fp)
        try:
            sheet = load_workbook('user-data.xlsx').active
            if sheet.cell(1, 1).value:
                chrome_options = webdriver.ChromeOptions()
                chrome_options.add_argument('--no-sandbox')
                chrome_options.add_argument('--disable-gpu')
                chrome_options.add_argument(f"--user-data-dir={sheet.cell(1, 1).value}")
                self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
            else:
                0>'00'
        except:

            try:
                subprocess.Popen(
                    '"C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"'
                    r' --user-data-dir="C:\selenium\AutomationProfile" --profile-directory="Profile 1"'
                    ' --remote-debugging-port=9222')
            except:
                subprocess.Popen(
                    '"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"'
                    r' --user-data-dir="C:\selenium\AutomationProfile" --profile-directory="Profile 1"'
                    '  --remote-debugging-port=9222')

            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
            self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)

    def safe_find_element_by(self, by, elem):
        try:
            return self.driver.find_element(by, elem)
        except NoSuchElementException:
            return None

    def check_login(self):
        self.driver.get('https://mbasic.facebook.com/login')
        return self.safe_find_element_by(By.NAME, 'email')

    def login(self):
        try:

            self.driver.get("https://mbasic.facebook.com/language.php")
            sleep(1)
            self.driver.find_element_by_css_selector('[value="English (US)"]').click()
            sleep(1)
            self.driver.get("https://mbasic.facebook.com/login")
            self.driver.find_element_by_name('email').send_keys(email)
            self.driver.find_element_by_name('pass').send_keys(password)
            self.driver.find_element_by_css_selector('[name="login"]').click()
            not_now = self.safe_find_element_by(By.LINK_TEXT, "Not Now")
            if not_now:
                not_now.click()

        except Exception as e:
            print(e)
            print("There was some error while logging in.")

    def start_collect(self, url):
        self.file_save = 'face -' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + '.xlsx'
        self.is_group = None
        url = url.replace('m.facebook.com', 'mbasic.facebook.com').replace('www.facebook.com', 'mbasic.facebook.com')
        if '?' in url:
            url = url.split('?')[0]
        url += '?v=timeline'
        if '/groups/' in url:
            self.is_group = True
        self.driver.get(url)
        sleep(10)
        self.GetPosts()

    def mobile_start(self, url):
        self.file_save = 'face -' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + '.xlsx'
        url = url.replace('www.facebook.com', 'm.facebook.com')
        if '?' in url:
            url = url.split('?')[0]
        # url += '/posts/?ref=page_internal&mt_nav=0'
        self.driver.get(url)
        sleep(10)
        self.GetPostsNew()

    def GetPostsNew(self):
        lists = []
        for i in range(numb):
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            sleep(4)
            for element in self.driver.find_elements_by_css_selector('[data-sigil="m-feed-voice-subtitle"] > a'):
                u = element.get_attribute('href')
                if u:
                    u = u.replace('m.facebook.com', 'mbasic.facebook.com')
                if u not in lists:
                    lists.append(u)
            print(len(lists))
        for url in lists:
            self.Post(url)

    def Post(self, url):
        if self.is_group:
            url = url.replace('mbasic.', 'm.')
        self.driver.get(url)
        sleep(10)
        link = self.driver.find_element(By.CSS_SELECTOR, "#add_comment_switcher_placeholder ~ div > a").get_attribute(
            'href')
        likes_url = link.replace('m.facebook', 'mbasic.facebook')
        comment_url = self.driver.current_url
        if self.is_group:
            lc = comment_url.split('?')
            if len(lc):
                comment_url = lc[0]
        else:
            lc = comment_url.split('&refid=')
            if len(lc):
                comment_url = lc[0]

        self.get_comments(comment_url)
        self.get_likes(likes_url)

    def GetPosts(self):
        urls = []
        total = 0
        break_ = False
        next_url = ''
        while True:
            for element in self.driver.find_elements_by_css_selector('section > article'):
                try:
                    try:
                        u = element.find_element(By.LINK_TEXT, "Full Story").get_attribute('href')
                    except:
                        u = element.find_element(By.LINK_TEXT, "الحدث بالكامل").get_attribute('href')

                    if u not in urls:
                        total += 1
                        urls.append(u)
                except Exception as e:
                    print(e)

            try:
                next_url = self.driver.find_element_by_xpath(
                    '//div[a/span[contains(text(),"See more stories")]]/a | '
                    '//div[a/span[contains(text(),"See More Stories")]]/a | '
                    '//div/a[contains(text(),"Show more")] | '
                    '//div[a/span[contains(text(),"See more")]]/a').get_attribute('href')
            except:
                print('- ***************** Finish ***************')
                break_ = True

            if len(urls) > 4:
                for index, url in enumerate(urls):
                    try:
                        print('- Number :', len(urls), '-', index + 1, 'All :', total)
                        self.Post(url)
                    except Exception as e:
                        print(e)
                urls = []
            if total >= numb or break_:
                break
            else:
                print('- Next .......')
                self.driver.get(next_url)
                sleep(10)

    def get_likes(self, url):
        self.driver.get(url)
        sleep(10)
        while True:
            for element in self.driver.find_elements(By.CSS_SELECTOR, "ul > li table table a"):
                name = element.text
                link = element.get_attribute('href')
                user_id = element.get_attribute('href').split('/')[-1] if '?id=' not in element.get_attribute(
                    'href') else \
                    element.get_attribute('href').split('?id=')[-1]
                if '?' in user_id:
                    user_id = user_id.split('?')[0]
                if '&rc=' in user_id:
                    user_id = user_id.split('&rc=')[0]
                if '?ref=' in user_id:
                    user_id = user_id.split('&rc=')[0]

                if user_id in self.data_likes:
                    print('- found')
                    continue

                self.data_likes.add(user_id)
                self.sheet.cell(self.current_cell_likes, 1).value = name
                self.sheet.cell(self.current_cell_likes, 2).value = user_id
                self.sheet.cell(self.current_cell_likes, 3).value = link
                print(self.current_cell_likes, user_id)
                self.current_cell_likes += 1

            next_ = self.safe_find_element_by(By.LINK_TEXT, "See more")
            if not next_:
                next_ = self.safe_find_element_by(By.LINK_TEXT, "عرض المزيد")

            if next_:
                url = next_.get_attribute('href').replace('limit=10', 'limit=3000')
                self.driver.get(url)
                print('- wait...')
                sleep(10)

            else:
                break

        self.book.save(self.file_save)

    def get_comments(self, url):
        # self.driver.get(url)
        # sleep(5)
        if self.is_group:
            by = By.XPATH
            value = '//div[div[@data-sigil="comment-body"]] //a'
        else:
            by = By.CSS_SELECTOR
            value = 'h3 > a'

        while True:
            for element in self.driver.find_elements(by, value):
                name = element.text
                user_id = element.get_attribute('href').split('/')[-1] if '?id=' not in element.get_attribute(
                    'href') else \
                    element.get_attribute('href').split('?id=')[-1]
                if '?' in user_id:
                    user_id = user_id.split('?')[0]
                if '&rc=' in user_id:
                    user_id = user_id.split('&rc=')[0]
                if '?ref=' in user_id:
                    user_id = user_id.split('&rc=')[0]

                link = element.get_attribute('href')
                if '/albums/' in link or 'story.php' in link:
                    continue
                if user_id in self.data_likes:
                    print('- found')
                    continue
                if 'facebook.com/' not in link or 'facebook.com/hashtag' in link:
                    continue

                self.data_likes.add(user_id)
                self.sheet.cell(self.current_cell_likes, 1).value = name
                self.sheet.cell(self.current_cell_likes, 2).value = user_id
                self.sheet.cell(self.current_cell_likes, 3).value = link
                print(self.current_cell_likes, user_id)
                self.current_cell_likes += 1

            more = self.safe_find_element_by(By.LINK_TEXT, 'View more comments…')
            if more:
                self.driver.get(more.get_attribute('href'))
                print('- wait...')
                sleep(10)

            else:
                break

        self.book.save(self.file_save)

    def get_number(self, file):
        book_ = load_workbook(file)
        sheet = book_.active
        n = 0
        for i in range(2, sheet.max_row + 1):
            url = sheet.cell(i, 3).value
            if url:
                if n == 50:
                    sleep(1000)
                self.driver.get(url.replace('m.face', 'mbasic.face'))
                n += 1
                sleep(10)
                phone = self.safe_find_element_by(
                    By.XPATH,
                    '//div[@id="contact-info"] //tr[td/div/span[contains(text(),"WhatsApp")]]/td[2] |'
                    '//div[@id="contact-info"] //tr[td/div/span[contains(text(),"Mobile")]]/td[2]'
                )
                gender = self.safe_find_element_by(
                    By.XPATH,
                    '//div[@id="basic-info"] //tr[td/div/span[contains(text(),"Gender")]]/td[2]')

                if phone:
                    sheet.cell(i, 4).value = phone.text
                if gender:
                    sheet.cell(i, 5).value = gender.text
                error = self.safe_find_element_by(
                    By.XPATH,
                    '//div[contains(@title(),"You can\'t use this feature at the moment")]')
            print(i, sheet.cell(i, 4).value, sheet.cell(i, 5).value)
        book_.save(file)


if __name__ == '__main__':
    email = input('- Enter your email : ')
    password = input('- Enter your password : ')

    app = CollectPosts()
    app.login()
    while True:
        input_res = input('- 1 >>> for extract data Old'
                          '\n- 2 >>> for extract data New only page'
                          '\n- 3 >>> get phone number'
                          ': ')
        if input_res == '1':
            link = str(input('''- Enter URL: '''))
            numb = input('- Max Number : ')
            try:
                numb = int(numb)
            except:
                numb = 50

            app.start_collect(link)
            break
        elif input_res == '2':
            link = str(input('''- Enter URL: '''))
            numb = input('- Max Number : ')
            try:
                numb = int(numb)
            except:
                numb = 50

            app.mobile_start(link)
            break
        elif input_res == '3':
            while True:
                try:
                    file_name = input('- Enter file name : ')  # 'Book1.xlsx'
                    book = load_workbook(file_name)
                    book.close()
                    break
                except FileNotFoundError:
                    print('- No such file or directory')
                except Exception as e:
                    print(e)
            app.get_number(file_name)

    # self = app
r'C:\Users\3mora\AppData\Local\Google\Chrome\User Data\Profile 4'
r'C:\Users\3mora\AppData\Local\Google\Chrome\User Data\Profile 3'
'https://www.facebook.com/HorusAndHorAinGifts'