from selenium import webdriver
from time import sleep
import os
from requests_html import HTMLSession

s = HTMLSession()
from openpyxl import load_workbook

# # i = 1
# path = 'desktop.xlsx'
# book = load_workbook(path)
# sheet = book.active
# end_cell = sheet.max_row
# i = 2
# links = [sheet[f'N{o}'].value for o in range(i, end_cell + 1) if
#          sheet[f'N{o}'].value is not None]
# print(len(links))
# driver = webdriver.Chrome()
# for ur in links:
#     driver.get(ur)  # https://www.extra.com/ar-sa/small-appliances/c/5
#     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#     sleep(2)
#     try:
#         cat1 = driver.find_element_by_css_selector('div.c_breadcrumb > ul > li:nth-child(2) > a').text
#     except:
#         cat1 = None
#     try:
#         cat2 = driver.find_element_by_css_selector('div.c_breadcrumb > ul > li:nth-child(3) > a').text
#     except:
#         cat2 = None
#     try:
#         cat3 = driver.find_element_by_css_selector('div.c_breadcrumb > ul > li:nth-child(4) > a').text
#     except:
#         cat3 = None
#     try:
#         cat4 = driver.find_element_by_css_selector('div.c_breadcrumb > ul > li:nth-child(5) > a').text
#     except:
#         cat4 = None
#     try:
#         r = s.get(ur.replace('com/ar-sa/', 'com/en-sa/'))
#         try:
#             name_en = r.html.find(
#                 'div.c_product-sidebar > div.c_product-details-title span.promotion-title-text')[0].text
#         except:
#             name_en = r.html.find(
#                 'div.c_product-sidebar > div.c_product-details-title')[0].text
#     except Exception as e:
#         print(e)
#         name_en = None
#     try:
#         name = driver.find_element_by_css_selector(
#             'div.c_product-sidebar > div.c_product-details-title span.promotion-title-text').text
#     except:
#         name = driver.find_element_by_css_selector(
#             'div.c_product-sidebar > div.c_product-details-title').text
#     prev_price = None
#     try:
#         price = ' '.join(driver.find_element_by_css_selector(
#             'div.section-price div.c_product-price-prices  div.c_product-price-current').text.split('\n')[0:-1])
#         prev_price = driver.find_element_by_css_selector(
#             '#content > div:nth-child(4) > div > div.col-flex-lg-5 > div.c_product-sidebar > div.section-price > div.simple-flex.product-details-price-section > div > div.c_product-price-prices > div.c_product-price-was-is > span.c_product-price-previous').text
#
#     except:
#         try:
#             price = driver.find_element_by_css_selector(
#                 'div.section-price > div.simple-flex.product-details-price-section > div > div.c_product-price-current').text.replace(
#                 '\n', ' ')
#
#         except:
#             price = None
#     driver.execute_script("window.scrollTo(0,800);")
#     try:
#         main_mm = driver.find_element_by_css_selector(
#             '#content > div:nth-child(4) > div > div.col-flex-lg-7 > div.c_keyfeatures-content.hidden-xs.hidden-sm > div > div > div').text
#     except:
#         main_mm = None
#     try:
#         inf = driver.find_element_by_css_selector(
#             '#product_specification > div.row-flex.c_product-information-attrirubutes').text
#     except:
#         inf = None
#     try:
#         des = driver.find_element_by_css_selector('#product_information_specification > div').text
#     except:
#         des = None
#     driver.execute_script("window.scrollTo(0,1500);")
#     try:
#         mm = driver.find_element_by_css_selector('#highlightcont').text
#     except Exception as e:
#         mm = None
#     try:
#         if mm is None or mm == '':
#             mm = driver.find_element_by_css_selector('#inpage_container').text
#     except:
#         pass
#     driver.execute_script("window.scrollTo(0, document.body.scrollHeight-1900);")
#     try:
#         prod = '\n'.join([ele.text for ele in driver.find_elements_by_css_selector(
#             '#item_page-rank1  div.c_item.swiper-slide.c_item--grid.js-gtm-button.swiper-slide-visible> a > div.c_item--grid--detail > div.c_item--grid--title > h3')])
#     except:
#         prod = None
#
#     try:
#         prods = '\n'.join([ele.text for ele in driver.find_elements_by_css_selector(
#             '#item_page-rank2  div.c_item.swiper-slide.c_item--grid.js-gtm-button.swiper-slide-visible> a > div.c_item--grid--detail > div.c_item--grid--title > h3')])
#     except:
#         prods = None
#     try:
#         img = ' \n '.join([ele.get_attribute('src') for ele in driver.find_elements_by_css_selector(
#             'div.main-container > ul > div > li > div > img')])
#     except:
#         img = None
#     sheet[f'A{i}'].value = name
#     sheet[f'B{i}'].value = price
#     sheet[f'C{i}'].value = prev_price
#     sheet[f'D{i}'].value = cat1
#     sheet[f'E{i}'].value = cat2
#     sheet[f'F{i}'].value = cat3
#     sheet[f'G{i}'].value = cat4
#     sheet[f'H{i}'].value = main_mm
#     sheet[f'I{i}'].value = des
#     sheet[f'J{i}'].value = inf
#     sheet[f'K{i}'].value = mm
#     sheet[f'L{i}'].value = img
#     sheet[f'M{i}'].value = name_en
#     sheet[f'O{i}'].value = prod
#     sheet[f'P{i}'].value = prods
#     print(i, name, price)
#     i += 1
#     book.save(path)
#############
# try:
#     k = 2
#     book = Workbook()
#     sheet = book.active
#     links = []
#     page = 0
#     while True:
#         driver.get(
#             f'https://www.extra.com/ar-sa/computer/laptops/c/3-303?q=%3Arelevance%3AinStock%3Atrue&pg={}&pageSize=48')
#         page += 1
#         sleep(5)
#         if len(driver.find_elements_by_css_selector('div.row-flex.c_product-listing > div > div > a')) == 0:
#             break
#         for u in driver.find_elements_by_css_selector('div.row-flex.c_product-listing > div > div > a'):
#             link = u.get_attribute('href')
#             if link not in links:
#                 links.append(link)
#                 print(len(links))
#         sleep(3)
#         try:
#             driver.find_element_by_css_selector('.c_pagination-next a').click()
#         except:
#             pass
#     try:
#         for i in links:
#             sheet[f'N{k}'].value = i
#             k += 1
#         book.save('laptops.xlsx')
#     except Exception as e:
#         print(e)
# except Exception as r:
#     print(r)
driver = webdriver.Firefox()
for file in os.listdir():
    u = []
    if file.endswith('.xlsx'):
        print(file)
        book = load_workbook(file)
        sheet = book.active
        end_cell = sheet.max_row

        for o in range(2, end_cell + 1):
            if sheet[f'M{o}'].value is None:
                u.append((sheet[f'N{o}'].value, o))
        print(len(u))
        continue
        for ur, i in u:
            driver.get(ur)  # https://www.extra.com/ar-sa/small-appliances/c/5
            sleep(5)
            # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            # sleep(2)
            # try:
            #     cat1 = driver.find_element_by_css_selector('div.c_breadcrumb > ul > li:nth-child(2) > a').text
            # except:
            #     cat1 = None
            # try:
            #     cat2 = driver.find_element_by_css_selector('div.c_breadcrumb > ul > li:nth-child(3) > a').text
            # except:
            #     cat2 = None
            # try:
            #     cat3 = driver.find_element_by_css_selector('div.c_breadcrumb > ul > li:nth-child(4) > a').text
            # except:
            #     cat3 = None
            # try:
            #     cat4 = driver.find_element_by_css_selector('div.c_breadcrumb > ul > li:nth-child(5) > a').text
            # except:
            #     cat4 = None
            # try:
            #     r = s.get(ur.replace('com/ar-sa/', 'com/en-sa/'))
            #     try:
            #         name_en = r.html.find(
            #             'div.c_product-sidebar > div.c_product-details-title span.promotion-title-text')[0].text
            #     except:
            #         name_en = r.html.find(
            #             'div.c_product-sidebar > div.c_product-details-title')[0].text
            # except Exception as e:
            #     print(e)
            #     name_en = None
            # try:
            #     name = driver.find_element_by_css_selector(
            #         'div.c_product-sidebar > div.c_product-details-title span.promotion-title-text').text
            # except:
            #     name = driver.find_element_by_css_selector(
            #         'div.c_product-sidebar > div.c_product-details-title').text
            # prev_price = None
            # try:
            #     price = ' '.join(driver.find_element_by_css_selector(
            #         'div.section-price div.c_product-price-prices  div.c_product-price-current').text.split('\n')[0:-1])
            #     prev_price = driver.find_element_by_css_selector(
            #         '#content > div:nth-child(4) > div > div.col-flex-lg-5 > div.c_product-sidebar > div.section-price > div.simple-flex.product-details-price-section > div > div.c_product-price-prices > div.c_product-price-was-is > span.c_product-price-previous').text
            #
            # except:
            #     try:
            #         price = driver.find_element_by_css_selector(
            #             'div.section-price > div.simple-flex.product-details-price-section > div > div.c_product-price-current').text.replace(
            #             '\n', ' ')
            #
            #     except:
            #         price = None
            # driver.execute_script("window.scrollTo(0,800);")
            # try:
            #     main_mm = driver.find_element_by_css_selector(
            #         '#content > div:nth-child(4) > div > div.col-flex-lg-7 > div.c_keyfeatures-content.hidden-xs.hidden-sm > div > div > div').text
            # except:
            #     main_mm = None
            # try:
            #     inf = driver.find_element_by_css_selector(
            #         '#product_specification > div.row-flex.c_product-information-attrirubutes').text
            # except:
            #     inf = None
            # try:
            #     des = driver.find_element_by_css_selector('#product_information_specification > div').text
            # except:
            #     des = None
            # driver.execute_script("window.scrollTo(0,1500);")
            # try:
            #     mm = driver.find_element_by_css_selector('#highlightcont').text
            # except Exception as e:
            #     mm = None
            # try:
            #     if mm is None or mm == '':
            #         mm = driver.find_element_by_css_selector('#inpage_container').text
            # except:
            #     pass
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight-1900);")
            try:
                prod = '\n'.join([ele.text for ele in driver.find_elements_by_css_selector(
                    '#item_page-rank1  div.c_item.swiper-slide.c_item--grid.js-gtm-button.swiper-slide-visible> a > div.c_item--grid--detail > div.c_item--grid--title > h3')])
            except:
                prod = None
            try:
                prods = '\n'.join([ele.text for ele in driver.find_elements_by_css_selector(
                    '#item_page-rank2  div.c_item.swiper-slide.c_item--grid.js-gtm-button.swiper-slide-visible> a > div.c_item--grid--detail > div.c_item--grid--title > h3')])
            except:
                prods = None
            try:
                img = ' \n '.join([ele.get_attribute('src') for ele in driver.find_elements_by_css_selector(
                    'div.main-container > ul > div > li > div > img')])
            except:
                img = None
            # sheet[f'A{i}'].value = name
            # sheet[f'B{i}'].value = price
            # sheet[f'C{i}'].value = prev_price
            # sheet[f'D{i}'].value = cat1
            # sheet[f'E{i}'].value = cat2
            # sheet[f'F{i}'].value = cat3
            # sheet[f'G{i}'].value = cat4
            # sheet[f'H{i}'].value = main_mm
            # sheet[f'I{i}'].value = des
            # sheet[f'J{i}'].value = inf
            # sheet[f'K{i}'].value = mm
            sheet[f'L{i}'].value = img
            # sheet[f'M{i}'].value = name_en
            sheet[f'O{i}'].value = prod
            sheet[f'P{i}'].value = prods
            print(i, ' - ', img)
            book.save(file)

# 2646
