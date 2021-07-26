import os
import sys
import threading
import warnings
from idlelib.searchbase import _searchbase

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium import webdriver

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from time import sleep

session = requests.session()


def print_percent_done(index, total, bar_len=50):
    percent_done = index / total * 100
    percent_done = round(percent_done, 1)

    done = round(percent_done / (100 / bar_len))
    togo = bar_len - done

    done_str = '█' * int(done)
    togo_str = '░' * int(togo)

    print(f'- ⏳ {total}\\{index} - [{done_str}{togo_str}] ({percent_done} %)', end='\r')


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 's':
                break


class Main:
    def __init__(self):
        self.urls = set()
        self.user_data = os.path.expanduser("~") + r'\AppData\Local\Google\Chrome\User Data'
        self.book = Workbook()
        self.sheet = self.book.active
        try:
            chrome_options = Options()
            chrome_options.add_argument("--disable-infobars")
            chrome_options.add_argument("--disable-notifications")
            chrome_options.add_argument(f"--user-data-dir={self.user_data}")
            self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
        except:
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
        ##############################################################################################################
        self.driver.maximize_window()
        self.driver.get('https://www.google.com/preferences#languages')

    def open(self, txt):
        thr = Thread()
        thr.start()
        current_n = 0
        current_n_n = 0
        if save_img == '1':
            self.sheet['A1'] = 'Image'
            self.sheet.column_dimensions['A'].width = 13

        self.sheet['B1'] = 'Name'
        self.sheet['C1'] = 'Type'
        self.sheet['D1'] = 'Rate'
        self.sheet['E1'] = 'Comments'
        self.sheet['F1'] = 'Address'
        self.sheet['G1'] = 'Plus Code'
        self.sheet['H1'] = 'Phone'
        self.sheet['I1'] = 'Website'
        self.sheet['J1'] = "IMG"
        self.sheet['K1'] = "URL"

        self.sheet.column_dimensions['B'].width = 50
        self.sheet.column_dimensions['C'].width = 30
        self.sheet.column_dimensions['D'].width = 10
        self.sheet.column_dimensions['E'].width = 10
        self.sheet.column_dimensions['F'].width = 100
        self.sheet.column_dimensions['G'].width = 70
        self.sheet.column_dimensions['H'].width = 25
        self.sheet.column_dimensions['I'].width = 30
        self.sheet.column_dimensions['J'].width = 50
        self.sheet.column_dimensions['K'].width = 120
        for column in range(1, 11):
            try:
                self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except:
                pass

        self.driver.get(f'https://www.google.com/maps/search/{txt}')
        print('Program is loading,please wait!')
        sleep(5)
        input('- Enter...')

        while True:
            num_scroll = 0
            while True:
                try:
                    _ = self.driver.find_element_by_css_selector('.wo1ice-loading').location_once_scrolled_into_view
                    sleep(5)
                    num_scroll += 1

                    if num_scroll > 5:
                        print('- Break Scroll..')
                except:
                    break

            sleep(3)
            num_pl = len(self.driver.find_elements_by_css_selector('div.section-layout.section-scrollbox div > a'))
            if num_pl == 0:
                break

            for element in self.driver.find_elements_by_css_selector('div.section-layout.section-scrollbox div>a'):
                self.urls.add(element.get_attribute('href'))
            print(len(self.urls))

            if thr.p == 's':
                break

            try:
                if self.driver.find_element_by_css_selector(
                        'button[jsaction="pane.paginationSection.nextPage"]').get_attribute('disabled') == 'true':
                    break

                try:
                    self.driver.execute_script('document.querySelector(\'*[aria-label="الصفحة التالية"]\').click()')
                except:
                    self.driver.execute_script(
                        'document.querySelector(\'button[jsaction="pane.paginationSection.nextPage"]\').click()')
                sleep(5)
            except:
                break

            if len(self.urls) != current_n:
                current_n = len(self.urls)
                current_n_n = 0

            else:
                current_n_n += 1
            if current_n_n > 3:
                break

        thr.p = 's'

    def extract(self):
        thr = Thread()
        thr.start()
        num = 2
        for url in self.urls:
            if thr.p == 's':
                break

            try:
                self.driver.get(url)
                sleep(4)
                try:
                    main_photo = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((
                        By.CSS_SELECTOR,
                        'button[jsaction="pane.heroHeaderImage.click"] > img'))).get_attribute(
                        'src')
                except:
                    main_photo = None

                try:
                    rat = ''.join(set(rat.get_attribute('aria-label') for rat in
                                      self.driver.find_elements_by_css_selector('ol')
                                      if rat.get_attribute('aria-label') is not None)).replace('stars', '').strip()
                except:
                    try:
                        _ = self.driver.find_element_by_css_selector(
                            '[jsaction="pane.reviewChart.moreReviews"] .gm2-display-2').location_once_scrolled_into_view
                        rat = self.driver.find_element_by_css_selector(
                            '[jsaction="pane.reviewChart.moreReviews"] .gm2-display-2').text
                    except:
                        rat = None
                try:
                    comment = list(set(rat.get_attribute('aria-label') for rat in
                                       self.driver.find_elements_by_css_selector('[jsaction="pane.rating.moreReviews"]')
                                       if rat.get_attribute('aria-label') is not None))[0]
                except:
                    comment = None

                name = WebDriverWait(self.driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, '//h1/span[1]'))).text
                try:
                    ty = self.driver.find_element_by_css_selector('[jsaction="pane.rating.category"]').text
                except:
                    ty = None
                try:
                    phone = \
                        self.driver.find_element_by_xpath(
                            "//div/button[contains(@data-item-id, 'phone:')]").get_attribute(
                            'data-item-id').split('tel:')[-1]
                except:
                    phone = None
                try:
                    address = \
                        self.driver.find_element_by_xpath(
                            "//div/button[contains(@data-item-id, 'address')]").get_attribute(
                            'aria-label').split(":")[-1]
                except:
                    address = None
                try:
                    web = \
                        self.driver.find_element_by_xpath(
                            "//div/button[contains(@data-item-id, 'authority')]").get_attribute(
                            'aria-label').split(":")[-1]
                except:
                    web = None
                try:
                    cont = \
                        self.driver.find_element_by_css_selector('[data-item-id="oloc"]').get_attribute(
                            'aria-label').split(
                            ':')[-1]
                except:
                    cont = None

                if save_web != '1':
                    try:
                        if len(self.driver.find_elements_by_xpath(
                                "//div/button[contains(@data-item-id, 'authority')]")):
                            self.driver.execute_script(
                                'document.querySelector("[data-item-id=\'authority\'] [dir=\'ltr\']").click();')
                            sleep(3)
                            if len(self.driver.window_handles) > 1:
                                self.driver.close()
                                self.driver.switch_to.window(self.driver.window_handles[-1])
                                self.sheet[f'I{num}'] = self.driver.current_url
                            else:
                                pass
                    except:
                        pass

                try:
                    if main_photo is not None and save_img == '1':
                        res = session.get(main_photo)
                        image_file = BytesIO(res.content)
                        img = Image(image_file)
                        img.width = 90
                        img.height = 75
                        self.sheet.row_dimensions[num].height = 56
                        self.sheet.add_image(img, f'A{num}')
                except:
                    pass
                self.sheet[f'B{num}'] = name
                self.sheet[f'C{num}'] = ty
                self.sheet[f'D{num}'] = rat
                self.sheet[f'E{num}'] = comment
                self.sheet[f'F{num}'] = address
                self.sheet[f'G{num}'] = cont
                self.sheet[f'H{num}'] = phone
                if save_web == '1':
                    self.sheet[f'I{num}'] = web
                self.sheet[f'J{num}'] = main_photo
                self.sheet[f'K{num}'] = url

                for column in range(2, 11):
                    try:
                        self.sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True)
                        self.sheet.cell(num + 1, column).alignment = Alignment(horizontal='center',
                                                                               vertical='center', wrap_text=True)
                    except:
                        pass
                print_percent_done(num - 1, len(self.urls))
                num += 1
            except Exception as e:
                pass
            self.book.save('final_all.xlsx')
        print('- Done....', end='\r')
        try:
            self.book.save(f'{sear}.xlsx')
        except:
            self.book.save('Result.xlsx')
        self.driver.quit()


if __name__ == '__main__':
    self = Main()
    while True:
        sear = input('- Search :')
        if sear != '':
            break
    save_img = input('- 1 >> Download image : ')
    save_web = input('- 1 >> Extract complete web : ')

    search = sear.replace(' ', '+')
    self.open(search)
    while True:
        init = input('- 1 >> Repeat \n'
                     '- 2 >> Add \n'
                     '- 3 >> Start Extract \n'
                     ': ')

        if init == '1':
            self.open(search)

        elif init == '2':
            while True:
                sear = input('- Search :')
                if sear != '':
                    break
            search = sear.replace(' ', '+')
            self.open(search)

        elif init == '3':
            self.extract()
            break
