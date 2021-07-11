import os
import traceback
import uuid
from time import sleep

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


def print_percent_done(index, total, bar_len=50):
    percent_done = (index-1)/total*100
    percent_done = round(percent_done, 1)

    done = round(percent_done/(100/bar_len))
    togo = bar_len-done

    done_str = '█'*int(done)
    togo_str = '░'*int(togo)

    print(f'- ⏳ {total}\\{index - 1} - [{done_str}{togo_str}] ({percent_done} %)', end='\r')


class Main:
    def __init__(self):
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.cell(1, 1).value = 'Title'
        self.sheet.cell(1, 2).value = 'Price'
        self.sheet.cell(1, 3).value = 'Quantity'
        self.sheet.cell(1, 4).value = 'Available'
        self.sheet.cell(1, 5).value = 'URL'
        self.sheet.cell(1, 6).value = 'Code'
        self.sheet.cell(1, 7).value = 'SKU'
        self.sheet.cell(1, 8).value = 'BarCode'
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--no-sandbox')  # Solve the error that the DevToolsActivePort file does not exist
        chrome_options.add_argument('--disable-gpu')  # Google documentation mentions that this attribute needs to be added to avoid bugs
        chrome_options.add_argument('blink-settings=imagesEnabled=false')  # Don't load pictures, speed up
        chrome_options.add_argument('--disable-plugins')
        chrome_options.add_argument('--disable-popup-blocking')

        self.driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
        self.driver.get('https://sa.iherb.com/')

    def next_page(self):
        try:
            url = self.driver.find_element_by_css_selector('.pagination-next').get_attribute('href')
            self.get(url)

            return True
        except:
            return False

    def get(self, url):
        try:
            self.driver.get(url)
            sleep(4)
        except:
            print(traceback.print_exc())

    def language(self):
        if 'EN' not in self.driver.find_element_by_css_selector('div.language-select').text:
            self.driver.find_element_by_css_selector('div.language-select').click()
            sleep(2)
            # self.driver.find_element_by_css_selector('div.language-menu.language-menu-universal div.select-language.ui.fluid.search.selection.dropdown').click()
            self.driver.find_element_by_css_selector('.select-language').click()
            sleep(1)
            # self.driver.find_element_by_css_selector('div.select-language.ui.fluid.search.selection.dropdown > div.menu.search-list.open > div:nth-child(2)').click()
            self.driver.find_element_by_css_selector('[data-val="en-US"]').click()
            sleep(4)
            self.driver.find_element_by_css_selector('button.save-selection').click()
            sleep(2)
            self.driver.find_element_by_css_selector(
                'div.panel-content.border.show-page.text-right > div > div > select').send_keys('48')
            sleep(2)

    def save(self, ele, cont):
        try:
            if 'InStock' in ele.find_element_by_css_selector('meta+link').get_attribute('href'):
                available = True
            else:
                available = False

            title = ele.find_element_by_css_selector('div.absolute-link-wrapper > a').get_attribute('title')
            href = ele.find_element_by_css_selector('div.absolute-link-wrapper > a').get_attribute('href')
            code = ele.find_element_by_css_selector('div.absolute-link-wrapper > a').get_attribute('data-part-number')
            price = ele.find_element_by_css_selector('meta[itemprop="price"]').get_attribute('content')
            price = float(price) + float(price) * (float(pres) / 100)

            self.sheet.cell(cont, 1).value = title
            self.sheet.cell(cont, 2).value = price
            self.sheet.cell(cont, 3).value = quantity
            self.sheet.cell(cont, 4).value = available
            self.sheet.cell(cont, 5).value = href
            self.sheet.cell(cont, 6).value = code
        except:
            print(traceback.print_exc())

    def extract(self, t):
        cont = 2
        next_pag = True
        self.language()
        while next_pag:

            for ele in self.driver.find_elements_by_css_selector('div.product-inner.product-inner-wide'):
                try:
                    if t == 1:
                        pass
                    else:
                        if 'InStock' not in ele.find_element_by_css_selector('meta+link').get_attribute('href'):
                            continue
                    self.save(ele, cont)
                    print(f'- {cont - 1}', end='\r')
                except:
                    print(traceback.print_exc())
                cont += 1

            sleep(2)

            self.book.save('sa-iherb.xlsx')
            next_pag = self.next_page()
            sleep(1)

    def get_code(self):
        for i in range(2, self.sheet.max_row+1):
            try:
                url = self.sheet.cell(i, 5).value
                if url:
                    self.driver.get(url)
                    code = self.driver.find_element_by_css_selector('[itemprop="gtin12"]').text
                    self.sheet.cell(i, 8).value = code
                    print_percent_done(i, self.sheet.max_row)
            except:
                pass
        self.book.save('sa-iherb.xlsx')
        self.driver.quit()


class Upload:
    driver: WebDriver
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-sa/noon-catalog'

    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'a[href*="/en-sa/noon-catalog/preview/"]')

    SKU_TEXT_SELECTOR = (By.CSS_SELECTOR, 'div.tabsWrapper > div')
    SKU_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[name="partner_sku"]')
    SKU_BUTTON_SELECTOR = (By.CSS_SELECTOR, '.solid')

    PRICE_MAX_SELECTOR = (
        By.XPATH,
        "//div[contains(@class, 'inputInfo') and contains(text(), 'Noon SKU Price')] / *[@class='priceRange']")

    PRICE_INPUT_SELECTOR = (By.NAME, 'price')

    QUANTITY_SITE_SELECTOR = (By.CSS_SELECTOR, 'div.placeholder')
    QUANTITY_SITE_CHOSE_SELECTOR = (By.CSS_SELECTOR, 'div.active > div > div.countryLabel')
    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')

    BARCODE_INSIDE_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="text"][placeholder="Enter Barcode"]')
    BUTTON_ADD_BARCODE_INSIDE_SELECTOR = (By.CSS_SELECTOR, 'div.eachInput > div > div.solid')

    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="offerActive"]\').checked == false){document.querySelector(\'input[id="offerActive"]\').click()};'

    # SAVE_CHANGE_SELECTOR = (By.CSS_SELECTOR, 'div.tabsWrapper ~ div  div.solid')
    SAVE_CHANGE_SELECTOR = (By.XPATH, '//*[contains(text(),"Save Changes")]')
    SUBMIT_SELECTOR = (By.XPATH, '//div[class="btnWrapper"]/div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')

    def __init__(self):
        self.book = load_workbook(file_name)
        self.sheet = self.book.active
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.login()

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def add_new(self, name):
        try:
            self.driver.get(self.URL_CAT)
            sleep(2)
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))
                                                 ).send_keys(name, Keys.ENTER)
            sleep(2)
        except:
            if 'login.noon.partners' in self.driver.current_url:
                self.login()

            sleep(30)
            try:
                self.driver.get(self.URL_CAT)
                sleep(2)
                WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))
                                                     ).send_keys(name, Keys.ENTER)
                sleep(2)
            except:
                sleep(10)
                print(traceback.print_exc())
                return False

    def add_serial(self, serial):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
            sleep(4)
        except:
            print('- Not found')
            return 'False'
        sleep(1)
        try:

            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.SKU_INPUT_SELECTOR)).send_keys(serial)
            sleep(5)
            while True:
                if not len(self.driver.find_elements_by_css_selector('.kHWBRM')):
                    break
            if len(self.driver.find_elements_by_css_selector('.inputError ')):
                print('SKU FOUND')
                return 'SKU FOUND'
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_BUTTON_SELECTOR)).click()
            return serial
        except:
            print('- Exist, Continue')
            return False

    def add_price(self, price):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a', Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(str(price))
        except:
            print(traceback.print_exc())
            print('- Cant Change Price')

    def add_qu(self, q):
        sleep(1)
        try:
            WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
        except:
            try:
                WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(self.QUANTITY_SITE_SELECTOR))).click()
                sleep(1)
                WebDriverWait(self.driver, 6).until((ec.visibility_of_all_elements_located(self.QUANTITY_SITE_CHOSE_SELECTOR)))[site].click()
                sleep(1)
                WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
                sleep(1)
            except:
                print(traceback.print_exc())
                print('- Cant Change quantity')
                return False
        try:
            self.driver.find_element_by_css_selector('.stockInputValues > .processingTimeInpurCtr').click()
            sleep(1)
            self.driver.find_elements_by_css_selector('.stockInputValues > .processingTimeInpurCtr .countryLabel')[site_time].click()
        except:
            print(traceback.print_exc())
            print('- Cant add time')

    def add_bar(self, sk):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.BARCODE_INSIDE_INPUT_SELECTOR)
                                                 ).send_keys(Keys.CONTROL, 'a', Keys.BACKSPACE)
            WebDriverWait(self.driver, 10).until(
                ec.visibility_of_element_located(self.BARCODE_INSIDE_INPUT_SELECTOR)).send_keys(sk)
            sleep(1)
            WebDriverWait(self.driver, 10).until(
                ec.visibility_of_element_located(self.BUTTON_ADD_BARCODE_INSIDE_SELECTOR)).click()
            sleep(1)
        except:
            print(traceback.print_exc())
            print('- Cant Change Barcode')

    def add_active(self):
        sleep(1)
        try:
            self.driver.execute_script(self.ACTIVE_SCRIPT)
        except:
            print(traceback.print_exc())

    def upload(self):
        barcodes = list()
        for i in range(2, self.sheet.max_row + 1):
            names_way = False
            try:
                name = self.sheet.cell(i, 1).value
                price = self.sheet.cell(i, 2).value
                _quantity = self.sheet.cell(i, 3).value
                available = self.sheet.cell(i, 4).value
                code = self.sheet.cell(i, 6).value
                bar_code = self.sheet.cell(i, 8).value
                print('- ', i, name, price, code, bar_code, available)

                # first check by code
                if self.add_new(code) is False:
                    self.sheet.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor='ff0000')
                    continue
                else:
                    res = self.add_serial(code)
                    # not found code
                    if res == 'False':
                        names_way = True
                        if self.add_new(name) is False:
                            self.sheet.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor='ff0000')
                            continue

                        res = self.add_serial(code)
                        if res == 'False':
                            print('- Try remove ')
                            # name = name[:name.rfind(',')]
                            try:
                                name = ','.join(name.split(',')[:-end_word])
                                print('- ', name, price)
                                add_res = self.add_new(name)
                                if add_res is False:
                                    self.sheet.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor='ff0000')
                                    continue
                                res = self.add_serial(code)
                            except:
                                print(traceback.print_exc())
                        if res == 'False':
                            self.sheet.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor='ff0000')
                            continue

                    if res == 'SKU FOUND':
                        self.sheet.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor='0000FF')
                        continue

                    if res is False:
                        try:
                            sku = self.driver.find_element_by_css_selector('.labelSku').text
                            if sku not in barcodes:
                                barcodes.append(sku)
                                self.sheet.cell(i, 7).value = sku
                                print(sku)
                        except:
                            pass
                        try:
                            if price and self.driver.find_element_by_name('price').get_attribute('value') == '':
                                self.add_price(price)
                            try:
                                WebDriverWait(self.driver, 10).until(
                                    (ec.element_to_be_clickable(self.SAVE_CHANGE_SELECTOR))).click()
                                sleep(1)
                                WebDriverWait(self.driver, 10).until(
                                    (ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                                sleep(1)
                            except:
                                print(traceback.print_exc())
                                print('Cant Save')
                        except:
                            pass
                        self.sheet.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor='00FF00')
                        continue

                sleep(1)
                if price:
                    self.add_price(price)

                sleep(2)
                if available not in ['FALSE', False, 'False', 'false']:
                    self.add_active()

                if _quantity is not None and names_way is False:
                    self.add_qu(_quantity)

                try:
                    title = self.driver.find_element_by_css_selector('input[placeholder="Title"]').get_attribute('value')
                    self.sheet.cell(i, 9).value = title
                    self.sheet.cell(i, 10).value = name
                except:
                    pass

                sleep(1)
                self.add_bar(bar_code)
                sleep(2)
                sku = self.driver.find_element_by_css_selector('.labelSku').text
                if sku not in barcodes:
                    barcodes.append(sku)
                    self.sheet.cell(i, 7).value = sku
                    print(sku)

                try:
                    WebDriverWait(self.driver, 10).until((ec.element_to_be_clickable(self.SAVE_CHANGE_SELECTOR))).click()
                    sleep(1)
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                    sleep(1)
                except:
                    print(traceback.print_exc())
                    print('Cant Save')
                sleep(2)

            except:
                print(traceback.print_exc())
        self.book.save(file_name)


def add_mac():
    try:
        os.remove(str(uuid.getnode()))
    except:
        pass
    try:
        with open(str(uuid.getnode()), 'w+') as file:
            file.write(str(uuid.getnode()))
        os.popen('attrib +h ' + str(uuid.getnode()))
        return True
    except:
        print('delete old file')
        return False


def check_mac():
    try:
        with open(str(uuid.getnode()), 'r') as file:
            text = file.readlines()
        if str(uuid.getnode()) in text:
            return True
        else:
            print('ref')
            return False
    except FileNotFoundError:
        print('ref')
        return False


if __name__ == "__main__":
    pr = check_mac()
    while True:
        input_res = str(input('''- Enter 1 ---> Extract   or 2 ----> Upload  or 3 ---> All '''))
        if input_res == '1' and pr is True:
            rs = str(input('''- Enter 1 ---> Extract All  or 2 ----> Extract Only'''))
            if rs == '1':
                ty = 1
            else:
                ty = 2
            quantity = float(input('- Enter Quint : '))
            pres = float(input('- Enter prc : '))
            ex = Main()
            input('- Enter : ')
            ex.extract(ty)
            ex.get_code()

        elif input_res == '2' and pr is True:

            while True:
                try:
                    file_name = input('- Enter file name : ')  # 'Book1.xlsx'
                    book = load_workbook(file_name)
                    book.close()
                    break
                except FileNotFoundError:
                    print('- No such file or directory')
                except Exception as e:
                    print(e)

            email = input('- Enter your email : ')
            password = input('- Enter your password : ')
            site = input('- Enter Number countryLabel : ')
            site_time = input('- Enter Time : ')
            try:
                site_time = int(site_time) - 1
                if site_time >= 0:
                    pass
                else:
                    site_time = 0
            except:
                site_time = 0
            try:
                site = int(site) - 1
                if site >= 0:
                    pass
                else:
                    site = 0
            except:
                site = 0

            end_word = input('- Enter Number Label : ')
            try:
                end_word = int(end_word)
                if end_word > 0:
                    pass
                else:
                    end_word = 1
            except:
                end_word = 1

            print(site)
            u = Upload()
            u.upload()
        elif input_res == '3' and pr is True:
            rs = str(input('''- Enter 1 ---> Extract All  or 2 ----> Extract Only'''))
            if rs == '1':
                ty = 1
            else:
                ty = 2

            quantity = float(input('- Enter Quint : '))
            pres = float(input('- Enter prc : '))
            email = input('- Enter your email : ')
            password = input('- Enter your password : ')
            site = input('- Enter Number countryLabel : ')
            site_time = input('- Enter Time : ')
            try:
                site_time = int(site_time) - 1
                if site_time >= 0:
                    pass
                else:
                    site_time = 0
            except:
                site_time = 0
            try:
                site = int(site) - 1
                if site >= 0:
                    pass
                else:
                    site = 0
            except:
                site = 0

            end_word = input('- Enter Number Label : ')
            try:
                end_word = int(end_word)
                if end_word > 0:
                    pass
                else:
                    end_word = 1
            except:
                end_word = 1

            ex = Main()
            input('- Enter : ')
            ex.extract(ty)
            file_name = 'sa-iherb.xlsx'
            u = Upload()
            u.upload()
        elif input_res == 'scrt*python*add*mac*address':
            pr = add_mac()
