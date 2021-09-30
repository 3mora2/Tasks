import os
import sys
import pandas as pd
from PySide2 import QtCore
from PySide2.QtCore import Qt, QMetaObject, QCoreApplication
from PySide2.QtGui import QIcon, QCursor
from PySide2.QtWidgets import QMainWindow, QApplication, QHeaderView, QFileDialog, QTableWidgetItem, QPushButton, \
    QDialog, QGridLayout, QFrame, QLineEdit, QComboBox, QDialogButtonBox, QSizePolicy, QMessageBox
from openpyxl import load_workbook, Workbook
from mainwindow import Ui_MainWindow
from dialog import Ui_Dialog


class FileNameObject:
    def __init__(self, path):
        self.name = path.split('/')[-1]
        self.full_path = path

    def __repr__(self):
        return str(self.name)

    def __str__(self):
        return str(self.name)


class MainWindow(QMainWindow, Ui_MainWindow, Ui_Dialog):
    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)
        self.setupUi(self)
        self.Handel_Button()
        self.Handel_Ui()

    def Handel_Ui(self):
        self.setWindowIcon(QIcon('logo.png'))
        self.setWindowTitle(" ")

        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget_2.verticalHeader().setVisible(False)
        self.tableWidget_3.verticalHeader().setVisible(False)

        # self.tableWidget.horizontalHeader().setStretchLastSection(True)
        # self.tableWidget_2.horizontalHeader().setStretchLastSection(True)
        self.tableWidget_3.horizontalHeader().setStretchLastSection(True)

        # self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.tableWidget_2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_3.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.tableWidget.setColumnWidth(0, 250)
        self.tableWidget.setColumnWidth(1, 100)
        self.tableWidget.setColumnWidth(2, 70)
        self.tableWidget.setColumnWidth(3, 5)

        self.tableWidget_2.setColumnWidth(0, 250)
        self.tableWidget_2.setColumnWidth(1, 100)
        self.tableWidget_2.setColumnWidth(2, 70)
        self.tableWidget_2.setColumnWidth(3, 5)

    def Handel_Button(self):
        self.pushButton_2.clicked.connect(self.clear)
        self.pushButton_4.clicked.connect(self.clear)
        self.pushButton.clicked.connect(self.Dialog)
        self.pushButton_3.clicked.connect(self.Dialog)
        self.pushButton_5.clicked.connect(self.start)
        self.pushButton_6.clicked.connect(self.save)
        self.pushButton_7.clicked.connect(self.add_folder)

    def Dialog(self):
        Dialog = QDialog()
        if not Dialog.objectName():
            Dialog.setObjectName(u"Dialog")
        # Dialog.resize(654, 134)
        Dialog.resize(654, 134)
        self.gridLayout_2 = QGridLayout(Dialog)
        self.gridLayout_2.setObjectName(u"gridLayout_2")
        self.frame = QFrame(Dialog)
        self.frame.setObjectName(u"frame")
        self.frame.setFrameShape(QFrame.StyledPanel)
        self.frame.setFrameShadow(QFrame.Raised)
        self.gridLayout = QGridLayout(self.frame)
        self.gridLayout.setObjectName(u"gridLayout")
        self.lineEdit = QLineEdit(self.frame)
        self.lineEdit.setObjectName(u"lineEdit")

        self.gridLayout.addWidget(self.lineEdit, 1, 1, 1, 1)

        self.comboBox = QComboBox(self.frame)
        self.comboBox.setObjectName(u"comboBox")
        sizePolicy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.comboBox.sizePolicy().hasHeightForWidth())
        self.comboBox.setSizePolicy(sizePolicy)

        self.gridLayout.addWidget(self.comboBox, 1, 0, 1, 1)

        self.pushButton = QPushButton(self.frame)
        self.pushButton.setObjectName(u"pushButton")

        self.gridLayout.addWidget(self.pushButton, 1, 2, 1, 1)

        self.gridLayout.setColumnStretch(0, 1)
        self.gridLayout.setColumnStretch(1, 2)

        self.gridLayout_2.addWidget(self.frame, 0, 0, 1, 1)

        self.buttonBox = QDialogButtonBox(Dialog)
        self.buttonBox.setObjectName(u"buttonBox")
        self.buttonBox.setOrientation(Qt.Horizontal)
        self.buttonBox.setStandardButtons(QDialogButtonBox.Cancel | QDialogButtonBox.Ok)

        self.gridLayout_2.addWidget(self.buttonBox, 1, 0, 1, 1)

        self.retranslateUi(Dialog)
        self.buttonBox.accepted.connect(Dialog.accept)
        self.buttonBox.rejected.connect(Dialog.reject)

        QMetaObject.connectSlotsByName(Dialog)

        self.pushButton.clicked.connect(self.add_file)
        ret = Dialog.exec_()
        if ret == QDialogButtonBox.Ok or ret == 1:
            path = self.lineEdit.text()
            cell = self.comboBox.currentText()
            if path and cell:
                # cell = int(cell.split('- (')[-1].replace(')', '').strip())
                self.add(path, cell, self.max_row)

    def add_file(self):
        open_file = QFileDialog.getOpenFileName(self, "اختار", '', '*.xlsx')[0]
        if open_file != '':
            sheet = load_workbook(open_file).active
            self.max_row = sheet.max_row
            for i in range(self.comboBox.count()):
                self.comboBox.removeItem(i)
            _ = [self.comboBox.addItem(str(f'{sheet.cell(1, i).value} - ({i})')) for i in range(1, sheet.max_column+1) if sheet.cell(1, i).value]
            self.lineEdit.setText(open_file)

    def add_folder(self):
        folder_path = QFileDialog().getExistingDirectory(self, 'اختار المسار ')
        for file in os.listdir(folder_path):
            if file.endswith('.xlsx'):
                open_file = folder_path+'/'+file
                sheet = load_workbook(open_file).active
                max_row = sheet.max_row
                try:
                    cell = [str(f'{sheet.cell(1, i).value} - ({i})') for i in range(1, sheet.max_column+1) if sheet.cell(1, i).value][0]
                except:
                    cell = '- (1)'
                self.add(open_file, cell, max_row)

        Box = QMessageBox()
        Box.setIcon(QMessageBox.Information)
        Box.setWindowTitle("تم")
        Box.setText("تم اضافة الملفات")
        Box.setStandardButtons(QMessageBox.Ok)
        Box.exec()

    def clear(self):
        sender = self.sender().objectName()
        if sender == 'pushButton_2':
            self.tableWidget.clearContents()
            self.tableWidget.setRowCount(0)
        elif sender == 'pushButton_4':
            self.tableWidget_2.clearContents()
            self.tableWidget_2.setRowCount(0)

    def add(self, openfile, cell, max_row):
        sender = self.sender().objectName()
        if sender == 'pushButton' or sender == 'pushButton_7':
            tableWidget = self.tableWidget
            delete_clicked = self.delete_clicked
        elif sender == 'pushButton_3':
            tableWidget = self.tableWidget_2
            delete_clicked = self.delete_clicked_2
        else:
            return

        r = tableWidget.rowCount()
        tableWidget.insertRow(r)
        tableWidget.setItem(r, 0, QTableWidgetItem(str(openfile)))
        tableWidget.setItem(r, 1, QTableWidgetItem(str(cell)))
        tableWidget.setItem(r, 2, QTableWidgetItem(str(max_row)))
        deleteButton = QPushButton("حذف")
        deleteButton.setStyleSheet('border-radius: 20px;')
        deleteButton.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        deleteButton.setMaximumWidth(50)
        deleteButton.setStyleSheet(u"""QPushButton{background-color: rgb(244, 243, 242, 0);}""")

        deleteButton.clicked.connect(delete_clicked)
        tableWidget.setCellWidget(r, 3, deleteButton)

    def delete_clicked(self):
        button = self.sender()
        if button:
            row = self.tableWidget.indexAt(button.pos()).row()
            self.tableWidget.removeRow(row)

    def delete_clicked_2(self):
        button = self.sender()
        if button:
            row = self.tableWidget_2.indexAt(button.pos()).row()
            self.tableWidget_2.removeRow(row)

    def start(self):
        all_data = []
        remove_data = []
        for i in range(self.tableWidget.rowCount()):
            path = self.tableWidget.item(i, 0).text()
            cell = int(self.tableWidget.item(i, 1).text().split('- (')[-1].replace(')', '').strip())
            sheet_remove = load_workbook(path).active
            remove_data += [sheet_remove.cell(i, cell).value for i in range(2, sheet_remove.max_row+1) if sheet_remove.cell(i, cell).value]

        for i in range(self.tableWidget_2.rowCount()):
            path = self.tableWidget_2.item(i, 0).text()
            cell = int(self.tableWidget_2.item(i, 1).text().split('- (')[-1].replace(')', '').strip())
            sheet = load_workbook(path).active
            all_data += [sheet.cell(i, cell).value for i in range(2, sheet.max_row+1) if sheet.cell(i, cell).value]

        r = 0
        new_number = []
        for number in all_data:
            if number not in remove_data and number not in new_number:
                new_number.append(number)
                r = self.tableWidget_3.rowCount()
                self.tableWidget_3.insertRow(r)
                self.tableWidget_3.setItem(r, 0, QTableWidgetItem(str(number)))
        self.label_4.setText(str(r))

        Box = QMessageBox()
        Box.setIcon(QMessageBox.Information)
        Box.setWindowTitle("تم")
        Box.setText("تم الانتهاء")
        Box.setStandardButtons(QMessageBox.Ok)
        Box.exec()

    def save(self):
        save_file = QFileDialog.getSaveFileName(self, "حفظ ك", 'New.xlsx', '*.xlsx')[0]
        if save_file != '':
            book = Workbook()
            sheet = book.active
            sheet.cell(1, 1).value = 'رقم الهاتف'
            for i in range(self.tableWidget_3.rowCount()):
                sheet.cell(i+2, 1).value = self.tableWidget_3.item(i, 0).text().strip()
            book.save(save_file)
            Box = QMessageBox()
            Box.setIcon(QMessageBox.Information)
            Box.setWindowTitle("تم")
            Box.setText("تم ")
            Box.setStandardButtons(QMessageBox.Ok)
            Box.exec()


def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec_()

