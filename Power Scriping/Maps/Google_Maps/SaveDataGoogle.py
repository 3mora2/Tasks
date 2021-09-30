from io import BytesIO

import requests
from PySide2.QtCore import QThread, Signal
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

session = requests.session()


class SaveAsGoogle(QThread):
    sheet: Worksheet
    book: Workbook
    data: dict
    final = Signal()

    def run(self):
        try:
            self.book = Workbook()
            self.sheet = self.book.active
            if self.download_img:
                self.sheet['A1'] = 'Image'
                self.sheet.column_dimensions['A'].width = 13

            self.sheet['B1'] = 'Name'
            self.sheet['C1'] = 'Type'
            self.sheet['D1'] = 'Rate'
            self.sheet['E1'] = 'Comments'
            self.sheet['F1'] = 'Address'
            self.sheet['G1'] = 'Plus Code'
            self.sheet['H1'] = 'Phone'
            self.sheet['I1'] = 'Website'
            self.sheet['J1'] = "IMG"
            self.sheet['K1'] = "Opened"
            self.sheet['L1'] = "Coordinates"
            self.sheet['M1'] = "URL"

            self.sheet.column_dimensions['B'].width = 50
            self.sheet.column_dimensions['C'].width = 30
            self.sheet.column_dimensions['D'].width = 10
            self.sheet.column_dimensions['E'].width = 10
            self.sheet.column_dimensions['F'].width = 100
            self.sheet.column_dimensions['G'].width = 70
            self.sheet.column_dimensions['H'].width = 25
            self.sheet.column_dimensions['I'].width = 30
            self.sheet.column_dimensions['J'].width = 50
            self.sheet.column_dimensions['K'].width = 50
            self.sheet.column_dimensions['L'].width = 25
            self.sheet.column_dimensions['M'].width = 120
            for column in range(1, 14):
                try:
                    self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                     wrap_text=True)
                except:
                    pass

            for num in self.data.keys():
                main_photo = self.data[num]['main_photo']
                try:
                    if main_photo and self.download_img:
                        res = session.get(main_photo)
                        image_file = BytesIO(res.content)
                        img = Image(image_file)
                        img.width = 90
                        img.height = 75
                        self.sheet.row_dimensions[num].height = 56
                        self.sheet.add_image(img, f'A{num}')
                except Exception as e:
                    print(e)
                    pass
                self.sheet[f'B{num}'] = self.data[num]['name']
                self.sheet[f'C{num}'] = self.data[num]['ty']
                self.sheet[f'D{num}'] = self.data[num]['rat']
                self.sheet[f'E{num}'] = self.data[num]['comment']
                self.sheet[f'F{num}'] = self.data[num]['address']
                self.sheet[f'G{num}'] = self.data[num]['cont']
                self.sheet[f'H{num}'] = self.data[num]['phone']
                self.sheet[f'J{num}'] = self.data[num]['main_photo']
                self.sheet[f'K{num}'] = self.data[num]['opened']
                self.sheet[f'I{num}'] = self.data[num]['web']
                self.sheet[f'L{num}'] = self.data[num]['address_num']
                self.sheet[f'M{num}'] = self.data[num]['url']

                for column in range(2, 13):
                    try:
                        self.sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True)
                    except:
                        pass

            self.book.save(self.save_file)
            self.final.emit()
        except Exception as e:
            print(e)
