from io import BytesIO

import requests
from PySide2.QtCore import QThread, Signal
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

session = requests.session()


class SaveAs(QThread):
    sheet: Worksheet
    book: Workbook
    data: dict
    final = Signal()

    def run(self):
        try:
            self.book = Workbook()
            self.sheet = self.book.active
            if self.download_img:
                self.sheet['A1'] = 'Image'
                self.sheet.column_dimensions['A'].width = 13

            self.sheet.cell(1, 2).value = 'Company Name'
            self.sheet.cell(1, 3).value = 'Phone Number'
            self.sheet.cell(1, 4).value = 'Whatsapp'
            self.sheet.cell(1, 5).value = 'Show Review'
            self.sheet.cell(1, 6).value = 'Business Description'
            self.sheet.cell(1, 7).value = 'Category'
            self.sheet.cell(1, 8).value = 'Show Directions'
            self.sheet.cell(1, 9).value = 'Keywords'
            self.sheet.cell(1, 10).value = 'Company Address'
            self.sheet.cell(1, 11).value = 'Website'
            self.sheet.cell(1, 12).value = 'Branches URL'
            self.sheet.cell(1, 13).value = 'About Us'
            self.sheet.cell(1, 14).value = 'Company URL'

            self.sheet.column_dimensions['B'].width = 35
            self.sheet.column_dimensions['C'].width = 16
            self.sheet.column_dimensions['D'].width = 57
            self.sheet.column_dimensions['E'].width = 15
            self.sheet.column_dimensions['F'].width = 22
            self.sheet.column_dimensions['G'].width = 25
            self.sheet.column_dimensions['H'].width = 65
            self.sheet.column_dimensions['I'].width = 26
            self.sheet.column_dimensions['J'].width = 70
            self.sheet.column_dimensions['K'].width = 70
            self.sheet.column_dimensions['L'].width = 50
            self.sheet.column_dimensions['M'].width = 120
            self.sheet.column_dimensions['N'].width = 120
            for column in range(1, 14):
                try:
                    self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                     wrap_text=True)
                except:
                    pass

            for num in self.data.keys():
                img = self.data[num]['img']
                try:
                    if img and self.download_img:
                        res = requests.get(img)
                        image_file = BytesIO(res.content)
                        img = Image(image_file)
                        img.width = 90
                        img.height = 75
                        self.sheet.row_dimensions[num].height = 56
                        self.sheet.add_image(img, f'A{num}')
                except:
                    pass
                self.sheet.cell(num, 2).value = self.data[num]['companyName']
                self.sheet.cell(num, 3).value = self.data[num]['mob']
                self.sheet.cell(num, 4).value = self.data[num]['whatsapp']
                self.sheet.cell(num, 5).value = self.data[num]['show_review']
                self.sheet.cell(num, 6).value = self.data[num]['business_description']
                self.sheet.cell(num, 7).value = self.data[num]['category']
                self.sheet.cell(num, 8).value = self.data[num]['show_directions']
                self.sheet.cell(num, 9).value = self.data[num]['keywords']
                self.sheet.cell(num, 10).value = self.data[num]['company_address']
                self.sheet.cell(num, 11).value = self.data[num]['website']
                self.sheet.cell(num, 12).value = self.data[num]['show_branches']
                self.sheet.cell(num, 13).value = self.data[num]['aboutUs']
                self.sheet.cell(num, 14).value = self.data[num]['companyURL']
                for column in range(2, 15):
                    try:
                        self.sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True)
                    except:
                        pass

            self.book.save(self.save_file)
            self.final.emit()
        except Exception as e:
            print(e)
