from typing import List
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import re
from webdriver_manager.chrome import ChromeDriverManager
import os

os.environ['WDM_LOG_LEVEL'] = '0'


class Main:
    barcode_exit: List[str]
    links: set

    def __init__(self):
        pass

    def get_url(self):
        while True:
            try:
                file_name = input('- Enter file name : ')  # 'data.xlsx'
                book = load_workbook(file_name)
                sheet = book.active
                self.links = {sheet.cell(i, 2).value for i in range(2, sheet.max_row + 1) if
                              sheet.cell(i, 2).value is not None}
                self.Get_Barcode_From_Excel(sheet)
                break
            except FileNotFoundError:
                print('- No such file or directory')
            except Exception as e:
                print(e)

    def Get_Barcode_From_Excel(self, sheet):
        barcode_exit = []
        end = sheet.max_row
        for row in range(2, end + 1):
            if sheet.cell(row, 1).value is not None:
                barcode_exit.append(str(sheet.cell(row, 1).value))

        self.barcode_exit = list(set(barcode_exit))

    def Get_url(self):
        self.get_url()
        book = Workbook()
        sheet = book.active
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.maximize_window()
        Links = set()
        n = 2
        for link in self.links:
            while True:
                if 'section' not in link:
                    if '?' in link:
                        rr = link.split('?')
                        if '&' in rr[1]:
                            link = rr[0] + '?' + '&section=2&'.join(rr[1].split('&'))
                        else:
                            link = rr[0] + '?section=2&' + rr[1]
                    else:
                        link = link + '?section=2'

                # link = link.split('&')[0] + '&section=2&' + link.split('&')[1]
                # print(link)
                driver.get(link)
                sleep(1)
                if len(driver.find_elements_by_css_selector('ul > li > h6.itemTitle > a.itemLink')) == 0:
                    sleep(8)
                    driver.refresh()
                    sleep(3)
                    if len(driver.find_elements_by_css_selector('ul > li > h6.itemTitle > a.itemLink')) == 0:
                        break

                for e in driver.find_elements_by_css_selector('div.single-item'):
                    bar = e.get_attribute('data-ean')
                    if bar in self.barcode_exit:
                        print(f'- {n} - {bar} - Found')
                        continue

                    Links.add(e.find_element_by_css_selector('.sPrimaryLink').get_attribute('href'))
                    sheet.cell(n, 1).value = e.find_element_by_css_selector('.sPrimaryLink').get_attribute('href')
                    book.save('first.xlsx')
                    print(f'- {n} - {bar}')
                    n += 1
                # print(len(Links))
                try:
                    link = driver.find_element_by_css_selector('.pagination-next a').get_attribute('href')
                except:
                    break

            # r = s.get(link)
            # for element in r.html.find('ul > li > h6.itemTitle > a.itemLink'):
            #     Links.add(element.attrs['href'])
            # print(len(Links))
            # link = r.html.find('.pagination-next a')[0].attrs['href']
            # sleep(4)
        driver.quit()


def Show():
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.maximize_window()

    n = 1
    book = Workbook()
    sheet = book.active
    sheet[f'A{n}'] = 'رقم السلعه'
    sheet[f'D{n}'] = 'اسم السلعه'
    sheet[f'H{n}'] = 'السعر الحالى'
    sheet[f'L{n}'] = 'اسم البائع الحالى'
    sheet[f'M{n}'] = 'تقييم البائع الحالى'
    sheet[f'N{n}'] = 'fbs'
    sheet[f'O{n}'] = 'الحد الادنى للسعر'
    sheet[f'P{n}'] = 'الحد الاقصى للسعر'

    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['D'].width = 130
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['L'].width = 15
    sheet.column_dimensions['M'].width = 15
    sheet.column_dimensions['N'].width = 14
    sheet.column_dimensions['O'].width = 13
    sheet.column_dimensions['P'].width = 13

    try:
        sheet.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, 12).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, 13).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, 14).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, 15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(1, 16).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    except:
        pass

    book = load_workbook('first.xlsx')
    sheet = book.active
    n = 2
    for u in range(2, sheet.max_row + 1):
        url = sheet.cell(u, 1).value
        if url is not None:
            try:
                driver.get(url.replace('/eg-en/', '/eg-ar/'))
                sleep(3)
                ans = input('- Enter (y) if you want add Or (Enter) : ')
                if ans.lower() == 'y':
                    while True:
                        try:
                            Max = float(input('- Enter Maximum Price : '))
                            break
                        except ValueError:
                            print('Please Enter Number')

                    while True:
                        try:
                            Min = float(input('- Enter Minimum Price : '))
                            break
                        except ValueError:
                            print('Please Enter Number')

                    barcode = driver.find_element_by_css_selector('#productTrackingParams').get_attribute(
                        'data-ean')
                    name = driver.find_element_by_css_selector('#productTrackingParams').get_attribute('data-title')
                    try:
                        price = driver.find_element_by_css_selector('#productTrackingParams').get_attribute(
                            'data-price')
                    except:
                        price = None
                    # name = driver.find_element_by_css_selector('.product-title h1').text
                    # price = driver.find_element_by_css_selector('.price-container .price').text.strip()
                    # price = driver.find_element_by_css_selector('.price-container .price').text.strip().split()[0]
                    # rat = driver.find_element_by_css_selector('.seller-rating').text
                    # seller = driver.find_element_by_css_selector('.unit-seller-link a').text
                    try:
                        rat = re.findall(r'\d+', driver.find_element_by_css_selector('.seller-rating').text)[0]
                    except:
                        rat = None
                    try:
                        seller = driver.find_element_by_css_selector('#productTrackingParams').get_attribute(
                            'data-seller-name')
                    except:
                        seller = None
                    try:
                        _ = driver.find_element_by_class_name('header-product-fulfilled').get_attribute('src')
                        fbs = 'Fulfilled by SOUK'
                    except:
                        fbs = None
                    sheet[f'A{n}'] = barcode
                    sheet[f'D{n}'] = name
                    sheet[f'H{n}'] = price
                    sheet[f'L{n}'] = seller
                    sheet[f'M{n}'] = rat
                    sheet[f'N{n}'] = fbs
                    sheet[f'O{n}'] = Min
                    sheet[f'P{n}'] = Max
                    try:
                        sheet.cell(n, 1).alignment = Alignment(horizontal='center', vertical='center',
                                                               wrap_text=True)
                        sheet.cell(n, 4).alignment = Alignment(horizontal='center', vertical='center',
                                                               wrap_text=True)
                        sheet.cell(n, 8).alignment = Alignment(horizontal='center', vertical='center',
                                                               wrap_text=True)
                        sheet.cell(n, 12).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrap_text=True)
                        sheet.cell(n, 13).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrap_text=True)
                        sheet.cell(n, 14).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrap_text=True)
                        sheet.cell(n, 15).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrap_text=True)
                        sheet.cell(n, 16).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrap_text=True)
                    except:
                        pass

                    book.save('upload.xlsx')
                    print(f'- {n} - {barcode} ')
                    n += 1
            except Exception as e:
                print(e)
    driver.quit()


def Souk_dash():
    user = 'infinity.unlimited.co@gmail.com'
    pas = 'NewEra2020'
    # driver = webdriver.Chrome(ChromeDriverManager().install())
    driver = webdriver.Chrome()
    # driver.maximize_window()
    #
    driver.get('https://uae.souq.com/ae-en/auth_portal.php')
    sleep(5)
    driver.find_element_by_css_selector('input#email').send_keys(user)
    driver.find_element_by_css_selector('input#siteLogin').click()
    sleep(5)
    driver.find_element_by_css_selector('input#continue').click()
    sleep(5)
    driver.find_element_by_css_selector('input#ap_password').send_keys(pas)
    driver.find_element_by_css_selector('input#signInSubmit').click()
    sleep(5)
    driver.get('https://sell.souq.com/inventory/search-results/casio%20fx/')
    print(len(driver.find_elements_by_css_selector('.product')))
    for element in driver.find_elements_by_css_selector('.product'):
        print(element.find_element_by_css_selector('.item-title').text)
        print(element.find_element_by_css_selector('.item-ean').text)
        print(element.find_elements_by_css_selector('.item-price span')[-1].text)
        for ele in element.find_elements_by_css_selector('.product-actions p'):
            if 'ng-hide' not in ele.get_attribute('class'):
                print(ele.get_attribute('ng-show').replace('item.', ''))

    driver.find_element_by_css_selector('.product').find_element_by_css_selector('.item-price a').get_attribute('href')
# if __name__ == "__main__":
#     s = Souk_dash()

#     while True:
#         answer = str(input(
# '''
# -----------------------------------------------------
# - This script to :-                                 -
# - extract information from Souk                     -
# - then show to save them                            -
# -----------------------------------------------------
# - 1 - To Extract >>> Enter 1                        -
# - 2 - To Show    >>> Enter 2                        -
# - 3 - To All     >>> Enter 3 or (Enter)             -
# -----------------------------------------------------
# >>> '''))
#         if answer == '1':
#             m = Main()
#             m.Get_url()
#             break
#         elif answer == '2':
#             Show()
#             break
#         elif answer == '3':
#             m = Main()
#             m.Get_url()
#             Show()
#             break
