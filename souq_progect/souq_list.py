from datetime import datetime

from selenium.webdriver.chrome.webdriver import WebDriver
# from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium import webdriver

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from time import sleep
import threading

'''
*\ Start in 
** 2020 - 10 - 13
** 2020 - 10 - 24
\*
'''


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class SouCode:
    driver: WebDriver

    def __init__(self):
        self.url = 'https://egypt.souq.com//eg-ar/'
        while True:
            try:
                self.file_name = input('- Enter file name : ')  # 'souq text.xlsx'
                self.book = load_workbook(self.file_name)
                self.sheet = self.book.active
                break
            except FileNotFoundError:
                print('- No such file or directory')
            except Exception as e:
                print(e)

        self.main_seller = input('- Enter Main Seller : ')  # 'Prime.Shop'
        self.user = input('- Enter Email : ')  # 'infinity.unlimited.co@gmail.com'
        self.pas = input('- Enter Password : ')  # 'NewEra2020'

        self.sheet.cell(2, 9).value = self.user
        self.sheet.cell(3, 9).value = self.pas
        self.sheet.cell(4, 9).value = self.main_seller
        try:
            self.sheet.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet.cell(1, 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet.cell(1, 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet.cell(1, 9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except:
            pass

        self.driver_search = webdriver.Chrome()#ChromeDriverManager().install())
        self.driver_code = webdriver.Chrome()#ChromeDriverManager().install())
        self.driver_code.implicitly_wait(10)
        self.login()

    def login(self):
        self.driver_code.get('https://sell.souq.com/inventory/inventory-management?tab=live')
        if '/inventory/' not in self.driver_code.current_url:
            self.driver_code.get('https://uae.souq.com/ae-en/auth_portal.php')
            sleep(4)
            sleep(5)
            self.driver_code.find_element_by_css_selector('input#email').send_keys(self.user)
            self.driver_code.find_element_by_css_selector('input#siteLogin').click()
            sleep(5)
            self.driver_code.find_element_by_css_selector('input#continue').click()
            sleep(5)
            self.driver_code.find_element_by_css_selector('input#ap_password').send_keys(self.pas)
            self.driver_code.find_element_by_css_selector('input#signInSubmit').click()
            sleep(5)
            while True:
                if '/signin/' not in self.driver_code.current_url:
                    self.driver_code.get('https://sell.souq.com/inventory/inventory-management?tab=live')
                    break

    def Search(self):
        thr = Thread()
        file_save = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + ' - ' + self.file_name
        exception = [self.sheet.cell(o, 4).value for o in range(2, self.sheet.max_row + 1) if self.sheet.cell(o, 4).value is not None]

        u = False
        for i in range(2, self.sheet.max_row + 1):
            try:
                self.sheet.cell(i, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                self.sheet.cell(i, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                self.sheet.cell(i, 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                self.sheet.cell(i, 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                self.sheet.cell(i, 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                self.sheet.cell(i, 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                self.sheet.cell(i, 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                self.sheet.cell(i, 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                self.sheet.cell(i, 9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except:
                pass
            self.book.save(file_save)

            if thr.p == 'pause' or thr.p == 'p':
                start = input('(Type start or s and press Enter)>>>')
                while True:
                    if start == 'start' or start == 's':
                        thr = Thread()
                        thr.start()
                        break
                    elif start == 'exit' or start == 'e':
                        self.driver_code.quit()
                        self.driver_search.quit()
                        u = True
                        break

                    else:
                        start = input()
            elif thr.p == 'exit' or thr.p == 'e':
                self.driver_code.quit()
                self.driver_search.quit()
                break

            if u:
                break

            bar = self.sheet.cell(i, 1).value
            max_p = float(self.sheet.cell(i, 2).value)
            min_p = float(self.sheet.cell(i, 3).value)
            print(f'- {i} - {bar}')

            self.driver_search.get(f'{self.url}{bar}/s/')
            sleep(5)
            try:
                _ = self.driver_search.find_element_by_class_name('zero-results').text
                static = 'Not Found'
                print('- ', static)
                continue
            except:
                pass

            try:
                _ = self.driver_search.find_element_by_class_name('notice').text
                static = 'Not Available'
                print(f'- {i} - {bar} ', static)
                continue
            except:
                pass

            # Number Full Filled
            if len(self.driver_search.find_elements_by_class_name('header-product-fulfilled')) != 0:
                fulfilled = True
            else:
                fulfilled = False

            price = float(self.driver_search.find_element_by_css_selector('#productTrackingParams').get_attribute('data-price'))
            self.sheet.cell(i, 6).value = price

            seller = self.driver_search.find_element_by_css_selector('#productTrackingParams').get_attribute('data-seller-name')
            self.sheet.cell(i, 5).value = seller

            # Seller Fulled By SOUQ
            if fulfilled:
                static = 'Fulled By Souq'
                self.sheet.cell(i, 7).value = static
                print('- ', static)
                continue
            else:
                self.sheet.cell(i, 7).value = 'Not Fulled By Souq'

            # Seller in Exception
            if seller in exception:
                static = 'Seller Exception'
                print('- ', static)
                continue
            try:
                # Edit From In
                # Search Bar
                self.driver_code.get('https://sell.souq.com/inventory/inventory-management?tab=live')
                sleep(4)
                self.driver_code.find_element_by_css_selector('form > fieldset > div > div.advanced-search > ul input.advanced-searchTxt').send_keys(bar, Keys.ENTER)
                sleep(3)

                # Open Product
                self.driver_code.find_element_by_css_selector('#table-inventory > tbody > tr > td:nth-child(3)').click()
                sleep(2)
            except:
                print('- Barcode Not Found')
                continue
            # Set Final Price = Product Price
            final_price = price

            # You Main Seller
            if seller.strip() == self.main_seller:

                # Your Price = Max Price
                if float(price) >= float(max_p):
                    print('- Your Price = Max Price')
                    print(f'- Final price : {final_price}')
                    self.sheet.cell(i, 8).value = final_price
                    continue

                # Your Price != Max Price
                else:

                    column = 0
                    # 10
                    # Generate Many Price
                    for i1 in range(10, 100, 10):
                        new_price = price + (price / 100 * i1)
                        self.sheet.cell(i, 10 + column).value = new_price
                        column += 1

                        # New Price <= Max Price
                        if new_price <= max_p:
                            self.driver_code.find_element_by_css_selector(
                                'div.ng-scope > label.price > input[name="offerPrice"]').clear()
                            self.driver_code.find_element_by_css_selector(
                                'div.ng-scope > label.price > input[name="offerPrice"]').send_keys(str(new_price), Keys.ENTER)
                            sleep(3)
                            self.driver_search.refresh()

                            # Main Seller Still Seller
                            if self.main_seller == self.driver_search.find_element_by_css_selector(
                                    '#productTrackingParams').get_attribute('data-seller-name'):
                                if final_price == max_p:
                                    print('- New Price = Max and Stop')
                                    self.sheet.cell(i, 8).value = final_price
                                    break

                                final_price = new_price
                                print(f'- price : {final_price} Is OK')
                                continue

                            # Main Seller Change, So Change To Final Price
                            else:
                                self.driver_code.find_element_by_css_selector(
                                    'div.ng-scope > label.price > input[name="offerPrice"]').clear()
                                self.driver_code.find_element_by_css_selector(
                                    'div.ng-scope > label.price > input[name="offerPrice"]').send_keys(str(final_price), Keys.ENTER)
                                print(f'Final price : {final_price}')
                                self.sheet.cell(i, 8).value = final_price
                                break

                        # New Price > Max Price, And Return Final Price To Max Price
                        else:
                            print('New price > Max Price ')
                            final_price = max_p
                            self.driver_code.find_element_by_css_selector(
                                'div.ng-scope > label.price > input[name="offerPrice"]').clear()
                            self.driver_code.find_element_by_css_selector(
                                'div.ng-scope > label.price > input[name="offerPrice"]').send_keys(str(final_price), Keys.ENTER)
                            print(f'Final price : {final_price}')
                            self.sheet.cell(i, 8).value = final_price
                            break

            # for i2 in range(5, 100, 5):
            #     price_list.append(price - (price / 100 * i2))

            # Edit From Out
            # self.driver_code.find_element_by_css_selector('#table-inventory > tbody > tr').click()
            # self.driver_code.find_element_by_css_selector('input#editableInput').clear()
            # self.driver_code.find_element_by_css_selector('input#editableInput').send_keys('50', Keys.ENTER)

            # Not Main Seller
            else:
                # Open Offer
                try:
                    # self.driver_search.find_element_by_css_selector('div.other-sellers-container div.show-for-medium  a').click()
                    of_url = self.driver_search.find_element_by_css_selector('div.other-sellers-container div.show-for-medium  a').get_attribute('href')
                    self.driver_search.get(of_url)
                    sleep(4)
                except:
                    print('- No Offer')
                    continue

                try:
                    price_main = None

                    # Get Main Price
                    for element in self.driver_search.find_elements_by_css_selector('#condition-all > div '):
                        name = element.find_element_by_css_selector(
                            'div.large-2.medium-2.small-6.columns.seller-field > div.field.seller-name > span.value > a').text
                        if name == self.main_seller:
                            price_main = float(element.find_element_by_css_selector(
                                'div.large-3.medium-3.small-6.columns > div.field.price-field').text.split(' ')[0])
                            break

                    if price_main is None:
                        print('- Your Price Not Found')
                        continue
                    else:
                        self.driver_search.get(f'{self.url}{bar}/s/')
                        sleep(4)

                        column = 0
                        # Generate Many Price
                        for i1 in range(1, 100, 2):
                            new_price = price_main - (price_main / 100 * i1)
                            self.sheet.cell(i, 10 + column).value = new_price
                            column += 1

                            # New Price <= Max Price
                            if new_price >= min_p:
                                self.driver_code.find_element_by_css_selector(
                                    'div.ng-scope > label.price > input[name="offerPrice"]').clear()
                                self.driver_code.find_element_by_css_selector(
                                    'div.ng-scope > label.price > input[name="offerPrice"]').send_keys(str(new_price), Keys.ENTER)
                                sleep(3)
                                self.driver_search.refresh()
                                sleep(2)

                                # Main Seller Become Seller
                                if self.main_seller == self.driver_search.find_element_by_css_selector(
                                        '#productTrackingParams').get_attribute('data-seller-name'):
                                    final_price = new_price
                                    print(f'- Final Price : {final_price}')
                                    self.sheet.cell(i, 8).value = final_price
                                    break

                                # Main Seller Not Seller
                                else:
                                    if new_price < min_p:
                                        final_price = max_p
                                        self.driver_code.find_element_by_css_selector(
                                            'div.ng-scope > label.price > input[name="offerPrice"]').clear()
                                        self.driver_code.find_element_by_css_selector(
                                            'div.ng-scope > label.price > input[name="offerPrice"]').send_keys(str(final_price), Keys.ENTER)
                                        self.driver_search.refresh()
                                        print('- New Price < Min Price')
                                        print(f'- Final price : {final_price}')
                                        self.sheet.cell(i, 8).value = final_price
                                        break

                            # New Price < Min Price
                            else:
                                print('New price < Min Price ')
                                final_price = max_p
                                self.driver_code.find_element_by_css_selector(
                                    'div.ng-scope > label.price > input[name="offerPrice"]').clear()
                                self.driver_code.find_element_by_css_selector(
                                    'div.ng-scope > label.price > input[name="offerPrice"]').send_keys(str(final_price), Keys.ENTER)
                                print(f'- Final price : {final_price}')
                                self.sheet.cell(i, 8).value = final_price
                                break

                except Exception as e:
                    print(e)


if __name__ == "__main__":
    my_bot = SouCode()

    my_bot.Search()

'''
لو انا البائع الرئيسى:
حاله 1 - لو السعر بتاعى على الماكسيمم ميعملش حاجه

حاله 2 - لو السعر بتاعى فرضا 60 والمنافس 55 والماكسيمم 100 فيبتدى عندى محاولة تحسين العرض
بانه ياخد سعر 65 ويرجع يقرا النتيجه,يلاقيك لسه مسيطر,يحسن السعر ل 70
,يرجع يقرا النتيجه,وهكذا لغاية ميفقد ميزة البيع فى حركة منهم فيعكسها
,ويبقى مسيطر باحسن سعر,,فيخرج,,او يوصل الحد الاقصى وهو مسيطر على العرض فيخرج برضه ويبدا عمليه جديده.

لو مش انا البائع الرئيسى
يبتدى يحرك السعر عن السعر الحالى بتاعنا, تنازليا بنسبة 2فى الميه كل مره لحين استجابة التغيير على الموقع لصالح العرض بتاعنا
,وانهاء العمليه,,,او الوصول للحد الادنى بدون استجابه فيتم الرجوع للحد الاقصى للسعر تدريجيا وانهاء العمليه بالحد الاقصى للسعر,,

لو البائع  تشحن من سوق تسيبه من غير اى اكشن
لو البائع اكسيبشن تسيبه من غير اى اكشن
'''

# infinity.unlimited.co@gmail.com
# NewEra2020
# Prime.Shop
