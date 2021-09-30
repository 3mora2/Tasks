# pip install PySide2 selenium webdriver_manager
import traceback
from io import BytesIO
from time import sleep
from typing import Optional, Any

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

import psutil
from PySide2.QtCore import QThread, Signal
from PySide2.QtGui import QIcon
from PySide2.QtWidgets import QMainWindow, QApplication, QMessageBox, QTabWidget, QHeaderView, QTableWidgetItem, QCheckBox, QFileDialog
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import sys
import warnings
import chromedriver_autoinstaller
from mainwindow import Ui_MainWindow
import datetime
import os
import start
session = requests.session()

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")

SCRIPT_DIRECTORY = os.path.dirname(os.path.abspath(__file__))


def error_Text(error):
    with open('log.log', 'a+', encoding="utf-8") as f:
        f.write(f'{datetime.datetime.now()}: {str(error)} \n')


def close_chrome_driver():
    for proc in psutil.process_iter():
        try:
            processName = proc.name()
            if 'chromedriver' in processName.lower():
                proc.kill()

        except:
            error_Text(traceback.format_exc())


# noinspection PyUnresolvedReferences
class MainWindow(QMainWindow, Ui_MainWindow):
    book: object
    tabWidget: QTabWidget
    checkBox: QCheckBox
    limit = True

    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)
        self.setupUi(self)
        self.Handel_Ui()
        self.Handel_Button()
        self.Handel_Default()
        self.Handel_check()
        self.OpenDriver = OpenDriver()
        self.OpenDriver.final.connect(self.final_)
        self.OpenDriver.wait.connect(self.Driver_Download)

    def Handel_Ui(self):
        scriptDir = os.path.dirname(os.path.realpath(__file__))
        self.setWindowIcon(QIcon(scriptDir + os.path.sep + 'logo.ico'))
        self.setWindowTitle("Google Map")
        self.tabWidget.tabBar().setVisible(False)

        # self.tableWidget.verticalHeader().setVisible(False)
        # self.tableWidget_2.verticalHeader().setVisible(False)

        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        self.tableWidget_2.horizontalHeader().setStretchLastSection(True)

        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def Handel_Default(self):
        self.tabWidget.setCurrentIndex(0)

        self.lineEdit_6.setDisabled(False)
        self.pushButton.setDisabled(False)
        self.pushButton_4.setDisabled(True)
        self.pushButton_3.setDisabled(True)
        self.pushButton_2.setDisabled(True)
        self.pushButton_8.setDisabled(True)
        self.pushButton_10.setDisabled(True)

    def Handel_Button(self):
        self.pushButton_6.clicked.connect(self.Main_Tab)
        self.pushButton_7.clicked.connect(self.Google_Tab)
        self.pushButton_49.clicked.connect(self.check_serial)
        self.pushButton.clicked.connect(self.Start_Google)
        self.pushButton_4.clicked.connect(self.Pause_OpenDriver)
        self.pushButton_3.clicked.connect(self.Restart_OpenDriver)
        self.pushButton_2.clicked.connect(self.Stop_OpenDriver)
        self.pushButton_8.clicked.connect(self.Skip_OpenDriver)
        self.pushButton_10.clicked.connect(self.Save_File)

    def Handel_check(self):
        if os.path.isfile(start.path_file):
            res = start.open_file()
            try:
                if res[0] and res[1]:
                    self.lineEdit_5.setText(
                        f' تاريخ انتهاء الاشتراك :{res[0].year}/{res[0].month}/{res[0].day}')
                    self.lineEdit_4.hide()
                    self.pushButton_49.hide()
                    self.limit = False
                elif res[0] is False and res[1]:
                    self.lineEdit_5.setText('انتهت صلاحية السيريال هذه النسخة تجريبية')
                    self.lineEdit_4.show()
                    self.pushButton_49.show()
                    self.limit = True

                else:
                    self.lineEdit_5.setText('هذه النسخة تجريبية')
                    self.lineEdit_4.show()
                    self.pushButton_49.show()
                    self.limit = True

            except Exception as e:
                print(e)
        else:
            self.lineEdit_4.show()
            self.pushButton_49.show()
            self.lineEdit_5.setText('هذه النسخة تجريبية')

    def check_serial(self):
        text = self.lineEdit_4.text()
        if text == '' or text is None:
            Box = QMessageBox()
            Box.setIcon(QMessageBox.Warning)
            Box.setWindowTitle("تحذير")
            Box.setText("من فضلك ادخل السيريال")
            Box.setStandardButtons(QMessageBox.Ok)
            Box.exec()
        else:
            if start.add_new(text):
                Box = QMessageBox()
                Box.setIcon(QMessageBox.Information)
                Box.setWindowTitle("تنبيه")
                Box.setText("تم تفعيل السريال بنجاح")
                Box.setStandardButtons(QMessageBox.Ok)
                Box.exec()
                self.Handel_check()
            else:
                Box = QMessageBox()
                Box.setIcon(QMessageBox.Warning)
                Box.setWindowTitle("خطا")
                Box.setText(" حدث خطأ \nهذا السيريال غير صحيح.")
                Box.setStandardButtons(QMessageBox.Ok)
                Box.exec()

    def Main_Tab(self):
        self.tabWidget.setCurrentIndex(0)

    def Google_Tab(self):
        if self.limit:
            pass
        else:
            self.tabWidget.setCurrentIndex(1)

    def show_limit(self):
        self.Box_Limit.exec()

    @staticmethod
    def Driver_Download():
        Box = QMessageBox()
        Box.setIcon(QMessageBox.Information)
        Box.setWindowTitle("تنبيه")
        Box.setText(" جاري تنزيل بعض الملفات \n قد يستغرق ذالك بضع دقائق")
        Box.setStandardButtons(QMessageBox.Ok)
        Box.exec()

    def Start_Google(self):
        text = self.lineEdit_6.text()
        if text == '' or text is None:
            Box = QMessageBox()
            Box.setIcon(QMessageBox.Warning)
            Box.setWindowTitle("تحذير")
            Box.setText("من فضلك ادخل كلمات للبحث")
            Box.setStandardButtons(QMessageBox.Ok)
            Box.exec()
        else:
            if self.checkBox.isChecked():
                self.OpenDriver.download_img = True
            else:
                self.OpenDriver.download_img = False

            if self.checkBox_2.isChecked():
                self.OpenDriver.open_url = True
            else:
                self.OpenDriver.open_url = False

            self.OpenDriver.tableWidget = self.tableWidget
            self.OpenDriver.tableWidget_2 = self.tableWidget_2
            self.OpenDriver.text = text
            self.OpenDriver.start()
            self.lineEdit_6.setDisabled(True)
            self.pushButton.setDisabled(True)
            self.pushButton_4.setDisabled(False)
            self.pushButton_3.setDisabled(True)
            self.pushButton_2.setDisabled(False)
            self.pushButton_8.setDisabled(False)
            self.pushButton_10.setDisabled(True)

    def Save_File(self):
        save_file = QFileDialog.getSaveFileName(self, "حفظ ك", 'result.xlsx', '*.xlsx')[0]
        if save_file != '':
            self.book.save(save_file)

    def Pause_OpenDriver(self):
        self.OpenDriver.pause = True
        self.pushButton_4.setDisabled(True)
        self.pushButton_3.setDisabled(False)

    def Restart_OpenDriver(self):
        self.OpenDriver.pause = False
        self.pushButton_4.setDisabled(False)
        self.pushButton_3.setDisabled(True)

    def Skip_OpenDriver(self):
        self.OpenDriver.skip = True
        self.pushButton_8.setDisabled(True)

    def Stop_OpenDriver(self):
        self.OpenDriver.stop = True
        self.pushButton_4.setDisabled(True)
        self.pushButton_2.setDisabled(True)
        self.pushButton_3.setDisabled(True)

    def final_(self, book):
        self.book = book
        self.lineEdit_6.setDisabled(False)
        self.pushButton.setDisabled(False)
        self.pushButton_4.setDisabled(True)
        self.pushButton_3.setDisabled(True)
        self.pushButton_2.setDisabled(True)
        self.pushButton_8.setDisabled(True)
        self.pushButton_10.setDisabled(False)
        Box = QMessageBox()
        Box.setIcon(QMessageBox.Information)
        Box.setWindowTitle("تنبيه")
        Box.setText("تم الانتهاء")
        Box.setStandardButtons(QMessageBox.Ok)
        Box.exec()


# noinspection PyUnresolvedReferences
class OpenDriver(QThread):
    sheet: object
    book: object
    urls: set[Optional[Any]]
    driver: WebDriver = None
    final = Signal(object)
    wait = Signal()
    pause = False
    stop = False
    skip = False

    def run(self):
        self.close()
        try:
            try:
                self.driver = webdriver.Chrome()
            except WebDriverException:
                self.wait.emit()
                scriptDir = os.path.dirname(os.path.realpath(__file__))
                path = chromedriver_autoinstaller.install(cwd=scriptDir)
                if 'Traceback' not in path:
                    try:
                        os.popen('attrib +h chromedriver.exe')
                    except:
                        with open('log.log', 'a+', encoding="utf-8") as f:
                            f.write(
                                f'{datetime.datetime.now()}: {str(traceback.format_exc())}\n')

                    self.driver = webdriver.Chrome(path)

                else:
                    error_Text(path)
                    raise Exception
            except:
                error_Text(traceback.format_exc())

            print('- done')
            # self.driver.get('https://www.google.com/preferences#languages')
            # input('- Enter: ')

            if not self.stop:
                self.open(self.text)
                if not self.stop:
                    self.extract()
        except Exception as a:
            print(a)
            error_Text(traceback.format_exc())

        self.final.emit(self.book)

    def close(self):
        try:
            self.driver.quit()
        except:
            pass
        close_chrome_driver()

    def open(self, txt):
        current_n = 0
        self.urls = set()
        current_n_n = 0
        self.driver.get(f'https://www.google.com/maps/search/{txt}')
        print('Program is loading,please wait!')
        sleep(5)
        while True:
            while self.pause:
                if self.stop:
                    break
                if self.skip:
                    break
                sleep(1)

            if self.stop:
                break

            if self.skip:
                break

            num_scroll = 0
            while True:
                try:
                    _ = self.driver.find_element_by_css_selector('.wo1ice-loading').location_once_scrolled_into_view
                    sleep(5)
                    num_scroll += 1

                    if num_scroll > 5:
                        print('- Break Scroll..')
                except Exception as e:
                    # print(e)
                    break

            sleep(3)
            num_pl = len(self.driver.find_elements_by_css_selector('div.section-layout.section-scrollbox div > a'))
            if num_pl == 0:
                break

            old_number = [self.tableWidget.item(i, 0).text() for i in range(self.tableWidget.rowCount())]
            for element in self.driver.find_elements_by_css_selector('div.section-layout.section-scrollbox div>a'):
                self.urls.add(element.get_attribute('href'))
                if element.get_attribute('href') not in old_number:
                    r = self.tableWidget.rowCount()
                    self.tableWidget.insertRow(r)
                    self.tableWidget.setItem(r, 0, QTableWidgetItem(str(element.get_attribute('href'))))

            print(len(self.urls))

            try:
                if self.driver.find_element_by_css_selector(
                        'button[jsaction="pane.paginationSection.nextPage"]').get_attribute('disabled') == 'true':
                    break

                try:
                    self.driver.execute_script('document.querySelector(\'*[aria-label="الصفحة التالية"]\').click()')
                except:
                    self.driver.execute_script(
                        'document.querySelector(\'button[jsaction="pane.paginationSection.nextPage"]\').click()')
                sleep(5)
            except:
                break

            if len(self.urls) != current_n:
                current_n = len(self.urls)
                current_n_n = 0

            else:
                current_n_n += 1
            if current_n_n > 3:
                break

    def extract(self):
        self.book = Workbook()
        self.sheet = self.book.active
        if self.download_img:
            self.sheet['A1'] = 'Image'
            self.sheet.column_dimensions['A'].width = 13

        self.sheet['B1'] = 'Name'
        self.sheet['C1'] = 'Type'
        self.sheet['D1'] = 'Rate'
        self.sheet['E1'] = 'Comments'
        self.sheet['F1'] = 'Address'
        self.sheet['G1'] = 'Plus Code'
        self.sheet['H1'] = 'Phone'
        self.sheet['I1'] = 'Website'
        self.sheet['J1'] = "IMG"
        self.sheet['K1'] = "Opened"
        self.sheet['L1'] = "Coordinates"
        self.sheet['M1'] = "URL"

        self.sheet.column_dimensions['B'].width = 50
        self.sheet.column_dimensions['C'].width = 30
        self.sheet.column_dimensions['D'].width = 10
        self.sheet.column_dimensions['E'].width = 10
        self.sheet.column_dimensions['F'].width = 100
        self.sheet.column_dimensions['G'].width = 70
        self.sheet.column_dimensions['H'].width = 25
        self.sheet.column_dimensions['I'].width = 30
        self.sheet.column_dimensions['J'].width = 50
        self.sheet.column_dimensions['K'].width = 50
        self.sheet.column_dimensions['L'].width = 25
        self.sheet.column_dimensions['M'].width = 120
        for column in range(1, 14):
            try:
                self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except:
                pass
        num = 2
        for url in self.urls:
            while self.pause:
                if self.stop:
                    break
                sleep(1)

            if self.stop:
                break
            try:
                self.driver.get(url)
                sleep(4)
                try:
                    main_photo = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((
                        By.CSS_SELECTOR,
                        'button[jsaction="pane.heroHeaderImage.click"] > img'))).get_attribute(
                        'src')
                except:
                    main_photo = None

                try:
                    rat = ''.join(set(rat.get_attribute('aria-label') for rat in
                                      self.driver.find_elements_by_css_selector('ol')
                                      if rat.get_attribute('aria-label') is not None)).replace('stars', '').strip()
                except:
                    try:
                        _ = self.driver.find_element_by_css_selector(
                            '[jsaction="pane.reviewChart.moreReviews"] .gm2-display-2').location_once_scrolled_into_view
                        rat = self.driver.find_element_by_css_selector(
                            '[jsaction="pane.reviewChart.moreReviews"] .gm2-display-2').text
                    except:
                        rat = None
                try:
                    comment = list(set(rat.get_attribute('aria-label') for rat in
                                       self.driver.find_elements_by_css_selector('[jsaction="pane.rating.moreReviews"]')
                                       if rat.get_attribute('aria-label') is not None))[0]
                except:
                    comment = None

                name = WebDriverWait(self.driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, '//h1/span[1]'))).text
                try:
                    ty = self.driver.find_element_by_css_selector('[jsaction="pane.rating.category"]').text
                except:
                    ty = None
                try:
                    phone = \
                        self.driver.find_element_by_xpath(
                            "//div/button[contains(@data-item-id, 'phone:')]").get_attribute(
                            'data-item-id').split('tel:')[-1]
                except:
                    phone = None
                try:
                    address = \
                        self.driver.find_element_by_xpath(
                            "//div/button[contains(@data-item-id, 'address')]").get_attribute(
                            'aria-label').split(":")[-1]
                except:
                    address = None
                try:
                    web = \
                        self.driver.find_element_by_xpath(
                            "//div/button[contains(@data-item-id, 'authority')]").get_attribute(
                            'aria-label').split(":")[-1]
                except:
                    web = None
                try:
                    cont = \
                        self.driver.find_element_by_css_selector('[data-item-id="oloc"]').get_attribute(
                            'aria-label').split(
                            ':')[-1]
                except:
                    cont = None
                try:
                    opened = self.driver.find_element_by_xpath("//div[contains(@jsaction, 'pane.openhours')]").text
                except:
                    opened = None
                try:
                    address_num = '@'+','.join(self.driver.find_element_by_css_selector('link[rel="shortcut icon"]+script').get_attribute('innerText').split('/@')[1].split(',')[:2])
                except Exception as e:
                    print(e)
                    address_num = None

                if self.open_url:
                    try:
                        if len(self.driver.find_elements_by_xpath(
                                "//div/button[contains(@data-item-id, 'authority')]")):
                            self.driver.execute_script(
                                'document.querySelector("[data-item-id=\'authority\'] [dir=\'ltr\']").click();')
                            sleep(3)
                            if len(self.driver.window_handles) > 1:
                                self.driver.close()
                                self.driver.switch_to.window(self.driver.window_handles[-1])
                                self.sheet[f'I{num}'] = self.driver.current_url
                            else:
                                pass
                    except:
                        pass
                else:
                    self.sheet[f'I{num}'] = web

                try:
                    if main_photo and self.download_img:
                        res = session.get(main_photo)
                        image_file = BytesIO(res.content)
                        img = Image(image_file)
                        img.width = 90
                        img.height = 75
                        self.sheet.row_dimensions[num].height = 56
                        self.sheet.add_image(img, f'A{num}')
                except Exception as e:
                    print(e)
                    pass
                self.sheet[f'B{num}'] = name
                self.sheet[f'C{num}'] = ty
                self.sheet[f'D{num}'] = rat
                self.sheet[f'E{num}'] = comment
                self.sheet[f'F{num}'] = address
                self.sheet[f'G{num}'] = cont
                self.sheet[f'H{num}'] = phone
                self.sheet[f'J{num}'] = main_photo
                self.sheet[f'K{num}'] = opened
                self.sheet[f'L{num}'] = address_num
                self.sheet[f'M{num}'] = url

                for column in range(2, 13):
                    try:
                        self.sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True)
                        self.sheet.cell(num + 1, column).alignment = Alignment(horizontal='center',
                                                                               vertical='center', wrap_text=True)
                    except:
                        pass
                print(num - 1, len(self.urls))

                r = self.tableWidget_2.rowCount()
                self.tableWidget_2.insertRow(r)
                self.tableWidget_2.setItem(r, 0, QTableWidgetItem(str(name)))
                self.tableWidget_2.setItem(r, 1, QTableWidgetItem(str(phone)))

                num += 1
            except Exception as e:
                print(e)

        print('- Done....', end='\r')
        self.driver.quit()

    def Stop(self):
        pass


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec_()
'''
pyside2-uic ui.ui -o mainwindow.py
pyside6-rcc icons.rc -o rc_icons.py
pyinstaller --windowed --icon "C:/Users/3mora/Dropbox/All/google map/logo.ico" --add-data "C:/Users/3mora/Dropbox/All/google map/logo.ico;."  "C:/Users/3mora/Dropbox/All/google map/main.py"
'''
