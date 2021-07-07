import mysql.connector
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
import requests
from openpyxl import Workbook, load_workbook
from time import sleep


class Main_DB:
    def __init__(self):
        self.database = mysql.connector.connect(host="localhost", user="root", passwd="", database="lawyers")
        # self.driver = webdriver.Chrome(ChromeDriverManager().install())

    def Test(self):
        self.driver.get('https://lawyers.findlaw.com/lawyer/firm/motor-vehicle-accidents-plaintiff/new-york/new-york')
        urls = []
        while True:
            for element in self.driver.find_elements_by_css_selector('div#serp_results div.row'):
                url = element.find_element_by_css_selector('.listing-ctas  a.directory_profile').get_attribute('href')
                urls.append(url)
            print(len(urls), end='\r')
                # try:
                #     web = element.find_element_by_css_selector('.listing-ctas  a.directory_website').get_attribute('href').split('*')[-1]
                # except:
                #     web = None
                # try:
                #     phone = element.find_element_by_css_selector('.directory_phone').text
                # except:
                #     phone = None
                # name = element.find_element_by_css_selector('.listing-details-header').text
                # try:
                #     address = element.find_element_by_css_selector('.listing-details-review').text
                # except:
                #     address = None
                # img = element.find_element_by_css_selector('img').get_attribute('src')
                # res = requests.get(img)
                # img_b = res.content
                # sql = f"INSERT INTO location (name, phone, address, website, image) VALUES  (%s,%s,%s,%s,%s)"
                # cursor.execute(sql, (name, phone, address, web, img_b))
                # self.database.commit()
                # print(n, name, phone, address, web)
                # n += 1
            # break
            try:
                self.driver.find_element_by_css_selector('a[rel="next"]').click()
            except:
                break
        self.get_d(urls)

    def get_d(self, urls=None):
        book = load_workbook('layer.xlsx')
        sheet = book.active
        n = 1
        cursor = self.database.cursor()
        for num in range(2, sheet.max_row+1):
            name = sheet.cell(num, 1).value
            if name is None:
                continue
            phone = sheet.cell(num, 2).value
            address = sheet.cell(num, 3).value
            web = sheet.cell(num, 4).value
            img = sheet.cell(num, 5).value
            res = requests.get(img)
            img_b = res.content
            cont = sheet.cell(num, 8).value
            sql = f"INSERT INTO location (name, phone, address, country, website, image) VALUES  (%s,%s,%s,%s,%s,%s)"
            cursor.execute(sql, (name, phone, address, cont, web, img_b))
            self.database.commit()
            print(num)
            # try:
            #     url = sheet.cell(num, 7).value
            #     if url is not None and sheet.cell(num, 1).value is None:
            #         self.driver.get(url)
            #         img = self.driver.find_element_by_css_selector('div.media-object-section img').get_attribute('src')
            #         name = self.driver.find_element_by_css_selector('.listing-details-header').text
            #         try:
            #             web = self.driver.find_element_by_css_selector('a.profile-website-header').get_attribute('href')
            #         except:
            #             web = None
            #         address = self.driver.find_element_by_css_selector('.pp_card_street').text
            #         try:
            #             phone = self.driver.find_element_by_css_selector('.listing-desc-phone.profile-phone-header').get_attribute('href').split(':')[-1]
            #         except:
            #             phone = None
            #         # res = requests.get(img)
            #         # img_b = res.content
            #         # sql = f"INSERT INTO location (name, phone, address, website, image) VALUES  (%s,%s,%s,%s,%s)"
            #         # cursor.execute(sql, (name, phone, address, web, img_b))
            #         # self.database.commit()
            #
            #         sheet.cell(num, 1).value = name
            #         sheet.cell(num, 2).value = phone
            #         sheet.cell(num, 3).value = address
            #         sheet.cell(num, 4).value = web
            #         sheet.cell(num, 5).value = img
            #         print(n, name, phone, address, web)
            #         n += 1
            # except Exception as e:
            #     print(e)
        # book.save('layer.xlsx')


if __name__ == '__main__':
    app = Main_DB()
    app.get_d()

'''
DROP DATABASE IF EXISTS lawyers;

CREATE DATABASE lawyers;

USE lawyers;

CREATE TABLE location (
    id INT NOT NULL PRIMARY KEY AUTO_INCREMENT,
    name VARCHAR(50),
    phone  VARCHAR(20),
    address VARCHAR(200),
    country VARCHAR(100),
    website  VARCHAR(200),
    image   longblob 
);

'''
# driver = webdriver.Chrome(ChromeDriverManager().install())
# driver.get('https://lawyers.findlaw.com/lawyer/firm/motor-vehicle-accidents-plaintiff/new-york/new-york')
# sleep(5)
# l = [(el.get_attribute('href'), el.text) for el in driver.find_elements_by_css_selector('#sg_browse_for_a_lawyer > div > div > div:nth-child(1) > div > ul > li > a')]
# # 3885
# tt = 0
# book = Workbook()
# sheet = book.active
# n = 2
# for u, t in l:
#     driver.get(u)
#     urls = []
#     while True:
#         for element in driver.find_elements_by_css_selector('div#serp_results div.row'):
#             url = element.find_element_by_css_selector('.listing-ctas  a.directory_profile').get_attribute('href')
#             if url not in urls:
#                 urls.append(url)
#                 sheet.cell(n, 7).value = url
#                 sheet.cell(n, 8).value = t
#                 n += 1
#         try:
#             driver.find_element_by_css_selector('a[rel="next"]').click()
#         except:
#             break
#     tt += len(urls)
#     print(t, len(urls), tt)
#
# book.save('layer.xlsx')

# https://warehouse.noon.partners/en-eg/stock?tab=readyforpickup&limits=100&page=1
# https://catalog.noon.partners/en-eg/noon-catalog