from PySide2.QtWidgets import QTableWidget
from PySide2.QtCore import Signal, QThread
from openpyxl import Workbook


class SaveAqarExcel(QThread):
    final = Signal()
    error = Signal()
    tableWidget_2: QTableWidget

    def run(self) -> None:
        book = Workbook()
        sheet = book.active
        row = 1
        for c in range(self.tableWidget_2.columnCount()):
            sheet.cell(row, c+1).value = self.tableWidget_2.horizontalHeaderItem(c).text()

        row += 1
        for i in range(self.tableWidget_2.rowCount()):
            for c in range(self.tableWidget_2.columnCount()):
                try:
                    text = self.tableWidget_2.item(i, c).text()
                except:
                    text = None
                sheet.cell(row, c + 1).value = text if text not in ['None', 'UNKNOWN'] else None
            row += 1

        book.save(self.filename)
        self.final.emit()
