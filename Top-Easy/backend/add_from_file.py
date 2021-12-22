import json
import sqlite3

from PySide2.QtCore import QThread, Signal
from openpyxl import load_workbook
from aqar_way import FULL_NAMES_RE
from settings import DB_URL


class AddFromFile(QThread):
    final = Signal()

    def run(self):
        book = load_workbook(self.file)
        sheet = book.active
        for r in range(2, sheet.max_row+1):
            data = dict()
            for c in range(1, sheet.max_column+1):
                key = sheet.cell(1, c).value
                value = sheet.cell(r, c).value
                if key in FULL_NAMES_RE:
                    key = FULL_NAMES_RE.get(key)

                data[key] = value
            try:
                data.pop('url')
            except:
                pass
            try:
                user_info_id = data.pop('user_id')
            except:
                user_info_id = None
            try:
                user_name = data.pop('user_name')
            except:
                user_name = None
            try:
                phone = data.pop('phone')
            except:
                phone = None
            try:
                extra_phone = data.pop('extra_phone')
            except:
                extra_phone = None
            try:
                documents = data.pop('documents')
            except:
                documents = None
            try:
                email = data.pop('email')
            except:
                email = None
            try:
                title = data.pop('title')
            except:
                title = None
            try:
                id_ = data.pop('id')
            except:
                id_ = None
            sql_data = '''INSERT INTO REAL_OFFERS(NAME, PHONE, EXTRA_NUMBER, ID_NUMBER, EMAIL , OFFER_NAME, DOCUMENTS, OFFER_ID, DATA)
                    SELECT ?, ?, ?, ?, ?, ?, ?, ?, ?
                    WHERE NOT EXISTS(SELECT 1 FROM REAL_OFFERS WHERE NAME = ? AND DATA = ?)'''
            sql_user = '''INSERT INTO REAL_USER(NAME, ID_NUMBER, PHONE, EXTRA_NUMBER, EMAIL) 
                                SELECT ?, ?, ?, ?, ?
                                WHERE NOT EXISTS(SELECT 1 FROM REAL_USER WHERE NAME = ? AND ID_NUMBER = ?)'''
            if user_info_id or user_name or phone:
                data = json.dumps(data)
                db = sqlite3.connect(DB_URL)
                cur = db.cursor()
                # Add offer to db
                cur.execute(sql_data, (
                    user_name, phone, extra_phone, user_info_id, email, title, documents, id_, data, user_name, data))
                db.commit()
                # Add User to db

                cur.execute(sql_user, (user_name, user_info_id, phone, extra_phone, email, user_name, user_info_id))
                db.commit()

                cur.close()
                db.close()
                print('- add')

        self.final.emit()
