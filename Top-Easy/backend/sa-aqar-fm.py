import json
import threading
from time import sleep
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import smtplib
from email.message import EmailMessage


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 's':
                break


def send_mail_with_excel():
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = SENDER_EMAIL
    msg['To'] = Recipient_email
    msg.set_content(topic)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(SENDER_EMAIL, APP_PASSWORD)
        smtp.send_message(msg)


def select_number(n, url):
    if '?' in url:
        url = url.split('?')[0]

    list_url = url.split('/')
    if list_url[-1].isnumeric():
        list_url[-1] = str(n)
        url = '/'.join(list_url)
    else:
        url = '/'.join(list_url)+'/'+str(n)
    return url


def print_percent_done(index, total, bar_len=50):
    percent_done = index / total * 100
    percent_done = round(percent_done, 1)

    done = round(percent_done / (100 / bar_len))
    togo = bar_len - done

    done_str = '█' * int(done)
    togo_str = '░' * int(togo)

    print(f'- ⏳ {total}\\{index} - [{done_str}{togo_str}] ({percent_done} %)', end='\r')


class CollectPosts(object):
    def __init__(self):
        self.data_set = list()
        self.links = []
        self.current_cell = 1
        self.book = Workbook()
        self.sheet = self.book.active

        self.sheet.cell(self.current_cell, 1).value = 'الاسم'
        self.sheet.cell(self.current_cell, 2).value = 'النوع'
        self.sheet.cell(self.current_cell, 3).value = 'البلد'
        self.sheet.cell(self.current_cell, 4).value = 'الحي'
        self.sheet.cell(self.current_cell, 5).value = 'رقم الهاتف'
        self.sheet.cell(self.current_cell, 6).value = 'السعر'
        self.sheet.cell(self.current_cell, 7).value = 'المساحه'
        self.sheet.cell(self.current_cell, 8).value = 'الوقت'
        self.sheet.cell(self.current_cell, 9).value = 'سعر المتر'
        self.sheet.cell(self.current_cell, 10).value = 'الوجهه'
        self.sheet.cell(self.current_cell, 11).value = 'الغرض'
        self.sheet.cell(self.current_cell, 12).value = 'عرض الشارع'
        self.sheet.cell(self.current_cell, 13).value = 'البائع'
        self.sheet.cell(self.current_cell, 14).value = 'الوصف'
        self.sheet.cell(self.current_cell, 15).value = 'الصور'
        self.sheet.cell(self.current_cell, 16).value = 'الرابط'
        self.current_cell += 1

        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get('https://sa.aqar.fm/login')

    def close(self):
        try:
            self.driver.find_element_by_css_selector('.national-closeBtn').click()
            sleep(1)
        except:
            pass

    def login(self):
        try:
            sleep(1)
            self.driver.find_element_by_css_selector('input[name="phone"]').send_keys(email)
            sleep(1)
            self.driver.find_element_by_css_selector('input[name="password"]').send_keys(password)
            sleep(1)
            # self.driver.find_element_by_css_selector('a.submitBtn').click()
            self.close()
            self.driver.execute_script("document.querySelector('a.submitBtn').click()")
            sleep(2)
        except:
            pass

    def safe_find_element_by(self, by, elem):
        try:
            return self.driver.find_element(by, elem)
        except NoSuchElementException:
            return None

    def selenium_posts(self):
        # self.driver.get('https://sa.aqar.fm/')
        input('- Enter: ')
        thr = Thread()
        thr.start()
        if init > 1:
            self.driver.get(select_number(init, self.driver.current_url))
        page = 1
        while True:
            for element in self.driver.find_elements(By.CSS_SELECTOR, "h4 a.listTitle"):
                link = element.get_attribute('href')
                if link in self.links:
                    print('- found')
                    continue
                self.links.append(link)
                print(len(self.links), page)
            if page == num or thr.p == 's':
                break
            try:
                next_url = self.driver.find_element_by_css_selector('#next-page').get_attribute('href')
                self.driver.get(next_url)
                sleep(1)
                page += 1
            except:
                break

        self.get_posts(self.links)

    def get_posts(self, links):
        thr = Thread()
        thr.start()

        for link in links:
            self.driver.get(link)
            sleep(2)
            try:
                name = self.driver.find_element_by_css_selector('a.listingPageTitleLink').text
                tpe = self.driver.find_elements_by_css_selector('.tree.treeMargin li')[0].text
                country = self.driver.find_elements_by_css_selector('.tree.treeMargin li')[1].text
                hy = ' '.join([cat.text for cat in self.driver.find_elements_by_css_selector('.tree.treeMargin li')[2:-1]])
                img = '\n'.join([cat.get_attribute('src') for cat in
                                 self.driver.find_elements_by_css_selector('.listingImgContainer img') if
                                 cat.get_attribute('src')])
                price = self.driver.find_element_by_css_selector('.listingPagePrice').text
                content = self.driver.find_element_by_css_selector('.listingPageContent').text
                try:
                    area = self.driver.find_element_by_xpath('//tr[td[contains(text(),"المساحة")]] /td[div]').text
                except:
                    area = None
                try:
                    price_m = self.driver.find_element_by_xpath('//tr[td[contains(text(),"سعر المتر")]] /td[div]').text
                except:
                    price_m = None

                try:
                    direct = self.driver.find_element_by_xpath('//tr[td[contains(text(),"الواجهة")]] /td[div]').text
                except:
                    direct = None

                try:
                    props = self.driver.find_element_by_xpath('//tr[td[contains(text(),"الغرض")]] /td[div]').text
                except:
                    props = None

                try:
                    wh = self.driver.find_element_by_xpath('//tr[td[contains(text(),"عرض الشارع")]] /td[div]').text
                except:
                    wh = None

                try:
                    time = self.driver.find_element_by_css_selector('[data-icon="clock"]~ span').text
                except:
                    time = None

                try:
                    user = self.driver.find_element_by_css_selector('.userName span').text
                except:
                    user = None

                try:
                    last = self.driver.find_element_by_css_selector('span.userViewPhone').text
                    self.driver.find_element_by_css_selector('span.userViewPhone').click()
                    while True:
                        phone = self.driver.find_element_by_css_selector('span.userViewPhone').text
                        if phone != last:
                            break

                except:
                    phone = None

                self.sheet.cell(self.current_cell, 1).value = name
                self.sheet.cell(self.current_cell, 2).value = tpe
                self.sheet.cell(self.current_cell, 3).value = country
                self.sheet.cell(self.current_cell, 4).value = hy
                self.sheet.cell(self.current_cell, 5).value = phone
                self.sheet.cell(self.current_cell, 6).value = price
                self.sheet.cell(self.current_cell, 7).value = area
                self.sheet.cell(self.current_cell, 8).value = time
                self.sheet.cell(self.current_cell, 9).value = price_m
                self.sheet.cell(self.current_cell, 10).value = direct
                self.sheet.cell(self.current_cell, 11).value = props
                self.sheet.cell(self.current_cell, 12).value = wh
                self.sheet.cell(self.current_cell, 13).value = user
                self.sheet.cell(self.current_cell, 14).value = content
                self.sheet.cell(self.current_cell, 15).value = img
                self.sheet.cell(self.current_cell, 16).value = link

                print_percent_done(self.current_cell-1, len(links))
                self.current_cell += 1
            except:
                pass

            if thr.p == 's':
                break

        self.book.save('Data.xlsx')
        self.driver.quit()
        if SENDER_EMAIL and APP_PASSWORD and Recipient_email:
            send_mail_with_excel()


if __name__ == '__main__':
    while True:
        try:
            num = int(input('- Enter Page Number: '))
            if num > 0:
                break
        except:
            pass
    while True:
        try:
            init = int(input('- Enter Start Page : '))
            if init > 0:
                break
        except:
            pass
    try:
        with open('data.json') as f:
            text = f.read()
        text = json.loads(text)
        SENDER_EMAIL = text['SENDER_EMAIL']
        APP_PASSWORD = text['APP_PASSWORD']
        Recipient_email = text['Recipient_email']
        topic = text['topic']
        subject = text['subject']
        email = text['email']
        password = text['password']
    except Exception as e:
        print(e)
        print('- Cant Found Json file')
        SENDER_EMAIL = None
        APP_PASSWORD = None
        Recipient_email = None
        email = input('- Enter your Phone : ')
        password = input('- Enter your password : ')

    self = CollectPosts()
    self.login()
    input('- Enter: ')
    self.selenium_posts()

'''
0550344513
afaf4321
'''
