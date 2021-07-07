from datetime import datetime
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.keys import Keys
from selenium import webdriver

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


from time import sleep
from requests_html import HTML
import threading


class Thread(threading.Thread):
    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        self.p = None
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class SouCode:
    driver: WebDriver

    def __init__(self, file_name):
        self.file_save = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + file_name
        self.url = 'https://www.noon.com/egypt-en/'
        try:
            self.book = load_workbook(file_name)
            self.sheet = self.book.active
            self.sheet[f'A1'] = 'noon_sku'
            self.sheet[f'B1'] = 'Minimum Price'
            self.sheet[f'C1'] = 'Maximum Price'
            self.sheet[f'D1'] = 'Name Seller'
            self.sheet[f'E1'] = 'Price'
            self.sheet[f'F1'] = 'Express'
            self.sheet[f'G1'] = 'Only'
            self.sheet[f'H1'] = 'New price'
            self.sheet[f'I1'] = 'Stat'
            self.sheet[f'J1'] = 'Account'
            self.sheet[f'K1'] = 'Pass'
            self.sheet[f'L1'] = 'Change Value +'
            self.sheet[f'M1'] = 'Change Value -'
            self.sheet[f'N1'] = 'Except'
            self.sheet[f'O1'] = 'Numbers'
            self.sheet[f'P1'] = 'Note'
            self.sheet[f'Q1'] = 'Note change'
            self.sheet.column_dimensions['A'].width = 14
            self.sheet.column_dimensions['B'].width = 15
            self.sheet.column_dimensions['C'].width = 15
            self.sheet.column_dimensions['D'].width = 15
            self.sheet.column_dimensions['E'].width = 10
            self.sheet.column_dimensions['F'].width = 14
            self.sheet.column_dimensions['G'].width = 10
            self.sheet.column_dimensions['H'].width = 11
            self.sheet.column_dimensions['I'].width = 18
            self.sheet.column_dimensions['J'].width = 30
            self.sheet.column_dimensions['K'].width = 20
            self.sheet.column_dimensions['L'].width = 16
            self.sheet.column_dimensions['M'].width = 16
            self.sheet.column_dimensions['N'].width = 16
            self.sheet.column_dimensions['O'].width = 13
            self.sheet.column_dimensions['P'].width = 70
            self.sheet.column_dimensions['Q'].width = 50
            for column in range(1, 18):
                try:
                    self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                except:
                    pass
        except Exception as e:
            print(e)
            exit()

    def Search(self):
        avp = avrp
        avn = avrn
        Note = None
        self.sheet[f'J2'] = str(email)
        self.sheet[f'K2'] = str(password)

        self.sheet[f'L2'] = avrp
        self.sheet[f'M2'] = avrn
        end_cell = self.sheet.max_row
        exception_seller = [self.sheet[f'N{i}'].value for i in range(2, end_cell) if
                            self.sheet[f'N{i}'].value is not None]

        print(len(exception_seller))
        print(end_cell)
        self.thr = Thread()
        self.thr.start()
        self.driver = webdriver.Chrome()
        for i in range(2, end_cell + 1):
            bar = self.sheet[f'A{i}'].value
            if bar is None:
                continue
            max_pr = float(self.sheet[f'C{i}'].value)
            min_pr = float(self.sheet[f'B{i}'].value)
            try:
                self.driver.get(f'{self.url}{bar}/p/')
                sleep(2)
                r = HTML(html=self.driver.page_source)
                price = float(r.find('span.sellingPrice span.value', first=True).text)
                seller = r.find('p.sellerName a', first=True).text
                express = r.find('div.bottomRow div.container img', first=True).attrs['alt']

                Offer_number = len(
                    self.driver.find_elements_by_css_selector('ul.offersList > li > div.sellerDetails > p > strong'))
                ############################################################################################################
                if 'Infinity' in seller:
                    if Offer_number == 0:
                        new_price = max_pr
                        only = 'Only'
                        static = 'Maximum Price'
                        print('You are Only Maximum price is ', max_pr)
                        Note = 'You are Only Maximum price is ' + str(max_pr)
                    else:
                        only = 'No'
                        best = []
                        self.driver.find_element_by_css_selector('span.cta').click()
                        sleep(4)
                        offer = self.driver.find_elements_by_css_selector('ul.offersList > li')
                        try:
                            for ELEMENT in offer:
                                if offer.index(ELEMENT) == 4:
                                    break
                                if 'selectedItem' not in ELEMENT.get_attribute('class') \
                                        and ELEMENT.find_element_by_css_selector(
                                    'div.sellerDetails > p > strong').text not in exception_seller \
                                        and 'express' not in ELEMENT.find_element_by_css_selector(
                                    '.container img').get_attribute('alt'):
                                    best.append(
                                        (float(ELEMENT.find_element_by_css_selector(
                                            'span.sellingPrice > span > span.value').text),
                                         ELEMENT.find_element_by_css_selector('div.sellerDetails > p > strong').text,
                                         ELEMENT.find_element_by_css_selector('.container img').get_attribute('alt')))
                        except:
                            offer = self.driver.find_elements_by_css_selector('ul.offersList > li')
                            for ELEMENT in offer:
                                if offer.index(ELEMENT) == 4:
                                    break
                                if 'selectedItem' not in ELEMENT.get_attribute('class') \
                                        and ELEMENT.find_element_by_css_selector(
                                    'div.sellerDetails > p > strong').text not in exception_seller \
                                        and 'express' not in ELEMENT.find_element_by_css_selector(
                                    '.container img').get_attribute('alt'):
                                    best.append(
                                        (float(ELEMENT.find_element_by_css_selector(
                                            'span.sellingPrice > span > span.value').text),
                                         ELEMENT.find_element_by_css_selector('div.sellerDetails > p > strong').text,
                                         ELEMENT.find_element_by_css_selector('.container img').get_attribute('alt')))
                        best.sort()
                        if len(best) == 0:
                            new_price = max_pr
                            print('All except or express, price Maximum....... ')
                            Note = 'All except or express, price Maximum. '
                            static = 'Maximum price'
                        else:
                            y = float(best[0][0])
                            new_price = y - y * (avp / 100)
                            if new_price < price:
                                Note = 'Warning New price is smaller than you price i will keep your price or minimum' + str(
                                    new_price)
                                print(Note, new_price)
                                new_price = None if price > min_pr else min_pr

                            print('new price after +', new_price)
                            static = 'change price to +'

                elif seller in exception_seller or 'express' in express:
                    only = 'No'
                    static = 'Except'
                    new_price = max_pr
                    Note = 'Seller Except or Express Maximum price '
                    print(Note)
                else:
                    only = 'No'
                    static = 'change price to -'
                    y = float(price)
                    new_price = y - y * (avn / 100)
                    if new_price < min_pr:
                        Note = 'Warning New price is smaller than Minimum price i will keep minimum price'
                        print(Note)
                        new_price = min_pr
                        static = 'Minimum price'

                sleep(5)
            except Exception as e:
                seller = price = express = only = new_price = None
                Note = f'{e}'
                self.sheet[f'P{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                static = "not found"
            try:
                if new_price is not None:
                    # self.driver.get('https://login.noon.partners/en/')
                    self.driver.get('https://catalog.noon.partners/en-eg/catalog')
                    sleep(5)
                    if 'login.noon' in self.driver.current_url:
                        self.driver.find_element_by_name('email').send_keys(email)
                        self.driver.find_element_by_name('password').send_keys(password)
                        self.driver.find_element_by_css_selector('#formContainer > button').click()
                        sleep(5)
                        while True:
                            if 'login.noon' not in self.driver.current_url:
                                break
                        self.driver.get('https://catalog.noon.partners/en-eg/catalog')
                        sleep(4)
                    self.driver.find_element_by_css_selector('input.searchInput').send_keys(bar)
                    self.driver.find_element_by_css_selector('input.searchInput').send_keys(Keys.ENTER)
                    sleep(6)
                    try:
                        self.driver.find_element_by_css_selector('td:nth-child(1)').click()
                    except Exception as e:
                        print(e)
                        self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                        self.sheet[f'Q{i}'] = 'cant find'
                        print('cant find')
                    try:
                        sleep(4)
                        self.driver.find_element_by_name('price_eg').clear()
                        sleep(1)
                        self.driver.find_element_by_name('price_eg').send_keys(str(new_price))
                        sleep(2)
                        content = self.driver.find_element_by_css_selector('div.stockNet').text.split('\n')[1]
                        print(content)
                        self.sheet[f'O{i}'] = content
                        if int(content) < int(content_inp):
                            self.sheet[f'Q{i}'] = 'warning number is smell'+str(content)
                        self.driver.find_element_by_css_selector('div.fixedBottom > button').click()
                        sleep(5)
                    except Exception as e:
                        print(e)
                        self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                        self.sheet[f'Q{i}'] = 'cant change'
                        print('cant change')

            except Exception as e:
                print(e)
            for column in range(1, 18):
                try:
                    self.sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                except:
                    pass
            self.sheet[f'D{i}'] = seller
            self.sheet[f'E{i}'] = price
            if 'express' in express:
                self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
            self.sheet[f'F{i}'] = express
            self.sheet[f'G{i}'] = only
            self.sheet[f'H{i}'] = new_price
            self.sheet[f'I{i}'] = static
            self.sheet[f'P{i}'] = Note
            self.book.save(self.file_save)
            print(
                f'{i} - {bar} > seller : {seller}, price : {price}, new price : {new_price}, u only : {only}, static : {static}')
            #########################################
            if self.thr.p == 'pause' or self.thr.p == 'p':
                start = input('(put start)>>>')
                while True:
                    if start == 'start' or start == 's':
                        self.thr = Thread()
                        self.thr.start()
                        break
                    elif start == 'exit' or start == 'e':
                        self.driver.quit()
                        exit()
                    else:
                        start = input()
            elif self.thr.p == 'exit' or self.thr.p == 'e':
                self.driver.quit()
                exit()
            #########################################

        self.driver.quit()


if __name__ == "__main__":
    avrp = int(input('enter averdge + like 1'))
    avrn = int(input('enter averdge + like 1'))
    content_inp = int(input('enter content'))
    email = input('enter email')  # 'mohabamr157@gmail.com'
    password = input('enter pass')  # 'NewEra2020'
    file = input('name file like TEST.xlsx').strip()  # 'TEST.xlsx'  # str(input('Enter the name of file, For example( P.update-2.xlsx ) : ')).strip()
    while True:
        my_bot = SouCode(file_name=file)
        my_bot.Search()

# def None_re(self):
#     print('Please Wait While Detect None Value')
#     try:
#         self.book = load_workbook(file)
#         self.sheet = self.book.active
#     except Exception as e:
#         print(e)
#         exit()
#     for i in range(2, self.sheet.max_row):
#         if self.sheet[f'AA{i}'].value == 'Not Found' or \
#                 self.sheet[f'AB{i}'].value == 'Not Found' or self.sheet[f'AC{i}'].value == 'Not Found':
#             if self.sheet[f'F{i}'].value not in self.bar_code_re:
#                 self.bar_code_re.append((self.sheet[f'F{i}'].value, i))
#     print('Number of Non Found value is : ', len(self.bar_code_re))
#     if len(self.bar_code_re) != 0:
#         self.driver = webdriver.Chrome()
#         for elem in self.bar_code_re:
#             self.index = elem[1]
#             self.driver.get(f'{self.url}{elem[0]}/p/')
#             sleep(3)
#             r = HTML(html=self.driver.page_source)
#             try:
#                 price = r.find('span.sellingPrice span.value', first=True).text
#             except:
#                 price = None
#             try:
#                 seller = r.find('p.sellerName a', first=True).text
#             except:
#                 seller = None
#             try:
#                 express = r.find('div.bottomRow div.container img', first=True).attrs['alt']
#             except:
#                 express = None
#             print(self.index - 1, price, seller, express)
#             self.sheet[f'AA{self.index}'] = price
#             self.sheet[f'AB{self.index}'] = seller
#             self.sheet[f'AC{self.index}'] = express
#             self.book.save(self.file)
#
#         self.driver.quit()
# my_bot.driver.find_element_by_css_selector('span.cta').click()
# offer_number = len(my_bot.driver.find_elements_by_css_selector('ul.offersList > li > div.sellerDetails > p > strong'))
# best_seller_name = my_bot.driver.find_elements_by_css_selector('ul.offersList > li.selectedItem > div.sellerDetails > p > strong')
# best_price = my_bot.driver.find_element_by_css_selector('ul.offersList > li.selectedItem span.sellingPrice > span > span.value').text
# best_express = my_bot.driver.find_element_by_css_selector('ul.offersList > li.selectedItem .container img').get_attribute('alt')
# for element in my_bot.driver.find_elements_by_css_selector('ul.offersList > li'):
#     if 'selectedItem' not in element.get_attribute('class'):
#         print(element.find_element_by_css_selector('div.sellerDetails > p > strong').text,
#               element.find_element_by_css_selector('span.sellingPrice > span > span.value').text,
#               element.find_element_by_css_selector('.container img').get_attribute('alt'))

# try:
#     price = r.find('span.sellingPrice span.value', first=True).text
# except:
#     price = 'Not Found'
# try:
#     seller = r.find('p.sellerName a', first=True).text
# except:
#     seller = 'Not Found'
# try:
#     express = r.find('div.bottomRow div.container img', first=True).attrs['alt']
# except:
#     express = 'Not Found'
"""
        for i in range(2, end_cell):
            bar = self.sheet[f'A{i}'].value
            self.driver.get(f'{self.url}{bar}/p/')
            sleep(2)
            Offer_number = len(
                self.driver.find_elements_by_css_selector('ul.offersList > li > div.sellerDetails > p > strong'))
            if Offer_number == 0:
                only = 'Only'
                r = HTML(html=self.driver.page_source)
                price = float(r.find('span.sellingPrice span.value', first=True).text)
                seller = r.find('p.sellerName a', first=True).text
                express = r.find('div.bottomRow div.container img', first=True).attrs['alt']
                if 'Infinity' in seller:
                    static = 'Maximum Price'
                else:
                    only = 'you not found'
                    static = 'you not found'
            else:
                only = 'No'
                self.driver.find_element_by_css_selector('span.cta').click()
                sleep(3)
                seller = self.driver.find_element_by_css_selector(
                    'ul.offersList > li.selectedItem > div.sellerDetails > p > strong').text
                price = float(self.driver.find_element_by_css_selector(
                    'ul.offersList > li.selectedItem span.sellingPrice > span > span.value').text)
                express = self.driver.find_element_by_css_selector(
                    'ul.offersList > li.selectedItem .container img').get_attribute('alt')
                best = []
                me = None
                for ELEMENT in self.driver.find_elements_by_css_selector('ul.offersList > li'):
                    if 'selectedItem' not in ELEMENT.get_attribute('class'):

                        best.append((ELEMENT.find_element_by_css_selector('span.sellingPrice > span > span.value').text,
                                     ELEMENT.find_element_by_css_selector('div.sellerDetails > p > strong').text))
                        if me is None:
                            me = [(ELEMENT.find_element_by_css_selector('span.sellingPrice > span > span.value').text,
                                  ELEMENT.find_element_by_css_selector('div.sellerDetails > p > strong').text)] if 'nfinit' in ELEMENT.find_element_by_css_selector(
                                'div.sellerDetails > p > strong').text else None
                        # print(ELEMENT.find_element_by_css_selector('div.sellerDetails > p > strong').text,
                        #       ELEMENT.find_element_by_css_selector('span.sellingPrice > span > span.value').text,
                        #       ELEMENT.find_element_by_css_selector('.container img').get_attribute('alt'))
                best.sort()
                if 'Infinity' in seller:
                    static = 'Wait u best'
                    print(best)
                    print(f'price you is {price} other minimum price {best[0][0]}')
                elif me is not None:
                    if seller in exception_seller:
                        static = 'Except'
                    else:
                        static = 'u not best'
                        print(f'price is {price} you price is {me[0][0]}')
                        print(best)
                else:
                    static = '******************* you not found *****************'
                    print(seller, price)
                    print(best)

            print(i, '-', bar, '>', seller, price, only, static)
            self.sheet[f'D{i}'] = seller
            self.sheet[f'E{i}'] = price
            self.sheet[f'F{i}'] = express
            self.sheet[f'G{i}'] = only
            self.sheet[f'I{i}'] = static
            self.book.save(self.file_save)
"""
#
# driver = webdriver.Chrome()
# driver.get('https://login.noon.partners/en/')
# sleep(3)
# driver.find_element_by_name('email').send_keys('mohabamr157@gmail.com')
# driver.find_element_by_name('password').send_keys('NewEra2020')
# driver.find_element_by_css_selector('#formContainer > button').click()
# sleep(5)
# while True:
#     if 'login.noon' not in driver.current_url:
#         break
#
# driver.get('https://catalog.noon.partners/en-eg/catalog')
# sleep(3)
# driver.find_element_by_css_selector('input.searchInput').send_keys('N16521293A')
# driver.find_element_by_css_selector('input.searchInput').send_keys(Keys.ENTER)
# sleep(2)
# driver.find_elements_by_css_selector('td')[0].click()
# sleep(2)
#
# driver.find_element_by_name('price_eg').clear()
# driver.find_element_by_name('price_eg').send_keys('2020')
# driver.find_element_by_css_selector('div.fixedBottom > button').click()
