from selenium.webdriver.support import expected_conditions as ec
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from selenium import webdriver
from time import sleep


class Noon:
    def __init__(self):
        url = 'https://www.noon.com/saudi-en/'
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get(url)

    def Search(self):
        book = Workbook()
        sheet = book.active
        sheet.cell(1, 1).value = 'SKU'
        sheet.cell(1, 2).value = 'Price'
        sheet.cell(1, 3).value = 'New Price'
        # sheet.cell(1, 4).value = 'Old Price'

        row = 2
        l = []

        try:
            WebDriverWait(self.driver, 20).until((ec.visibility_of_element_located(
                (By.CSS_SELECTOR, 'header > div > div.container > div:nth-child(2) > div > button')))).click()
            sleep(2)
            self.driver.find_elements_by_css_selector('div.popDown > ul > li')[-1].click()
        except:
            pass
        while True:
            for element in WebDriverWait(self.driver, 20).until(
                    ec.visibility_of_all_elements_located((By.CSS_SELECTOR,
                                                           '.productListContainer > div > div.productListWrapper > div.productList.gridView > div > div > a.product'))):

                if element.get_attribute('href') not in l:
                    url = element.get_attribute('href')
                    l.append(url)
                    bar = \
                    url.replace('https://www.noon.com', '').replace('/saudi-en/', '').replace('/egypt-en/', '').replace(
                        '/saudi-ar/', '').replace('/egypt-ar/', '').split('/')[1]

                    price = element.find_element_by_css_selector('div.priceRow > p > span > span > span.value').text
                    # try:
                    #     price_2 = \
                    #         element.find_elements_by_css_selector('div.priceRow > p > span > span > span.value')[1].text
                    # except:
                    #     price_2 = None
                    new_price = float(price) + vol
                    sheet.cell(row, 1).value = bar
                    sheet.cell(row, 2).value = price
                    sheet.cell(row, 3).value = new_price
                    # sheet.cell(row, 4).value = price_2
                    # sheet.cell(row, 5).value = url

                    print('- ', row, ' - ', bar)
                    row += 1
            book.save('products_search.xlsx')
            try:
                if self.driver.find_element_by_css_selector('.nextLink').get_attribute('aria-disabled') != 'true':
                    self.driver.find_element_by_css_selector('.nextLink').click()
                    sleep(3)
                else:
                    break
            except:
                break
        self.driver.quit()


if __name__ == "__main__":
    vol = float(input('- Enter value (+) '))
    m = Noon()
    input('- Enter : ')
    m.Search()
