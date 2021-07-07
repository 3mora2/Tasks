from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from time import sleep
from random import randrange
import uuid, os
import traceback


class main:
    def __init__(self):
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.cell(1, 1).value = 'Title'
        self.sheet.cell(1, 2).value = 'Price'
        self.sheet.cell(1, 3).value = 'Quantity'
        self.sheet.cell(1, 4).value = 'available'
        self.sheet.cell(1, 5).value = 'url'

        self.driver = webdriver.Chrome()
        self.driver.get('https://sa.iherb.com/')

    def next_page(self):
        try:
            url = self.driver.find_element_by_css_selector('.pagination-next').get_attribute('href')
            self.get(url)

            return True
        except:
            return False

    def get(self, url):
        try:
            self.driver.get(url)
            sleep(4)
        except:
            print(traceback.print_exc())

    def language(self):
        if 'EN' not in self.driver.find_element_by_css_selector('div.language-select').text:
            self.driver.find_element_by_css_selector('div.language-select').click()
            sleep(2)
            self.driver.find_element_by_css_selector(
                'div.language-menu.language-menu-universal div.select-language.ui.fluid.search.selection.dropdown').click()
            sleep(1)
            self.driver.find_element_by_css_selector(
                'div.select-language.ui.fluid.search.selection.dropdown > div.menu.search-list.open > div:nth-child(2)').click()
            sleep(4)
            self.driver.find_element_by_css_selector(
                'div.language-menu.language-menu-universal div.ccl-btn > button.save-selection').click()
            sleep(2)
            self.driver.find_element_by_css_selector(
                'div.panel-content.border.show-page.text-right > div > div > select').send_keys('48')
            sleep(2)

    def save(self, ele, cont):
        try:
            if 'InStock' in ele.find_element_by_css_selector('meta+link').get_attribute('href'):
                available = True
            else:
                available = False

            title = ele.find_element_by_css_selector('div.absolute-link-wrapper > a').get_attribute('title')
            href = ele.find_element_by_css_selector('div.absolute-link-wrapper > a').get_attribute('href')
            price = ele.find_element_by_css_selector('meta[itemprop="price"]').get_attribute('content')
            price = float(price) + float(price) * (float(pres) / 100)

            self.sheet.cell(cont, 1).value = title
            self.sheet.cell(cont, 2).value = price
            self.sheet.cell(cont, 3).value = quantity
            self.sheet.cell(cont, 4).value = available
            # self.sheet.cell(cont, 5).value = available
            self.sheet.cell(cont, 5).value = href
        except:
            print(traceback.print_exc())

    def extract(self):
        cont = 2
        next_pag = True
        self.language()
        while next_pag:
            for ele in self.driver.find_elements_by_css_selector('div.product-inner.product-inner-wide'):
                try:
                    if 'InStock' not in ele.find_element_by_css_selector('meta+link').get_attribute('href'):
                        continue

                    self.save(ele, cont)
                    print(f'- {cont - 1}')
                except:
                    print(traceback.print_exc())
                cont += 1

            sleep(3)

            self.book.save('sa-iherb.xlsx')
            next_pag = self.next_page()

        self.driver.quit()


class Upload:
    driver: WebDriver
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-sa/noon-catalog'
    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'a[href*="/en-sa/noon-catalog/preview/"]')

    SKU_TEXT_SELECTOR = (By.CSS_SELECTOR, 'div.tabsWrapper > div')
    SKU_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[name="partner_sku"]')
    SKU_BUTTON_SELECTOR = (By.CSS_SELECTOR, '.solid')

    PRICE_MAX_SELECTOR = (
        By.XPATH,
        "//div[contains(@class, 'inputInfo') and contains(text(), 'Noon SKU Price')] / *[@class='priceRange']")
    PRICE_INPUT_SELECTOR = (By.NAME, 'price')

    QUANTITY_SITE_SELECTOR = (By.CSS_SELECTOR, 'div.placeholder')
    QUANTITY_SITE_CHOSE_SELECTOR = (By.CSS_SELECTOR, 'div.active > div > div.countryLabel')
    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')

    BARCODE_INSIDE_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="text"][placeholder="Enter Barcode"]')
    BUTTON_ADD_BARCODE_INSIDE_SELECTOR = (By.CSS_SELECTOR, 'div.eachInput > div > div.solid')
    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="active"]\').checked == false){document.querySelector(\'input[id="active"]\').click()};'

    SAVE_CHANGE_SELECTOR = (By.CSS_SELECTOR, 'div.transparent ~ div.solid')
    SUBMIT_SELECTOR = (
    By.XPATH, '//div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')

    def __init__(self):
        self.book = load_workbook(file_name)
        self.sheet = self.book.active
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.login()

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def add_new(self, name):
        try:
            self.driver.get(self.URL_CAT)
            sleep(2)
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))
                                                 ).send_keys(name, Keys.ENTER)
            sleep(2)
        except:
            print(traceback.print_exc())

    def add_serial(self):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
        except:
            print('- Not found')
            return 'False'
        sleep(1)
        try:
            serial = \
                WebDriverWait(self.driver, 5).until(
                    ec.visibility_of_element_located(self.SKU_TEXT_SELECTOR)).text.split('\n')[-1].replace('N', '').replace('A', '') + str(randrange(9990, 999999))

            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_INPUT_SELECTOR)).send_keys(serial)
            sleep(3)
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_BUTTON_SELECTOR)).click()
            return serial
        except:
            print('- Exist, Continue')
            return False

    def add_price(self, price):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a', Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(str(price))
        except:
            print(traceback.print_exc())
            print('- Cant Change Price')

    def add_qu(self, q):
        sleep(1)
        try:
            WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
        except:
            try:
                WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(self.QUANTITY_SITE_SELECTOR))).click()
                sleep(1)
                WebDriverWait(self.driver, 6).until((ec.visibility_of_all_elements_located(self.QUANTITY_SITE_CHOSE_SELECTOR)))[site].click()
                sleep(1)
                WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
                sleep(1)
            except:
                print(traceback.print_exc())
                print('- Cant Change quantity')

    def add_bar(self, sk):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.BARCODE_INSIDE_INPUT_SELECTOR)).send_keys(sk)
            sleep(1)
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.BUTTON_ADD_BARCODE_INSIDE_SELECTOR)).click()
            sleep(1)
        except:
            print(traceback.print_exc())
            print('- Cant Change Barcode')

    def add_active(self):
        sleep(1)
        try:
            self.driver.execute_script(self.ACTIVE_SCRIPT)
        except:
            print(traceback.print_exc())

    def upload(self):

        for i in range(2, self.sheet.max_row + 1):
            try:
                name = self.sheet.cell(i, 1).value
                price = self.sheet.cell(i, 2).value
                Quantity = self.sheet.cell(i, 3).value
                print('- ', name, price)
                self.add_new(name)
                res = self.add_serial()
                if res == 'False':
                    print('- Try remove ')
                    # name = name[:name.rfind(',')]
                    try:
                        name = ','.join(name.split(',')[:-end_word])
                        print('- ', name, price)
                        self.add_new(name)
                        res = self.add_serial()
                    except:
                        print(traceback.print_exc())
                if res is False or res == 'False':
                    continue
                sleep(1)
                if price is not None:
                    try:
                        pp = WebDriverWait(self.driver, 4).until((ec.visibility_of_element_located(
                            self.PRICE_MAX_SELECTOR))).text.replace('SAR', '').replace(',', '').strip().split('-')
                        p_n = float(pp[0].strip())
                        p_m = float(pp[1].strip())

                        print(f'- Max noon : {p_m}, Minimum Noon : {p_n}, Your New Price {price}')

                        # if new price small than noon mine, new price = noon min price
                        if float(price) < float(p_n):
                            price = p_n
                            print(f'- Edit To Noon Min {price}')

                        # if new price > noon max price, new price = noon max price
                        elif float(price) > float(p_m):
                            price = p_m
                            print(f'- Edit To Noon Max {price}')

                    except TimeoutError:
                        pass
                    except:
                        print(traceback.print_exc())

                    self.add_price(price)
                sleep(2)
                self.add_active()
                if Quantity is not None:
                    self.add_qu(Quantity)
                sleep(1)
                self.add_bar(res)
                sleep(2)
                try:
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SAVE_CHANGE_SELECTOR))).click()
                    sleep(1)
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                    sleep(1)
                except:
                    print(traceback.print_exc())
                    print('Cant Save')
                sleep(2)
            except:
                print(traceback.print_exc())


def add_mac():
    try:
        os.remove(str(uuid.getnode()))
    except:
        pass
    try:
        with open(str(uuid.getnode()), 'w+') as file:
            file.write(str(uuid.getnode()))
        os.popen('attrib +h ' + str(uuid.getnode()))
        return True
    except:
        print('delete old file')
        return False


def check_mac():
    try:
        with open(str(uuid.getnode()), 'r') as file:
            text = file.readlines()
        if str(uuid.getnode()) in text:
            return True
        else:
            print('ref')
            return False
    except FileNotFoundError:
        print('ref')
        return False


if __name__ == "__main__":
    pr = check_mac()
    while True:
        ress = str(input('''- Enter 1 ---> Extract   or 2 ----> Upload  or 3 ---> All '''))
        if ress == '1' and pr is True:
            quantity = float(input('- Enter Quint : '))
            pres = float(input('- Enter prc : '))
            ex = main()
            input('- Enter : ')
            ex.extract()
        elif ress == '2' and pr is True:

            while True:
                try:
                    file_name = input('- Enter file name : ')  # 'Book1.xlsx'
                    book = load_workbook(file_name)
                    book.close()
                    break
                except FileNotFoundError:
                    print('- No such file or directory')
                except Exception as e:
                    print(e)

            email = input('- Enter your email : ')
            password = input('- Enter your password : ')
            site = input('- Enter Number countryLabel : ')
            try:
                site = int(site) - 1
                if site >= 0:
                    pass
                else:
                    site = 0
            except:
                site = 0

            end_word = input('- Enter Number Label : ')
            try:
                end_word = int(end_word)
                if end_word > 0:
                    pass
                else:
                    end_word = 1
            except:
                end_word = 1

            print(site)
            u = Upload()
            u.upload()
        elif ress == '3' and pr is True:
            quantity = float(input('- Enter Quint : '))
            pres = float(input('- Enter prc : '))
            email = input('- Enter your email : ')
            password = input('- Enter your password : ')
            site = input('- Enter Number countryLabel : ')

            try:
                site = int(site) - 1
                if site >= 0:
                    pass
                else:
                    site = 0
            except:
                site = 0
            print(site)
            end_word = input('- Enter Number Label : ')
            try:
                end_word = int(end_word)
                if end_word > 0:
                    pass
                else:
                    end_word = 1
            except:
                end_word = 1

            ex = main()
            input('- Enter : ')
            ex.extract()
            file_name = 'sa-iherb.xlsx'
            u = Upload()
            u.upload()
        elif ress == 'scrt*python*add*mac*address':
            pr = add_mac()
