from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
from random import randrange
import uuid, os
import traceback

'''
* UPDATE ALL FIND_ELEMENT
'''


class Upload:
    driver: WebDriver
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-sa/noon-catalog'
    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'a[href*="/en-sa/noon-catalog/preview/"]')

    SKU_TEXT_SELECTOR = (By.CSS_SELECTOR, 'div.tabsWrapper > div')
    SKU_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[name="partner_sku"]')
    SKU_BUTTON_SELECTOR = (By.CSS_SELECTOR, '.solid')

    PRICE_INPUT_SELECTOR = (By.NAME, 'price')

    QUANTITY_SITE_SELECTOR = (By.CSS_SELECTOR, 'div.placeholder')
    QUANTITY_SITE_CHOSE_SELECTOR = (By.CSS_SELECTOR, 'div.active > div > div.countryLabel')
    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')

    BARCODE_INSIDE_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="text"][placeholder="Enter Barcode"]')
    BUTTON_ADD_BARCODE_INSIDE_SELECTOR = (By.CSS_SELECTOR, 'div.eachInput > div > div.solid')
    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="active"]\').checked == false){document.querySelector(\'input[id="active"]\').click()};'

    SAVE_CHANGE_SELECTOR = (By.CSS_SELECTOR, 'div.transparent ~ div.solid')
    SUBMIT_SELECTOR = (By.XPATH, '//div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')
    # SUBMIT_SELECTOR = (By.CSS_SELECTOR, 'div.showAlert > div.btnWrapper > div.jcTkjm.solid')

    def __init__(self):
        self.book = load_workbook(file_name)
        self.sheet = self.book.active
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.login()

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def add_new(self, name):
        sleep(1)
        try:
            self.driver.get(self.URL_CAT)
            sleep(2)
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))).send_keys(name, Keys.ENTER)
            sleep(2)
        except:
            print(traceback.print_exc())

    def add_serial(self):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
        except:
            print('- Not found')
            return 'False'
        try:
            serial = \
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_TEXT_SELECTOR)).text.split('\n')[
                -1].replace('N', '').replace('A', '') + str(randrange(9990, 999999))

            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_INPUT_SELECTOR)).send_keys(serial)
            sleep(3)
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_BUTTON_SELECTOR)).click()

            return serial
        except:
            print('- Continue')
            return False

    def add_price(self, price):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a', Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a', Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(str(price))
        except:
            print(traceback.print_exc())
            print('- Cant Change Price')

    def add_qu(self, q):
        sleep(1)
        try:
            WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(self.QUANTITY_SITE_SELECTOR))).click()
            sleep(1)
            WebDriverWait(self.driver, 6).until((ec.visibility_of_all_elements_located(self.QUANTITY_SITE_CHOSE_SELECTOR)))[site].click()
            sleep(1)
            WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
            sleep(1)
        except:
            print(traceback.print_exc())
            print('- Cant Change quantity')

    def add_bar(self, sk):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.BARCODE_INSIDE_INPUT_SELECTOR)).send_keys(sk)
            sleep(1)
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.BUTTON_ADD_BARCODE_INSIDE_SELECTOR)).click()
            sleep(1)
        except:
            print(traceback.print_exc())
            print('- Cant Change Barcode')

    def add_active(self):
        sleep(1)
        try:
            self.driver.execute_script(self.ACTIVE_SCRIPT)
        except:
            print(traceback.print_exc())
            print('Cant active')

    def upload(self):
        for i in range(2, self.sheet.max_row+1):
            try:
                bar = self.sheet.cell(i, 1).value
                price = self.sheet.cell(i, 2).value
                quantity = self.sheet.cell(i, 3).value
                serial = self.sheet.cell(i, 4).value
                if bar is None:
                    continue
                print('- ', bar, price)
                self.add_new(bar)
                sleep(1)
                res = self.add_serial()
                if res == 'False':
                    continue
                sleep(1)
                if price is not None:
                    self.add_price(price)
                sleep(1)
                self.add_active()
                sleep(1)
                if serial is not None:
                    self.add_bar(serial)
                sleep(1)
                if res is not False:
                    self.add_bar(res)
                sleep(1)
                if quantity is not None:
                    self.add_qu(quantity)
                try:
                    sleep(1)
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SAVE_CHANGE_SELECTOR))).click()
                    sleep(1)
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                    sleep(1)
                except:
                    print(traceback.print_exc())
            except:
                print(traceback.print_exc())


def add_mac():
    try:
        os.remove(str(uuid.getnode()))
    except:
        pass
    try:
        with open(str(uuid.getnode()), 'w+') as file:
            file.write(str(uuid.getnode()))
        os.popen('attrib +h ' + str(uuid.getnode()))
        return True
    except:
        print('delete old file')
        return False


def check_mac():
    try:
        with open(str(uuid.getnode()), 'r') as file:
            text = file.readlines()
        if str(uuid.getnode()) in text:
            return True
        else:
            print('ref')
            return False
    except FileNotFoundError:
        print('ref')
        return False


if __name__ == "__main__":
    file_name = None
    pr = check_mac()
    while True:
        try:
            file_name = input('- Enter file name : ')
            with open(file_name)as f:
                pass
            break
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)

        if file_name == 'scrt*python*add*mac*address':
            pr = add_mac()

    email = input('- Enter your email : ')
    password = input('- Enter your password : ')
    site = input('- Enter Number countryLabel : ')

    try:
        site = int(site)-1
        if site >= 0:
            pass
        else:
            site = 0
    except:
        site = 0
    print(site)

    if pr:
        u = Upload()
        u.upload()
