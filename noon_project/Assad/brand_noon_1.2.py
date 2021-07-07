from selenium.webdriver.support import expected_conditions as ec
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from selenium import webdriver
from time import sleep
import re


class Noon:
    def __init__(self):
        url = 'https://www.noon.com/saudi-en/'
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get(url)

    def Search(self):
        book = Workbook()
        sheet = book.active
        sheet.cell(1, 1).value = 'SKU'
        sheet.cell(1, 2).value = 'Price'
        sheet.cell(1, 3).value = 'New Price'
        sheet.cell(1, 4).value = 'Name'

        row = 2
        l = []

        try:
            WebDriverWait(self.driver, 20).until((ec.visibility_of_all_elements_located(
                (By.CSS_SELECTOR, 'span + div > span'))))[-1].click()
            sleep(2)
            self.driver.find_element_by_css_selector('li[data-value="150"]').click()
            sleep(3)
        except:
            pass

        while True:

            for element in WebDriverWait(self.driver, 20).until(
                    ec.visibility_of_all_elements_located((By.CSS_SELECTOR, 'div.productContainer > a'))):
                try:
                    if element.get_attribute('href') not in l:
                        url = element.get_attribute('href')
                        l.append(url)
                        bar = re.findall('N[0-9]{7,}A', url)[0]
                        # bar = url.replace('https://www.noon.com/saudi-en/', '').replace('/saudi-en/', '').replace(
                        #     '/saudi-ar/', '').split('/')[1]

                        nam = element.find_element_by_css_selector('div.grid > div:nth-child(2) > div:nth-child(1) > div')
                        try:
                            name = nam.text.replace(nam.find_element_by_css_selector('span').text, '')
                        except:
                            name = nam.text

                        price = element.find_element_by_css_selector('span.currency + strong').text
                        new_price = float(price) + vol
                        sheet.cell(row, 1).value = bar
                        sheet.cell(row, 2).value = price
                        sheet.cell(row, 3).value = new_price
                        sheet.cell(row, 4).value = name
                        # sheet.cell(row, 4).value = price_2
                        # sheet.cell(row, 5).value = url

                        print('- ', row-1, ' - ', bar)
                        row += 1
                except Exception as e:
                    print(e)

            book.save('products_search.xlsx')
            try:
                if 'disabled' not in self.driver.find_element_by_css_selector('.next').get_attribute('class'):
                    self.driver.find_element_by_css_selector('.next').click()
                    sleep(4)
                else:
                    break
            except:
                break
        self.driver.quit()


if __name__ == "__main__":
    vol = float(input('- Enter value (+) : '))
    m = Noon()
    input('- Enter : ')
    m.Search()
#https://www.noon.com/saudi-en/p-21099?limit=150
