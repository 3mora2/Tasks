from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import *
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from time import sleep
import threading

'''
* 2020 - 10 -25
* Change price and quantity
* Noon SA
'''


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class SouCode:
    driver: WebDriver

    def __init__(self):
        while True:
            try:
                file_name = input('- Enter file name : ')
                self.book = load_workbook(file_name)
                self.sheet = self.book.active
                break
            except FileNotFoundError:
                print('- No such file or directory')
            except Exception as e:
                print(e)

        self.email = input('- Enter your email : ')
        self.password = input('- Enter your password : ')
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.login()

    def login(self):
        try:
            self.driver.get('https://catalog.noon.partners/en-sa/catalog')
            sleep(5)
            if 'login.noon' in self.driver.current_url:
                self.driver.find_element_by_name('email').send_keys(self.email)
                self.driver.find_element_by_name('password').send_keys(self.password)
                self.driver.find_element_by_css_selector('#formContainer > button').click()
                sleep(5)
                while True:
                    if 'login.noon' not in self.driver.current_url:
                        break

        except Exception as e:
            print(e)
            print('Login Error')

    def Search(self):
        end_cell = self.sheet.max_row

        thread = Thread()
        thread.start()

        for i in range(2, end_cell + 1):
            bar = self.sheet[f'A{i}'].value
            price = self.sheet[f'B{i}'].value
            quantity = self.sheet[f'C{i}'].value
            sk = self.sheet[f'D{i}'].value

            if bar is None or price is None:
                continue
            print(f'- {i} - {bar} > price : {price}')

            try:
                self.driver.get('https://catalog.noon.partners/en-sa/catalog')

                WebDriverWait(self.driver, 8).until((ec.visibility_of_element_located(
                    (By.CSS_SELECTOR, 'input.searchInput')))).send_keys(bar)
                self.driver.find_element_by_css_selector('input.searchInput').send_keys(Keys.ENTER)
            except Exception as e:
                print(e)
                continue

            try:
                WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(
                    (By.CSS_SELECTOR, 'td:nth-child(1) > div > a')))).click()
                # self.driver.find_element_by_css_selector('td:nth-child(1) > div > a').click()
            except Exception as e:
                print(e)
                self.sheet[f'E{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                self.sheet[f'E{i}'] = 'Cant find no click'
                print('- no click')
                continue
            try:
                WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(
                    (By.NAME, 'price_sa')))).clear()
                WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(
                    (By.NAME, 'price_sa')))).clear()

                sleep(.5)
                self.driver.find_element_by_name('price_sa').send_keys(str(price))
            except:
                self.sheet[f'E{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                self.sheet[f'E{i}'] = '- Cant Change Price'
                print('- Cant Change Price')

            try:
                if quantity is not None:
                    WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(
                        (By.NAME, 'quantity')))).send_keys(str(quantity))
                    self.driver.find_element_by_css_selector('div.fixedBottom > button').click()
            except:
                self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                self.sheet[f'F{i}'] = '- Cant Change quantity'
                print('- Cant Change quantity')

            try:
                self.driver.execute_script(
                    'if (document.querySelector(\'input[name="is_active"]\').checked == false){document.querySelector(\'input[name="is_active"]\').click()};')
            except:
                self.sheet[f'J{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                self.sheet[f'J{i}'] = '- Cant Change Active'

            try:
                self.driver.find_element_by_css_selector('div.inputCtr > input').send_keys(sk)
                sleep(2)
                try:
                    self.driver.find_element_by_css_selector('div.inputCtr > div.btnCtr > button').click()
                except ElementClickInterceptedException:
                    _ = self.driver.find_element_by_css_selector(
                        'div.inputCtr > div.btnCtr').location_once_scrolled_into_view
                    sleep(2)
                    self.driver.find_element_by_css_selector('div.inputCtr > div.btnCtr > button').click()
                    sleep(1)

            except:
                self.sheet[f'I{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                self.sheet[f'I{i}'] = '- Cant Change Barcode'

            self.driver.find_element_by_css_selector('div.fixedBottom > button').click()
            sleep(2)

            self.book.save('final.xlsx')

            #########################################
            if thread.p == 'pause' or thread.p == 'p':
                start = input('(Type start or s and press Enter)>>>')
                while True:
                    if start == 'start' or start == 's':
                        thread = Thread()
                        thread.start()
                        break
                    elif start == 'exit' or start == 'e':
                        self.driver.quit()
                        exit()
                    else:
                        start = input()
            elif thread.p == 'exit' or thread.p == 'e':
                self.driver.quit()
                exit()

        self.driver.quit()


if __name__ == "__main__":
    my_bot = SouCode()
    my_bot.Search()

'''
import os    
from chardet import detect

# get file encoding type
def get_encoding_type(file):
    with open(file, 'rb') as f:
        rawdata = f.read()
    return detect(rawdata)['encoding']
'''