from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from time import sleep
from datetime import datetime
import threading
import uuid, os
import traceback


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class SouCode:
    driver: WebDriver
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-sa/catalog'
    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'tr:nth-child(1) > td > a')

    PRICE_MAX_SELECTOR = (
    By.XPATH, "//div[contains(@class, 'inputInfo') and contains(text(), 'Noon SKU Price')] / *[@class='priceRange']")
    PRICE_INPUT_SELECTOR = (By.NAME, 'price')

    QUANTITY_VALUE_SELECTOR = (By.CSS_SELECTOR, 'div.WhStockNet > div > div.WhItemValue')
    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')

    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="active"]\').checked == false){document.querySelector(\'input[id="active"]\').click()};'

    SAVE_CHANGE_SELECTOR = (By.CSS_SELECTOR, 'div.transparent ~ div.solid')
    SUBMIT_SELECTOR = (By.XPATH, '//div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')
    #############################################
    URL_SEARCH = 'https://www.noon.com/saudi-en/'
    RATE_SELECTOR = (By.CSS_SELECTOR, 'div.detail_percentage')
    BRAND_SELECTOR = (By.CSS_SELECTOR, 'a > div.MCFyV')
    PRICE_SCRIPT = "return parseFloat(document.querySelector('.priceNow').textContent.replace('SAR',''));"
    SELLER_TEXT_SELECTOR = (By.CSS_SELECTOR, 'a.storeLink')
    EXPRESS_SELECTOR = (By.CSS_SELECTOR, 'div.estimator_right > div > img')
    ALL_OFFER_SELECTOR = (By.CSS_SELECTOR, '.allOffers')
    S_PRICE_SELECTOR = (By.CSS_SELECTOR, 'div.regular > strong')
    S_SELLER_SELECTOR = (By.CSS_SELECTOR, '.soldby > strong')
    S_EXPRESS_SELECTOR = (By.CSS_SELECTOR, 'div[direction="row"] > div > img')

    def __init__(self, file_name):
        self.file_save = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + file_name

        try:
            self.book = load_workbook(file_name)
            self.sheet = self.book.active
            self.Load_File()
        except Exception as err:
            print(err)
            exit()
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.login()

    def Load_File(self):
        try:
            self.sheet[f'A1'] = 'Noon_SKU'
            self.sheet[f'B1'] = 'Minimum Price'
            self.sheet[f'C1'] = 'Maximum Price'
            self.sheet[f'D1'] = 'Seller Name'
            self.sheet[f'E1'] = 'Price'
            self.sheet[f'F1'] = 'Express'
            self.sheet[f'G1'] = 'Only'
            self.sheet[f'H1'] = 'New price'
            self.sheet[f'I1'] = 'Status'
            self.sheet[f'J1'] = 'Brand'
            self.sheet[f'K1'] = 'Account'
            # self.sheet[f'L1'] = 'Change Value +'
            self.sheet[f'L1'] = 'Current Price'
            self.sheet[f'M1'] = 'Change Value -'
            self.sheet[f'N1'] = 'Exception'
            self.sheet[f'O1'] = 'Quantity'
            self.sheet[f'P1'] = 'Note 1'
            self.sheet[f'Q1'] = 'Note 2'
            self.sheet.column_dimensions['A'].width = 14
            self.sheet.column_dimensions['B'].width = 15
            self.sheet.column_dimensions['C'].width = 15
            self.sheet.column_dimensions['D'].width = 15
            self.sheet.column_dimensions['E'].width = 10
            self.sheet.column_dimensions['F'].width = 14
            self.sheet.column_dimensions['G'].width = 10
            self.sheet.column_dimensions['H'].width = 11
            self.sheet.column_dimensions['I'].width = 18
            self.sheet.column_dimensions['J'].width = 30
            self.sheet.column_dimensions['K'].width = 20
            self.sheet.column_dimensions['L'].width = 16
            self.sheet.column_dimensions['M'].width = 16
            self.sheet.column_dimensions['N'].width = 16
            self.sheet.column_dimensions['O'].width = 13
            self.sheet.column_dimensions['P'].width = 85
            self.sheet.column_dimensions['Q'].width = 50
        except:
            pass

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def search_bar(self, code):
        try:
            self.driver.get(self.URL_CAT)
            sleep(1)
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))).send_keys(code, Keys.ENTER)
            sleep(2)
        except:
            print(traceback.print_exc())
            return False
        try:
            sleep(1)
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
            sleep(3)
            return True
        except:
            print(traceback.print_exc())
            print('- Not found')
            return False

    def change_price(self, price):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a',
                                                                                         Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a',
                                                                                         Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(str(price))
            sleep(1)
        except:
            print(traceback.print_exc())
            print('- Cant Change Price')

    def change_quantity(self, i):
        sleep(1)
        try:
            content = WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_VALUE_SELECTOR))).text
            print(content)
            self.sheet[f'O{i}'] = content
            if int(content) < int(content_inp):
                if int(content) == 0:
                    self.sheet[f'Q{i}'] = 'Warning: Quantity is ' + str(content)
                else:
                    WebDriverWait(self.driver, 6).until(
                        (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(content_new))
                    print('Change Quantity to ', content_new)

                    self.sheet[f'Q{i}'] = 'Warning: Quantity is ' + str(content)
                    self.sheet[f'Q{i}'] = 'Change Quantity to ' + str(content_new)
                try:
                    self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                    self.sheet[f'O{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                except:
                    pass
        except:
            print(traceback.print_exc())
            print('- Cant Change quantity')

    def add_active(self):
        sleep(1)
        try:
            self.driver.execute_script(self.ACTIVE_SCRIPT)
        except:
            print(traceback.print_exc())
            print('Cant active')

    def Search(self):
        Note = None
        self.sheet['K2'] = str(email)
        self.sheet['K4'] = str(main_seller)
        self.sheet['K3'] = str(password)

        end_cell = self.sheet.max_row

        exception_seller = [self.sheet[f'N{i}'].value for i in range(2, end_cell+1) if
                            self.sheet[f'N{i}'].value is not None]

        group1 = [self.sheet[f'R{i}'].value for i in range(2, end_cell+1) if self.sheet[f'R{i}'].value is not None]
        group2 = [self.sheet[f'S{i}'].value for i in range(2, end_cell+1) if self.sheet[f'S{i}'].value is not None]
        group3 = [self.sheet[f'T{i}'].value for i in range(2, end_cell+1) if self.sheet[f'T{i}'].value is not None]
        group4 = [self.sheet[f'U{i}'].value for i in range(2, end_cell+1) if self.sheet[f'U{i}'].value is not None]
        group5 = [self.sheet[f'V{i}'].value for i in range(2, end_cell+1) if self.sheet[f'V{i}'].value is not None]

        brand_group = {}
        for i in range(2, end_cell+1):
            if self.sheet[f'X{i}'].value is not None and self.sheet[f'Y{i}'].value is not None:
                brand_group[self.sheet[f'X{i}'].value] = self.sheet[f'Y{i}'].value

        # New
        brand_exception = [self.sheet[f'W{i}'].value for i in range(2, end_cell) if self.sheet[f'V{i}'].value is not None]

        thr = Thread()
        thr.start()

        for i in range(2, end_cell + 1):
            bar = self.sheet[f'A{i}'].value
            if bar is None:
                continue

            max_pr = float(self.sheet[f'C{i}'].value)
            min_pr = float(self.sheet[f'B{i}'].value)
            url = f'{self.URL_SEARCH}{bar}/p/'
            try:
                self.driver.get(url)
                sleep(2)
                price = float(self.driver.execute_script(self.PRICE_SCRIPT))
                seller = WebDriverWait(self.driver, 3).until(
                    (ec.visibility_of_element_located(self.SELLER_TEXT_SELECTOR))).text
                try:
                    express = WebDriverWait(self.driver, 3).until(
                        (ec.visibility_of_element_located(self.EXPRESS_SELECTOR))).get_attribute('alt')
                except:
                    express = None

                try:
                    brand = WebDriverWait(self.driver, 4).until((ec.visibility_of_element_located(self.BRAND_SELECTOR))).text
                    self.sheet[f'J{i}'] = brand
                except:
                    brand = None

                if brand in brand_exception:
                    self.sheet[f'D{i}'] = seller
                    self.sheet[f'E{i}'] = price
                    try:
                        self.sheet[f'I{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
                        if 'express' in express:
                            self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')

                    except:
                        pass
                    self.sheet[f'F{i}'] = express
                    self.sheet[f'I{i}'] = 'Exception Brand'
                    print(f'- {i} - {bar} > seller : {seller}, price : {price}, static : Exception Brand')
                    continue

                if seller in group1:
                    avp = avn = float(group1[0])
                elif seller in group2:
                    avp = avn = float(group2[0])
                elif seller in group3:
                    avp = avn = float(group3[0])
                elif seller in group4:
                    avp = avn = float(group4[0])
                elif seller in group5:
                    avp = avn = float(group5[0])

                elif brand in brand_group.keys():
                    avp = avn = float(brand_group[brand])

                else:
                    avp = avrpd
                    avn = avrnd

                try:
                    Offer_number = len(WebDriverWait(self.driver, 1).until((ec.visibility_of_all_elements_located(self.ALL_OFFER_SELECTOR))))
                except:
                    Offer_number = 0

                ########################################################################################################
                if main_seller in seller:
                    if Offer_number == 0:
                        new_price = max_pr
                        only = 'Only'
                        static = 'Maximum Price'
                        print(' Only me. So, maximum price is ', max_pr)
                        Note = ' Only me. So, maximum price is ' + str(max_pr)
                        self.sheet[f'P{i}'] = Note
                    else:
                        only = 'No'
                        WebDriverWait(self.driver, 2).until(
                            (ec.visibility_of_element_located(self.ALL_OFFER_SELECTOR))).click()
                        sleep(4)
                        s_seller = WebDriverWait(self.driver, 2).until(
                            (ec.visibility_of_all_elements_located(self.S_SELLER_SELECTOR)))[1].text.split(' ')[-1]
                        s_price = WebDriverWait(self.driver, 2).until(
                            (ec.visibility_of_all_elements_located(self.S_PRICE_SELECTOR)))[1].text
                        s_express = WebDriverWait(self.driver, 2).until(
                            (ec.visibility_of_all_elements_located(self.S_EXPRESS_SELECTOR)))[1].get_attribute('alt')
                        print(s_seller, s_price, s_express)

                        if s_seller in exception_seller:
                            new_price = None
                            print(' Exception. So, maximum price. ')
                            Note = ' Exception. So, maximum price. '
                            self.sheet[f'P{i}'] = Note
                            static = 'Maximum price'
                        elif float(s_price) < min_pr:
                            new_price = None
                            Note = 'next price < Minimum'
                            self.sheet[f'P{i}'] = Note
                            static = 'freeze'
                        else:
                            y = float(s_price)
                            new_price = y - avp
                            if new_price < price:
                                Note = ' Warning: New price is smaller than yours. So, price not change. '
                                self.sheet[f'P{i}'] = Note
                                print(Note)
                                new_price = None
                                static = 'freeze'
                            else:
                                print('New price after +', new_price)
                                static = 'Change price to +'

                    if new_price == price:
                        new_price = None

                elif seller in exception_seller:
                    only = 'No'
                    static = 'Exception'
                    new_price = None
                    Note = 'Exception, So maximum price'
                    self.sheet[f'P{i}'] = Note
                    print(Note)
                else:
                    only = 'No'
                    if price < min_pr:
                        new_price = min_pr
                        static = 'Minimum'
                        Note = 'Price seller < minimum'
                        self.sheet[f'P{i}'] = Note
                        try:
                            self.sheet[f'H{i}'].fill = PatternFill(fill_type='solid', fgColor='ffe700')
                            self.sheet[f'I{i}'].fill = PatternFill(fill_type='solid', fgColor='ffe700')
                        except:
                            pass
                    else:
                        static = 'Change price to -'
                        y = float(price)
                        new_price = y - avn
                        if new_price < min_pr:
                            Note = ' Warning: New price is smaller than minimum price. So, will keep minimum price. '
                            print(Note)
                            self.sheet[f'P{i}'] = Note
                            new_price = min_pr
                            static = 'Minimum price'
            
                            try:
                                self.sheet[f'H{i}'].fill = PatternFill(fill_type='solid', fgColor='ffe700')
                                self.sheet[f'I{i}'].fill = PatternFill(fill_type='solid', fgColor='ffe700')
                            except:
                                pass

            except:
                seller = price = express = only = new_price = None
                self.sheet[f'P{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                static = "Not found"
                avn = None

            if new_price is not None:
                try:
                    res = self.search_bar(bar)
                    sleep(1)
                    if res is True:
                        # try:
                        #     pp = WebDriverWait(self.driver, 3).until((ec.visibility_of_element_located(
                        #         self.PRICE_MAX_SELECTOR))).text.replace('SAR', '').replace(',', '').strip().split('-')
                        #     p_n = float(pp[0].strip())
                        #     p_m = float(pp[1].strip())
                        #
                        #     print(f'- Max noon : {p_m}, Minimum Noon : {p_n}, Your New Price {new_price}')
                        #
                        #     # if new price small than noon mine, new price = noon min price
                        #     if float(new_price) < float(p_n):
                        #         new_price = p_n
                        #         print(f'- Edit To Noon Min {new_price}')
                        #         static = 'Noon Minimum price'
                        #
                        #     # if new price > noon max price, new price = noon max price
                        #     elif float(new_price) > float(p_m):
                        #         new_price = p_m
                        #         print(f'- Edit To Noon Max {new_price}')
                        #         static = 'Noon Maximum price'
                        #
                        # except:
                        #     print(traceback.print_exc())

                        self.change_price(new_price)
                        sleep(1)
                        self.change_quantity(i)
                        sleep(2)
                        try:
                            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SAVE_CHANGE_SELECTOR))).click()
                            sleep(1)
                            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                            sleep(1)
                        except:
                            self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                            self.sheet[f'Q{i}'] = 'Cant Save'
                            print('Cant Save')
                            print(traceback.print_exc())

                        try:
                            va = self.driver.execute_script("return document.querySelector('input[name=\"price\"]').value")
                            print(f'Current Price : {va}')
                            self.sheet[f'L{i}'] = va
                        except:
                            print(traceback.print_exc())
                except:
                    print(traceback.print_exc())
            for column in range(1, 18):
                try:
                    self.sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                     wrap_text=True)
                except:
                    pass
            self.sheet[f'D{i}'] = seller
            self.sheet[f'E{i}'] = price
            try:
                if 'express' in express:
                    self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
            except:
                pass
            self.sheet[f'F{i}'] = express
            self.sheet[f'G{i}'] = only
            self.sheet[f'H{i}'] = new_price
            self.sheet[f'I{i}'] = static
            try:
                self.sheet[f'M{i}'] = avn
            except:
                pass
            self.sheet[f'P{i}'] = Note

            self.book.save(self.file_save)
            
            print(f'- {i} - {bar} > seller : {seller}, price : {price}, new price : {new_price}, u only : {only}, static : {static}')

            #########################################
            if thr.p == 'pause' or thr.p == 'p':
                start = input('(Type start or s and press Enter)>>>')
                while True:
                    if start == 'start' or start == 's':
                        thr = Thread()
                        thr.start()
                        break
                    elif start == 'exit' or start == 'e':
                        self.driver.quit()
                        exit()
                    else:
                        start = input()
            elif thr.p == 'exit' or thr.p == 'e':
                self.driver.quit()
                exit()
            sleep(1)

        self.driver.quit()


def add_mac():
    try:
        os.remove(str(uuid.getnode()))
    except:
        pass
    try:
        with open(str(uuid.getnode()), 'w+') as fil:
            fil.write(str(uuid.getnode()))
        os.popen('attrib +h ' + str(uuid.getnode()))
        return True
    except:
        print('delete old file')
        return False


def check_mac():
    try:
        with open(str(uuid.getnode()), 'r') as fil:
            text = fil.readlines()
        if str(uuid.getnode()) in text:
            return True
        else:
            print('ref')
            return False
    except FileNotFoundError:
        print('ref')
        return False


if __name__ == "__main__":
    file = None
    pr = check_mac()
    while True:
        try:
            file = input('- Enter file name : ').strip()
            load_workbook(file).close()
            break
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)
        if file == 'scrt*python*add*mac*address':
            pr = add_mac()

    main_seller = input('- Main seller : ')
    email = input('- Enter your email : ')
    password = input('- Enter your password : ')
    avrpd = avrnd = float(input('- Enter value (-) '))

    content_inp = int(input('- Enter minimum quantity : '))
    content_new = int(input('- Enter new quantity : '))
    if pr:
        while True:
            my_bot = SouCode(file_name=file)
            my_bot.Search()
