from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from time import sleep

'''
len(bot.driver.find_elements_by_css_selector('td > div > div > div > div > div.countryLabel > div > a'))

bot.driver.find_elements_by_css_selector('div.filterCtr  div.header')[0].click()
bot.driver.find_elements_by_css_selector('div.filterCtr > div > div.active > div:nth-child(6) > div')[0].click()

bot.driver.find_element_by_css_selector('div:nth-child(7) > div > div.attrVal').text
bot.driver.find_element_by_css_selector('div.solid').text
bot.driver.find_element_by_css_selector('div.spacer ~ div.value').text.replace('SAR', '').strip()
'''


# class Thread(threading.Thread):
#     p = None
#
#     def __init__(self):
#         super(Thread, self).__init__()
#
#     def run(self):
#         while True:
#             self.p = input()
#             if self.p == 'pause' or self.p == 'p':
#                 break
#             elif self.p == 'exit' or self.p == 'e':
#                 break


class SouCode:

    def __init__(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.login(self.driver)

    @staticmethod
    def login(driver):
        try:
            driver.get('https://core.noon.partners/en-sa/sales')
            sleep(5)
            if 'login.noon' in driver.current_url:
                driver.find_element_by_name('email').send_keys(email)
                driver.find_element_by_name('password').send_keys(password)
                driver.find_element_by_css_selector('#formContainer > button').click()
                sleep(5)
                while True:
                    if 'login.noon' not in driver.current_url:
                        driver.get('https://core.noon.partners/en-sa/sales')
                        break

        except Exception as error:
            print(error)
            print('- Login Error')

    def Get_URL(self):
        URL = []
        if all_type is False:
            # self.driver.find_element_by_css_selector('div.filter > div.dropdownCtr > div:nth-child(1) > div > button').click()
            # sleep(1)
            # self.driver.find_element_by_css_selector(
            #     'div.filter > div.dropdownCtr > div:nth-child(1) > div > div > ul > li:nth-child(5)').click()
            # sleep(5)
            #
            self.driver.find_element_by_css_selector('div.filterCtr  div.header').click()
            sleep(2)
            self.driver.find_element_by_css_selector('div.filterCtr > div > div.active > div:nth-child(5) > div').click()
            sleep(5)
        q = 1
        while True:
            for element in self.driver.find_elements_by_css_selector('td > div > div > div > div > div.countryLabel > div > a'):
                # if element.get_attribute('href') not in URL:
                URL.append(element.get_attribute('href'))
                print(q, element.get_attribute('href'))
                q += 1
            try:
                if 'disabled' not in self.driver.find_element_by_css_selector('.next').get_attribute('class'):
                    self.driver.execute_script("document.querySelector('li.next a').click();")
                    sleep(2)
                else:
                    break
            except:
                break
        return URL

        # while True:
        #     sleep(4)
        #
        #     for element in self.driver.find_elements_by_css_selector('table > tbody > tr '):
        #
        #         text = element.find_element_by_css_selector('span.badge').get_attribute('class')
        #         if 'shipped' in text:
        #             # static = 'shipped'
        #             static = ' تم شحنها'
        #             if all_type is False:
        #                 continue
        #
        #         elif 'delivered' in text:
        #             # static = 'delivered'
        #             static = 'مُستلمة'
        #             if all_type is False:
        #                 continue
        #
        #         elif 'returned' in text:
        #             # static = 'returned'
        #             static = 'مُرجعة'
        #             # if all_type:
        #             #     continue
        #
        #         else:
        #             static = None
        #             if all_type is False:
        #                 continue
        #
        #         try:
        #             element.find_element_by_css_selector('div.noonNowBadge')
        #             # express = 'Express'
        #             express = 'اكسبرس'
        #         except:
        #             express = None
        #         SKU = element.find_element_by_css_selector('td[data-label="SKU"]').text
        #
        #         link = element.find_element_by_css_selector('div.viewMoreBtn > a').get_attribute('href')
        #         URL.append((SKU, static, express, link))
        #         print(len(URL))
        #     try:
        #         if 'disabled' not in self.driver.find_element_by_css_selector('.next').get_attribute('class'):
        #             self.driver.find_element_by_css_selector('li.next').click()
        #
        #         else:
        #             break
        #     except:
        #         break
        # return URL

    def Get_Products(self):
        data = self.Get_URL()
        print(len(data))
        book = Workbook()
        sheet = book.active
        sheet.cell(1, 1).value = 'SKU'
        sheet.cell(1, 2).value = 'Price'
        sheet.cell(1, 3).value = 'New Price'
        # sheet.cell(1, 4).value = 'Express'
        sheet.cell(1, 5).value = 'Static'
        try:
            sheet.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except:
            pass

        cont = 2
        for url in data:
            print(url)
            u = True
            while u:
                self.driver.get(url)
                sleep(2)
                try:
                    price = WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(
                        (By.CSS_SELECTOR, 'div.spacer ~ div.value')))).text.replace('SAR', '').strip()
                    sleep(1)
                    u = False
                except:
                    sleep(90)
                    continue

                try:
                    # price = self.driver.find_element_by_css_selector('div.spacer ~ div.value').text.replace('SAR', '').strip()
                    SKU = self.driver.find_element_by_css_selector('div:nth-child(7) > div > div.attrVal').text
                    static = self.driver.find_element_by_css_selector('div.solid').text

                    try:
                        new_price = float(price) + vol
                    except:
                        new_price = None
                    sheet.cell(cont, 1).value = SKU
                    sheet.cell(cont, 2).value = price
                    sheet.cell(cont, 3).value = new_price
                    # sheet.cell(cont, 4).value = express
                    sheet.cell(cont, 5).value = static
                    sheet.cell(cont, 6).value = url

                    print(f'{cont - 1} - {SKU} - Price : {price}')
                    sleep(1)
                except Exception as e:
                    sheet.cell(cont, 6).value = url
                    print(e)
            cont += 1
            book.save('product.xlsx')
        # cont = 2
        #         for SKU, static, express, url in data:
        #             Next = False
        #             while Next is False:
        #                 self.Get(url)
        #                 Next = self.Get_Price()
        #                 sleep(1)
        #
        #             price = Next
        #             if SKU is None or SKU.strip() == '' or 'N' not in SKU:
        #                 SKU = self.driver.find_element_by_css_selector('div.box > div:nth-child(4) > p.value').text
        #
        #             try:
        #                 new_price = float(price) + vol
        #             except:
        #                 new_price = None
        #
        #             sheet.cell(cont, 1).value = SKU
        #             sheet.cell(cont, 2).value = price
        #             sheet.cell(cont, 3).value = new_price
        #             sheet.cell(cont, 4).value = express
        #             sheet.cell(cont, 5).value = static
        #
        #             print(f'{cont - 1} - {SKU} - Price : {price}')
        #             cont += 1
        #             book.save('product.xlsx')
        #             sleep(1)

        self.driver.quit()


if __name__ == "__main__":
    email = 'bekj.119@gmail.com'  # input('- Enter your email : ')
    password = 'bekj.119'  # input('- Enter your password : ')
    vol = float(input('- Enter value (+) : '))
    type_se = input('- Enter (1 ----> All) Or (2 ----> Returned) : ')

    if type_se == 2 or type_se.strip() == '2':
        all_type = False
    else:
        all_type = True

    bot = SouCode()
    input('- Enter : ')
    bot.Get_Products()

# bekj.119@gmail.com
# [10:29 AM, 1/14/2021] +966 55 747 0145: bekj.119@gmail.com
# [10:29 AM, 1/14/2021] +966 55 747 0145: bekj.119
