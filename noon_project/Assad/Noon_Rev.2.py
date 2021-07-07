from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from time import sleep
from datetime import datetime
import threading
'''
* Update 'div.sellerDetails > p > strong' to 'div.soldBy' 10-1--2020
* Update to Souq Saudi 
* rapid 
* change if price < min_pr:new_price = max_pr
         if price < min_pr:new_price = min_pr
'''


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class SouCode:
    driver: WebDriver

    def __init__(self, file_name):
        self.file_save = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + file_name

        self.url = 'https://www.noon.com/saudi-en/'
        try:

            self.book = load_workbook(file_name)
            self.sheet = self.book.active
            self.sheet[f'A1'] = 'Noon_SKU'
            self.sheet[f'B1'] = 'Minimum Price'
            self.sheet[f'C1'] = 'Maximum Price'
            self.sheet[f'D1'] = 'Seller Name'
            self.sheet[f'E1'] = 'Price'
            self.sheet[f'F1'] = 'Express'
            self.sheet[f'G1'] = 'Only'
            self.sheet[f'H1'] = 'New price'
            self.sheet[f'I1'] = 'Status'
            self.sheet[f'J1'] = 'Brand'
            self.sheet[f'K1'] = 'Account'
            # self.sheet[f'L1'] = 'Change Value +'
            self.sheet[f'L1'] = 'Current Price'
            self.sheet[f'M1'] = 'Change Value -'
            self.sheet[f'N1'] = 'Exception'
            self.sheet[f'O1'] = 'Quantity'
            self.sheet[f'P1'] = 'Note 1'
            self.sheet[f'Q1'] = 'Note 2'
            self.sheet.column_dimensions['A'].width = 14
            self.sheet.column_dimensions['B'].width = 15
            self.sheet.column_dimensions['C'].width = 15
            self.sheet.column_dimensions['D'].width = 15
            self.sheet.column_dimensions['E'].width = 10
            self.sheet.column_dimensions['F'].width = 14
            self.sheet.column_dimensions['G'].width = 10
            self.sheet.column_dimensions['H'].width = 11
            self.sheet.column_dimensions['I'].width = 18
            self.sheet.column_dimensions['J'].width = 30
            self.sheet.column_dimensions['K'].width = 20
            self.sheet.column_dimensions['L'].width = 16
            self.sheet.column_dimensions['M'].width = 16
            self.sheet.column_dimensions['N'].width = 16
            self.sheet.column_dimensions['O'].width = 13
            self.sheet.column_dimensions['P'].width = 85
            self.sheet.column_dimensions['Q'].width = 50
        except Exception as err:
            print(err)
            exit()
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.login()

    def login(self):
        try:
            self.driver.get('https://catalog.noon.partners/en-sa/catalog')
            sleep(5)
            if 'login.noon' in self.driver.current_url:
                self.driver.find_element_by_name('email').send_keys(email)
                self.driver.find_element_by_name('password').send_keys(password)
                self.driver.find_element_by_css_selector('#formContainer > button').click()
                sleep(5)
                while True:
                    if 'login.noon' not in self.driver.current_url:
                        break

        except Exception as err:
            print(err)
            print('Login Error')

    def Search(self):
        Note = None
        self.sheet['K2'] = str(email)
        self.sheet['K4'] = str(main_seller)
        self.sheet['K3'] = str(password)

        end_cell = self.sheet.max_row

        exception_seller = [self.sheet[f'N{i}'].value for i in range(2, end_cell+1) if
                            self.sheet[f'N{i}'].value is not None]

        group1 = [self.sheet[f'R{i}'].value for i in range(2, end_cell+1) if self.sheet[f'R{i}'].value is not None]
        group2 = [self.sheet[f'S{i}'].value for i in range(2, end_cell+1) if self.sheet[f'S{i}'].value is not None]
        group3 = [self.sheet[f'T{i}'].value for i in range(2, end_cell+1) if self.sheet[f'T{i}'].value is not None]
        group4 = [self.sheet[f'U{i}'].value for i in range(2, end_cell+1) if self.sheet[f'U{i}'].value is not None]
        group5 = [self.sheet[f'V{i}'].value for i in range(2, end_cell+1) if self.sheet[f'V{i}'].value is not None]

        brand_group = {}
        for i in range(2, end_cell+1):
            if self.sheet[f'X{i}'].value is not None and self.sheet[f'Y{i}'].value is not None:
                brand_group[self.sheet[f'X{i}'].value] = self.sheet[f'Y{i}'].value

        # New
        brand_exception = [self.sheet[f'W{i}'].value for i in range(2, end_cell) if self.sheet[f'V{i}'].value is not None]

        thr = Thread()
        thr.start()

        for i in range(2, end_cell + 1):
            bar = self.sheet[f'A{i}'].value
            if bar is None:
                continue

            max_pr = float(self.sheet[f'C{i}'].value)
            min_pr = float(self.sheet[f'B{i}'].value)
            try:
                self.driver.get(f'{self.url}{bar}/p/')
                sleep(2)
                try:
                    brand = WebDriverWait(self.driver, 4).until((ec.visibility_of_element_located((By.CSS_SELECTOR, 'a.brand')))).text
                    # brand = self.driver.find_element_by_css_selector('a.brand').text
                    self.sheet[f'J{i}'] = brand
                except AttributeError:
                    brand = None
                except:
                    brand = None

                price = float(self.driver.find_element_by_css_selector('span.sellingPrice span.value').text)
                seller = self.driver.find_element_by_css_selector('p.sellerName a').text
                try:
                    express = self.driver.find_element_by_css_selector('div.bottomRow div.container img').get_attribute('alt')
                except:
                    express = None

                if brand in brand_exception:
                    self.sheet[f'D{i}'] = seller
                    self.sheet[f'E{i}'] = price
                    try:
                        self.sheet[f'I{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
                        if 'express' in express:
                            self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')

                    except:
                        pass
                    self.sheet[f'F{i}'] = express
                    self.sheet[f'I{i}'] = 'Exception Brand'
                    print(f'- {i} - {bar} > seller : {seller}, price : {price}, static : Exception Brand')
                    continue

                if seller in group1:
                    avp = avn = float(group1[0])
                elif seller in group2:
                    avp = avn = float(group2[0])
                elif seller in group3:
                    avp = avn = float(group3[0])
                elif seller in group4:
                    avp = avn = float(group4[0])
                elif seller in group5:
                    avp = avn = float(group5[0])

                elif brand in brand_group.keys():
                    avp = avn = float(brand_group[brand])

                else:
                    avp = avrpd
                    avn = avrnd

                Offer_number = len(
                    self.driver.find_elements_by_css_selector('ul.offersList > li'))

                ########################################################################################################
                if main_seller in seller:
                    if Offer_number == 0:
                        new_price = max_pr
                        only = 'Only'
                        static = 'Maximum Price'
                        print(' Only me. So, maximum price is ', max_pr)
                        Note = ' Only me. So, maximum price is ' + str(max_pr)
                        self.sheet[f'P{i}'] = Note
                    else:
                        only = 'No'
                        best = []
                        self.driver.find_element_by_css_selector('span.cta').click()
                        # sleep(4)
                        # offer = self.driver.find_elements_by_css_selector('ul.offersList > li')
                        offer = WebDriverWait(self.driver, 5).until((ec.visibility_of_all_elements_located((By.CSS_SELECTOR, 'ul.offersList > li'))))
                        try:
                            for ELEMENT in offer:
                                if offer.index(ELEMENT) == 4:
                                    break
                                if 'selectedItem' not in ELEMENT.get_attribute('class'):

                                    best.append((float(ELEMENT.find_element_by_css_selector('span.sellingPrice > span > span.value').text),
                                         ELEMENT.find_element_by_css_selector('div.soldBy').text,
                                         ELEMENT.find_element_by_css_selector('.container img').get_attribute('alt')))

                        except:
                            offer = self.driver.find_elements_by_css_selector('ul.offersList > li')
                            for ELEMENT in offer:
                                if offer.index(ELEMENT) == 4:
                                    break
                                if 'selectedItem' not in ELEMENT.get_attribute('class'):

                                    best.append(
                                        (float(ELEMENT.find_element_by_css_selector(
                                            'span.sellingPrice > span > span.value').text),
                                         ELEMENT.find_element_by_css_selector('div.soldBy').text,
                                         ELEMENT.find_element_by_css_selector('.container img').get_attribute('alt')))

                        if best[0][1] in exception_seller:
                            # new_price = max_pr
                            new_price = None
                            print(' Exception. So, maximum price. ')
                            Note = ' Exception. So, maximum price. '
                            self.sheet[f'P{i}'] = Note
                            static = 'Maximum price'
                        elif float(best[0][0]) < min_pr:
                            new_price = None
                            Note = 'next price < Minimum'
                            self.sheet[f'P{i}'] = Note
                            static = 'freeze'
                        else:

                            y = float(best[0][0])
                            # new_price = y - y * (avp / 100)
                            new_price = y - avp
                            if new_price < price:
                                Note = ' Warning: New price is smaller than yours. So, price will be your price or minimum. ' + str(
                                    new_price)
                                self.sheet[f'P{i}'] = Note
                                print(Note, new_price)
                                new_price = None if price > min_pr else min_pr

                            print('New price after +', new_price)
                            static = 'Change price to +'

                    if new_price == price:
                        new_price = None

                elif seller in exception_seller:
                    only = 'No'
                    static = 'Exception'
                    # new_price = max_pr
                    new_price = None
                    Note = 'Exception, So maximum price'
                    self.sheet[f'P{i}'] = Note
                    print(Note)
                else:
                    only = 'No'
                    if price < min_pr:
                        # New change 2020-11-13
                        # new_price = max_pr
                        new_price = min_pr
                        static = 'Minimum'
                        Note = 'Price seller < minimum'
                        self.sheet[f'P{i}'] = Note
                        try:
                            self.sheet[f'H{i}'].fill = PatternFill(fill_type='solid', fgColor='ffe700')
                            self.sheet[f'I{i}'].fill = PatternFill(fill_type='solid', fgColor='ffe700')
                        except:
                            pass
                    else:
                        static = 'Change price to -'
                        y = float(price)
                        new_price = y - avn
                        if new_price < min_pr:
                            Note = ' Warning: New price is smaller than minimum price. So, will keep minimum price. '
                            print(Note)
                            self.sheet[f'P{i}'] = Note
                            new_price = min_pr
                            static = 'Minimum price'
            
                            try:
                                self.sheet[f'H{i}'].fill = PatternFill(fill_type='solid', fgColor='ffe700')
                                self.sheet[f'I{i}'].fill = PatternFill(fill_type='solid', fgColor='ffe700')
                            except:
                                pass
                # sleep(1)

            except Exception as err:
                seller = price = express = only = new_price = None
                print(f'{err}')
                self.sheet[f'P{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                static = "Not found"

            try:
                if new_price is not None:

                    self.driver.get('https://catalog.noon.partners/en-sa/catalog')
                    sleep(2)
                    WebDriverWait(self.driver, 8).until((ec.visibility_of_element_located(
                        (By.CSS_SELECTOR, 'input.searchInput')))).send_keys(bar)
                    self.driver.find_element_by_css_selector('input.searchInput').send_keys(Keys.ENTER)
                    sleep(3)
                    try:
                        WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located(
                            (By.CSS_SELECTOR, 'td:nth-child(1) > div > a')))).click()

                        # self.driver.find_element_by_css_selector('td:nth-child(1) > div > a').click()
                    except Exception as err:
                        print(err)
                        self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                        self.sheet[f'Q{i}'] = 'Cant find no click'
                        print('Cant find', 'no click')
                    try:
                        sleep(3)
                        # self.driver.find_element_by_name('price_sa').clear()
                        WebDriverWait(self.driver, 6).until((ec.visibility_of_element_located((By.NAME, 'price_sa')))).clear()
                        ########################################################################################################
                        sleep(1)
                        self.driver.execute_script(f"document.querySelector('input[name=\"price_sa\"]').value = '' ")
                        sleep(1)
                        self.driver.find_element_by_name('price_sa').clear()
                        sleep(1)
                        self.driver.find_element_by_name('price_sa').send_keys(str(new_price))
                        # va = self.driver.execute_script("return document.querySelector('input[name=\"price_sa\"]').value")
                        #
                        # if int(float(va)) == int(float(new_price)) or -3 < int(float(va)) - int(float(new_price)) < 3:
                        #     print('- Price Prefect')
                        # else:
                        #     print('- Re Price')
                        #     self.driver.execute_script(
                        #         f"document.querySelector('input[name=\"price_sa\"]').value = {new_price}")
                        #     sleep(2)
                        #     va = self.driver.execute_script(
                        #         "return document.querySelector('input[name=\"price_sa\"]').value")
                        #     if int(float(va)) == int(float(new_price)) or -3 < int(float(va)) - int(
                        #         float(new_price)) < 3:
                        #         pass
                        #     else:
                        #         continue

                        #########################################################################################################
                        # self.driver.find_element_by_name('price_sa').clear()
                        # sleep(1)
                        # self.driver.find_element_by_name('price_sa').send_keys(str(new_price))

                        content = self.driver.find_element_by_css_selector('div.stockNet').text.split('\n')[1]
                        print(content)
                        self.sheet[f'O{i}'] = content
                        if int(content) < int(content_inp):
                            if int(content) == 0:
                                self.sheet[f'Q{i}'] = 'Warning: Quantity is ' + str(content)
                            else:
                                self.driver.find_element_by_name('quantity').send_keys(str(content_new))
                                print('Change Quantity to ', content_new)
                                self.sheet[f'Q{i}'] = 'Warning: Quantity is ' + str(content)
                                self.sheet[f'R{i}'] = 'Change Quantity to ' + str(content_new)
                            try:
                                self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                                self.sheet[f'R{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                                self.sheet[f'O{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                            except:
                                pass
                        self.driver.find_element_by_css_selector('div.fixedBottom > button').click()
                        sleep(2)

                        try:
                            va = self.driver.execute_script("return document.querySelector('input[name=\"price_sa\"]').value")
                            print(f'Current Price : {va}')
                            self.sheet[f'L{i}'] = va
                        except Exception as err:
                            print(err)

                    except Exception as err:
                        print(err)
                        self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                        self.sheet[f'Q{i}'] = 'Cant change'
                        print('Cant change other')

            except Exception as err:
                print(err)

            for column in range(1, 18):
                try:
                    self.sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                     wrap_text=True)
                except:
                    pass
            self.sheet[f'D{i}'] = seller
            self.sheet[f'E{i}'] = price
            try:
                if 'express' in express:
                    self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
            except:
                pass
            self.sheet[f'F{i}'] = express
            self.sheet[f'G{i}'] = only
            self.sheet[f'H{i}'] = new_price
            self.sheet[f'I{i}'] = static
            try:
                # self.sheet[f'L{i}'] = avp
                self.sheet[f'M{i}'] = avn
            except:
                pass
            self.sheet[f'P{i}'] = Note

            self.book.save(self.file_save)
            
            print(f'- {i} - {bar} > seller : {seller}, price : {price}, new price : {new_price}, u only : {only}, static : {static}')

            #########################################
            if thr.p == 'pause' or thr.p == 'p':
                start = input('(Type start or s and press Enter)>>>')
                while True:
                    if start == 'start' or start == 's':
                        thr = Thread()
                        thr.start()
                        break
                    elif start == 'exit' or start == 'e':
                        self.driver.quit()
                        exit()
                    else:
                        start = input()
            elif thr.p == 'exit' or thr.p == 'e':
                self.driver.quit()
                exit()

        self.driver.quit()


if __name__ == "__main__":
    while True:
        try:
            file = input('- Enter file name : ').strip()
            book = load_workbook(file)
            book.close()
            break
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)

    main_seller = input('- Main seller : ')
    email = input('- Enter your email : ')
    password = input('- Enter your password : ')
    avrpd = avrnd = float(input('- Enter value (-) '))

    content_inp = int(input('- Enter minimum quantity : '))
    content_new = int(input('- Enter new quantity : '))

    while True:
        my_bot = SouCode(file_name=file)
        my_bot.Search()
