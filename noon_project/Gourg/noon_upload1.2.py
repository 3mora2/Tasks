from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as ec
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
import traceback


class Upload:
    driver: WebDriver
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-eg/noon-catalog'
    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'a[href*="/en-eg/noon-catalog/preview/"]')

    SKU_TEXT_SELECTOR = (By.CSS_SELECTOR, 'div.tabsWrapper > div')
    SKU_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[name="partner_sku"]')
    SKU_BUTTON_SELECTOR = (By.CSS_SELECTOR, '.solid')

    PRICE_MAX_SELECTOR = (
        By.XPATH,
        "//div[contains(@class, 'inputInfo') and contains(text(), 'Noon SKU Price')] / *[@class='priceRange']")
    PRICE_INPUT_SELECTOR = (By.NAME, 'price')

    QUANTITY_SITE_SELECTOR = (By.CSS_SELECTOR, 'div.placeholder')
    QUANTITY_SITE_CHOSE_SELECTOR = (By.CSS_SELECTOR, 'div.active > div > div.countryLabel')
    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')

    BARCODE_INSIDE_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="text"][placeholder="Enter Barcode"]')
    BUTTON_ADD_BARCODE_INSIDE_SELECTOR = (By.CSS_SELECTOR, 'div.eachInput > div > div.solid')
    # ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="active"]\').checked == false){document.querySelector(\'input[id="active"]\').click()};'
    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="offerActive"]\').checked == false){document.querySelector(\'input[id="offerActive"]\').click()};'
    # SAVE_CHANGE_SELECTOR = (By.CSS_SELECTOR, 'div.transparent ~ div.solid')
    # SAVE_CHANGE_SELECTOR = (By.CSS_SELECTOR, 'div.tabsWrapper ~ div > div.solid')
    # SUBMIT_SELECTOR = (
    # By.XPATH, '//div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')
    # SUBMIT_SELECTOR = (
    #     By.XPATH,
    #     '//div[calss="btnWrapper"]/div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')

    SAVE_CHANGE_SELECTOR = (By.XPATH, '//*[contains(text(),"Save Changes")]')
    SUBMIT_SELECTOR = (By.XPATH,
                       '//div[class="btnWrapper"]/div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')

    def __init__(self):
        self.book = load_workbook(file_name)
        self.sheet = self.book.active
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.login()

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def add_new(self, name):
        try:
            self.driver.get(self.URL_CAT)
            sleep(2)
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))).send_keys(name, Keys.ENTER)
            sleep(2)
        except:
            print(traceback.print_exc())

    def add_serial(self, serial):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
        except:
            print('- Not found')
            return 'False'
        sleep(1)
        try:
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_INPUT_SELECTOR)).send_keys(serial)
            sleep(3)
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_BUTTON_SELECTOR)).click()
            return serial
        except:
            print('- Exist, Continue')
            return False

    def add_price(self, price):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a', Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(str(price))
        except:
            print(traceback.print_exc())
            print('- Cant Change Price')

    def add_qu(self, q):
        sleep(1)
        try:
            WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
        except:
            print(traceback.print_exc())
            print('- Cant Change quantity')

    def add_bar(self, sk):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.BARCODE_INSIDE_INPUT_SELECTOR)).send_keys(sk)
            sleep(1)
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.BUTTON_ADD_BARCODE_INSIDE_SELECTOR)).click()
            sleep(1)
        except:
            print(traceback.print_exc())
            print('- Cant Change Barcode')

    def add_active(self):
        sleep(1)
        try:
            self.driver.execute_script(self.ACTIVE_SCRIPT)
        except:
            print(traceback.print_exc())

    def upload(self):

        for i in range(2, self.sheet.max_row + 1):
            try:
                bar = self.sheet.cell(i, 1).value
                serial = self.sheet.cell(i, 2).value
                price = self.sheet.cell(i, 3).value
                Quantity = self.sheet.cell(i, 4).value
                if bar is None:
                    continue

                self.add_new(bar)
                res = self.add_serial(serial)

                if res == 'False':
                    continue
                sleep(1)
                if price is not None:
                    self.add_price(price)

                sleep(2)
                if Quantity is not None:
                    self.add_qu(Quantity)

                sleep(1)
                self.add_bar(serial)
                sleep(2)
                self.add_active()
                sleep(3)
                try:
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SAVE_CHANGE_SELECTOR))).click()
                    sleep(1)
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                    sleep(6)
                    WebDriverWait(self.driver, 10).until((ec.element_to_be_clickable(self.SAVE_CHANGE_SELECTOR))).click()
                    sleep(1)
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                    sleep(8)
                except:
                    print(traceback.print_exc())
                    print('Cant Save')

                try:
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SAVE_CHANGE_SELECTOR))).click()
                    sleep(1)
                    WebDriverWait(self.driver, 10).until((ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                    sleep(3)
                except:
                    print(traceback.print_exc())
                    print('Cant Save')
            except:
                print(traceback.print_exc())

        self.driver.quit()


if __name__ == "__main__":

    while True:
        try:
            file_name = input('- Enter file name : ')
            load_workbook(file_name).close()
            break
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)

    email = input('- Enter your email : ')
    password = input('- Enter your password : ')

    u = Upload()
    u.upload()
'''
Trevi@trevi-trade.com
maged777
'''