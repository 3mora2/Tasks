import warnings
warnings.filterwarnings("ignore")
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from time import sleep
from datetime import datetime
from requests_html import HTMLSession
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

mail = input('Enter your Email : ')  # 'samcoeg@yahoo.com'
# gogo1500@gmail.com maged777
password = input('Enter password : ')  # 'maged777'
book = Workbook()
sheet = book.active
i = 2
sheet[f'A{i}'].value = 'الصور'
sheet[f'B{i}'].value = 'الاسم'
sheet[f'C{i}'].value = 'كود'
sheet[f'D{i}'].value = 'الكمية'
sheet[f'E{i}'].value = 'السعر'
sheet[f'F{i}'].value = 'القيمة'
sheet.column_dimensions['A'].width = 13
sheet.column_dimensions['B'].width = 70
sheet.column_dimensions['C'].width = 13
sheet.column_dimensions['D'].width = 5
sheet.column_dimensions['E'].width = 7
sheet.column_dimensions['F'].width = 7
for column in range(1, 7):
    try:
        sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    except:
        pass

session = HTMLSession()


class Main:
    def __init__(self):
        self.driver = webdriver.Chrome()
        # self.start()

    def start(self):
        try:
            self.driver.get('https://login.noon.partners/en/register')
            sleep(2)
            WebDriverWait(self.driver, 30).until(ec.visibility_of_element_located((By.NAME, "email"))).send_keys(mail)
            WebDriverWait(self.driver, 30).until(ec.visibility_of_element_located((By.NAME, "password"))).send_keys(
                password)
            self.driver.find_element_by_css_selector('div > div.btnWrapper > button').click()
            sleep(5)
            WebDriverWait(self.driver, 30).until(ec.visibility_of_element_located(
                (By.CSS_SELECTOR, "div > div.summaryCtr > a:nth-child(3) > div > span")))
            sleep(4)
            num = WebDriverWait(self.driver, 30).until(ec.visibility_of_element_located(
                    (By.CSS_SELECTOR, "div > div.summaryCtr > a:nth-child(3) > div > span"))).text
            if num != "0":

                WebDriverWait(self.driver, 30).until(
                    ec.visibility_of_element_located((By.CSS_SELECTOR, "div > div.summaryCtr > a:nth-child(3)"))).click()
                sleep(4)

                WebDriverWait(self.driver, 30).until(
                    ec.visibility_of_element_located((By.CSS_SELECTOR, "div.filterWrapper > div:nth-child(1)")))
                sleep(2)
                ASN = WebDriverWait(self.driver, 30).until(
                    ec.visibility_of_element_located((By.CSS_SELECTOR, "div.filterWrapper > div:nth-child(1)"))).text
                ASN = ASN if 'ASN' in ASN else ''
                print(ASN)
                # .replace('ASN:', '').strip()

                da = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
                sheet['A1'].value = ASN
                sheet['B1'].value = da

                list_Q = 0
                list_v = 0
                i = 3

                elements = WebDriverWait(self.driver, 30).until(
                    ec.visibility_of_all_elements_located((By.CSS_SELECTOR, "table > tbody > tr")))
                for element in elements:
                    bar_code = WebDriverWait(element, 30).until(
                        ec.visibility_of_element_located((By.CSS_SELECTOR, "td:nth-child(4) > div"))).text
                    name = element.find_element_by_css_selector(
                        'td:nth-child(1) > a > div > div > div.desc > div.title > span > span').text
                    QTY = element.find_element_by_css_selector('td:nth-child(5) > div').text
                    img = element.find_element_by_css_selector(
                        'td:nth-child(1) > a > div > div > div.proImage > img').get_attribute('src')
                    url = element.find_element_by_css_selector('td:nth-child(1) > a').get_attribute('href')

                    self.driver.execute_script("window.open('')")
                    self.driver.switch_to_window(self.driver.window_handles[1])
                    sleep(1)
                    self.driver.get(url)
                    sleep(3)
                    price = WebDriverWait(self.driver, 30).until(
                        ec.visibility_of_element_located((By.NAME, "price_eg"))).get_attribute('value')
                    self.driver.close()
                    self.driver.switch_to_window(self.driver.window_handles[0])
                    print(i-1, ' - ', name)
                    try:
                        if img is not None:
                            res = session.get(img)
                            image_file = BytesIO(res.content)
                            img = Image(image_file)
                            img.width = 90
                            img.height = 75
                            sheet.row_dimensions[i].height = 60
                            sheet.add_image(img, f'A{i}')
                    except:
                        pass
                    sheet[f'B{i}'].value = name
                    sheet[f'C{i}'].value = bar_code
                    sheet[f'D{i}'].value = QTY
                    sheet[f'E{i}'].value = price
                    sheet[f'F{i}'].value = float(QTY) * float(price)
                    for column in range(1, 7):
                        try:
                            sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                        wrap_text=True)
                        except:
                            pass
                    list_v += float(QTY) * float(price)
                    list_Q += float(QTY)
                    i += 1

                sheet[f'D{i}'].value = list_Q
                sheet[f'F{i}'].value = list_v
                try:
                    sheet.cell(i, 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet.cell(i, 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                except:
                    pass

                book.save(str(ASN.replace('ASN:', 'ASN -').strip()) + ' ' + str(da) + '.xlsx')
            else:
                print('Not Found')
        except Exception as e:
            print(e)
        self.driver.quit()


if __name__ == '__main__':
    m = Main()
    m.start()
