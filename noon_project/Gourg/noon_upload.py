from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import *
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep


class Upload:
    def __init__(self):
        while True:
            try:
                file_name = input('- Enter file name : ')  # 'Book1.xlsx'
                self.book = load_workbook(file_name)
                self.sheet = self.book.active
                break
            except FileNotFoundError:
                print('- No such file or directory')
            except Exception as e:
                print(e)
        self.Bar_Price = self.Get_Barcode_And_Price()
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()

    def Get_Barcode_And_Price(self):
        Barcode = []
        l = []
        end = self.sheet.max_row
        for row in range(2, end + 1):
            if self.sheet.cell(row, 1).value is not None and self.sheet.cell(row, 2).value is not None:
                if self.sheet.cell(row, 1).value not in l:
                    l.append(self.sheet.cell(row, 1).value)
                    Barcode.append((self.sheet.cell(row, 1).value, self.sheet.cell(row, 2).value,self.sheet.cell(row, 3).value, self.sheet.cell(row, 4).value))
        print(len(Barcode))
        return Barcode

    def upload(self):
        for bar, serial, price, quantity in self.Bar_Price:
            print('- ', bar, price)
            self.driver.get('https://catalog.noon.partners/en-eg/noon-catalog')
            sleep(5)
            if 'login.noon' in self.driver.current_url:
                self.driver.find_element_by_name('email').send_keys(email)
                sleep(1)
                self.driver.find_element_by_name('password').send_keys(password)
                sleep(1)
                self.driver.find_element_by_css_selector('#formContainer > button').click()
                sleep(5)
                while True:
                    if 'login.noon' not in self.driver.current_url:
                        break
                if '/noon-catalog' not in self.driver.current_url:
                    self.driver.get('https://catalog.noon.partners/en-eg/noon-catalog')
                sleep(4)
            self.driver.find_element_by_css_selector('input.searchInput').send_keys(bar)
            sleep(1)
            self.driver.find_element_by_css_selector('input.searchInput').send_keys(Keys.ENTER)
            sleep(4)
            try:
                WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located((By.CSS_SELECTOR, 'div.content div.ctr > div.result > div.resultCtr > a'))).click()
                sleep(3)
                WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(
                    (By.CSS_SELECTOR, 'div.eachForm > div.inputCtr > div > input.input'))).send_keys(serial)
                sleep(2)
                # self.driver.find_element_by_css_selector('div.eachForm > div.inputCtr > div > input.input').send_keys(bar[::-1])
                self.driver.find_element_by_css_selector('div.rightView > div.btnCtr > button').click()
            except TimeoutException:
                print('- Continue')
                continue
            sleep(4)
            WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located(
                (By.CSS_SELECTOR, 'div.headerCtr > div.mapButtonCtr > div > div > input'))).click()

            # self.driver.find_element_by_css_selector('div.headerCtr > div.mapButtonCtr > div > div > input').click()
            sleep(1)
            self.driver.find_element_by_name('price_eg').clear()
            sleep(1)
            self.driver.find_element_by_name('price_eg').send_keys(price)
            sleep(1)
            self.driver.find_element_by_name('quantity').send_keys(quantity)
            sleep(1)
            self.driver.find_element_by_css_selector('div.inputCtr > input').send_keys(serial)
            sleep(2)
            try:
                self.driver.find_element_by_css_selector('div.inputCtr > div.btnCtr > button').click()
            except ElementClickInterceptedException:
                _ = self.driver.find_element_by_css_selector('div.inputCtr > div.btnCtr').location_once_scrolled_into_view
                sleep(1)
                self.driver.find_element_by_css_selector('div.inputCtr > div.btnCtr > button').click()
            sleep(2)
            self.driver.find_element_by_css_selector('div.fixedBottom > button').click()
            sleep(4)

        self.driver.quit()


if __name__ == "__main__":
    email = input('- Enter your email : ')
    password = input('- Enter your password : ')
    u = Upload()
    u.upload()

# samcoeg@yahoo.com
# maged777
# gogo1500@gmail.com
# maged777