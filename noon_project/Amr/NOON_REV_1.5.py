from datetime import datetime
from selenium.webdriver.chrome.webdriver import WebDriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from time import sleep
import threading
import traceback


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class SouCode:
    driver: WebDriver
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-eg/catalog'
    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'tr:nth-child(1) > td > a')

    PRICE_MAX_SELECTOR = (
        By.XPATH,
        "//div[contains(@class, 'inputInfo') and contains(text(), 'Noon SKU Price')] / *[@class='priceRange']")
    PRICE_INPUT_SELECTOR = (By.NAME, 'price')

    QUANTITY_VALUE_SELECTOR = (By.CSS_SELECTOR, 'div.WhStockNet > div > div.WhItemValue')
    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')

    # update
    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="offerActive"]\').checked == false){document.querySelector(\'input[id="offerActive"]\').click()};'
    SAVE_CHANGE_SELECTOR = (By.XPATH, '//*[contains(text(),"Save Changes")]')
    SUBMIT_SELECTOR = (
        By.XPATH,
        '//div[calss="btnWrapper"]/div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')
    #############################################
    URL_SEARCH = 'https://www.noon.com/egypt-en/'
    RATE_SELECTOR = (By.CSS_SELECTOR, 'div.detail_percentage')
    PRICE_SCRIPT = "return parseFloat(document.querySelector('.priceNow').textContent.replace('EGP',''));"

    # update
    SELLER_TEXT_SELECTOR = (By.CSS_SELECTOR, '[alt="seller"] + div a.storeLink')

    EXPRESS_SELECTOR = (By.CSS_SELECTOR, 'div.estimator_right > div > img')
    ALL_OFFER_SELECTOR = (By.CSS_SELECTOR, '.allOffers')
    S_PRICE_SELECTOR = (By.CSS_SELECTOR, 'div.regular > strong')
    S_SELLER_SELECTOR = (By.CSS_SELECTOR, '.soldby > strong')
    S_EXPRESS_SELECTOR = (By.CSS_SELECTOR, 'div[direction="row"] > div > img')

    def __init__(self, file_name):
        self.file_save = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + file_name
        self.book = load_workbook(file_name)
        self.sheet = self.book.active
        self.Load_File()

    def Load_File(self):
        try:
            self.sheet[f'A1'] = 'Noon_SKU'
            self.sheet[f'B1'] = 'Minimum Price'
            self.sheet[f'C1'] = 'Maximum Price'
            self.sheet[f'D1'] = 'Seller Name'
            self.sheet[f'E1'] = 'Price'
            self.sheet[f'F1'] = 'Express'
            self.sheet[f'G1'] = 'Only'
            self.sheet[f'H1'] = 'New price'
            self.sheet[f'I1'] = 'Status'
            self.sheet[f'J1'] = 'Account'
            self.sheet[f'K1'] = 'Pass'
            self.sheet[f'L1'] = 'Change Value +'
            self.sheet[f'M1'] = 'Change Value -'
            self.sheet[f'N1'] = 'Exception'
            self.sheet[f'O1'] = 'Quantity'
            self.sheet[f'P1'] = 'Note 1'
            self.sheet[f'Q1'] = 'Note 2'
            self.sheet.column_dimensions['A'].width = 14
            self.sheet.column_dimensions['B'].width = 15
            self.sheet.column_dimensions['C'].width = 15
            self.sheet.column_dimensions['D'].width = 15
            self.sheet.column_dimensions['E'].width = 10
            self.sheet.column_dimensions['F'].width = 14
            self.sheet.column_dimensions['G'].width = 10
            self.sheet.column_dimensions['H'].width = 11
            self.sheet.column_dimensions['I'].width = 18
            self.sheet.column_dimensions['J'].width = 30
            self.sheet.column_dimensions['K'].width = 20
            self.sheet.column_dimensions['L'].width = 16
            self.sheet.column_dimensions['M'].width = 16
            self.sheet.column_dimensions['N'].width = 16
            self.sheet.column_dimensions['O'].width = 13
            self.sheet.column_dimensions['P'].width = 85
            self.sheet.column_dimensions['Q'].width = 50
        except Exception as e:
            print(e)
            exit()

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            sleep(3)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until((ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def search_bar(self, code):
        try:
            sleep(1)
            self.driver.get(self.URL_CAT)
            WebDriverWait(self.driver, 10).until(
                (ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))).send_keys(
                code, Keys.ENTER)
            sleep(1)
        except:
            print(traceback.print_exc())

            return False
        sleep(2)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
            sleep(3)
            return True
        except:
            print('- Not found')
            return False

    def change_price(self, price):
        try:
            sleep(1)
            WebDriverWait(self.driver, 10).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a',
                                                                                         Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a',
                                                                                         Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(str(price))
            sleep(1)
        except:
            print('- Cant Change Price')

    def change_quantity(self, i):
        sleep(1)
        try:
            content = WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_VALUE_SELECTOR))).text
            print(content)
            self.sheet[f'O{i}'] = content
            if int(content) < int(content_inp):
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(content_new))
                print('Change Quantity to ', content_new)

                self.sheet[f'Q{i}'] = 'Warning: Quantity is ' + str(content)
                self.sheet[f'Q{i}'] = 'Change Quantity to ' + str(content_new)
                try:
                    self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                    self.sheet[f'R{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                    self.sheet[f'O{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                except:
                    pass
            sleep(1)
        except:
            print(traceback.print_exc())

            print('- Cant Change quantity')

    def add_active(self):
        sleep(1)
        try:
            self.driver.execute_script(self.ACTIVE_SCRIPT)
        except:
            print(traceback.print_exc())
            print('- Cant active ')
        sleep(1)

    def Search(self):
        Note = None
        self.sheet['J2'] = str(email)
        self.sheet['J3'] = str(main_seller)
        self.sheet['K2'] = str(password)

        end_cell = self.sheet.max_row

        exception_seller = [self.sheet[f'N{i}'].value for i in range(2, end_cell) if
                            self.sheet[f'N{i}'].value is not None]
        group1 = [self.sheet[f'R{i}'].value for i in range(2, end_cell) if self.sheet[f'R{i}'].value is not None]
        group2 = [self.sheet[f'S{i}'].value for i in range(2, end_cell) if self.sheet[f'S{i}'].value is not None]
        group3 = [self.sheet[f'T{i}'].value for i in range(2, end_cell) if self.sheet[f'T{i}'].value is not None]
        group4 = [self.sheet[f'U{i}'].value for i in range(2, end_cell) if self.sheet[f'U{i}'].value is not None]
        group5 = [self.sheet[f'V{i}'].value for i in range(2, end_cell) if self.sheet[f'V{i}'].value is not None]

        thread = Thread()
        thread.start()
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.login()
        for i in range(2, end_cell + 1):
            bar = self.sheet[f'A{i}'].value
            if bar is None:
                continue
            max_f = False
            max_pr = float(self.sheet[f'C{i}'].value)
            min_pr = float(self.sheet[f'B{i}'].value)
            try:
                self.driver.get(f'{self.URL_SEARCH}{bar}/p/')
                sleep(2)
                price = float(self.driver.execute_script(self.PRICE_SCRIPT))
                seller = WebDriverWait(self.driver, 3).until(
                    (ec.visibility_of_element_located(self.SELLER_TEXT_SELECTOR))).text
                try:
                    express = WebDriverWait(self.driver, 3).until(
                        (ec.visibility_of_element_located(self.EXPRESS_SELECTOR))).get_attribute('alt')
                except:
                    express = None
                if seller in group1:
                    avp = avn = float(group1[0])
                elif seller in group2:
                    avp = avn = float(group2[0])
                elif seller in group3:
                    avp = avn = float(group3[0])
                elif seller in group4:
                    avp = avn = float(group4[0])
                elif seller in group5:
                    avp = avn = float(group5[0])
                else:
                    try:
                        rat1 = float(WebDriverWait(self.driver, 2).until(
                            (ec.visibility_of_all_elements_located(self.RATE_SELECTOR)))[0].text.replace('%', ''))
                        rat2 = float(WebDriverWait(self.driver, 2).until(
                            (ec.visibility_of_all_elements_located(self.RATE_SELECTOR)))[1].text.replace('%', ''))
                        print(rat1 + rat2)
                        if rat1 + rat2 >= percent:
                            avp = avrpu
                            avn = avrnu
                        else:
                            avp = avrpd
                            avn = avrnd
                    except:
                        rat2 = None
                        rat1 = None
                        avp = avrpd
                        avn = avrnd
                        print(rat1, rat2)
                try:
                    Offer_number = len(WebDriverWait(self.driver, 1).until((ec.visibility_of_all_elements_located(self.ALL_OFFER_SELECTOR))))
                except:
                    Offer_number = 0
                ########################################################################################################
                if main_seller in seller:
                    if Offer_number == 0:
                        new_price = max_pr
                        max_f = True
                        only = 'Only'
                        static = 'Maximum Price'
                        print(' Only me. So, maximum price is ', max_pr)
                        Note = ' Only me. So, maximum price is ' + str(max_pr)
                    else:
                        only = 'No'
                        WebDriverWait(self.driver, 2).until((ec.visibility_of_element_located(self.ALL_OFFER_SELECTOR))).click()
                        sleep(4)
                        s_seller = WebDriverWait(self.driver, 2).until((ec.visibility_of_all_elements_located(self.S_SELLER_SELECTOR)))[1].text.split(' ')[-1]
                        s_price = WebDriverWait(self.driver, 2).until((ec.visibility_of_all_elements_located(self.S_PRICE_SELECTOR)))[1].text
                        s_express = WebDriverWait(self.driver, 2).until((ec.visibility_of_all_elements_located(self.S_EXPRESS_SELECTOR)))[1].get_attribute('alt')
                        print(s_seller, s_price, s_express)
                        if s_seller in exception_seller:
                            new_price = max_pr
                            max_f = True
                            print(' Exception. So, maximum price. ')
                            Note = ' Exception. So, maximum price. '
                            static = 'Maximum price'
                        elif float(s_price) < min_pr:
                            new_price = None
                            Note = 'next price < Minimum'
                            static = 'freeze'
                        else:
                            y = float(s_price)
                            new_price = y - y * (avp / 100)
                            if new_price < price:
                                Note = ' Warning: New price is smaller than yours. So, price will be your price or minimum. ' + str(
                                    new_price)
                                print(Note, new_price)
                                new_price = None if price > min_pr else min_pr

                            print('New price after +', new_price)
                            static = 'Change price to +'
                elif 'express' in express:
                    only = 'No'
                    new_price = None
                    Note = 'No change'
                    static = 'Express'
                elif seller in exception_seller:
                    only = 'No'
                    static = 'Exception'
                    new_price = max_pr
                    max_f = True
                    Note = 'Exception, So maximum price'
                    print(Note)
                else:
                    only = 'No'
                    if price < min_pr:
                        new_price = max_pr
                        max_f = True
                        static = 'Maximum'
                        Note = 'Price seller < minimum'
                    else:
                        static = 'Change price to -'
                        y = float(price)
                        new_price = y - y * (avn / 100)
                        if new_price < min_pr:
                            Note = ' Warning: New price is smaller than minimum price. So, will keep minimum price. '
                            print(Note)
                            new_price = min_pr
                            static = 'Minimum price'

                sleep(5)
            except Exception as e:
                seller = price = express = only = new_price = None
                Note = f'{e}'
                self.sheet[f'P{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                static = "Not found"

            if new_price is not None:
                try:
                    res = self.search_bar(bar)
                    if res is True:
                        if max_f is True:
                            try:
                                p_m = float(WebDriverWait(self.driver, 3).until(
                                    (ec.visibility_of_element_located(self.PRICE_MAX_SELECTOR))).text.replace('EGP',
                                                                                                              '').replace(
                                    ',',
                                    '').strip().split('-')[-1].strip())
                                print(f'- Max noon : {p_m} , Your New Price {new_price}')
                                if p_m < new_price:
                                    new_price = p_m
                                    print(f'- Edit To Noon Max {new_price}')
                                    static = 'Noon Maximum price'
                            except TimeoutError:
                                pass
                            except Exception as e:
                                print(e)
                        sleep(1)
                        self.change_price(new_price)
                        sleep(1)
                        self.change_quantity(i)
                        sleep(2)
                        try:
                            WebDriverWait(self.driver, 10).until(
                                (ec.visibility_of_element_located(self.SAVE_CHANGE_SELECTOR))).click()
                            sleep(1)
                            WebDriverWait(self.driver, 10).until(
                                (ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                            sleep(1)
                        except:
                            self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                            self.sheet[f'Q{i}'] = 'Cant Save'
                            print('Cant Save')
                            print(traceback.print_exc())

                    else:
                        self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                        self.sheet[f'Q{i}'] = "Can't find element"
                except:
                    print(traceback.print_exc())

            for column in range(1, 18):
                try:
                    self.sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                     wrap_text=True)
                except:
                    pass
            self.sheet[f'D{i}'] = seller
            self.sheet[f'E{i}'] = price
            try:
                if 'express' in express:
                    self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
            except:
                pass
            self.sheet[f'F{i}'] = express
            self.sheet[f'G{i}'] = only
            self.sheet[f'H{i}'] = new_price
            self.sheet[f'I{i}'] = static
            try:
                self.sheet[f'L{i}'] = avp
                self.sheet[f'M{i}'] = avn
            except:
                pass
            self.sheet[f'P{i}'] = Note
            self.book.save(self.file_save)
            print(
                f'{i} - {bar} > seller : {seller}, price : {price}, new price : {new_price}, u only : {only}, static : {static}')

            #########################################
            if thread.p == 'pause' or thread.p == 'p':
                start = input('(Type start or s and press Enter)>>>')
                while True:
                    if start == 'start' or start == 's':
                        thread = Thread()
                        thread.start()
                        break
                    elif start == 'exit' or start == 'e':
                        self.driver.quit()
                        exit()
                        break
                    else:
                        start = input()
            elif thread.p == 'exit' or thread.p == 'e':
                self.driver.quit()
                break
            sleep(1)
        self.driver.quit()


if __name__ == "__main__":
    file = input('Enter file name : ').strip()
    main_seller = input('Main seller : ')
    email = input('Enter your email : ')
    password = input('Enter your password : ')
    percent = float(input('Enter rate percentage : '))
    avrpu = avrnu = float(input('Enter Percentage >= : '))
    avrpd = avrnd = float(input('Enter Percentage < : '))
    content_inp = int(input('Enter minimum quantity : '))
    content_new = int(input('Enter new quantity : '))

    while True:
        my_bot = SouCode(file_name=file)
        my_bot.Search()
