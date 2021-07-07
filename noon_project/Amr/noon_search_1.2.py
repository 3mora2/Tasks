from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.webdriver import WebDriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook, Workbook
from selenium.webdriver.common.by import By
from openpyxl.styles import Alignment
from selenium import webdriver
from time import sleep
import traceback


class SouCode:
    driver: WebDriver

    def __init__(self):
        while True:
            try:
                file_name = input('- Enter file name : ')  # 'data.xlsx'
                book = load_workbook(file_name)
                sheet = book.active
                break
            except FileNotFoundError:
                print('- No such file or directory')
            except Exception as e:
                print(e)

        self.book_new = Workbook()
        self.sheet_new = self.book_new.active
        self.sheet_new.cell(1, 1).value = 'x_noon_sku'
        self.sheet_new.cell(1, 2).value = 'Price'
        self.sheet_new.cell(1, 3).value = 'Price 2'
        self.sheet_new.cell(1, 4).value = 'URL'

        self.barcode_exit = self.Get_Barcode_From_Excel(sheet)
        self.Links = self.Get_Links_Search(sheet)
        print('- Number Barcode : ', len(self.barcode_exit), ', Number Links', len(self.Links))

    def Get_Barcode_From_Excel(self, sheet):
        barcode_exit = []
        end = sheet.max_row
        for row in range(2, end + 1):
            if sheet.cell(row, 1).value is not None:
                barcode_exit.append(sheet.cell(row, 1).value.strip())

        barcode_exit = list(set(barcode_exit))
        return barcode_exit

    def Get_Links_Search(self, sheet):
        Links = []
        end = sheet.max_row
        for row in range(2, end + 1):
            if sheet.cell(row, 2).value is not None:
                Links.append(sheet.cell(row, 2).value)

        Links = list(set(Links))
        return Links

    def Search(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        # self.driver = webdriver.Chrome()
        row = 2
        l = []
        for link in self.Links:
            self.driver.get(link)
            sleep(2)
            try:
                WebDriverWait(self.driver, 20).until((ec.visibility_of_all_elements_located(
                    (By.CSS_SELECTOR, 'span + div > span'))))[-1].click()
                sleep(2)
                self.driver.find_element_by_css_selector('li[data-value="150"]').click()
                sleep(3)
            except:
                pass
            while True:
                for element in WebDriverWait(self.driver, 20).until(
                        ec.visibility_of_all_elements_located((By.CSS_SELECTOR, 'div.productContainer > a'))):
                    if element.get_attribute('href') not in l:
                        url = element.get_attribute('href')
                        l.append(url)
                        bar = url.replace('https://www.noon.com/egypt-en/', '').split('/')[1]
                        if bar.strip() in self.barcode_exit:
                            print('-  Barcode exist')
                            continue
                        price = element.find_element_by_css_selector('span.currency + strong').text
                        try:
                            price_2 = \
                                element.find_element_by_css_selector('span.oldPrice').text.replace('جنيه', '').replace('EGP', '').strip()
                        except:
                            price_2 = None
                        self.sheet_new.cell(row, 1).value = bar
                        self.sheet_new.cell(row, 2).value = price
                        self.sheet_new.cell(row, 3).value = price_2
                        self.sheet_new.cell(row, 4).value = url

                        print('- ', row, ' - ', bar)
                        row += 1
                self.book_new.save('new.xlsx')

                if 'disabled' not in self.driver.find_element_by_css_selector('.next').get_attribute('class'):
                    self.driver.find_element_by_css_selector('.next').click()
                    sleep(4)
                else:
                    break
        self.driver.quit()


class Show:
    driver: WebDriver

    def __init__(self):
        while True:
            try:
                file_name = 'new.xlsx'
                self.book = load_workbook(file_name)
                self.sheet = self.book.active
                break
            except FileNotFoundError:
                print('No such file or directory')
            except Exception as e:
                print(e)

        self.book_new = Workbook()
        self.sheet_new = self.book_new.active
        self.sheet_new[f'A1'] = 'Noon_SKU'
        self.sheet_new[f'B1'] = 'Price'
        self.sheet_new[f'C1'] = 'Price 2'
        self.sheet_new[f'D1'] = 'Maximum Price'
        self.sheet_new[f'E1'] = 'Minimum Price'
        self.sheet_new[f'F1'] = 'URL'
        self.sheet_new.column_dimensions['A'].width = 14
        self.sheet_new.column_dimensions['B'].width = 12
        self.sheet_new.column_dimensions['C'].width = 12
        self.sheet_new.column_dimensions['D'].width = 15
        self.sheet_new.column_dimensions['E'].width = 15
        self.sheet_new.column_dimensions['F'].width = 120
        try:
            self.sheet_new.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except:
            pass

        self.Links = self.Get_Links_Search()

    def Get_Links_Search(self):
        Links = []
        end = self.sheet.max_row
        for row in range(2, end + 1):
            if self.sheet.cell(row, 4).value is not None:
                Links.append((self.sheet.cell(row, 4).value, row))

        return Links

    def Show_Links(self):
        # self.driver = webdriver.Chrome()
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        row_new = 2
        for link, row in self.Links:
            self.driver.get(link)
            sleep(3)
            ans = input('- Enter (y) if you want add Or (Enter) : ')
            if ans.lower() == 'y':
                while True:
                    try:
                        Max = float(input('- Enter Maximum Price : '))
                        break
                    except ValueError:
                        print('Please Enter Number')

                while True:
                    try:
                        Min = float(input('- Enter Minimum Price : '))
                        break
                    except ValueError:
                        print('Please Enter Number')

                self.sheet_new.cell(row_new, 1).value = self.sheet.cell(row, 1).value
                self.sheet_new.cell(row_new, 2).value = self.sheet.cell(row, 2).value
                self.sheet_new.cell(row_new, 3).value = self.sheet.cell(row, 3).value
                self.sheet_new.cell(row_new, 4).value = Max
                self.sheet_new.cell(row_new, 5).value = Min
                self.sheet_new.cell(row_new, 6).value = self.sheet.cell(row, 4).value
                try:
                    self.sheet_new.cell(row_new, 1).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 2).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 3).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 4).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 5).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 6).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                except:
                    pass
                self.book_new.save('final.xlsx')
                print('- ', row_new, ' - ', self.sheet.cell(row, 1).value)
                row_new += 1
        self.driver.quit()


class Upload:
    URL_LOGIN = 'https://login.noon.partners/en/'
    URL_CAT = 'https://catalog.noon.partners/en-eg/noon-catalog'
    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    SEARCH_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="search"]')
    FIRST_ELEMENT_SELECTOR = (By.CSS_SELECTOR, 'a[href*="/en-eg/noon-catalog/preview/"]')

    SKU_TEXT_SELECTOR = (By.CSS_SELECTOR, 'div.tabsWrapper > div')
    SKU_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[name="partner_sku"]')
    SKU_BUTTON_SELECTOR = (By.CSS_SELECTOR, '.solid')

    PRICE_INPUT_SELECTOR = (By.NAME, 'price')

    QUANTITY_SITE_SELECTOR = (By.CSS_SELECTOR, 'div.placeholder')
    QUANTITY_SITE_CHOSE_SELECTOR = (By.CSS_SELECTOR, 'div.active > div > div.countryLabel')
    QUANTITY_INPUT_SELECTOR = (By.NAME, 'quantity')

    BARCODE_INSIDE_INPUT_SELECTOR = (By.CSS_SELECTOR, 'input[type="text"][placeholder="Enter Barcode"]')
    BUTTON_ADD_BARCODE_INSIDE_SELECTOR = (By.CSS_SELECTOR, 'div.eachInput > div > div.solid')
    ACTIVE_SCRIPT = 'if (document.querySelector(\'input[id="active"]\').checked == false){document.querySelector(\'input[id="active"]\').click()};'

    SAVE_CHANGE_SELECTOR = (By.CSS_SELECTOR, 'div.transparent ~ div.solid')
    SUBMIT_SELECTOR = (
        By.XPATH, '//div[text()="Submit"] | //div[contains(@class,"showAlert")]//div[contains(@class,"solid")][2]')

    def __init__(self):
        while True:
            try:
                file_name = 'final.xlsx'
                self.book = load_workbook(file_name)
                self.sheet = self.book.active
                break
            except FileNotFoundError:
                print('- No such file or directory')
            except Exception as e:
                print(e)
        self.Bar_Price = self.Get_Barcode_And_Price()
        # self.driver = webdriver.Chrome()
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.login()

    def Get_Barcode_And_Price(self):
        Barcode = []
        l = []
        end = self.sheet.max_row
        for row in range(2, end + 1):
            if self.sheet.cell(row, 1).value is not None and self.sheet.cell(row, 2).value is not None:
                if self.sheet.cell(row, 1).value not in l:
                    l.append(self.sheet.cell(row, 1).value)
                    Barcode.append((self.sheet.cell(row, 1).value, self.sheet.cell(row, 2).value))
        print(len(Barcode))
        return Barcode

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            sleep(3)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get(self.URL_CAT)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def add_new(self, name):
        sleep(1)
        try:
            self.driver.get(self.URL_CAT)
            sleep(2)
            WebDriverWait(self.driver, 10).until(
                (ec.visibility_of_element_located(self.SEARCH_INPUT_SELECTOR))).send_keys(name, Keys.ENTER)
            sleep(2)
        except:
            print(traceback.print_exc())

    def add_serial(self, bar):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(self.FIRST_ELEMENT_SELECTOR)).click()
        except:
            print('- Not found')
            return False
        try:
            sleep(1)
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_INPUT_SELECTOR)).send_keys(
                bar)
            sleep(3)
            WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(self.SKU_BUTTON_SELECTOR)).click()
            return True
        except:
            print('- Continue')
            return False

    def add_price(self, price):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a',
                                                                                         Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(Keys.CONTROL, 'a',
                                                                                         Keys.BACKSPACE)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PRICE_INPUT_SELECTOR))).send_keys(str(price))
        except:
            print(traceback.print_exc())
            print('- Cant Change Price')

    def add_qu(self, q):
        sleep(1)
        try:
            WebDriverWait(self.driver, 6).until(
                (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
        except:
            try:
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_element_located(self.QUANTITY_SITE_SELECTOR))).click()
                sleep(1)
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_all_elements_located(self.QUANTITY_SITE_CHOSE_SELECTOR)))[site].click()
                sleep(1)
                WebDriverWait(self.driver, 6).until(
                    (ec.visibility_of_element_located(self.QUANTITY_INPUT_SELECTOR))).send_keys(str(q))
                sleep(1)
            except:
                print(traceback.print_exc())
                print('- Cant Change quantity')

    def add_bar(self, sk):
        sleep(1)
        try:
            WebDriverWait(self.driver, 10).until(
                ec.visibility_of_element_located(self.BARCODE_INSIDE_INPUT_SELECTOR)).send_keys(sk)
            sleep(1)
            WebDriverWait(self.driver, 10).until(
                ec.visibility_of_element_located(self.BUTTON_ADD_BARCODE_INSIDE_SELECTOR)).click()
            sleep(1)
        except:
            print(traceback.print_exc())
            print('- Cant Change Barcode')

    def add_active(self):
        sleep(1)
        try:
            self.driver.execute_script(self.ACTIVE_SCRIPT)
        except:
            print(traceback.print_exc())
            print('Cant active')

    def upload(self):
        num = 1
        for bar, price in self.Bar_Price:
            try:
                print(num, ' - ', bar, price)
                num += 1
                self.add_new(bar)
                sleep(1)
                rr = self.add_serial(bar[::-1])
                if rr is False:
                    continue
                sleep(1)
                self.add_price(price)
                sleep(1)
                self.add_active()
                sleep(1)
                self.add_qu(quantity)
                sleep(1)
                self.add_bar(bar[::-1] + 'A')
                sleep(1)
                try:
                    sleep(1)
                    WebDriverWait(self.driver, 10).until(
                        (ec.visibility_of_element_located(self.SAVE_CHANGE_SELECTOR))).click()
                    sleep(1)
                    WebDriverWait(self.driver, 10).until(
                        (ec.visibility_of_element_located(self.SUBMIT_SELECTOR))).click()
                    sleep(1)
                except:
                    print(traceback.print_exc())
            except:
                print(traceback.print_exc())

        self.driver.quit()


if __name__ == "__main__":
    while True:
        answer = str(input('''
-----------------------------------------------------
- This script v1.2 to :-                            - 
- extract information from noon                     -
- then show to save them                            -
- then open noon and upload them                    -
----------------------------------------------------- 
- 1 - To Extract >>> Enter 1                        -
- 2 - To Show    >>> Enter 2                        -
- 3 - To upload  >>> Enter 3                        -
- 4 - To All     >>> Enter 4 or (Enter)             -
-----------------------------------------------------
>>> '''))
        if answer == '1':
            m = SouCode()
            m.Search()
            break
        elif answer == '2':
            s = Show()
            s.Show_Links()
            break
        elif answer == '3':
            email = input('Enter your email : ')  # 'mohabamr157@gmail.com'
            password = input('Enter your password : ')  # 'StarWar2020'
            quantity = 3
            site = input('- Enter Number countryLabel : ')
            try:
                site = int(site) - 1
                if site >= 0:
                    pass
                else:
                    site = 0
            except:
                site = 0
            print(site + 1)

            u = Upload()
            u.upload()
            break
        elif answer == '4' or answer == '':
            email = input('Enter your email : ')  # 'mohabamr157@gmail.com'
            password = input('Enter your password : ')  # 'StarWar2020'
            quantity = 3
            site = input('- Enter Number countryLabel : ')

            try:
                site = int(site) - 1
                if site >= 0:
                    pass
                else:
                    site = 0
            except:
                site = 0
            print(site + 1)
            m = SouCode()
            m.Search()
            s = Show()
            s.Show_Links()
            u = Upload()
            u.upload()
            break
