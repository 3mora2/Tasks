from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook, Workbook
from selenium.webdriver.common.by import By
from selenium.common.exceptions import *
from openpyxl.styles import Alignment
from selenium import webdriver
from time import sleep


class SouCode:
    driver: WebDriver

    def __init__(self):
        while True:
            try:
                file_name = input('- Enter file name : ')  # 'data.xlsx'
                book = load_workbook(file_name)
                sheet = book.active
                break
            except FileNotFoundError:
                print('- No such file or directory')
            except Exception as e:
                print(e)

        self.book_new = Workbook()
        self.sheet_new = self.book_new.active
        self.sheet_new.cell(1, 1).value = 'x_noon_sku'
        self.sheet_new.cell(1, 2).value = 'Name'
        self.sheet_new.cell(1, 3).value = 'Price'
        self.sheet_new.cell(1, 4).value = 'Price 2'
        self.sheet_new.cell(1, 5).value = 'URL'

        self.barcode_exit = self.Get_Barcode_From_Excel(sheet)
        self.Links = self.Get_Links_Search(sheet)
        print('- Number Barcode : ', len(self.barcode_exit), ', Number Links', len(self.Links))

    @staticmethod
    def Get_Barcode_From_Excel(sheet):
        barcode_exit = []
        end = sheet.max_row
        for row in range(2, end + 1):
            if sheet.cell(row, 1).value is not None:
                barcode_exit.append(sheet.cell(row, 1).value.strip())

        barcode_exit = list(set(barcode_exit))
        return barcode_exit

    @staticmethod
    def Get_Links_Search(sheet):
        Links = []
        end = sheet.max_row
        for row in range(2, end + 1):
            if sheet.cell(row, 2).value is not None:
                Links.append(sheet.cell(row, 2).value)

        Links = list(set(Links))
        return Links

    def Search(self):
        self.driver = webdriver.Chrome()
        row = 2
        l = []
        for link in self.Links:
            self.driver.get(link)
            sleep(2)
            try:
                WebDriverWait(self.driver, 20).until((ec.visibility_of_element_located(
                    (By.CSS_SELECTOR, 'header > div > div.container > div:nth-child(2) > div > button')))).click()
                sleep(2)
                self.driver.find_elements_by_css_selector('div.popDown > ul > li')[-1].click()
            except:
                pass
            while True:
                for element in WebDriverWait(self.driver, 20).until(
                        ec.visibility_of_all_elements_located((By.CSS_SELECTOR,
                                                               '.productListContainer > div > div.productListWrapper > div.productList.gridView > div > div > a.product'))):

                    if element.get_attribute('href') not in l:
                        url = element.get_attribute('href')
                        l.append(url)
                        bar = url.replace('https://www.noon.com/egypt-en/', '').split('/')[1]
                        if bar.strip() in self.barcode_exit:
                            print('-  Barcode exist')
                            continue
                        price = element.find_element_by_css_selector('div.priceRow > p > span > span > span.value').text
                        try:
                            price_2 = \
                                element.find_elements_by_css_selector('div.priceRow > p > span > span > span.value')[
                                    1].text
                        except:
                            price_2 = None
                        self.sheet_new.cell(row, 1).value = bar
                        self.sheet_new.cell(row, 2).value = name
                        self.sheet_new.cell(row, 3).value = price
                        self.sheet_new.cell(row, 4).value = price_2
                        self.sheet_new.cell(row, 5).value = url

                        print('- ', row, ' - ', bar)
                        row += 1
                self.book_new.save('new.xlsx')

                if self.driver.find_element_by_css_selector('.nextLink').get_attribute('aria-disabled') != 'true':
                    self.driver.find_element_by_css_selector('.nextLink').click()
                    sleep(3)
                else:
                    break
        self.driver.quit()


if __name__ == "__main__":
    m = SouCode()
    m.Search()
