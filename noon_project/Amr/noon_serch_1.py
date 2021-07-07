from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook, Workbook
from selenium.webdriver.common.by import By
from selenium.common.exceptions import *
from openpyxl.styles import Alignment
from selenium import webdriver
from time import sleep
import logging as log

log.basicConfig(filename='error.log', filemode='w', format='- %(asctime)s - %(message)s')


class SouCode:
    driver: WebDriver

    def __init__(self):
        while True:
            try:
                file_name = input('- Enter file name : ')  # 'data.xlsx'
                book = load_workbook(file_name)
                sheet = book.active
                break
            except FileNotFoundError:
                print('- No such file or directory')
            except Exception as e:
                print(e)

        self.book_new = Workbook()
        self.sheet_new = self.book_new.active
        self.sheet_new.cell(1, 1).value = 'x_noon_sku'
        self.sheet_new.cell(1, 2).value = 'Price'
        self.sheet_new.cell(1, 3).value = 'Price 2'
        self.sheet_new.cell(1, 4).value = 'URL'

        self.barcode_exit = self.Get_Barcode_From_Excel(sheet)
        self.Links = self.Get_Links_Search(sheet)
        print('- Number Barcode : ', len(self.barcode_exit), ', Number Links', len(self.Links))

    def Get_Barcode_From_Excel(self, sheet):
        barcode_exit = []
        end = sheet.max_row
        for row in range(2, end + 1):
            if sheet.cell(row, 1).value is not None:
                barcode_exit.append(sheet.cell(row, 1).value.strip())

        barcode_exit = list(set(barcode_exit))
        return barcode_exit

    def Get_Links_Search(self, sheet):
        Links = []
        end = sheet.max_row
        for row in range(2, end + 1):
            if sheet.cell(row, 2).value is not None:
                Links.append(sheet.cell(row, 2).value)

        Links = list(set(Links))
        return Links

    def Search(self):
        self.driver = webdriver.Chrome()
        row = 2
        l = []
        for link in self.Links:
            self.driver.get(link)
            sleep(2)
            try:
                WebDriverWait(self.driver, 20).until((ec.visibility_of_element_located(
                    (By.CSS_SELECTOR, 'header > div > div.container > div:nth-child(2) > div > button')))).click()
                sleep(2)
                self.driver.find_elements_by_css_selector('div.popDown > ul > li')[-1].click()
            except:
                pass
            while True:
                for element in WebDriverWait(self.driver, 20).until(
                        ec.visibility_of_all_elements_located((By.CSS_SELECTOR,
                                                               '.productListContainer > div > div.productListWrapper > div.productList.gridView > div > div > a.product'))):

                    if element.get_attribute('href') not in l:
                        url = element.get_attribute('href')
                        l.append(url)
                        bar = url.replace('https://www.noon.com/egypt-en/', '').split('/')[1]
                        if bar.strip() in self.barcode_exit:
                            print('-  Barcode exist')
                            continue
                        price = element.find_element_by_css_selector('div.priceRow > p > span > span > span.value').text
                        try:
                            price_2 = \
                                element.find_elements_by_css_selector('div.priceRow > p > span > span > span.value')[
                                    1].text
                        except:
                            price_2 = None
                        self.sheet_new.cell(row, 1).value = bar
                        self.sheet_new.cell(row, 2).value = price
                        self.sheet_new.cell(row, 3).value = price_2
                        self.sheet_new.cell(row, 4).value = url

                        print('- ', row, ' - ', bar)
                        row += 1
                self.book_new.save('new.xlsx')

                if self.driver.find_element_by_css_selector('.nextLink').get_attribute('aria-disabled') != 'true':
                    self.driver.find_element_by_css_selector('.nextLink').click()
                    sleep(3)
                else:
                    break
        self.driver.quit()


class Show:
    driver: WebDriver

    def __init__(self):
        while True:
            try:
                file_name = 'new.xlsx'
                self.book = load_workbook(file_name)
                self.sheet = self.book.active
                break
            except FileNotFoundError:
                print('No such file or directory')
            except Exception as e:
                print(e)

        self.book_new = Workbook()
        self.sheet_new = self.book_new.active
        self.sheet_new[f'A1'] = 'Noon_SKU'
        self.sheet_new[f'B1'] = 'Price'
        self.sheet_new[f'C1'] = 'Price 2'
        self.sheet_new[f'D1'] = 'Maximum Price'
        self.sheet_new[f'E1'] = 'Minimum Price'
        self.sheet_new[f'F1'] = 'URL'
        self.sheet_new.column_dimensions['A'].width = 14
        self.sheet_new.column_dimensions['B'].width = 12
        self.sheet_new.column_dimensions['C'].width = 12
        self.sheet_new.column_dimensions['D'].width = 15
        self.sheet_new.column_dimensions['E'].width = 15
        self.sheet_new.column_dimensions['F'].width = 120
        try:
            self.sheet_new.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet_new.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except:
            pass

        self.Links = self.Get_Links_Search()

    def Get_Links_Search(self):
        Links = []
        end = self.sheet.max_row
        for row in range(2, end + 1):
            if self.sheet.cell(row, 4).value is not None:
                Links.append((self.sheet.cell(row, 4).value, row))

        return Links

    def Show_Links(self):
        self.driver = webdriver.Chrome()
        row_new = 2
        for link, row in self.Links:
            self.driver.get(link)
            sleep(3)
            ans = input('- Enter (y) if you want add Or (Enter) : ')
            if ans.lower() == 'y':
                while True:
                    try:
                        Max = float(input('- Enter Maximum Price : '))
                        break
                    except ValueError:
                        print('Please Enter Number')

                while True:
                    try:
                        Min = float(input('- Enter Minimum Price : '))
                        break
                    except ValueError:
                        print('Please Enter Number')

                self.sheet_new.cell(row_new, 1).value = self.sheet.cell(row, 1).value
                self.sheet_new.cell(row_new, 2).value = self.sheet.cell(row, 2).value
                self.sheet_new.cell(row_new, 3).value = self.sheet.cell(row, 3).value
                self.sheet_new.cell(row_new, 4).value = Max
                self.sheet_new.cell(row_new, 5).value = Min
                self.sheet_new.cell(row_new, 6).value = self.sheet.cell(row, 4).value
                try:
                    self.sheet_new.cell(row_new, 1).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 2).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 3).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 4).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 5).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                    self.sheet_new.cell(row_new, 6).alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                except:
                    pass
                self.book_new.save('final.xlsx')
                print('- ', row_new, ' - ', self.sheet.cell(row, 1).value)
                row_new += 1
        self.driver.quit()


class Upload:
    def __init__(self):
        while True:
            try:
                file_name = 'final.xlsx'
                self.book = load_workbook(file_name)
                self.sheet = self.book.active
                break
            except FileNotFoundError:
                print('- No such file or directory')
            except Exception as e:
                print(e)
        self.Bar_Price = self.Get_Barcode_And_Price()
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()

    def Get_Barcode_And_Price(self):
        Barcode = []
        l = []
        end = self.sheet.max_row
        for row in range(2, end + 1):
            if self.sheet.cell(row, 1).value is not None and self.sheet.cell(row, 2).value is not None:
                if self.sheet.cell(row, 1).value not in l:
                    l.append(self.sheet.cell(row, 1).value)
                    Barcode.append((self.sheet.cell(row, 1).value, self.sheet.cell(row, 2).value))
        print(len(Barcode))
        return Barcode

    def upload(self):
        num = 1
        for bar, price in self.Bar_Price:
            try:
                print(num, ' - ', bar, price)
                num += 1
                self.driver.get('https://catalog.noon.partners/en-eg/noon-catalog')
                sleep(6)
                if 'login.noon' in self.driver.current_url:
                    try:
                        self.driver.find_element_by_name('email').send_keys(email)
                        sleep(1)
                        self.driver.find_element_by_name('password').send_keys(password)
                        sleep(1)
                        self.driver.find_element_by_css_selector('#formContainer > button').click()
                        sleep(5)
                    except:
                        log.error('login error')
                        break
                    while True:
                        if 'login.noon' not in self.driver.current_url:
                            break
                    if '/noon-catalog' not in self.driver.current_url:
                        self.driver.get('https://catalog.noon.partners/en-eg/noon-catalog')
                    sleep(6)
                self.driver.find_element_by_css_selector('input.searchInput').send_keys(bar)
                sleep(2)
                try:
                    self.driver.find_element_by_css_selector('input.searchInput').send_keys(Keys.ENTER)
                except:
                    pass
                sleep(4)
                try:
                    WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located(
                        (By.CSS_SELECTOR, 'div.content div.ctr > div.result > div.resultCtr > a'))).click()
                    sleep(3)
                except:
                    log.error(f'error when search : {bar} - {price} ')
                    continue

                try:
                    WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located(
                        (By.CSS_SELECTOR, 'div.eachForm > div.inputCtr > div > input.input'))).send_keys(bar[::-1])
                    sleep(2)
                    # self.driver.find_element_by_css_selector('div.eachForm > div.inputCtr > div > input.input').send_keys(bar[::-1])
                except TimeoutException:
                    print('- Continue')
                    continue

                try:
                    self.driver.find_element_by_css_selector('div.rightView > div.btnCtr > button').click()
                    sleep(2)
                except NoSuchElementException:
                    WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located(
                        (By.CSS_SELECTOR, 'div.rightView > div.btnCtr > button'))).click()
                except ElementClickInterceptedException:
                    _ = self.driver.find_element_by_css_selector(
                        'div.rightView > div.btnCtr > button').location_once_scrolled_into_view
                    self.driver.find_element_by_css_selector('div.rightView > div.btnCtr > button').click()
                except Exception as e:
                    log.error(f'error create bar : {bar} - {price} ')
                    continue


                sleep(4)
                WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located(
                    (By.CSS_SELECTOR, 'div.headerCtr > div.mapButtonCtr > div > div > input'))).click()

                # self.driver.find_element_by_css_selector('div.headerCtr > div.mapButtonCtr > div > div > input').click()
                sleep(1)
                self.driver.find_element_by_name('price_eg').clear()
                sleep(1)
                self.driver.find_element_by_name('price_eg').send_keys(price)
                sleep(1)
                self.driver.find_element_by_name('quantity').send_keys(quantity)
                sleep(1)
                self.driver.find_element_by_css_selector('div.inputCtr > input').send_keys(bar[::-1] + 'A')
                sleep(2)
                try:
                    self.driver.find_element_by_css_selector('div.inputCtr > div.btnCtr > button').click()
                except ElementClickInterceptedException:
                    _ = self.driver.find_element_by_css_selector(
                        'div.inputCtr > div.btnCtr').location_once_scrolled_into_view
                    sleep(2)
                    try:
                        self.driver.find_element_by_css_selector('div.inputCtr > div.btnCtr > button').click()
                    except Exception as e:
                        log.error(f'error add bar : {bar} - {price} ')
                sleep(2)
                self.driver.find_element_by_css_selector('div.fixedBottom > button').click()
                sleep(4)
            except Exception as e:
                log.error(f'{e} : {bar} - {price} ', exc_info=True)
                print(e)

        self.driver.quit()


if __name__ == "__main__":
    while True:
        answer = str(input('''
-----------------------------------------------------
- This script to :-                                 - 
- extract information from noon                     -
- then show to save them                            -
- then open noon and upload them                    -
----------------------------------------------------- 
- 1 - To Extract >>> Enter 1                        -
- 2 - To Show    >>> Enter 2                        -
- 3 - To upload  >>> Enter 3                        -
- 4 - To All     >>> Enter 4 or (Enter)             -
-----------------------------------------------------
>>> '''))
        if answer == '1':
            m = SouCode()
            m.Search()
            break
        elif answer == '2':
            s = Show()
            s.Show_Links()
            break
        elif answer == '3':
            email = input('Enter your email : ')  # 'mohabamr157@gmail.com'
            password = input('Enter your password : ')  # 'StarWar2020'
            quantity = 3
            u = Upload()
            u.upload()
            break
        elif answer == '4' or answer == '':
            email = input('Enter your email : ')  # 'mohabamr157@gmail.com'
            password = input('Enter your password : ')  # 'StarWar2020'
            quantity = 3
            m = SouCode()
            m.Search()
            s = Show()
            s.Show_Links()
            u = Upload()
            u.upload()
            break
