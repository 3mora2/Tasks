from datetime import datetime
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.keys import Keys
from selenium import webdriver

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from time import sleep
from requests_html import HTML
import threading
'''
* Update 'div.sellerDetails > p > strong' to 'div.soldBy' 10-1--2020

'''


class Thread(threading.Thread):
    p = None

    def __init__(self):
        super(Thread, self).__init__()

    def run(self):
        while True:
            self.p = input()
            if self.p == 'pause' or self.p == 'p':
                break
            elif self.p == 'exit' or self.p == 'e':
                break


class SouCode:
    driver: WebDriver
    URL_SEARCH = 'https://www.noon.com/egypt-en/'

    def __init__(self, file_name):
        self.file_save = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + file_name
        try:
            self.book = load_workbook(file_name)
            self.sheet = self.book.active
            self.sheet[f'A1'] = 'Noon_SKU'
            self.sheet[f'B1'] = 'Minimum Price'
            self.sheet[f'C1'] = 'Maximum Price'
            self.sheet[f'D1'] = 'Seller Name'
            self.sheet[f'E1'] = 'Price'
            self.sheet[f'F1'] = 'Express'
            self.sheet[f'G1'] = 'Only'
            self.sheet[f'H1'] = 'New price'
            self.sheet[f'I1'] = 'Status'
            self.sheet[f'J1'] = 'Account'
            self.sheet[f'K1'] = 'Pass'
            self.sheet[f'L1'] = 'Change Value +'
            self.sheet[f'M1'] = 'Change Value -'
            self.sheet[f'N1'] = 'Exception'
            self.sheet[f'O1'] = 'Quantity'
            self.sheet[f'P1'] = 'Note 1'
            self.sheet[f'Q1'] = 'Note 2'
            self.sheet.column_dimensions['A'].width = 14
            self.sheet.column_dimensions['B'].width = 15
            self.sheet.column_dimensions['C'].width = 15
            self.sheet.column_dimensions['D'].width = 15
            self.sheet.column_dimensions['E'].width = 10
            self.sheet.column_dimensions['F'].width = 14
            self.sheet.column_dimensions['G'].width = 10
            self.sheet.column_dimensions['H'].width = 11
            self.sheet.column_dimensions['I'].width = 18
            self.sheet.column_dimensions['J'].width = 30
            self.sheet.column_dimensions['K'].width = 20
            self.sheet.column_dimensions['L'].width = 16
            self.sheet.column_dimensions['M'].width = 16
            self.sheet.column_dimensions['N'].width = 16
            self.sheet.column_dimensions['O'].width = 13
            self.sheet.column_dimensions['P'].width = 85
            self.sheet.column_dimensions['Q'].width = 50
        except Exception as e:
            print(e)
            exit()

    def Search(self):
        Note = None
        self.sheet['J2'] = str(email)
        self.sheet['J3'] = str(main_seller)
        self.sheet['K2'] = str(password)

        end_cell = self.sheet.max_row

        exception_seller = [self.sheet[f'N{i}'].value for i in range(2, end_cell) if
                            self.sheet[f'N{i}'].value is not None]
        group1 = [self.sheet[f'R{i}'].value for i in range(2, end_cell) if self.sheet[f'R{i}'].value is not None]
        group2 = [self.sheet[f'S{i}'].value for i in range(2, end_cell) if self.sheet[f'S{i}'].value is not None]
        group3 = [self.sheet[f'T{i}'].value for i in range(2, end_cell) if self.sheet[f'T{i}'].value is not None]
        group4 = [self.sheet[f'U{i}'].value for i in range(2, end_cell) if self.sheet[f'U{i}'].value is not None]
        group5 = [self.sheet[f'V{i}'].value for i in range(2, end_cell) if self.sheet[f'V{i}'].value is not None]

        self.thr = Thread()
        self.thr.start()
        self.driver = webdriver.Chrome()

        for i in range(2, end_cell + 1):
            bar = self.sheet[f'A{i}'].value
            if bar is None:
                continue

            max_pr = float(self.sheet[f'C{i}'].value)
            min_pr = float(self.sheet[f'B{i}'].value)
            try:
                self.driver.get(f'{self.URL_SEARCH}{bar}/p/')
                sleep(2)

                r = HTML(html=self.driver.page_source)
                price = float(r.find('span.sellingPrice span.value', first=True).text)
                seller = r.find('p.sellerName a', first=True).text
                express = r.find('div.bottomRow div.container img', first=True).attrs['alt']

                if seller in group1:
                    avp = avn = float(group1[0])
                elif seller in group2:
                    avp = avn = float(group2[0])
                elif seller in group3:
                    avp = avn = float(group3[0])
                elif seller in group4:
                    avp = avn = float(group4[0])
                elif seller in group5:
                    avp = avn = float(group5[0])
                else:
                    try:
                        rat1 = float(r.find('div.container > div> p:nth-child(3)')[0].text.replace('%', ''))
                        rat2 = float(r.find('div.container > div> p:nth-child(3)')[1].text.replace('%', ''))
                        print(rat1 + rat2)
                        if rat1 + rat2 >= percent:
                            avp = avrpu
                            avn = avrnu
                        else:
                            avp = avrpd
                            avn = avrnd
                    except:
                        rat2 = None
                        rat1 = None
                        avp = avrpd
                        avn = avrnd
                        print(rat1, rat2)
                Offer_number = len(
                    self.driver.find_elements_by_css_selector('ul.offersList > li'))

                ########################################################################################################
                if main_seller in seller:
                    if Offer_number == 0:
                        new_price = max_pr
                        only = 'Only'
                        static = 'Maximum Price'
                        print(' Only me. So, maximum price is ', max_pr)
                        Note = ' Only me. So, maximum price is ' + str(max_pr)
                    else:
                        only = 'No'
                        best = []
                        self.driver.find_element_by_css_selector('span.cta').click()
                        sleep(4)
                        offer = self.driver.find_elements_by_css_selector('ul.offersList > li')
                        try:
                            for ELEMENT in offer:
                                if offer.index(ELEMENT) == 4:
                                    break
                                if 'selectedItem' not in ELEMENT.get_attribute('class'):

                                    best.append((float(ELEMENT.find_element_by_css_selector('span.sellingPrice > span > span.value').text),
                                         ELEMENT.find_element_by_css_selector('div.soldBy').text,
                                         ELEMENT.find_element_by_css_selector('.container img').get_attribute('alt')))

                        except:
                            offer = self.driver.find_elements_by_css_selector('ul.offersList > li')
                            for ELEMENT in offer:
                                if offer.index(ELEMENT) == 4:
                                    break
                                if 'selectedItem' not in ELEMENT.get_attribute('class'):

                                    best.append(
                                        (float(ELEMENT.find_element_by_css_selector(
                                            'span.sellingPrice > span > span.value').text),
                                         ELEMENT.find_element_by_css_selector('div.soldBy').text,
                                         ELEMENT.find_element_by_css_selector('.container img').get_attribute('alt')))

                        if best[0][1] in exception_seller:
                            new_price = max_pr
                            print(' Exception. So, maximum price. ')
                            Note = ' Exception. So, maximum price. '
                            static = 'Maximum price'
                        elif float(best[0][0]) < min_pr:
                            new_price = None
                            Note = 'next price < Minimum'
                            static = 'freeze'
                        else:

                            y = float(best[0][0])
                            new_price = y - y * (avp / 100)
                            if new_price < price:
                                Note = ' Warning: New price is smaller than yours. So, price will be your price or minimum. ' + str(
                                    new_price)
                                print(Note, new_price)
                                new_price = None if price > min_pr else min_pr

                            print('New price after +', new_price)
                            static = 'Change price to +'
                elif 'express' in express:
                    only = 'No'
                    new_price = None
                    Note = 'No change'
                    static = 'Express'
                elif seller in exception_seller:
                    only = 'No'
                    static = 'Exception'
                    new_price = max_pr
                    Note = 'Exception, So maximum price'
                    print(Note)
                else:
                    only = 'No'
                    if price < min_pr:
                        new_price = max_pr
                        static = 'Maximum'
                        Note = 'Price seller < minimum'
                    else:
                        static = 'Change price to -'
                        y = float(price)
                        new_price = y - y * (avn / 100)
                        if new_price < min_pr:
                            Note = ' Warning: New price is smaller than minimum price. So, will keep minimum price. '
                            print(Note)
                            new_price = min_pr
                            static = 'Minimum price'

                sleep(5)
            except Exception as e:
                seller = price = express = only = new_price = None
                Note = f'{e}'
                self.sheet[f'P{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                static = "Not found"

            try:
                if new_price is not None:
                    self.driver.get('https://catalog.noon.partners/en-eg/catalog')
                    sleep(5)
                    if 'login.noon' in self.driver.current_url:
                        self.driver.find_element_by_name('email').send_keys(email)
                        self.driver.find_element_by_name('password').send_keys(password)
                        self.driver.find_element_by_css_selector('#formContainer > button').click()
                        sleep(5)
                        while True:
                            if 'login.noon' not in self.driver.current_url:
                                break
                        self.driver.get('https://catalog.noon.partners/en-eg/catalog')
                        sleep(4)
                    self.driver.find_element_by_css_selector('input.searchInput').send_keys(bar)
                    self.driver.find_element_by_css_selector('input.searchInput').send_keys(Keys.ENTER)
                    sleep(6)
                    try:
                        self.driver.find_element_by_css_selector('td:nth-child(1) > div > a').click()
                    except Exception as e:
                        print(e)
                        self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                        self.sheet[f'Q{i}'] = 'Cant find no click'
                        print('Cant find', 'no click')
                    try:
                        sleep(4)
                        self.driver.find_element_by_name('price_eg').clear()
                        sleep(1)
                        self.driver.find_element_by_name('price_eg').send_keys(str(new_price))

                        sleep(2)
                        content = self.driver.find_element_by_css_selector('div.stockNet').text.split('\n')[1]
                        print(content)
                        self.sheet[f'O{i}'] = content
                        if int(content) < int(content_inp):
                            self.driver.find_element_by_name('quantity').send_keys(str(content_new))
                            print('Change Quantity to ', content_new)
                            self.sheet[f'Q{i}'] = 'Warning: Quantity is ' + str(content)
                            self.sheet[f'Q{i}'] = 'Change Quantity to ' + str(content_new)
                            try:
                                self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                                self.sheet[f'R{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                                self.sheet[f'O{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                            except:
                                pass
                        self.driver.find_element_by_css_selector('div.fixedBottom > button').click()
                        sleep(5)
                    except Exception as e:
                        print(e)
                        self.sheet[f'Q{i}'].fill = PatternFill(fill_type='solid', fgColor='ff0000')
                        self.sheet[f'Q{i}'] = 'Cant change'
                        print('Cant change other')

            except Exception as e:
                print(e)

            for column in range(1, 18):
                try:
                    self.sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                     wrap_text=True)
                except:
                    pass
            self.sheet[f'D{i}'] = seller
            self.sheet[f'E{i}'] = price
            try:
                if 'express' in express:
                    self.sheet[f'F{i}'].fill = PatternFill(fill_type='solid', fgColor='86E507')
            except:
                pass
            self.sheet[f'F{i}'] = express
            self.sheet[f'G{i}'] = only
            self.sheet[f'H{i}'] = new_price
            self.sheet[f'I{i}'] = static
            try:
                self.sheet[f'L{i}'] = avp
                self.sheet[f'M{i}'] = avn
            except:
                pass
            self.sheet[f'P{i}'] = Note
            self.book.save(self.file_save)
            print(
                f'{i} - {bar} > seller : {seller}, price : {price}, new price : {new_price}, u only : {only}, static : {static}')

            #########################################
            if self.thr.p == 'pause' or self.thr.p == 'p':
                start = input('(Type start or s and press Enter)>>>')
                while True:
                    if start == 'start' or start == 's':
                        self.thr = Thread()
                        self.thr.start()
                        break
                    elif start == 'exit' or start == 'e':
                        self.driver.quit()
                        exit()
                    else:
                        start = input()
            elif self.thr.p == 'exit' or self.thr.p == 'e':
                self.driver.quit()
                exit()

        self.driver.quit()


if __name__ == "__main__":
    file = input('Enter file name : ').strip()
    main_seller = input('Main seller : ')
    email = input('Enter your email : ')
    password = input('Enter your password : ')
    percent = float(input('Enter rate percentage : '))
    avrpu = avrnu = float(input('Enter Percentage >= : '))
    avrpd = avrnd = float(input('Enter Percentage < : '))
    content_inp = int(input('Enter minimum quantity : '))
    content_new = int(input('Enter new quantity : '))

    while True:
        my_bot = SouCode(file_name=file)
        my_bot.Search()
