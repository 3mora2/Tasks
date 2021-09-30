from selenium.webdriver import ActionChains
from selenium.webdriver.firefox.webdriver import WebDriver
import random
from webdriver_manager.firefox import GeckoDriverManager
from time import sleep
from PySide2.QtCore import QThread, Signal
from PySide2.QtWidgets import QMainWindow, QApplication, QFileDialog, QCheckBox, QRadioButton, QListWidgetItem
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import os
import sys
import warnings
from main import Ui_MainWindow
from openpyxl import load_workbook, Workbook

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.Handel_Button()
        self.Main = Main()
        self.Main.finish.connect(self.addlist)
        self.Main.end.connect(self.end)

    def Handel_Button(self):
        self.ui.pushButton.clicked.connect(self.Open_File_Test)
        self.ui.pushButton_2.clicked.connect(self.start)
        self.ui.pushButton_3.clicked.connect(self.open)
        self.ui.pushButton_4.clicked.connect(self.Open_File_Add)
        path = os.path.expanduser("~") + '\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles'
        for file in os.listdir(path):
            prof = file
            box = QRadioButton(prof)
            item = QListWidgetItem()
            self.ui.listWidget_2.addItem(item)
            self.ui.listWidget_2.setItemWidget(item, box)

    def Open_File_Test(self):
        save_file = QFileDialog.getOpenFileName(self, "اختار", '', '*.xlsx')[0]
        if save_file != '':
            sheet = load_workbook(save_file).active
            self.Main.sheet = sheet
            self.ui.pushButton.setDisabled(True)
            self.ui.lineEdit.setText(str(save_file))
            # with open(save_file, 'r', encoding="utf-8") as f:
            #     self.Main.data = f.read()

    def Open_File_Add(self):
        save_file = self.ui.textEdit.toPlainText().strip()
        if save_file != '':
            self.ui.pushButton_4.setDisabled(True)
            self.Main.extra_text = save_file

    def addlist(self):
        print('- add')
        sleep(1)
        for i, element in enumerate(self.Main.driver.find_elements_by_css_selector('[aria-label="تصنيفات"] div label')):
            box = QCheckBox(element.text.strip())
            item = QListWidgetItem()
            self.ui.listWidget.addItem(item)
            self.ui.listWidget.setItemWidget(item, box)

    def start(self):
        cb_list = [self.ui.listWidget.itemWidget(self.ui.listWidget.item(i))
                   for i in range(self.ui.listWidget.count())]
        chooses = []
        for cb in cb_list:
            if cb.isChecked():
                chooses.append(cb.text())
        print(chooses)
        self.Main.cat = chooses
        self.Main.start()
        self.ui.listWidget.setDisabled(True)
        self.ui.pushButton.setDisabled(True)
        self.ui.pushButton_2.setDisabled(True)
        self.ui.pushButton_3.setDisabled(True)

    def open(self):
        path = os.path.expanduser("~") + '\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles'
        cb_list = [self.ui.listWidget_2.itemWidget(self.ui.listWidget_2.item(i))
                   for i in range(self.ui.listWidget_2.count())]

        for cb in cb_list:
            if cb.isChecked():
                self.Main.prof = path + '\\' + cb.text()

        self.Main.open()
        self.ui.pushButton_3.setDisabled(True)

    def end(self):
        self.ui.listWidget.setDisabled(False)
        self.ui.pushButton.setDisabled(False)
        self.ui.pushButton_2.setDisabled(False)
        self.ui.pushButton_3.setDisabled(False)


class Main(QThread):
    driver: WebDriver
    finish = Signal()
    end = Signal()

    def __init__(self):
        super().__init__()
        try:
            self.old_book = load_workbook('old.xlsx')
            self.old_sheet = self.old_book.active
        except:
            self.old_book = Workbook()
            self.old_sheet = self.old_book.active

        self.old = [self.old_sheet.cell(i, 1).value.strip() for i in range(1, self.old_sheet.max_row + 1) if
                    self.old_sheet.cell(i, 1).value]

        self.current = self.old_sheet.max_row + 2
        # self.f = open('Done.txt', 'a+', encoding="utf-8")
        # self.old = self.f.read()
        # path = os.path.expanduser("~") + '\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles'
        # for file in os.listdir(path):
        #     if file.endswith('release'):
        #         self.prof = path + '\\' + file

    def open(self):
        print('- open browser')
        # self.prof = r'C:\Users\3mora\AppData\Roaming\Mozilla\Firefox\Profiles\63z4wsqh.default-release'
        user_data = self.prof  # r'C:\Users\3mora\AppData\Roaming\Mozilla\Firefox\Profiles\63z4wsqh.default-release'
        fp = webdriver.FirefoxProfile(user_data)
        options = webdriver.FirefoxOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-setuid-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        try:
            self.driver = webdriver.Firefox(firefox_profile=fp, options=options)
        except:
            self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install(), firefox_profile=fp,
                                            options=options)
        self.driver.get('https://badayih.com/wp-admin/post-new.php?wp-post-new-reload=true')
        input('- Enter: ')
        try:
            self.driver.find_element_by_css_selector('[role="dialog"] [aria-label="إغلاق النافذة"]').click()
        except:
            pass
        self.open_setting()
        sleep(1)
        self.get_cat()
        sleep(1)
        self.finish.emit()
        print('- end')

    def open_setting(self):
        if 'true' not in self.driver.find_element_by_css_selector('[aria-label="الإعدادات"]').get_attribute(
                'aria-pressed'):
            self.driver.find_element_by_css_selector('[aria-label="الإعدادات"]').click()

        self.driver.find_element_by_css_selector('[data-label="مقالة"]').click()

    def get_cat(self):
        if 'false' in self.driver.find_element_by_xpath("//button[contains(text(), 'تصنيفات')]").get_attribute(
                'aria-expanded'):
            self.driver.find_element_by_xpath("//button[contains(text(), 'تصنيفات')]").click()
        while True:
            if self.driver.find_elements_by_css_selector('[aria-label="تصنيفات"] div label').__len__() > 0:
                break
            else:
                sleep(2)

    def run(self):
        for i in range(2, self.sheet.max_row + 1):
            print(i)
            post = self.sheet.cell(i, 1).value
            if post:
                post = post.strip()
                if post in self.old:
                    print('- found')
                    continue
                if post.replace('*', '').strip() == '':
                    print('- ****')
                    continue
                if post or post.replace('\n', '').replace(' ', '') != '':
                    self.post(post + '\n' + self.extra_text)
                    self.old_sheet.cell(self.current, 1).value = post
                    self.current += 1
                    self.old_book.save('old.xlsx')
                    ss = random.choice(range(60 * 10, 60 * 40))
                    print('Wait:', ss)
                    sleep(ss)
        self.end.emit()

    def do_action(self, element, text=None):
        actions = ActionChains(self.driver)
        if text:
            actions.send_keys_to_element(element, text).perform()
        else:
            actions.click(element).perform()

    def post(self, post):
        # print(post)
        self.driver.get('https://badayih.com/wp-admin/post-new.php?wp-post-new-reload=true')
        sleep(6)

        if len(self.driver.find_elements_by_css_selector('textarea[aria-label="HTML"]')) > 0:
            self.driver.find_element_by_css_selector('textarea[aria-label="HTML"]').send_keys(
                post.replace('\n', '<br>'))
        else:
            try:
                self.driver.find_element_by_css_selector('button[aria-label="إضافة مكوّن"]').click()
            except:
                WebDriverWait(self.driver, 20).until((ec.presence_of_element_located(
                    (By.CSS_SELECTOR, '.edit-post-visual-editor .block-editor-writing-flow div p')))).click()
                sleep(1)
                self.driver.find_element_by_css_selector('button[aria-label="إضافة مكوّن"]').click()
            sleep(1)
            self.driver.find_element_by_css_selector('.block-editor-inserter__search-input').send_keys('html')
            sleep(1)
            self.driver.find_element_by_css_selector(
                '.components-button.block-editor-block-types-list__item.editor-block-list-item-html').click()
            sleep(1)
            self.driver.find_element_by_css_selector('textarea[aria-label="HTML"]').send_keys(
                post.replace('\n', '<br>'))

        # element = WebDriverWait(self.driver, 20).until((ec.presence_of_element_located(
        #                 (By.CSS_SELECTOR, '.edit-post-visual-editor .block-editor-writing-flow div p'))))
        # self.do_action(element, post)
        # self.driver.execute_script("""var elm = arguments[0],
        # txt = arguments[1];
        # elm.innerHTML = txt;
        # elm.dispatchEvent(new Event('keydown', {bubbles: true}));
        # elm.dispatchEvent(new Event('keypress', {bubbles: true}));
        # elm.dispatchEvent(new Event('input', {bubbles: true}));
        # elm.dispatchEvent(new Event('keyup', {bubbles: true}));""", element, post)
        sleep(3)
        title = post.split('\n')[0]
        if len(title) < 10:
            title = post[:50]

        self.driver.find_element_by_css_selector('[placeholder="إضافة عنوان"]').send_keys(title)
        sleep(3)
        self.open_setting()
        sleep(1)
        # social check
        if 'closed' in self.driver.find_element_by_css_selector('#rop_publish_now_metabox').get_attribute('class'):
            self.driver.find_element_by_css_selector('#rop_publish_now_metabox').click()

        if not self.driver.find_element_by_css_selector('[name="publish_now"]').is_selected():
            self.driver.find_element_by_css_selector('[name="publish_now"]').click()
            sleep(2)

        if not self.driver.find_element_by_css_selector('input[name="publish_now_accounts[]"]').is_selected():
            self.driver.find_element_by_css_selector('input[name="publish_now_accounts[]"]').click()
            sleep(2)

        if not self.driver.find_elements_by_css_selector('input[name="publish_now_accounts[]"]')[1].is_selected():
            self.driver.find_elements_by_css_selector('input[name="publish_now_accounts[]"]')[1].click()
            sleep(2)

        # send keyword Input
        if 'closed' in self.driver.find_element_by_css_selector('#ilj_linkdefinition').get_attribute('class'):
            self.driver.find_element_by_css_selector('#ilj_linkdefinition').click()
            sleep(1)
        self.driver.find_element_by_css_selector('#Keywords input.keywordInput').send_keys(title)
        sleep(1)
        self.driver.find_element_by_css_selector('#Keywords input.keywordInput + a').click()

        try:
            # وسوم
            if 'false' in self.driver.find_element_by_xpath("//button[contains(text(), 'وسوم')]").get_attribute(
                    'aria-expanded'):
                self.driver.find_element_by_xpath("//button[contains(text(), 'وسوم')]").click()
            # self.driver.find_element_by_css_selector('#components-form-token-input-0').send_keys(title, Keys.ENTER)
            sleep(2)
            self.driver.find_element_by_css_selector('.components-form-token-field__input').send_keys(title, Keys.ENTER)
        except:
            print('- error components')
        # category
        self.get_cat()
        for element in self.driver.find_elements_by_css_selector('[aria-label="تصنيفات"] div'):
            if element.text.strip() in self.cat:
                if element.find_element_by_tag_name('input').is_selected():
                    pass
                try:
                    element.find_element_by_tag_name('input').click()
                except:
                    element.find_element_by_tag_name('input').click()
                sleep(3)
            else:
                if element.find_element_by_tag_name('input').is_selected():
                    element.find_element_by_tag_name('input').click()

        # add-focus-keyphrase
        if 'closed' in self.driver.find_element_by_css_selector('#aioseo-settings').get_attribute('class'):
            self.driver.find_element_by_css_selector('#aioseo-settings').click()
        sleep(2)
        self.driver.find_element_by_css_selector('.aioseo-editor-description > div').send_keys(title)
        sleep(2)
        self.driver.find_element_by_css_selector(
            '.aioseo-input.add-focus-keyphrase-metabox-input > input.medium').send_keys(title)
        sleep(3)
        self.driver.find_element_by_css_selector('#add-focus-keyphrase').click()
        sleep(3)
        try:
            self.driver.execute_script(
                "document.evaluate('//button[text()=\"نشر\"]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.click();")
        except:
            self.driver.find_element_by_xpath('//button[text()="نشر"]').click()
        sleep(1)
        try:
            self.driver.find_element_by_xpath('//div[@class="editor-post-publish-panel"]//button[text()="نشر"]').click()
        except:
            self.driver.find_element_by_xpath('//button[text()="نشر"]').click()
            self.driver.find_element_by_xpath('//div[@class="editor-post-publish-panel"]//button[text()="نشر"]').click()

        sleep(5)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec_()

'''
https://badayih.com/wp-admin/post-new.php?wp-post-new-reload=true
https://badayih.com/wp-admin/
user

qSVTZzzM#o9gi*q6cJxza02I
-)ULl48W-mQ;
pyside2-uic untitled.ui -o main.py

'''
p = '''

✳️ عااااجل مطلوب لمستشفى مرموقه #بالرياض نقل كفاله فوري 👇

-مطلوي اخصائي مخ واعصاب

-مطلوب اخصائي/اول طوارئ

-اخصائيه نسا وولاده مصريه


    اقرأ أيضًا:
<a href="https://badayih.com/category/وظائف-الاردن-بدايه-الخبر/" rel="bookmark" title="وظائف الاردن">وظائف الاردن</a>
<a href="https://badayih.com/category/وظائف-الامارات-بدايه-الخبر/" rel="bookmark" title="وظائف الامارات">وظائف الامارات</a>

'''
