import traceback
from time import sleep
import os
import sys
import warnings

from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from requests_html import HTMLSession
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class Main:
    URL_LOGIN = 'https://login.noon.partners/ar/'

    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    def __init__(self):
        try:
            self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())
        except:
            self.driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())
        self.session = HTMLSession()

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    self.driver.get('https://core.noon.partners/en-sa/sales')
                    sleep(2)
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def update_cookies(self):
        self.driver.refresh()
        sleep(1)
        cookies = self.driver.get_cookies()
        for cookie in cookies:
            self.session.cookies.set(str(cookie['name']), str(cookie['value']))

    def send_session(self, url, body):
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.5",
            "Content-Type": "application/json",
            "Cache-Control": "no-cache, no-store",
            "X-Locale": "ar-sa",
            "X-Platform": "web",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin"
        }
        error = 0
        while True:
            try:
                r = self.session.post(url, headers=headers, data=body)
                if r.ok:
                    return r
                else:
                    print(url, body)
                    print(r.status_code)
                    sleep(60)
                    self.update_cookies()
            except Exception as e:
                print(url, body, e)
                sleep(60)
                self.update_cookies()
            if error == 10:
                return False
            error += 1

    def get_noon_product(self, file):
        book = load_workbook(file)
        sheet = book.active
        sheet.cell(1, 2).value = 'Sku'
        sheet.cell(1, 3).value = 'Partner SKU'
        sheet.cell(1, 4).value = 'Price'
        sheet.cell(1, 5).value = 'New Price'
        sheet.cell(1, 6).value = 'Current Status Code'
        sheet.column_dimensions['A'].width = 22
        sheet.column_dimensions['B'].width = 13
        sheet.column_dimensions['C'].width = 16
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 18
        sheet.column_dimensions['F'].width = 10
        for column in range(1, 7):
            try:
                sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
            except:
                pass

        self.update_cookies()
        for el in self.driver.current_url.split('?')[-1].split('&'):
            if 'from' in el:
                from_ = el.split('=')[-1]
            elif 'to=' in el:
                to_ = el.split('=')[-1]

        for i in range(2, sheet.max_row + 1):
            print(i)
            num = sheet.cell(i, 1).value
            if not num:
                continue
            body = '{"itemNr":"' + num + '"}'
            body_m = '{' + f'"fromDate":"{from_}","toDate":"{to_}","perPage":20,"page":1,"search":"{num}"' + '}'

            r = self.send_session("https://core.noon.partners/_svc/partners-statement/sales/list", body_m)
            if r.json()['list']:
                pskuCanonical = r.json()['list'][0]['pskuCanonical']
                r = self.send_session("https://core.noon.partners/_svc/partners-statement/sales/get", body)
                price = r.json()['itemDeliveredScenario']['deliveredNetDue']
                sku = r.json()['itemGlobalStatus']['sku']
                currentStatusCode = r.json()['itemGlobalStatus']['currentStatusCode']
                try:
                    new_price = float(price) + vol
                except:
                    new_price = None
                sheet.cell(i, 2).value = sku
                sheet.cell(i, 3).value = pskuCanonical
                sheet.cell(i, 4).value = price
                sheet.cell(i, 5).value = new_price
                sheet.cell(i, 6).value = currentStatusCode
                for column in range(1, 7):
                    try:
                        sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                    wrap_text=True)
                    except:
                        pass
        book.save(file)


if __name__ == "__main__":
    'NSAD90057523370-2'
    # while True:
    #     try:
    #         file_name = input('- Enter file_name name : ').strip()
    #         load_workbook(file_name).close()
    #         break
    #     except FileNotFoundError:
    #         print('- No such file_name or directory')
    #     except Exception as e:
    #         print(e)

    email = input('- Enter your email : ')  # 'bekj.119@gmail.com'  #
    password = input('- Enter your password : ')  # 'bekj.119'  #
    vol = float(input('- Enter value (+) : '))
    self = Main()
    self.login()
    input('- Enter: ')
    # self.get_noon_product(file_name)

# pyinstaller --add-data C:/Users/3mora/AppData/Local/Programs/Python/Python39/Lib/site-packages/pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info
'''
bekj.119@gmail.com

bh4708251@gmail.com
Assad203064

perfect203070@gmail.com
Mm0987654
'''
