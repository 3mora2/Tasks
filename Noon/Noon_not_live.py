import traceback
from time import sleep
import os
import sys
import warnings

from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from requests_html import HTMLSession
from openpyxl import Workbook
from webdriver_manager.chrome import ChromeDriverManager


os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class Main:
    URL_LOGIN = 'https://login.noon.partners/ar/'

    MAIL_ELEMENT_SELECTOR = (By.NAME, 'email')
    PASS_ELEMENT_SELECTOR = (By.NAME, 'password')
    BUTTON_LOGIN_ELEMENT_SELECTOR = (By.CSS_SELECTOR, '#formContainer > button')

    def __init__(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.session = HTMLSession()

        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.cell(1, 1).value = 'SkU'
        # self.sheet.cell(1, 2).value = 'Message 1'
        # self.sheet.cell(1, 3).value = 'Message 2'
        # self.sheet.cell(1, 4).value = 'PartnerSku'
        # self.sheet.cell(1, 5).value = 'SkUParent'
        # self.sheet.cell(1, 6).value = 'PskuCode'
        # self.sheet.cell(1, 7).value = 'PskuCanonical'
        # self.sheet.cell(1, 8).value = 'URL'

        self.sheet.column_dimensions['A'].width = 11
        # self.sheet.column_dimensions['B'].width = 32
        # self.sheet.column_dimensions['C'].width = 32
        # self.sheet.column_dimensions['D'].width = 15
        # self.sheet.column_dimensions['E'].width = 26
        # self.sheet.column_dimensions['F'].width = 33
        # self.sheet.column_dimensions['G'].width = 15
        # self.sheet.column_dimensions['H'].width = 115
        for column in range(1, 9):
            try:
                self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                 wrap_text=True)
            except:
                pass

    def login(self):
        try:
            self.driver.get(self.URL_LOGIN)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.MAIL_ELEMENT_SELECTOR))).send_keys(email)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.PASS_ELEMENT_SELECTOR))).send_keys(password)
            WebDriverWait(self.driver, 5).until(
                (ec.visibility_of_element_located(self.BUTTON_LOGIN_ELEMENT_SELECTOR))).click()
            while True:
                if 'login.noon' not in self.driver.current_url:
                    break

        except:
            print(traceback.print_exc())
            print('- Login Error')

    def update_cookies(self):
        self.driver.refresh()
        sleep(1)
        cookies = self.driver.get_cookies()
        for cookie in cookies:
            self.session.cookies.set(str(cookie['name']), str(cookie['value']))

    def get_healthy(self):
        state = self.driver.find_element_by_css_selector('.topWrapper+div').text

    def get_noon_product(self):
        n = 2
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.5",
            "Content-Type": "application/json",
            "Cache-Control": "no-cache, no-store",
            "X-Locale": "ar-sa",
            "X-Platform": "web",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin"
        }

        headers_1 = {
            "accept": "application/json, text/plain, */*",
            "accept-language": "ar-US,en;q=0.9",
            "cache-control": "no-cache, no-store",
            "content-type": "application/json",
            "sec-ch-ua": "\" Not;A Brand\";v=\"99\", \"Google Chrome\";v=\"91\", \"Chromium\";v=\"91\"",
            "sec-ch-ua-mobile": "?0",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "x-locale": "ar-sa",
            "x-platform": "web"
        }
        body = ['{"productList":{"page":', ',"perPage":100, "status": "not_live"}}']
        page = 1
        self.update_cookies()
        while True:
            try:
                r = self.session.post("https://catalog.noon.partners/_svc/mp-partner-catalog/catalog/product/list",
                                      headers=headers, data=f'{page}'.join(body))
                if r.ok:
                    sleep(4)
                    page += 1
                    try:
                        products = r.json()['products']
                        if len(products) == 0:
                            break
                        for product in products:
                            sku = product['offers'][0]['sku']
                            partnerSku = product['psku']['partnerSku']
                            skuParent = product['psku']['skuParent']
                            pskuCode = product['psku']['pskuCode']
                            pskuCanonical = product['psku']['pskuCanonical']
                            url = 'https://catalog.noon.partners/ar-sa/catalog/' + skuParent + '/p?code=' + pskuCode + '&tab=health'
                            body_health = '{"pskuCanonical":"' + pskuCanonical + '","marketplaceCodes":["noon"],"withSalesData":true,"pskuCode":"' + pskuCode + '"}'
                            while True:
                                try:
                                    u = self.session.post(
                                        "https://catalog.noon.partners/_svc/mp-partner-catalog/psku/marketplace/health/get",
                                        headers=headers, data=body_health)
                                    if u.ok:
                                        break
                                except Exception as e:
                                    print(e)
                                    sleep(60)

                            state = u.json()['status'][0]
                            self.sheet.cell(n, 1).value = sku
                            c = 2
                            for i in state['offerStatusReason']:
                                self.sheet.cell(1, c).value = i
                                if state['offerStatusReason'][i]:
                                    self.sheet.cell(n, c).value = state['offerStatusReason'][i]
                                c += 1
                            if state['stockStatusReason']:
                                for i in state['stockStatusReason'][0]['stockStatusReason']:
                                    self.sheet.cell(1, c).value = i
                                    if state['stockStatusReason'][0]['stockStatusReason'][i]:
                                        self.sheet.cell(n, c).value = state['stockStatusReason'][0]['stockStatusReason'][i]
                                    c += 1

                            # error1 = [':'.join([i, state['offerStatusReason'][i]]) for i in
                            #           state['offerStatusReason'] if
                            #           state['offerStatusReason'][i]]
                            # if state['stockStatusReason']:
                            #     error2 = [':'.join([i, state['stockStatusReason'][0]['stockStatusReason'][i]])
                            #               for i in state['stockStatusReason'][0]['stockStatusReason'] if
                            #               state['stockStatusReason'][0]['stockStatusReason'][i]]
                            # else:
                            #     error2 = []
                            # self.sheet.cell(n, 2).value = '\n'.join(error1)
                            # self.sheet.cell(n, 3).value = '\n'.join(error2)
                            # self.sheet.cell(n, 4).value = partnerSku
                            # self.sheet.cell(n, 5).value = skuParent
                            # self.sheet.cell(n, 6).value = pskuCode
                            # self.sheet.cell(n, 7).value = pskuCanonical
                            # self.sheet.cell(n, 8).value = url
                            for column in range(1, 30):
                                try:
                                    self.sheet.cell(n, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                                     wrap_text=True)
                                except:
                                    pass
                            for column in range(1, 30):
                                try:
                                    self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                                     wrap_text=True)
                                except:
                                    pass
                            print(n)
                            n += 1
                            self.book.save('Final_Not_Live.xlsx')

                    except:
                        print(traceback.print_exc())
                        sleep(60)
                        self.update_cookies()

                else:
                    print(r.status_code)
                    sleep(60)
                    self.update_cookies()

            except:
                print(traceback.print_exc())
                sleep(60)
                self.update_cookies()


if __name__ == "__main__":
    email = input('- Enter your email : ')
    password = input('- Enter your password : ')
    self = Main()
    self.login()
    self.get_noon_product()

# pyinstaller --add-data C:/Users/3mora/AppData/Local/Programs/Python/Python39/Lib/site-packages/pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info
'''
bekj.119@gmail.com

perfect203070@gmail.com
Mm0987654
'''

