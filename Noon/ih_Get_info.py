from openpyxl import load_workbook
from requests_html import HTMLSession


def print_percent_done(index, total, bar_len=50):
    percent_done = (index-1)/total*100
    percent_done = round(percent_done, 1)

    done = round(percent_done/(100/bar_len))
    togo = bar_len-done

    done_str = '█'*int(done)
    togo_str = '░'*int(togo)

    print(f'- ⏳ {total}\\{index - 1} - [{done_str}{togo_str}] ({percent_done} %)', end='\r')


class Main:
    def __init__(self, file_name):
        self.file_name = file_name
        self.book = load_workbook(file_name)
        self.sheet = self.book.active
        self.sheet.cell(1, 1).value = 'Code'
        self.sheet.cell(1, 2).value = 'SKU'
        self.sheet.cell(1, 3).value = 'Price'
        self.sheet.cell(1, 4).value = 'Available'
        self.sheet.cell(1, 5).value = 'Name'
        self.sheet.cell(1, 6).value = 'Brand'

    def extract(self):
        for i in range(2, self.sheet.max_row+1):
            code = self.sheet.cell(i, 1).value
            if code and not self.sheet.cell(i, 2).value:
                s = HTMLSession()
                cookie = {'name': 'iher-pref1',
                          'value': 'storeid=0&sccode=SA&lan=en-US&scurcode=SAR&wp=2&lchg=1&ifv=1&accsave=0'}
                s.cookies.set(cookie['name'], cookie['value'])
                try:
                    r = s.get('https://sa.iherb.com/search?kw=' + code)
                    price = r.html.find('#price', first=True).text.replace('SAR', '')
                    # static = r.html.xpath('//script[contains(text(), "window.PRODUCT_DETAILS")]',
                    # first=True).text.split(';')[0].split('availableToPurchase: "')[-1].replace('" }', '')
                    static = r.html.find('#stock-status')[0].text
                    if 'متوفر حاليا' in static or 'In Stock' in static:
                        static = 'InStock'
                    else:
                        static = 'OutStock'
                    sku = r.html.find('[itemprop="gtin12"]', first=True).text
                    try:
                        name = r.html.find('#name', first=True).text
                        brand = r.html.find('#brand a', first=True).text
                    except:
                        name = brand = None
                    self.sheet.cell(i, 2).value = sku
                    self.sheet.cell(i, 3).value = price
                    self.sheet.cell(i, 4).value = static
                    self.sheet.cell(i, 5).value = name
                    self.sheet.cell(i, 6).value = brand

                    self.book.save(self.file_name)
                    print_percent_done(i, self.sheet.max_row)
                except:
                    pass

        print('- finish')


if __name__ == '__main__':
    while True:
        try:
            file = input('- Enter file name : ').strip()
            load_workbook(file).close()
            break
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)
    self = Main(file)
    self.extract()
# pyinstaller --add-data C:/Users/3mora/AppData/Local/Programs/Python/Python39/Lib/site-packages/pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info