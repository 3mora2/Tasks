from io import BytesIO
from time import sleep

from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from selenium import webdriver
import requests
from webdriver_manager.firefox import GeckoDriverManager
from requests_html import HTMLSession
import json

session = HTMLSession()


def headers():
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.5",
        "x-platform": "web",
        "x-cms": "v2",
        "x-mp": "noon",
        "Cache-Control": "no-cache, max-age=0, must-revalidate, no-store",
        # "x-Locale": "ar-eg",
        "x-content": "desktop",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin"
    }


class NOON:
    def __init__(self):
        # self.dict_pro = dict()
        # self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())
        self.book = Workbook()
        self.sheet = self.book.active
        i = 1
        self.sheet.cell(i, 1).value = 'SKU'
        self.sheet.cell(i, 3).value = 'Name'
        self.sheet.cell(i, 4).value = 'Price'
        self.sheet.cell(i, 5).value = 'Old Price'
        self.sheet.cell(i, 6).value = 'Brand'
        self.sheet.cell(i, 7).value = 'Seller'
        self.sheet.cell(i, 8).value = 'Seller URL'
        self.sheet.cell(i, 9).value = "URL"
        self.sheet.column_dimensions['B'].width = 25
        self.sheet.column_dimensions['A'].width = 12
        self.sheet.column_dimensions['C'].width = 40
        self.sheet.column_dimensions['D'].width = 8
        self.sheet.column_dimensions['E'].width = 8
        self.sheet.column_dimensions['F'].width = 20
        self.sheet.column_dimensions['G'].width = 15
        self.sheet.column_dimensions['H'].width = 20
        self.sheet.column_dimensions['I'].width = 60

        for column in range(1, 10):
            try:
                self.sheet.cell(i, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                 wrap_text=True)
            except Exception as e:
                print(e)

    def start_h(self, link, file):
        list_n = []
        i = 2
        if 'limit' in link:
            n = int(link.split('limit=')[-1].split('&')[0])
            link = link.replace('limit=' + str(n), 'limit=150')
        else:
            if '?' in link:
                link += '&limit=150'
            else:
                link += '?limit=150'
        while True:
            print(link)
            try:
                r = session.get(link, headers=headers())
                if not r.ok:
                    sleep(60)
                    continue

                if 'page' in link:
                    n = int(link.split('page=')[-1].split('&')[0])
                    link = link.replace('page=' + str(n), 'page=' + str(n + 1))
                else:
                    if '?' in link:
                        link += '&page=2'
                    else:
                        link += '?page=2'

                data = json.loads(r.html.find('#__NEXT_DATA__', first=True).text)
                if len(data['props']['pageProps']['catalog']['hits']) == 0:
                    break

                for element in data['props']['pageProps']['catalog']['hits']:
                    sku = element.get('sku')
                    if sku not in list_n:
                        list_n.append(sku)
                    else:
                        print('- found')
                        continue

                    brand = element.get('brand')
                    name = element['name']
                    price = element.get('price')
                    sale_price = element['sale_price']
                    image_key = element['image_key']
                    img = f'https://z.nooncdn.com/products/{image_key}.jpg'
                    url = element['url']
                    try:
                        d = session.get(f'https://www.noon.com/egypt-ar/{sku}/p?', headers=headers())
                        d = json.loads(d.html.find('#__NEXT_DATA__', first=True).text)
                        seller_url = 'https://www.noon.com/saudi-ar/' + \
                                     d['props']['pageProps']['catalog']['product']['variants'][0]['offers'][0]['store_code']
                        seller = d['props']['pageProps']['catalog']['product']['variants'][0]['offers'][0]['store_name']
                    except:
                        seller = None
                        seller_url = None

                    try:
                        res = session.get(img, headers=headers())
                        image_file = BytesIO(res.content)
                        img = Image(image_file)

                        img.width = 200
                        img.height = 270

                        self.sheet.row_dimensions[i].height = 206
                        self.sheet.add_image(img, f'B{i}')
                    except:
                        print('- error', img)
                    self.sheet.cell(i, 1).value = sku
                    self.sheet.cell(i, 3).value = name
                    self.sheet.cell(i, 4).value = price if not sale_price else sale_price
                    self.sheet.cell(i, 5).value = sale_price if not sale_price else price
                    self.sheet.cell(i, 6).value = brand
                    self.sheet.cell(i, 7).value = seller
                    self.sheet.cell(i, 8).value = seller_url
                    self.sheet.cell(i, 9).value = url
                    print(i)
                    i += 1
            except Exception as e:
                print(e)

        self.book.save(str(file) + '.xlsx')


self = NOON()

self.start_h(input('- Enter Url: '), input('- Enter Filename: '))
# https://www.noon.com/egypt-ar/beauty-and-health/beauty/fragrance/eau-de-parfum?limit=150
# pyinstaller --add-data c:\users\3mora\anaconda3\envs\python3.6\lib\site-packages\pyppeteer-0.2.5.dist-info;pyppeteer-0.2.5.dist-info
