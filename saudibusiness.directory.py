from time import sleep

from requests_html import HTMLSession
from openpyxl import Workbook, load_workbook


session = HTMLSession()
# urls = []
# page = 1
# while True:
#     r = session.get('https://saudibusiness.directory/category/restaurants/?page='+str(page))
#     page += 1
#     for element in r.html.find('.panel-body>h4>a'):
#         urls.append(element.attrs['href'])
#     if len(r.html.find('.panel-body>h4>a')) == 0:
#         break
#     print(len(urls))
# urls = list(set(urls))
# book = Workbook()
# sheet = book.active
# n = 2
# for url in urls:
#     sheet.cell(n, 10).value = url
#     n += 1
# print(n)
# book.save('saudibusiness.xlsx')

book = load_workbook('saudibusiness-final.xlsx')
sheet = book.active
i = 1
sheet.cell(i, 1).value = 'title'
sheet.cell(i, 2).value = 'keywords'
sheet.cell(i, 3).value = 'street_address'
sheet.cell(i, 4).value = 'locality'
sheet.cell(i, 5).value = 'region'
sheet.cell(i, 6).value = 'postal_code'
sheet.cell(i, 7).value = 'country_name'
sheet.cell(i, 8).value = 'phone_number'
sheet.cell(i, 9).value = 'fax_number'
sheet.cell(i, 11).value = 'website'
sheet.cell(i, 12).value = 'position'
sheet.cell(i, 13).value = 'ratingValue'
sheet.cell(i, 14).value = 'ratingCount'
sheet.cell(i, 15).value = 'type_'
sheet.cell(i, 16).value = 'vcard'
sheet.cell(i, 17).value = 'map_'
sheet.cell(i, 18).value = 'img'
sheet.cell(i, 19).value = 'desk'
sheet.cell(i, 20).value = 'cat1'
sheet.cell(i, 21).value = 'cat2'
sheet.cell(i, 22).value = 'cat3'
sheet.cell(i, 23).value = 'name'
sheet.cell(i, 24).value = 'city'
try:
    r = session.get('https://saudibusiness.directory/')
except:
    pass
for i in range(2, sheet.max_row+1):
    url = sheet.cell(i, 10).value
    if url and not sheet.cell(i, 1).value:
        # try:
            r = session.get(url)
            name = r.html.find('h1', first=True).text.split('في')[0]
            city = r.html.find('h1', first=True).text.split('في')[1]
            keywords = r.html.find('meta[name="keywords"]', first=True).attrs['content']
            title = r.html.find('meta[property="og:title"]', first=True).attrs['content']
            street_address = r.html.find('meta[property="business:contact_data:street_address"]', first=True).attrs['content']
            locality = r.html.find('meta[property="business:contact_data:locality"]', first=True).attrs['content']
            region = r.html.find('meta[property="business:contact_data:region"]', first=True).attrs['content']
            postal_code = r.html.find('meta[property="business:contact_data:postal_code"]', first=True).attrs['content']
            country_name = r.html.find('meta[property="business:contact_data:country_name"]', first=True).attrs['content']
            try:
                phone_number = r.html.find('meta[property="business:contact_data:phone_number"]', first=True).attrs['content']
            except:
                phone_number = None
            fax_number = r.html.find('meta[property="business:contact_data:fax_number"]', first=True).attrs['content']
            website = r.html.find('meta[property="business:contact_data:website"]', first=True).attrs['content']
            try:
                position = r.html.find('meta[name="geo.position"]', first=True).attrs['content']
            except:
                position = None
            try:
                ratingValue = r.html.find('[itemprop="ratingValue"]', first=True).text
            except:
                ratingValue = None
            try:
                ratingCount = r.html.find('[itemprop="ratingCount"]', first=True).text
            except:
                ratingCount = None
            try:
                type_ = r.html.find('h4', first=True).text
            except:
                type_ = None
            try:
                vcard = r.html.xpath('//*[contains(@href, "&action=vcard")]', first=True).attrs['href']
            except:
                vcard = None
            try:
                map_ = r.html.xpath( '//*[contains(@href, "maps.google.com/maps?q=")]', first=True).attrs['href']
            except:
                map_ = None

            img = '\n'.join([element.attrs['src'] for element in r.html.find('div > a > img')])
            desk = '\n'.join([element.text for element in r.html.find('#listing_rating ~ p')])
            cat1 = r.html.find('li[itemprop="itemListElement"]')[1].text
            cat2 = r.html.find('li[itemprop="itemListElement"]')[2].text
            cat3 = r.html.find('li[itemprop="itemListElement"]')[-1].text

            sheet.cell(i, 1).value = title
            sheet.cell(i, 2).value = keywords
            sheet.cell(i, 3).value = street_address
            sheet.cell(i, 4).value = locality
            sheet.cell(i, 5).value = region
            sheet.cell(i, 6).value = postal_code
            sheet.cell(i, 7).value = country_name
            sheet.cell(i, 8).value = phone_number
            sheet.cell(i, 9).value = fax_number
            sheet.cell(i, 11).value = website
            sheet.cell(i, 12).value = position
            sheet.cell(i, 13).value = ratingValue
            sheet.cell(i, 14).value = ratingCount
            sheet.cell(i, 15).value = type_
            sheet.cell(i, 16).value = vcard
            sheet.cell(i, 17).value = map_
            sheet.cell(i, 18).value = img
            sheet.cell(i, 19).value = desk
            sheet.cell(i, 20).value = cat1
            sheet.cell(i, 21).value = cat2
            sheet.cell(i, 22).value = cat3
            sheet.cell(i, 23).value = name
            sheet.cell(i, 24).value = city
            print(i)
            book.save('saudibusiness-final.xlsx')
        # except :
        #     sleep(100)