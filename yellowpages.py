import os
import sys
import warnings
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from time import sleep
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from webdriver_manager.firefox import GeckoDriverManager

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class Main:
    def __init__(self):
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet['A1'] = 'Image'
        self.sheet.cell(1, 2).value = 'Company Name'
        self.sheet.cell(1, 3).value = 'Phone Number'
        self.sheet.cell(1, 4).value = 'Whatsapp'
        self.sheet.cell(1, 5).value = 'Business Description'
        self.sheet.cell(1, 6).value = 'Category'
        self.sheet.cell(1, 7).value = 'Show Directions'
        self.sheet.cell(1, 8).value = 'Keywords'
        self.sheet.cell(1, 9).value = 'Company Address'
        self.sheet.cell(1, 10).value = 'Website'
        self.sheet.cell(1, 11).value = 'Branches URL'
        self.sheet.cell(1, 12).value = 'About Us'
        self.sheet.cell(1, 13).value = 'Company URL'

        self.sheet.column_dimensions['A'].width = 13
        self.sheet.column_dimensions['B'].width = 35
        self.sheet.column_dimensions['C'].width = 16
        self.sheet.column_dimensions['D'].width = 57
        self.sheet.column_dimensions['E'].width = 22
        self.sheet.column_dimensions['F'].width = 25
        self.sheet.column_dimensions['G'].width = 65
        self.sheet.column_dimensions['H'].width = 26
        self.sheet.column_dimensions['I'].width = 70
        self.sheet.column_dimensions['J'].width = 70
        self.sheet.column_dimensions['K'].width = 50
        self.sheet.column_dimensions['L'].width = 120
        self.sheet.column_dimensions['M'].width = 120
        for column in range(1, 14):
            try:
                self.sheet.cell(1, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                 wrap_text=True)
            except:
                pass
        self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())
        self.driver.get("https://www.yellowpages.com.eg/ar/")
        input('- Enter : ')

    @staticmethod
    def safe_find_element_by(driver, by, element):
        try:
            return driver.find_element(by, element)
        except NoSuchElementException:
            return None

    def start(self):
        num = 2
        while True:
            for element in self.driver.find_elements_by_css_selector('.searchResultsDiv'):
                companyName = self.safe_find_element_by(element, By.CSS_SELECTOR, '.companyName')
                if companyName:
                    companyName = companyName.text
                companyURL = self.safe_find_element_by(element, By.CSS_SELECTOR, '.companyName')
                if companyURL:
                    companyURL = companyURL.get_attribute('href')

                business_description = self.safe_find_element_by(element, By.CSS_SELECTOR, '.business-description')
                if business_description:
                    business_description = business_description.text

                img = self.safe_find_element_by(element, By.CSS_SELECTOR, 'a>img')
                if img:
                    img = img.get_attribute('src')

                company_address = self.safe_find_element_by(element, By.CSS_SELECTOR, '.company_address')
                if company_address:
                    company_address = company_address.text

                show_branches = self.safe_find_element_by(element, By.CSS_SELECTOR, '.show_branches_SRP')
                if show_branches:
                    show_branches = show_branches.get_attribute('href')

                category = self.safe_find_element_by(element, By.CSS_SELECTOR, '.category')
                if category:
                    category = category.text
                otherCategories = self.safe_find_element_by(element, By.CSS_SELECTOR, '.otherCategories')
                if otherCategories:
                    category += ',' + ','.join(otherCategories.get_attribute('data-content').split('<br>'))

                keywords = [el.text for el in element.find_elements_by_css_selector('.search-keywords .two-words span')]
                otherKeywords = self.safe_find_element_by(element, By.CSS_SELECTOR, '.otherKeywords')
                if otherKeywords:
                    keywords += otherKeywords.get_attribute('data-content').split('<br>')
                keywords = '\n'.join(keywords)

                aboutUs = self.safe_find_element_by(element, By.CSS_SELECTOR, '.aboutUs')
                if aboutUs:
                    aboutUs = aboutUs.text

                website = self.safe_find_element_by(element, By.CSS_SELECTOR, '.website > a')
                if website:
                    website = website.get_attribute('href')

                show_directions = self.safe_find_element_by(element, By.CSS_SELECTOR, '#show_directions.latlng-mob>a')
                if show_directions:
                    show_directions = show_directions.get_attribute('href')

                whatsapp = self.safe_find_element_by(element, By.CSS_SELECTOR, '.whatsapp > a')
                if whatsapp:
                    whatsapp = whatsapp.get_attribute('href')

                mob = self.safe_find_element_by(element, By.CSS_SELECTOR, '.search-call-mob')
                if mob:
                    mob = mob.get_attribute('href').split(":")[-1].strip()

                show_review = self.safe_find_element_by(element, By.CSS_SELECTOR, '.show-review')
                if show_review:
                    show_review = show_review.text

                try:
                    if img:
                        res = requests.get(img)
                        image_file = BytesIO(res.content)
                        img = Image(image_file)
                        img.width = 90
                        img.height = 75
                        self.sheet.row_dimensions[num].height = 56
                        self.sheet.add_image(img, f'A{num}')
                except:
                    pass
                self.sheet.cell(num, 2).value = companyName
                self.sheet.cell(num, 3).value = mob
                self.sheet.cell(num, 4).value = whatsapp
                self.sheet.cell(num, 5).value = business_description
                self.sheet.cell(num, 6).value = category
                self.sheet.cell(num, 7).value = show_directions
                self.sheet.cell(num, 8).value = keywords
                self.sheet.cell(num, 9).value = company_address
                self.sheet.cell(num, 10).value = website
                self.sheet.cell(num, 11).value = show_branches
                self.sheet.cell(num, 12).value = aboutUs
                self.sheet.cell(num, 13).value = companyURL
                for column in range(2, 14):
                    try:
                        self.sheet.cell(num, column).alignment = Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True)
                    except:
                        pass
                print(num)
                num += 1

            try:
                self.driver.find_element_by_css_selector('li + .waves-effect > a').click()
                sleep(4)
            except:
                break
        self.book.save('result.xlsx')


if __name__ == '__main__':
    self = Main()
    self.start()
'https://www.yellowpages.com.eg/ar/condensed-category/%D9%83%D9%87%D8%B1%D8%A8%D8%A7%D8%A1-%D9%88%D8%A7%D9%84%D9%83%D8%AA%D8%B1%D9%88%D9%86%D9%8A%D8%A7%D8%AA-%D8%A7%D8%AC%D9%87%D8%B2%D8%A9-%D9%88%D9%85%D8%B9%D8%AF%D8%A7%D8%AA'
"""
from requests_html import HTMLSession
s = HTMLSession()

r = s.get(
    "https://www.yellowpages.com.eg/ar/profile/%D9%85%D8%B9%D8%A7%D9%85%D9%84-%D8%AA%D8%B1%D8%B3%D8%AA/511853?position=60&key=%D9%85%D8%B9%D8%A7%D9%85%D9%84-%D8%B7%D8%A8%D9%8A%D8%A9&mod=category&categoryId=2305#show_branches",
    headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:92.0) Gecko/20100101 Firefox/92.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Alt-Used": "www.yellowpages.com.eg",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-User": "?1"
    }
    )
r.html.find('#show_branches')[0].text
"""
'''companyName = element.find_element_by_css_selector('.companyName').text
 companyURL = element.find_element_by_css_selector('.companyName').get_attribute('href')
 business_description = element.find_element_by_css_selector('.business-description').text
 img = element.find_element_by_css_selector('a>img').get_attribute('src')
 company_address = element.find_element_by_css_selector('.company_address').text
 show_branches = element.find_element_by_css_selector('.show_branches_SRP').get_attribute('href')
 category = element.find_element_by_css_selector('.category').text
 keywords = [el.text for el in element.find_elements_by_css_selector('.search-keywords .two-words span')]
 keywords += element.get_attribute('data-content').split('<br>')
 keywords = '\n'.join(keywords)
 aboutUs = element.find_element_by_css_selector('.aboutUs').text
 website = element.find_element_by_css_selector('.website > a').get_attribute('href')
 show_directions = element.find_element_by_css_selector('#show_directions.latlng-mob>a').get_attribute(
     'href')
 whatsapp = element.find_element_by_css_selector('.whatsapp > a').get_attribute('href')'''
