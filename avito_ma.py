from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
import json
from openpyxl import Workbook

book = Workbook()
sheet = book.active
sheet.cell(1, 1).value = 'Name'
sheet.cell(1, 2).value = 'Phone Number'
sheet.cell(1, 3).value = 'URL'
# 'https://www.avito.ma/fr/sefrou/immobilier'


def print_percent_done(index, total, bar_len=50):
    percent_done = (index-1)/total*100
    percent_done = round(percent_done, 1)

    done = round(percent_done/(100/bar_len))
    togo = bar_len-done

    done_str = '█'*int(done)
    togo_str = '░'*int(togo)

    # print(f'\t⏳{title}: [{done_str}{togo_str}] {percent_done}% done', end='\r')
    print(f'- {total}\\{i - 1}-Phone : {phone} [{done_str}{togo_str}]({percent_done} %)', end='\r')


while True:
    url = input('- Enter URL : ')
    if url != '':
        break
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get(url)
links = []
while True:
    for element in driver.find_elements_by_css_selector('div.sc-1nre5ec-0 > div > a'):
        if element.get_attribute('href') not in links:
            links.append(element.get_attribute('href'))
            print('- Number :', len(links), end='\r')
    try:
        driver.find_element_by_css_selector(
            '[data-testid="PaginationDesktopWrapper"] [aria-labelledby="ChevronRightTitleID"]').click()
    except:
        break
i = 2
print('- Number :', len(links))
print('- Start Get Data....')
for link in links:
    driver.get(link)
    j = driver.find_element_by_css_selector("#__NEXT_DATA__").get_attribute('innerHTML')
    js = json.loads(j)['props']['pageProps']['initialReduxState']['ad']['view']['adInfo']
    name = js['seller']['name']
    url = js['friendlyUrl']['url']
    if not js['isPhoneHidden']:
        phone = js['phone']
    else:
        phone = None

    sheet.cell(i, 1).value = name
    sheet.cell(i, 2).value = phone
    sheet.cell(i, 3).value = url
    print_percent_done(i, len(links))
    i += 1

book.save('Avito-ma.xlsx')
