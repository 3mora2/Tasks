import os
import subprocess
from time import sleep
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
import re


class CollectPosts(object):
    def __init__(self):
        self.data = dict()
        self.total = 0
        self.file_save = 'face.xlsx'
        self.book = Workbook()
        self.sheet = self.book.active
        self.old_data = []
        self.lists = []
        chrome_options = Options()
        chrome_options.add_argument("--user-data-dir={}".format(
            r'C:\Users\3mora\AppData\Local\Google\Chrome\User Data'))
        chrome_options.add_argument('--profile-directory=Profile 5')
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
        self.file_save = 'face -' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + '.xlsx'

    def login(self):
        try:
            self.driver.get("https://mbasic.facebook.com/language.php")
            sleep(1)
            self.driver.find_element_by_css_selector('[value="English (US)"]').click()
        except Exception as e:
            print(e)
            print("There was some error while logging in.")

    def load_old_data(self, file):
        sheet = load_workbook(file).active
        for i in range(2, sheet.max_row+1):
            url = sheet.cell(i, 2).value
            text = sheet.cell(i, 1).value
            if url:
                self.old_data.append(url)
            if text:
                self.old_data.append(text)

    def safe_find_element_by(self, by, elem):
        try:
            return self.driver.find_element(by, elem)
        except NoSuchElementException:
            return None

    def mobile_start(self, url):
        url = url.replace('www.facebook.com', 'm.facebook.com')
        if '?' in url:
            url = url.split('?')[0]
        self.driver.get(url)

    def GetPostsNew(self, n, e):
        for i in range(n):
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            sleep(4)
            print(len(self.driver.find_elements_by_css_selector('[data-sigil="m-feed-voice-subtitle"] > a')))
            if len(self.driver.find_elements_by_css_selector('[data-sigil="m-feed-voice-subtitle"] > a')) > e:
                break

    def GetPosts_(self):
        for element in self.driver.find_elements_by_css_selector('article'):  # [class="ds"]'):
            try:
                url = element.find_element_by_css_selector('[data-sigil="m-feed-voice-subtitle"] > a').get_attribute('href')
                if url in self.lists:
                    print('- found..')
                    continue
                self.lists.append(url)

                if len(element.find_elements_by_css_selector('article')) > 0:
                    element = element.find_element_by_css_selector('article')
                element_p = element.find_element_by_css_selector('header + div')
                text = element_p.text
                if text in self.lists:
                    print('- found..')
                    continue
                self.lists.append(text)

                if len(element_p.find_elements_by_css_selector('span[data-sigil="more"] > a')) > 0:
                    self.data[self.total] = {'opened': True, 'url': url, 'data': None}
                    self.total += 1
                elif len(re.findall(r'\d{10,15}', text)) > 0:
                    self.data[self.total] = {'opened': False,
                                             'url': url, 'data': text}
                    self.total += 1
                print(self.total)
            except Exception as e:
                print(e)

        for element in self.data.keys():
            if self.data[element]['opened']:
                try:
                    self.driver.get(self.data[element]['url'])
                    sleep(5)
                    if len(self.driver.find_elements_by_css_selector('article')) > 0:
                        text = self.driver.find_element_by_css_selector('article header + div').text
                    else:
                        text = self.driver.find_element_by_css_selector('header + div').text
                    if text in self.lists:
                        self.data[element]['opened'] = False
                        print('- found..')
                        continue
                    self.lists.append(text)

                    if len(re.findall(r'\d{10,15}', text)) > 0:
                        self.data[element]['opened'] = False
                        self.data[element]['data'] = text

                    self.data[element]['opened'] = False
                    print(element)
                except Exception as e:
                    print(e)

    def GetPosts(self):

            break_ = False
            next_url = ''
            while True:
                if len(self.driver.find_elements_by_css_selector('[title="You’re Temporarily Blocked"]')) > 0:
                    break

                for element in self.driver.find_elements_by_css_selector('section > article'):  # [class="ds"]'):
                    try:
                        url = element.find_element(By.LINK_TEXT, "Full Story").get_attribute('href')
                        if url in self.lists:
                            print('- found..')
                            continue
                        self.lists.append(url)
                        if len(element.find_elements_by_css_selector('article')) > 0:
                            element = element.find_element_by_css_selector('article')
                        element_p = element.find_element_by_css_selector('[class="ds"]')
                        text = element_p.text
                        if text in self.lists:
                            print('- found..')
                            continue
                        self.lists.append(text)

                        if len(element_p.find_elements_by_xpath('div/span/a[text()="More"]')) > 0:
                            self.data[self.total] = {'opened': True, 'url': element_p.find_element_by_xpath(
                                'div/span/a[text()="More"]').get_attribute('href'), 'data': None}
                            self.total += 1
                        elif len(re.findall(r'\d{10,15}', text)) > 0:
                            self.data[self.total] = {'opened': False,
                                                     'url': url, 'data': text}
                            self.total += 1
                        print(self.total)
                    except Exception as e:
                        print(e)

                try:
                    next_url = self.driver.find_element_by_xpath(
                        '//div[a/span[contains(text(),"See more stories")]]/a | '
                        '//div[a/span[contains(text(),"See More Stories")]]/a | '
                        '//div[a/span[contains(text(),"See More Posts")]]/a | '
                        '//div/a[contains(text(),"Show more")] | '
                        '//div[a/span[contains(text(),"See more")]]/a').get_attribute('href')
                except:
                    print('- ***************** Finish ***************')
                    break_ = True

                for element in self.data.keys():
                    if self.data[element]['opened']:
                        try:
                            self.driver.get(self.data[element]['url'])
                            sleep(5)
                            if len(self.driver.find_elements_by_css_selector('article')) > 0:
                                text = self.driver.find_element_by_css_selector('article header + div').text
                            else:
                                text = self.driver.find_element_by_css_selector('header + div').text
                            if text in self.lists:
                                self.data[element]['opened'] = False
                                print('- found..')
                                continue
                            self.lists.append(text)

                            if len(re.findall(r'\d{10,15}', text)) > 0:
                                self.data[element]['opened'] = False
                                self.data[element]['data'] = text

                            self.data[element]['opened'] = False
                            print(element)
                        except Exception as e:
                            print(e)

                if len(self.data.keys()) >= numb or break_:
                    break
                else:
                    print('- Next .......')
                    self.driver.get(next_url)
                    sleep(10)

    def save(self):
        with open(self.file_save.replace('xlsx', 'txt'), 'w', encoding="utf-8") as f:
            for num, key in enumerate(self.data.keys()):
                if not self.data[key]['data']:
                    continue
                if len(re.findall(r'\d{10,15}', self.data[key]['data'])) > 0:
                    f.write(self.data[key]['data'])
                    f.write('\n\n\n*******************************************\n\n\n')
                    self.sheet.cell(num + 2, 1).value = self.data[key]['data']
                    self.sheet.cell(num + 2, 2).value = self.data[key]['url']
                    print(num)

        self.book.save(self.file_save)


numb = 300
# self = CollectPosts()
# self.login()
# self.load_old_data('1.xlsx')
# self.load_old_data('2.xlsx')
# self.load_old_data('3.xlsx')
# self.load_old_data('4.xlsx')
# self.load_old_data('5.xlsx')
# self.load_old_data('6.xlsx')
# self.load_old_data('7.xlsx')
# self.load_old_data('8.xlsx')
# self.mobile_start('https://www.facebook.com/groups/Jobs.KS/')
# with open(r"C:\Users\3mora\Downloads\وظائف.txt", 'r', encoding="utf-8") as f:
#     data = f.readlines()
#
# l = []
# for d in data:
#     if 'https://' in d:
#         self.driver.get(d)
#         if input('- e : ') == 'a':
#             l.append(d)
old_data = []


def load_old_data(file):
    file = 'final/'+file
    book = load_workbook(file)
    sheet = book.active
    with open(file.replace('xlsx', 'txt'), 'w', encoding="utf-8") as f:
        for i in range(2, sheet.max_row + 1):
            print(i)
            url = sheet.cell(i, 2).value
            text = sheet.cell(i, 1).value
            if url and url not in old_data:
                old_data.append(url)
            else:
                print(file, 'found url')
                continue
            if text and text not in old_data:
                old_data.append(text)
                f.write(text)
                f.write('\n\n\n*******************************************\n\n\n')
            else:
                print(file, 'found text')
                continue

for file in os.listdir(r'.\final'):
    if file.endswith('xlsx'):
        load_old_data(file)



_=['https://www.facebook.com/groups/829476503746911\n', 'https://www.facebook.com/groups/653263421499678/\n',
   'https://www.facebook.com/groups/1655045174720945/\n', 'https://www.facebook.com/groups/1125761584159082/\n',
   'https://www.facebook.com/groups/1188893261197953/\n', 'https://www.facebook.com/groups/540505022985304/\n',
   'https://www.facebook.com/groups/ksa.jobs4all/\n', 'https://www.facebook.com/groups/Jobs.KS/\n',
   'https://www.facebook.com/groups/2050726501727817/\n', 'https://www.facebook.com/groups/869371139782049/\n', 'https://www.facebook.com/groups/1994985183854697/\n', 'https://www.facebook.com/groups/wazaefmo3almeen/\n', 'https://www.facebook.com/groups/491910867881457/\n', 'https://www.facebook.com/groups/776776469074255/\n', 'https://www.facebook.com/groups/2190718414514517/\n', 'https://www.facebook.com/groups/1505761466346039/\n', 'https://www.facebook.com/groups/119943318067122/\n', 'https://www.facebook.com/groups/750188792017629/\n', 'https://www.facebook.com/groups/339909412700733/\n', 'https://www.facebook.com/groups/1785730755065598/\n', 'https://www.facebook.com/groups/shokl/\n', 'https://www.facebook.com/groups/1469223106719399/\n', 'https://www.facebook.com/groups/405429739664059/\n', 'https://www.facebook.com/groups/FreeEngJobs/\n', 'https://www.facebook.com/groups/2211798935811006/\n', 'https://www.facebook.com/groups/940593089319329/\n', 'https://www.facebook.com/groups/3ayzjob/\n', 'https://www.facebook.com/groups/463806890638541/\n', 'https://www.facebook.com/groups/883932161738858/\n', 'https://www.facebook.com/groups/1131218980268502/\n', 'https://www.facebook.com/groups/344525992604065/\n', 'https://www.facebook.com/groups/108936883121842/\n', 'https://www.facebook.com/groups/1133353990105214/\n', 'https://www.facebook.com/groups/202838976819649/\n', 'https://www.facebook.com/groups/1212363249096928/\n', 'https://www.facebook.com/groups/213392439806523/\n', 'https://www.facebook.com/groups/qatarpoints/\n', 'https://www.facebook.com/groups/2801527159878645/\n', 'https://www.facebook.com/groups/1726517807372182/\n', 'https://www.facebook.com/groups/719467948515762/\n', 'https://www.facebook.com/groups/865890933824146/\n', 'https://www.facebook.com/groups/545538345548136/\n', 'https://www.facebook.com/groups/1004482309598283/\n', 'https://www.facebook.com/groups/1226543437544580/\n', 'https://www.facebook.com/groups/128766357915364/\n', 'https://www.facebook.com/groups/224845334704500/\n', 'https://www.facebook.com/groups/110311749059673/\n', 'https://www.facebook.com/groups/377902672406555/\n', 'https://www.facebook.com/groups/OLXtwzef/\n', 'https://www.facebook.com/groups/1609865872570528/\n', 'https://www.facebook.com/groups/October.City.Jobs/\n', 'https://www.facebook.com/groups/1951227068540112/\n', 'https://www.facebook.com/groups/822795718066389/\n', 'https://www.facebook.com/groups/187042181420723/\n', 'https://www.facebook.com/groups/366592933521857/\n', 'https://www.facebook.com/groups/shogl/\n', 'https://www.facebook.com/groups/1481065522201019/\n', 'https://www.facebook.com/groups/1016579245039862/\n', 'https://www.facebook.com/groups/1935661843373577/\n', 'https://www.facebook.com/groups/358423868046401/\n', 'https://www.facebook.com/groups/937612443271210/\n', 'https://www.facebook.com/groups/132454010100626/\n', 'https://www.facebook.com/groups/214033595455327/\n', 'https://www.facebook.com/groups/357259967749870/\n', 'https://www.facebook.com/groups/694583217358601/\n', 'https://www.facebook.com/groups/829476503746911\n']
