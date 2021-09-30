from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.webdriver import WebDriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
import os
import sys
import warnings
from random import choice

os.environ['WDM_LOG_LEVEL'] = '0'
if not sys.warnoptions:
    warnings.simplefilter("ignore")


class Send:
    driver: WebDriver

    def open(self, path, profile):
        chrome_options = Options()
        chrome_options.add_argument("--user-data-dir={}".format(path))
        chrome_options.add_argument('--profile-directory='+profile)
        self.driver = webdriver.Chrome(ChromeDriverManager(log_level=0).install(), chrome_options=chrome_options)
        try:
            self.driver.get("https://mbasic.facebook.com/language.php")
            sleep(1)
            self.driver.find_element_by_css_selector('[value="English (US)"]').click()
            sleep(1)
        except Exception as e:
            print(e)

    def close(self):
        try:
            self.driver.quit()
        except:
            pass

    def send(self, message, url):
        try:
            self.driver.get(url)
            sleep(1)
            self.driver.find_element_by_link_text('Message').click()
            sleep(2)
            self.driver.find_element_by_css_selector('textarea').send_keys(message)
            sleep(1)
            self.driver.find_element_by_css_selector('input[value="Send"]').click()
            sleep(3)
            try:
                result = self.driver.find_element_by_css_selector('head > title').text
            except:
                result = None

            return result
        except:
            return False


def get_message(path):
    messages = []
    try:
        book = load_workbook(path)
        sheet = book.active
        for i in range(1, sheet.max_row+1):
            message = sheet.cell(i, 1).value
            if message:
                messages.append(message)
    except:
        pass

    return messages


def user_data():
    list_path = []

    book = load_workbook(file_name)
    sheet = book.active
    for i in range(1, sheet.max_row+1):
        path = sheet.cell(i, 1).value
        pro = sheet.cell(i, 2).value
        if path and pro:
            list_path.append((path, pro))

    return list_path


def main(users_path, path_file, messages):
    book = load_workbook(path_file)
    sheet = book.active
    num = 2
    for path, pro in users_path:
        self.open(path, pro)
        for i in range(max_n):
            url = sheet.cell(num, 3).value
            user_id = sheet.cell(num, 2).value
            if user_id:
                result = self.send(choice(messages), url)
                if result is False:
                    sheet.cell(num, 5).value = 'Error False'
                else:
                    sheet.cell(num, 5).value = result
                book.save(path_file)
                print(num, i, user_id)
            num += 1
            sleep(time_sleep)
        self.close()


if __name__ == '__main__':
    path_message = 'messages.xlsx'
    file_name = 'user_data.xlsx'
    try:
        time_sleep = float(input('- Enter Time Sleep: '))
    except:
        time_sleep = 60
    try:
        max_n = int(input('- Enter Max Message For Each Account: '))
    except:
        max_n = 50

    while True:
        try:
            file_name_user = input('- Enter file name : ')  # 'Book1.xlsx'
            load_workbook(file_name_user).close()
            break
        except FileNotFoundError:
            print('- No such file or directory')
        except Exception as e:
            print(e)

    self = Send()
    lists = user_data()
    list_messages = get_message(path_message)
    if not len(list_messages):
        list_messages = [input('- Enter Massage: '), ]

    main(lists, file_name_user, list_messages)
